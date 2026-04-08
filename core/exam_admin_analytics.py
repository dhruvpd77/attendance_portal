"""
Exam Admin: multi-department mark analytics and Excel exports (uploaded marksheets only).
"""
import re
from collections import defaultdict
from io import BytesIO

from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import (
    Department,
    Batch,
    Student,
    Subject,
    ExamPhase,
    ExamPhaseSubject,
    StudentMark,
)
from .exam_phase_order import (
    exam_phase_header_short_name,
    exam_phase_name_sort_key,
    sort_exam_phases,
    sorted_phase_names,
)
from .student_marks_utils import normalize_student_mark
from .risk_policy import mark_fail_below_getter_for_departments, mark_fail_below_threshold

MARK_LOW_THRESHOLD = 9
LOW_FILL = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
HEADER_FILL = PatternFill(start_color='1e293b', end_color='1e293b', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')
SUBJECT_HEADER_FILL = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
PHASE_HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
SUBJECT_HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
HEADER_ALIGN_CC = Alignment(horizontal='center', vertical='center', wrap_text=True)
THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

# Rotating palette for phase row-1 bands (each Tn / SEE gets a distinct color where possible)
_PHASE_BAND_PALETTE = [
    PatternFill(start_color='0F766E', end_color='0F766E', fill_type='solid'),
    PatternFill(start_color='1D4ED8', end_color='1D4ED8', fill_type='solid'),
    PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid'),
    PatternFill(start_color='C2410C', end_color='C2410C', fill_type='solid'),
    PatternFill(start_color='0E7490', end_color='0E7490', fill_type='solid'),
    PatternFill(start_color='A16207', end_color='A16207', fill_type='solid'),
]
_SEE_BAND_FILL = PatternFill(start_color='BE123C', end_color='BE123C', fill_type='solid')
_PHASE_TITLE_RE = re.compile(r'^T(\d+)$', re.IGNORECASE)


def _phase_name_for_band_color(phase_title):
    """Row-1 title may be 'T1' or 'SY_1 — T1' (combined sheet)."""
    t = (phase_title or '').strip()
    if ' — ' in t:
        t = t.split(' — ')[-1].strip()
    return t


def _phase_band_fill(phase_title, sequence_index):
    """Background fill for one phase merged header — T1/T2/…/SEE each mapped to different colors."""
    base = _phase_name_for_band_color(phase_title).upper().replace(' ', '')
    if base == 'SEE':
        return _SEE_BAND_FILL
    m = _PHASE_TITLE_RE.match(base)
    if m:
        n = int(m.group(1))
        if 1 <= n <= len(_PHASE_BAND_PALETTE):
            return _PHASE_BAND_PALETTE[n - 1]
        return _PHASE_BAND_PALETTE[(n - 1) % len(_PHASE_BAND_PALETTE)]
    return _PHASE_BAND_PALETTE[sequence_index % len(_PHASE_BAND_PALETTE)]


def roll_sort_key(s):
    r = str(getattr(s, 'roll_no', '') or '').strip()
    try:
        return (0, int(r)) if r.isdigit() else (1, r)
    except ValueError:
        return (1, r)


def parse_department_ids(request):
    """
    From GET: dept_ids=1&dept_ids=2 or dept_ids=1,2.
    Returns None if parameter absent → caller uses all departments.
    Returns [] if parameter present but no valid ids.
    """
    raw = request.GET.getlist('dept_ids')
    if not raw and request.GET.get('dept_ids'):
        raw = [x.strip() for x in str(request.GET.get('dept_ids', '')).split(',') if x.strip()]
    if not raw and 'dept_ids' not in request.GET and not request.GET.get('dept_ids'):
        return None
    out = []
    for x in raw:
        try:
            out.append(int(x))
        except (TypeError, ValueError):
            continue
    return out


def selected_departments(request):
    from core.semester_scope import departments_for_institute_semester, get_active_institute_semester

    ids = parse_department_ids(request)
    qs = departments_for_institute_semester(get_active_institute_semester(request)).order_by('name')
    if ids is None:
        return list(qs)
    if not ids:
        return []
    return list(qs.filter(pk__in=ids))


def dept_ids_query_param(depts):
    return '&'.join(f'dept_ids={d.id}' for d in depts)


def build_student_phase_data(students, dept):
    """Same shape as _mark_analytics_build_data_from_students in views."""
    if not students:
        return []
    exam_phases = sort_exam_phases(ExamPhase.objects.filter(department=dept))
    phase_subjects = {}
    for ep in exam_phases:
        phase_subjects[ep.id] = list(
            ExamPhaseSubject.objects.filter(exam_phase=ep).select_related('subject').order_by('subject__name')
        )
    marks_qs = StudentMark.objects.filter(
        student__in=students,
        exam_phase__department=dept,
    ).select_related('subject', 'exam_phase')
    marks_by_student_phase = defaultdict(lambda: defaultdict(dict))
    for m in marks_qs:
        marks_by_student_phase[m.student_id][m.exam_phase_id][m.subject.name] = normalize_student_mark(
            m.marks_obtained
        )
    fail_lb = mark_fail_below_getter_for_departments([dept])
    phase_cut_by_ep = {ep.id: fail_lb(dept, ep.name) for ep in exam_phases}
    result = []
    for s in students:
        phase_wise = []
        for ep in exam_phases:
            subs = phase_subjects.get(ep.id, [])
            subject_marks = []
            phase_cut = phase_cut_by_ep[ep.id]
            for eps in subs:
                marks_val = marks_by_student_phase[s.id][ep.id].get(eps.subject.name)
                is_low = marks_val is not None and marks_val < phase_cut
                subject_marks.append({'name': eps.subject.name, 'marks': marks_val, 'is_low': is_low})
            phase_wise.append({'phase_name': ep.name, 'subjects': subject_marks})
        result.append({'student': s, 'phase_wise': phase_wise, 'department': dept})
    return result


def risk_rows_to_subject_groups(risk_rows):
    m = defaultdict(list)
    for r in risk_rows:
        m[r['subject_name']].append({
            'student': r['student'],
            'phase_name': r['phase_name'],
            'marks': r['marks'],
            'department': r['department'],
        })
    return [{'subject_name': k, 'at_risk': m[k]} for k in sorted(m.keys())]


def build_risk_from_student_data(student_data, threshold=MARK_LOW_THRESHOLD):
    risk_rows = []
    for item in student_data:
        st = item['student']
        dept = item.get('department')
        dname = dept.name if dept else ''
        for phase in item['phase_wise']:
            thr = float(mark_fail_below_threshold(dept, phase['phase_name'])) if dept else float(threshold)
            for s in phase['subjects']:
                v = normalize_student_mark(s.get('marks'))
                if v is None:
                    continue
                if v < thr:
                    risk_rows.append({
                        'department': dname,
                        'student': st,
                        'phase_name': phase['phase_name'],
                        'subject_name': s['name'],
                        'marks': v,
                    })
    risk_rows.sort(key=lambda x: (x['department'], x['subject_name'], x['student'].batch.name if x['student'].batch else '', roll_sort_key(x['student'])))
    return risk_rows


def collect_students(depts, batch_id, roll_search):
    """batch_id: None/'all' = all batches in selected depts; else int pk."""
    all_items = []
    for dept in depts:
        qs = Student.objects.filter(department=dept).select_related('batch', 'department')
        if batch_id and str(batch_id) != 'all':
            qs = qs.filter(batch_id=batch_id)
        if roll_search:
            q = (
                Q(roll_no__icontains=roll_search)
                | Q(name__icontains=roll_search)
                | Q(enrollment_no__icontains=roll_search)
            )
            qs = qs.filter(q)
        studs = list(qs)
        studs.sort(key=lambda s: (s.batch.name if s.batch else '', roll_sort_key(s)))
        all_items.extend(build_student_phase_data(studs, dept))
    return all_items


def filter_phase(student_data, phase_name):
    if not phase_name:
        return student_data
    out = []
    for item in student_data:
        pw = [p for p in item['phase_wise'] if p['phase_name'] == phase_name]
        if pw:
            c = {**item, 'phase_wise': pw}
            out.append(c)
    return out


def student_overall_avg(item):
    vals = []
    for phase in item['phase_wise']:
        for s in phase['subjects']:
            v = normalize_student_mark(s.get('marks'))
            if v is not None:
                vals.append(v)
    return sum(vals) / len(vals) if vals else None


def top_students_per_department(student_data, phase_name=None, n=10):
    by_dept = defaultdict(list)
    data = filter_phase(student_data, phase_name) if phase_name else student_data
    for item in data:
        dept = item.get('department')
        if not dept:
            continue
        avg = student_overall_avg(item)
        if avg is not None:
            by_dept[dept.name].append((avg, item))
    result = {}
    for dname, pairs in by_dept.items():
        pairs.sort(key=lambda x: (-x[0], x[1]['student'].batch.name if x[1]['student'].batch else '', roll_sort_key(x[1]['student'])))
        result[dname] = pairs[:n]
    return result


def top_students_all_departments(student_data, phase_name=None, n=10):
    data = filter_phase(student_data, phase_name) if phase_name else student_data
    pairs = []
    for item in data:
        avg = student_overall_avg(item)
        if avg is not None:
            pairs.append((avg, item))
    pairs.sort(key=lambda x: (-x[0], x[1].get('department').name if x[1].get('department') else '', roll_sort_key(x[1]['student'])))
    return pairs[:n]


def batches_for_departments(depts):
    return list(Batch.objects.filter(department__in=depts).select_related('department').order_by('department__name', 'name'))


def all_phase_names(depts):
    if not depts:
        return []
    names = ExamPhase.objects.filter(department__in=depts).values_list('name', flat=True)
    return sorted_phase_names(names)


def _apply_low_fill(cell, value, *, low_below=None):
    if value is None or value == '':
        return
    try:
        lim = float(low_below) if low_below is not None else float(MARK_LOW_THRESHOLD)
        if float(value) < lim:
            cell.fill = LOW_FILL
    except (TypeError, ValueError):
        pass


def _sanitize_sheet_title(name):
    invalid = '[]:*?/\\'
    s = ''.join('_' if c in invalid else c for c in str(name or '')).strip()
    return s[:31] or 'Sheet'


def _unique_sheet_title(wb, base):
    base = _sanitize_sheet_title(base)
    used = {ws.title for ws in wb.worksheets}
    if base not in used:
        return base
    for i in range(1, 999):
        cand = _sanitize_sheet_title(f'{base[:25]}_{i}')
        if cand not in used:
            return cand
    return _sanitize_sheet_title(f'sheet_{len(used)}')


# SY_1_A1, SY-1-a1, "SY 1 A1" → stream key SY_1 (one sheet for all SY_1_* batches)
_BATCH_STREAM_SY_RE = re.compile(r'^(SY)_(\d+)_(.+)$', re.IGNORECASE)


def _batch_tab_group_key(batch_name):
    """
    One Excel sheet per stream: all batches SY_1_A1, SY_1_A2, … share sheet **SY_1**;
    all SY_2_* share **SY_2**; same for SY_3, SY_4.

    Other names: if there are 3+ underscore pieces, group on first two (e.g. FE_3_X);
    else the whole name is one group.
    """
    name = re.sub(r'\s+', '_', (batch_name or '').strip())
    name = name.replace('-', '_')
    if not name:
        return '_NoBatch'
    m = _BATCH_STREAM_SY_RE.match(name)
    if m:
        return f'{m.group(1).upper()}_{m.group(2)}'
    parts = [p for p in name.split('_') if p]
    if len(parts) >= 3:
        return f'{parts[0]}_{parts[1]}'


def _group_key_sort_tuple(group_key):
    """Sort SY_1, SY_2, … SY_10 in numeric order."""
    if group_key == '_NoBatch':
        return ('\uffff',)
    return tuple(int(p) if p.isdigit() else p.lower() for p in str(group_key).split('_'))


def _col_phase_subject_for_dept(dept):
    col_phase_subject = []
    for ep in sort_exam_phases(ExamPhase.objects.filter(department=dept)):
        for eps in ExamPhaseSubject.objects.filter(exam_phase=ep).select_related('subject').order_by('subject__name'):
            col_phase_subject.append((ep.name, eps.subject.name))
    return col_phase_subject


def _col_phase_subject_union_across_depts(depts):
    """
    For the 'All departments' wide sheet: each phase (T1, T2, …, SEE) appears once in the header;
    subjects under a phase are the union of subject names across selected departments.
    DB phase names like "SY_I - T1" map to short header "T1" so all students share one band per test.
    Returns ordered (short_phase_name, subject_name) pairs for data columns.
    """
    phase_subjects = defaultdict(set)
    for d in depts:
        for pn, sn in _col_phase_subject_for_dept(d):
            short = exam_phase_header_short_name(pn)
            key = short if short else pn
            phase_subjects[key].add(sn)
    ordered_phases = sorted(phase_subjects.keys(), key=exam_phase_name_sort_key)
    col_list = []
    for pn in ordered_phases:
        for sn in sorted(phase_subjects[pn], key=str.lower):
            col_list.append((pn, sn))
    return col_list


def _group_mark_columns_single_dept(col_phase_subject):
    """[(phase_name, [subject_name, ...]), ...] in column order."""
    groups = []
    i = 0
    n = len(col_phase_subject)
    while i < n:
        pn, sn = col_phase_subject[i]
        subs = [sn]
        i += 1
        while i < n and col_phase_subject[i][0] == pn:
            subs.append(col_phase_subject[i][1])
            i += 1
        groups.append((pn, subs))
    return groups


def _group_mark_columns_union(union_cols, multi_dept):
    """union_cols: list of (dept, phase_name, subject_name)."""
    groups = []
    i = 0
    n = len(union_cols)
    while i < n:
        d, pn, sn = union_cols[i]
        subs = [sn]
        i += 1
        while i < n:
            d2, pn2, sn2 = union_cols[i]
            if d2.id != d.id or pn2 != pn:
                break
            subs.append(sn2)
            i += 1
        label = f'{d.name} — {pn}' if multi_dept else pn
        groups.append((label, subs))
    return groups


def _write_two_row_matrix_headers(ws, fixed_labels, phase_subject_groups):
    """
    Row 1–2: fixed columns merged vertically; mark area = merged phase row + subject names row.
    fixed_labels: ['Roll No', 'Name', ...] possibly with 'Department' first.
    phase_subject_groups: [(phase_title, [subj, ...]), ...]
    """
    nfix = len(fixed_labels)
    for c, label in enumerate(fixed_labels, 1):
        cell = ws.cell(1, c, label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN_CC
        cell.border = THIN
        ws.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)
        b2 = ws.cell(2, c)
        b2.border = THIN
    col = nfix + 1
    for seq_idx, (phase_title, subjects) in enumerate(phase_subject_groups):
        nsub = len(subjects)
        pc = ws.cell(1, col, phase_title)
        pc.fill = _phase_band_fill(phase_title, seq_idx)
        pc.font = PHASE_HEADER_FONT
        pc.alignment = HEADER_ALIGN_CC
        pc.border = THIN
        if nsub > 1:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + nsub - 1)
        for k, sub_name in enumerate(subjects):
            sc = ws.cell(2, col + k, sub_name)
            sc.fill = SUBJECT_HEADER_FILL
            sc.font = SUBJECT_HEADER_FONT
            sc.alignment = HEADER_ALIGN_CC
            sc.border = THIN
        col += nsub


def _matrix_column_widths(ws, fixed_labels, phase_subject_groups):
    """Set reasonable column widths from labels."""
    nfix = len(fixed_labels)
    for c, lab in enumerate(fixed_labels, 1):
        ws.column_dimensions[get_column_letter(c)].width = min(22, max(12, len(str(lab)) + 3))
    col = nfix + 1
    for phase_title, subjects in phase_subject_groups:
        for sub_name in subjects:
            w = max(len(str(phase_title)), len(str(sub_name))) + 3
            ws.column_dimensions[get_column_letter(col)].width = min(22, max(11, w))
            col += 1


def _fill_all_departments_wide_sheet(ws, depts, all_dept_student_pairs, *, fail_below_getter=None):
    """
    One matrix sheet: Department + student cols, then T1 / T2 / T3 / SEE (each once), subjects =
    union per phase across selected depts; one row per student; marks from that student's dept only.
    """
    if fail_below_getter is None:
        fail_below_getter = mark_fail_below_getter_for_departments(depts)
    col_phase_subject = _col_phase_subject_union_across_depts(depts)
    fixed = ['Department', 'Roll No', 'Name', 'Enrollment', 'Batch']
    groups = _group_mark_columns_single_dept(col_phase_subject)
    _write_two_row_matrix_headers(ws, fixed, groups)
    sorted_pairs = sorted(
        all_dept_student_pairs,
        key=lambda x: (
            x[0].name,
            _group_key_sort_tuple(
                _batch_tab_group_key(x[1]['student'].batch.name if x[1]['student'].batch else '')
            ),
            (x[1]['student'].batch.name if x[1]['student'].batch else '').lower(),
            roll_sort_key(x[1]['student']),
        ),
    )
    nfix = len(fixed)
    data_start = 3
    for r, (dept, item) in enumerate(sorted_pairs, data_start):
        s = item['student']
        ws.cell(r, 1, dept.name)
        ws.cell(r, 2, s.roll_no)
        ws.cell(r, 3, s.name)
        ws.cell(r, 4, getattr(s, 'enrollment_no', '') or '')
        ws.cell(r, 5, s.batch.name if s.batch else '')
        for i, (pn, sn) in enumerate(col_phase_subject):
            val = _mark_in_item(item, pn, sn)
            cell = ws.cell(r, nfix + 1 + i, '' if val is None else val)
            cell.border = THIN
            _apply_low_fill(cell, val, low_below=fail_below_getter(dept, pn))
        for cc in range(1, nfix + len(col_phase_subject) + 1):
            ws.cell(r, cc).border = THIN
    _matrix_column_widths(ws, fixed, groups)


def _phase_banner_label(phase_names):
    """Single phase → name; multiple → joined with middle dot."""
    if not phase_names:
        return ''
    u = sorted_phase_names(phase_names)
    if len(u) == 1:
        return u[0]
    return ' · '.join(u)


def _write_ranked_marks_sheet_phase_banner(ws, rows, *, with_department, fail_below_getter=None):
    """
    Subject-wise ranked sheet: row 1 = merged phase banner (same style as compiled export);
    row 2 = headers; data from row 3. If all rows share one phase, Phase column is omitted.
    rows: list of (mark_value, StudentMark) sorted by mark descending.
    fail_below_getter: optional (dept, phase_name) -> float; avoids per-row DB hits on large exports.
    """
    if fail_below_getter is None:

        def fail_below_getter(d, p):
            return float(mark_fail_below_threshold(d, p))

    phase_names = {m.exam_phase.name for _, m in rows}
    single_phase = len(phase_names) == 1
    if with_department:
        if single_phase:
            headers = ['Rank', 'Department', 'Roll No', 'Name', 'Enrollment', 'Batch', 'Marks']
        else:
            headers = ['Rank', 'Department', 'Roll No', 'Name', 'Enrollment', 'Batch', 'Phase', 'Marks']
    else:
        if single_phase:
            headers = ['Rank', 'Roll No', 'Name', 'Enrollment', 'Batch', 'Marks']
        else:
            headers = ['Rank', 'Roll No', 'Name', 'Enrollment', 'Batch', 'Phase', 'Marks']
    ncols = len(headers)
    banner = _phase_banner_label(phase_names)
    top = ws.cell(1, 1, banner)
    if single_phase:
        top.fill = _phase_band_fill(next(iter(phase_names)), 0)
    else:
        ordered = sorted_phase_names(phase_names)
        top.fill = _phase_band_fill(ordered[0], 0) if ordered else _PHASE_BAND_PALETTE[0]
    top.font = PHASE_HEADER_FONT
    top.alignment = HEADER_ALIGN_CC
    top.border = THIN
    if ncols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(2, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN_CC
        cell.border = THIN
    for rank, (v, m) in enumerate(rows, 1):
        r = rank + 2
        st = m.student
        c = 1
        ws.cell(r, c, rank)
        c += 1
        if with_department:
            ws.cell(r, c, st.department.name if st.department else '')
            c += 1
        ws.cell(r, c, st.roll_no)
        c += 1
        ws.cell(r, c, st.name)
        c += 1
        ws.cell(r, c, getattr(st, 'enrollment_no', '') or '')
        c += 1
        ws.cell(r, c, st.batch.name if st.batch else '')
        c += 1
        if not single_phase:
            ws.cell(r, c, m.exam_phase.name)
            c += 1
        c_m = ws.cell(r, c, v)
        c_m.border = THIN
        st_dept = st.department
        lb = fail_below_getter(st_dept, m.exam_phase.name) if st_dept else float(MARK_LOW_THRESHOLD)
        _apply_low_fill(c_m, v, low_below=lb)
        for cc in range(1, ncols + 1):
            ws.cell(r, cc).border = THIN
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        sample = str(headers[col - 1])
        ws.column_dimensions[letter].width = min(24, max(10, len(sample) + 2))


def _fill_compiled_matrix_sheet(ws, col_phase_subject, data_rows, *, fail_below_getter=None):
    """Two header rows + data: Roll No, Name, Enrollment, Batch + marks matrix."""
    fixed = ['Roll No', 'Name', 'Enrollment', 'Batch']
    groups = _group_mark_columns_single_dept(col_phase_subject)
    _write_two_row_matrix_headers(ws, fixed, groups)
    nfix = len(fixed)
    data_start = 3
    if fail_below_getter is None and data_rows:
        idpt = data_rows[0].get('department')
        fail_below_getter = mark_fail_below_getter_for_departments([idpt] if idpt else [])
    elif fail_below_getter is None:
        fail_below_getter = lambda dept, pn: float(MARK_LOW_THRESHOLD)
    for r, item in enumerate(data_rows, data_start):
        s = item['student']
        ws.cell(r, 1, s.roll_no)
        ws.cell(r, 2, s.name)
        ws.cell(r, 3, getattr(s, 'enrollment_no', '') or '')
        ws.cell(r, 4, s.batch.name if s.batch else '')
        for i, (pn, sn) in enumerate(col_phase_subject):
            ph = [p for p in item['phase_wise'] if p['phase_name'] == pn]
            val = None
            if ph:
                for sub in ph[0]['subjects']:
                    if sub['name'] == sn:
                        val = sub['marks']
                        break
            col = nfix + 1 + i
            cell = ws.cell(r, col, val if val is not None else '')
            cell.border = THIN
            idpt = item.get('department')
            lb = fail_below_getter(idpt, pn) if idpt else float(MARK_LOW_THRESHOLD)
            _apply_low_fill(cell, val, low_below=lb)
    _matrix_column_widths(ws, fixed, groups)


def _mark_in_item(item, phase_name, subject_name):
    """phase_name may be short (T1, SEE) or full DB name; matches student row by normalized phase."""
    want = exam_phase_header_short_name(phase_name)
    if not want:
        want = (phase_name or '').strip()
    for p in item['phase_wise']:
        got = exam_phase_header_short_name(p['phase_name'])
        if not got:
            got = (p['phase_name'] or '').strip()
        if got != want:
            continue
        for sub in p['subjects']:
            if sub['name'] == subject_name:
                return sub['marks']
    return None


def _fill_combined_sheet(ws, depts, all_dept_student_pairs, *, include_department=None, fail_below_getter=None):
    """
    Single sheet: all batches / departments. Union of mark columns; two-row headers (phase merge, subjects).
    all_dept_student_pairs: list of (dept, item) with item from build_student_phase_data for that dept.
    include_department: if True, always show Department column; if False, never; if None, only when len(depts) > 1.
    Phase row labels use 'Dept — Phase' only when len(depts) > 1; otherwise plain phase name (e.g. T1).
    """
    if include_department is None:
        dept_col = len(depts) > 1
    else:
        dept_col = include_department
    if fail_below_getter is None:
        fail_below_getter = mark_fail_below_getter_for_departments(depts)
    multi_labels = len(depts) > 1
    union_cols = []
    for d in depts:
        for pn, sn in _col_phase_subject_for_dept(d):
            union_cols.append((d, pn, sn))
    if dept_col:
        fixed = ['Department', 'Roll No', 'Name', 'Enrollment', 'Batch']
    else:
        fixed = ['Roll No', 'Name', 'Enrollment', 'Batch']
    groups = _group_mark_columns_union(union_cols, multi_labels)
    _write_two_row_matrix_headers(ws, fixed, groups)
    sorted_pairs = sorted(
        all_dept_student_pairs,
        key=lambda x: (
            x[0].name,
            _group_key_sort_tuple(
                _batch_tab_group_key(x[1]['student'].batch.name if x[1]['student'].batch else '')
            ),
            (x[1]['student'].batch.name if x[1]['student'].batch else '').lower(),
            roll_sort_key(x[1]['student']),
        ),
    )
    data_start = 3
    for r, (dept, item) in enumerate(sorted_pairs, data_start):
        s = item['student']
        c0 = 1
        if dept_col:
            ws.cell(r, c0, dept.name)
            c0 += 1
        ws.cell(r, c0, s.roll_no)
        ws.cell(r, c0 + 1, s.name)
        ws.cell(r, c0 + 2, getattr(s, 'enrollment_no', '') or '')
        ws.cell(r, c0 + 3, s.batch.name if s.batch else '')
        base_col = c0 + 4
        for j, (ud, pn, sn) in enumerate(union_cols):
            val = _mark_in_item(item, pn, sn) if ud.id == dept.id else None
            cell = ws.cell(r, base_col + j, val if val is not None else '')
            cell.border = THIN
            _apply_low_fill(cell, val, low_below=fail_below_getter(ud, pn))
    _matrix_column_widths(ws, fixed, groups)


def excel_compiled_per_department(depts):
    """
    Workbook layout:
    - Sheet **SY_1** = every student in any batch SY_1_A1, SY_1_A2, … (all SY_1 sections together).
    - Sheet **SY_2** = all SY_2_* batches, then SY_3, SY_4, …
    - **No_Batch** if needed.
    - Last sheet **Combined** = all streams (SY_1 + SY_2 + …) in one table.
    """
    wb = Workbook()
    wb.remove(wb.active)
    all_combined_pairs = []
    fail_lb = mark_fail_below_getter_for_departments(depts)

    for dept in depts:
        col_phase_subject = _col_phase_subject_for_dept(dept)

        dept_students = list(
            Student.objects.filter(department=dept).select_related('batch').order_by('batch__name', 'roll_no')
        )
        if not dept_students:
            continue
        full_data = build_student_phase_data(dept_students, dept)

        sheet_prefix = f'{_sanitize_sheet_title(dept.name)}_' if len(depts) > 1 else ''
        batches = list(Batch.objects.filter(department=dept).order_by('name'))

        group_to_batch_ids = defaultdict(set)
        for batch in batches:
            gkey = _batch_tab_group_key(batch.name)
            group_to_batch_ids[gkey].add(batch.id)

        for gkey in sorted(group_to_batch_ids.keys(), key=_group_key_sort_tuple):
            batch_ids = group_to_batch_ids[gkey]
            batch_data = [row for row in full_data if row['student'].batch_id in batch_ids]
            if not batch_data:
                continue
            batch_data.sort(
                key=lambda it: (
                    (it['student'].batch.name if it['student'].batch else '').lower(),
                    roll_sort_key(it['student']),
                )
            )
            title = _unique_sheet_title(wb, f'{sheet_prefix}{gkey}'.replace(' ', '_'))
            ws = wb.create_sheet(title=title)
            _fill_compiled_matrix_sheet(ws, col_phase_subject, batch_data, fail_below_getter=fail_lb)

        no_batch = [row for row in full_data if not row['student'].batch_id]
        if no_batch:
            no_batch.sort(key=lambda it: roll_sort_key(it['student']))
            title = _unique_sheet_title(wb, f'{sheet_prefix}No_Batch')
            ws = wb.create_sheet(title=title)
            _fill_compiled_matrix_sheet(ws, col_phase_subject, no_batch, fail_below_getter=fail_lb)

        for row in full_data:
            all_combined_pairs.append((dept, row))

    if not wb.worksheets:
        ws = wb.create_sheet('Empty')
        ws.cell(1, 1, 'No departments selected.')
    else:
        tcomb = _unique_sheet_title(wb, 'Combined')
        ws_c = wb.create_sheet(title=tcomb)
        _fill_combined_sheet(ws_c, depts, all_combined_pairs, fail_below_getter=fail_lb)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def excel_compiled_all_departments(depts):
    """
    Single sheet 'All departments': Department + student info, then header row with **T1, T2, T3, SEE**
    only (once each), subjects under each phase = union across selected depts; one row per student.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'All departments'
    if not depts:
        ws.cell(1, 1, 'No departments selected.')
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio

    all_dept_student_pairs = []
    for dept in depts:
        dept_students = list(
            Student.objects.filter(department=dept).select_related('batch').order_by('batch__name', 'roll_no')
        )
        if not dept_students:
            continue
        full_data = build_student_phase_data(dept_students, dept)
        for row in full_data:
            all_dept_student_pairs.append((dept, row))

    if not all_dept_student_pairs:
        ws.cell(1, 1, 'No students in selected departments.')
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio

    fail_lb = mark_fail_below_getter_for_departments(depts)
    _fill_all_departments_wide_sheet(ws, depts, all_dept_student_pairs, fail_below_getter=fail_lb)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def excel_subject_wise_per_department(depts):
    """Workbook: one sheet per (department, subject) with marks listed descending."""
    wb = Workbook()
    wb.remove(wb.active)
    used_titles = set()
    for dept in depts:
        fail_lb = mark_fail_below_getter_for_departments([dept])
        subject_ids = (
            StudentMark.objects.filter(student__department=dept, exam_phase__department=dept)
            .values_list('subject_id', flat=True)
            .distinct()
        )
        for subj in Subject.objects.filter(pk__in=subject_ids, department=dept).order_by('name'):
            base = f'{dept.name[:12]}_{subj.name}'[:28]
            title = base
            n = 0
            while title in used_titles:
                n += 1
                title = f'{base[:25]}_{n}'
            used_titles.add(title)
            ws = wb.create_sheet(title=title)
            marks_rows = []
            for m in (
                StudentMark.objects.filter(subject=subj, student__department=dept, exam_phase__department=dept)
                .select_related('student', 'student__batch', 'exam_phase')
            ):
                v = normalize_student_mark(m.marks_obtained)
                if v is None:
                    continue
                marks_rows.append((v, m))
            marks_rows.sort(key=lambda x: -x[0])
            _write_ranked_marks_sheet_phase_banner(ws, marks_rows, with_department=False, fail_below_getter=fail_lb)
    if not wb.worksheets:
        ws = wb.create_sheet('Empty')
        ws.cell(1, 1, 'No marks.')
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def excel_subject_wise_all_departments(depts):
    """One sheet per subject name (global): all depts, marks descending."""
    wb = Workbook()
    wb.remove(wb.active)
    fail_lb = mark_fail_below_getter_for_departments(depts)
    marks_flat = []
    for m in StudentMark.objects.filter(
        student__department__in=depts,
        exam_phase__department__in=depts,
    ).select_related('student', 'student__batch', 'student__department', 'exam_phase', 'subject'):
        v = normalize_student_mark(m.marks_obtained)
        if v is None:
            continue
        marks_flat.append((m.subject.name, v, m))
    by_subject = defaultdict(list)
    for sn, v, m in marks_flat:
        by_subject[sn].append((v, m))
    used = set()
    for sn in sorted(by_subject.keys(), key=str.lower):
        rows = by_subject[sn]
        rows.sort(key=lambda x: -x[0])
        title = ''.join(c if c.isalnum() or c in ' _-' else '_' for c in sn)[:31] or 'Subject'
        t2 = title
        i = 0
        while t2 in used:
            i += 1
            t2 = f'{title[:28]}_{i}'
        used.add(t2)
        ws = wb.create_sheet(title=t2)
        _write_ranked_marks_sheet_phase_banner(ws, rows, with_department=True, fail_below_getter=fail_lb)
    if not wb.worksheets:
        ws = wb.create_sheet('Empty')
        ws.cell(1, 1, 'No marks.')
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def excel_phase_compile(depts):
    """
    One sheet per phase name (union): wide matrix — Department, Roll No, Name, Enrollment, Batch +
    one column per subject (that phase). Merged phase band + subject header row; rows sorted by
    department name then roll number ascending.
    """
    phase_sid_marks = defaultdict(lambda: defaultdict(dict))
    phase_sid_student = {}
    for m in StudentMark.objects.filter(
        student__department__in=depts,
        exam_phase__department__in=depts,
    ).select_related('student', 'student__batch', 'student__department', 'exam_phase', 'subject'):
        v = normalize_student_mark(m.marks_obtained)
        if v is None:
            continue
        pname = m.exam_phase.name
        sid = m.student_id
        phase_sid_marks[pname][sid][m.subject.name] = v
        phase_sid_student[(pname, sid)] = m.student

    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    fixed = ['Department', 'Roll No', 'Name', 'Enrollment', 'Batch']
    nfix = len(fixed)
    fail_lb = mark_fail_below_getter_for_departments(depts)

    for pname in sorted(phase_sid_marks.keys(), key=exam_phase_name_sort_key):
        sid_to_subj = phase_sid_marks[pname]
        if not sid_to_subj:
            continue
        subj_set = set()
        for smap in sid_to_subj.values():
            subj_set.update(smap.keys())
        subjects = sorted(subj_set, key=str.lower)
        groups = [(pname, subjects)]

        title = ''.join(c if c.isalnum() or c in ' _-' else '_' for c in pname)[:31] or 'Phase'
        t2 = title
        i = 0
        while t2 in used:
            i += 1
            t2 = f'{title[:28]}_{i}'
        used.add(t2)
        ws = wb.create_sheet(title=t2)
        _write_two_row_matrix_headers(ws, fixed, groups)

        sorted_sids = sorted(
            sid_to_subj.keys(),
            key=lambda sid: (
                (phase_sid_student[(pname, sid)].department.name or '').lower(),
                roll_sort_key(phase_sid_student[(pname, sid)]),
            ),
        )
        for offset, sid in enumerate(sorted_sids):
            r = 3 + offset
            st = phase_sid_student[(pname, sid)]
            ws.cell(r, 1, st.department.name if st.department else '')
            ws.cell(r, 2, st.roll_no)
            ws.cell(r, 3, st.name)
            ws.cell(r, 4, getattr(st, 'enrollment_no', '') or '')
            ws.cell(r, 5, st.batch.name if st.batch else '')
            for j, sub in enumerate(subjects):
                val = sid_to_subj[sid].get(sub)
                cell = ws.cell(r, nfix + 1 + j, '' if val is None else val)
                cell.border = THIN
                sd = st.department
                lb = fail_lb(sd, pname) if sd else float(MARK_LOW_THRESHOLD)
                _apply_low_fill(cell, val, low_below=lb)
            for cc in range(1, nfix + len(subjects) + 1):
                ws.cell(r, cc).border = THIN

        _matrix_column_widths(ws, fixed, groups)

    if not wb.worksheets:
        ws = wb.create_sheet('Empty')
        ws.cell(1, 1, 'No marks.')
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def excel_risk_students(depts, student_data=None, subject_name=None):
    wb = Workbook()
    ws = wb.active
    if subject_name:
        st = _sanitize_sheet_title(subject_name)
        ws.title = st if st else 'Risk_lt9'
    else:
        ws.title = 'Risk_lt9'
    headers = ['Department', 'Roll No', 'Name', 'Enrollment', 'Batch', 'Phase', 'Subject', 'Marks']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN
    rows = build_risk_from_student_data(student_data) if student_data else build_risk_from_student_data(collect_students(depts, 'all', ''))
    if subject_name:
        rows = [r for r in rows if r['subject_name'] == subject_name]
    row = 2
    for r in rows:
        st = r['student']
        ws.cell(row, 1, r['department'])
        ws.cell(row, 2, st.roll_no)
        ws.cell(row, 3, st.name)
        ws.cell(row, 4, getattr(st, 'enrollment_no', '') or '')
        ws.cell(row, 5, st.batch.name if st.batch else '')
        ws.cell(row, 6, r['phase_name'])
        ws.cell(row, 7, r['subject_name'])
        c_m = ws.cell(row, 8, r['marks'])
        c_m.fill = LOW_FILL
        c_m.border = THIN
        for c in range(1, 8):
            ws.cell(row, c).border = THIN
        row += 1
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def excel_top10_report(depts, phase_name=None):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Top10_per_dept'
    full_data = collect_students(depts, 'all', '')
    top_per = top_students_per_department(full_data, phase_name=phase_name, n=10)
    row = 1
    ws.cell(row, 1, 'Top 10 by average' + (f' — phase {phase_name}' if phase_name else ' — all phases'))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 2
    for dname in sorted(top_per.keys(), key=str.lower):
        ws.cell(row, 1, dname)
        ws.cell(row, 1).font = Font(bold=True)
        row += 1
        for c, h in enumerate(['Rank', 'Roll No', 'Name', 'Enrollment', 'Batch', 'Avg'], 1):
            cell = ws.cell(row, c, h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN
        row += 1
        for i, (avg, item) in enumerate(top_per[dname], 1):
            st = item['student']
            ws.cell(row, 1, i)
            ws.cell(row, 2, st.roll_no)
            ws.cell(row, 3, st.name)
            ws.cell(row, 4, getattr(st, 'enrollment_no', '') or '')
            ws.cell(row, 5, st.batch.name if st.batch else '')
            ws.cell(row, 6, round(avg, 2))
            row += 1
        row += 1
    # All depts sheet
    ws2 = wb.create_sheet('Top10_all')
    ws2.cell(1, 1, 'Top 10 overall (by average)')
    pairs = top_students_all_departments(full_data, phase_name=phase_name, n=10)
    row = 3
    for c, h in enumerate(['Rank', 'Department', 'Roll No', 'Name', 'Enrollment', 'Batch', 'Avg'], 1):
        cell = ws2.cell(row, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN
    row += 1
    for i, (avg, item) in enumerate(pairs, 1):
        st = item['student']
        dept = item.get('department')
        ws2.cell(row, 1, i)
        ws2.cell(row, 2, dept.name if dept else '')
        ws2.cell(row, 3, st.roll_no)
        ws2.cell(row, 4, st.name)
        ws2.cell(row, 5, getattr(st, 'enrollment_no', '') or '')
        ws2.cell(row, 6, st.batch.name if st.batch else '')
        ws2.cell(row, 7, round(avg, 2))
        row += 1
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
