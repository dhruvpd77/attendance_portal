"""
Exam section hierarchy: exam_section → dept parent → sub-unit (SY_1) and supervision uploads.
"""
from __future__ import annotations

import random
import re
from collections import Counter
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from urllib.parse import urlencode

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db import IntegrityError, transaction
from django.db.models import Prefetch, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_http_methods

from accounts.models import UserRole
from core.models import (
    Department,
    DepartmentExamProfile,
    Faculty,
    InstituteSemester,
    PaperCheckingAdjustedShare,
    PaperCheckingAllocation,
    PaperCheckingCompletionRequest,
    PaperCheckingDuty,
    PaperCheckingPhase,
    PaperSettingCompletionRequest,
    PaperSettingDuty,
    PaperSettingPhase,
    SupervisionDuty,
    SupervisionExamPhase,
)
from core.paper_checking_credits import (
    credit_for_completion_request,
    department_approved_paper_credit_rows,
    paper_count_for_completion,
    remuneration_for_completion_request,
)
from core.paper_setting_credits import (
    credit_for_paper_setting_request,
    department_paper_setting_credit_rows,
    remuneration_for_paper_setting_request,
    supervision_credit_for_phase,
    supervision_remuneration_for_phase,
)
from core.paper_duty_scope import (
    build_faculty_paper_checking_rows,
    build_faculty_paper_setting_rows,
    checking_phases_department,
    checking_phases_hub_user,
    checking_phases_institute_for_request,
    hub_managed_department_ids,
    paper_checking_duties_for_child_prof,
    paper_setting_duties_for_child_prof,
    pending_paper_setting_completion_requests_for_parent,
    history_paper_setting_completion_requests_for_parent,
    setting_phases_department,
    setting_phases_hub_user,
    setting_phases_institute_for_request,
)
from core.exam_faculty_portal_visibility import (
    paper_checking_completion_in_faculty_exam_history,
    paper_checking_duty_in_faculty_exam_history,
    paper_setting_completion_in_faculty_exam_history,
    paper_setting_duty_in_faculty_exam_history,
    supervision_duty_in_faculty_exam_history,
)
from core.semester_scope import (
    SESSION_KEY_COORD_CHILD_PROFILE_ID,
    SESSION_KEY_COORD_PARENT_PROFILE_ID,
    child_must_select_exam_context,
    coordinator_must_select_exam_context,
    coordinator_child_profiles_qs,
    departments_for_exam_coordination_request,
    exam_section_working_semester_ids,
    set_exam_section_working_semester_ids,
    get_active_child_exam_profile,
    get_active_parent_exam_profile,
    institute_semester_for_exam_portal,
    is_exam_section_operator,
    q_completion_duty_phase_in_semesters,
    q_supervision_duty_phase_in_semesters,
)
from core.exam_daily_dr_export import (
    _phase_supervision_bucket,
    build_exam_daily_dr_workbook,
    parse_dates_from_request,
)
from core.supervision_dr_export import build_supervision_dr_excel
from core.exam_upload_staging import (
    clear_staging,
    supervision_stage_deserialize_rows,
    supervision_stage_get,
    supervision_stage_put,
)
from core.supervision_excel import match_faculty_for_department, match_faculty_global, parse_combined_supervision_workbook
from core.exam_subunit_scope import duty_visible_to_subunit, subunit_supervision_duty_filter_q, phases_for_subunit_prof
from core.views import is_super_admin


def _analytics_excel_query(
    date_from: str,
    date_to: str,
    faculty_id: int | None = None,
    *,
    paper_phase_id: int | None = None,
    supervision_phase_id: int | None = None,
) -> str:
    q = {}
    if (date_from or '').strip():
        q['date_from'] = date_from.strip()
    if (date_to or '').strip():
        q['date_to'] = date_to.strip()
    if faculty_id:
        q['faculty_id'] = str(faculty_id)
    if paper_phase_id:
        q['paper_phase_id'] = str(paper_phase_id)
    if supervision_phase_id:
        q['supervision_phase_id'] = str(supervision_phase_id)
    return urlencode(q)


def _parse_optional_int_param(request, key: str) -> int | None:
    raw = (request.GET.get(key) or '').strip()
    if not raw:
        return None
    try:
        return int(raw)
    except (TypeError, ValueError):
        return None


def _child_can_access_duty(prof, duty: SupervisionDuty) -> bool:
    return duty_visible_to_subunit(prof, duty)


def _exam_section_portal_access(request):
    """Exam-section login or institute super admin (e.g. DVP / central admin)."""
    return _exam_section_only(request) or is_super_admin(request)


def _exam_section_only(request):
    try:
        return request.user.role_profile.role == 'exam_section'
    except Exception:
        return False


def _dept_parent_only(request):
    try:
        return request.user.role_profile.role == 'dept_exam_parent'
    except Exception:
        return False


def _dept_child_only(request):
    try:
        return request.user.role_profile.role == 'dept_exam_child'
    except Exception:
        return False


def _parent_profile(request):
    return get_active_parent_exam_profile(request)


def _child_profile(request):
    return get_active_child_exam_profile(request)


def _is_hub_coordinator(prof) -> bool:
    return bool(prof and prof.parent_id is None and getattr(prof, 'is_hub_coordinator', False))


def _credit_analytics_faculty_qs(request):
    """Faculty list for coordinator paper-check credit analytics (department / hub scope)."""
    if _dept_child_only(request):
        prof = _child_profile(request)
        if not prof or not prof.department_id:
            return Faculty.objects.none()
        return Faculty.objects.filter(department_id=prof.department_id).select_related(
            'department'
        ).order_by('full_name')
    if _dept_parent_only(request):
        prof = _parent_profile(request)
        if not prof:
            return Faculty.objects.none()
        if _is_hub_coordinator(prof):
            all_d = hub_managed_department_ids(prof)
            if not all_d:
                return Faculty.objects.none()
            return Faculty.objects.filter(department_id__in=all_d).select_related(
                'department'
            ).order_by('department__name', 'full_name')
        if not prof.department_id:
            return Faculty.objects.none()
        return Faculty.objects.filter(department_id=prof.department_id).select_related(
            'department'
        ).order_by('full_name')
    return Faculty.objects.none()


def _approved_credit_rows_for_faculty(faculty: Faculty, date_from_str: str, date_to_str: str):
    """Approved completion rows + total credits, optional decided-at date filter."""
    qs = (
        PaperCheckingCompletionRequest.objects.filter(
            faculty=faculty,
            status=PaperCheckingCompletionRequest.APPROVED,
        )
        .select_related('duty', 'duty__phase', 'decided_by')
        .prefetch_related(
            Prefetch(
                'duty__allocations',
                queryset=PaperCheckingAllocation.objects.select_related('department'),
            ),
            Prefetch(
                'duty__adjusted_shares',
                queryset=PaperCheckingAdjustedShare.objects.select_related('faculty'),
            ),
        )
    )
    if date_from_str:
        try:
            d0 = datetime.strptime(date_from_str.strip(), '%Y-%m-%d').date()
            qs = qs.filter(decided_at__date__gte=d0)
        except ValueError:
            pass
    if date_to_str:
        try:
            d1 = datetime.strptime(date_to_str.strip(), '%Y-%m-%d').date()
            qs = qs.filter(decided_at__date__lte=d1)
        except ValueError:
            pass
    qs = qs.order_by('decided_at', 'id')
    rows = []
    total = Decimal('0')
    running = Decimal('0')
    for r in qs:
        c = credit_for_completion_request(r)
        total += c
        running += c
        rows.append({'req': r, 'credit': c, 'running': running})
    return rows, total


def _paper_completion_prefetch():
    return [
        Prefetch(
            'duty__allocations',
            queryset=PaperCheckingAllocation.objects.select_related('department'),
        ),
        Prefetch(
            'duty__adjusted_shares',
            queryset=PaperCheckingAdjustedShare.objects.select_related('faculty'),
        ),
    ]


def _bulk_approved_paper_completions_qs(
    fac_qs,
    date_from_str: str,
    date_to_str: str,
    phase_id: int | None = None,
    *,
    exam_semester_ids: list[int] | None = None,
):
    qs = (
        PaperCheckingCompletionRequest.objects.filter(
            faculty__in=fac_qs,
            status=PaperCheckingCompletionRequest.APPROVED,
        )
        .select_related('duty', 'duty__phase', 'faculty', 'decided_by')
        .prefetch_related(*_paper_completion_prefetch())
    )
    if exam_semester_ids:
        qs = qs.filter(q_completion_duty_phase_in_semesters(exam_semester_ids))
    if phase_id:
        qs = qs.filter(duty__phase_id=phase_id)
    if date_from_str:
        try:
            d0 = datetime.strptime(date_from_str.strip(), '%Y-%m-%d').date()
            qs = qs.filter(decided_at__date__gte=d0)
        except ValueError:
            pass
    if date_to_str:
        try:
            d1 = datetime.strptime(date_to_str.strip(), '%Y-%m-%d').date()
            qs = qs.filter(decided_at__date__lte=d1)
        except ValueError:
            pass
    return qs.order_by('faculty_id', 'decided_at', 'id')


def _paper_summaries_and_detail_map(fac_ordered: list, completions: list):
    """One pass: totals per faculty + detail rows with running credit."""
    by_f: dict[int, list] = {}
    for r in completions:
        c = credit_for_completion_request(r)
        by_f.setdefault(r.faculty_id, []).append({'req': r, 'credit': c, 'running': Decimal('0')})
    summaries = []
    grand = Decimal('0')
    detail_map: dict[int, list] = {}
    for fac in fac_ordered:
        items = by_f.get(fac.pk, [])
        running = Decimal('0')
        for it in items:
            running += it['credit']
            it['running'] = running
        summaries.append({'faculty': fac, 'total_credit': running, 'n_rows': len(items)})
        grand += running
        detail_map[fac.pk] = items
    return summaries, grand, detail_map


def _supervision_completed_scope_qs(request, phase_id: int | None = None):
    """Completed supervision rows visible to this coordinator role."""
    base = SupervisionDuty.objects.filter(
        completion_status=SupervisionDuty.COMPLETED,
        faculty__isnull=False,
    ).select_related('faculty', 'faculty__department', 'phase', 'original_faculty')
    if phase_id:
        base = base.filter(phase_id=phase_id)
    if _dept_child_only(request):
        prof = _child_profile(request)
        if not prof or not prof.department_id:
            return SupervisionDuty.objects.none()
        return base.filter(subunit_supervision_duty_filter_q(prof))
    if _dept_parent_only(request):
        prof = _parent_profile(request)
        if not prof:
            return SupervisionDuty.objects.none()
        if _is_hub_coordinator(prof):
            fac_ids = list(_credit_analytics_faculty_qs(request).values_list('pk', flat=True))
            if not fac_ids:
                return SupervisionDuty.objects.none()
            return base.filter(phase__hub_coordinator=prof.user, faculty_id__in=fac_ids)
        return base.filter(phase__department=prof.department)
    return SupervisionDuty.objects.none()


def _supervision_completed_institute_qs(
    phase_id: int | None = None,
    *,
    exam_semester_ids: list[int] | None = None,
):
    qs = SupervisionDuty.objects.filter(
        completion_status=SupervisionDuty.COMPLETED,
        faculty__isnull=False,
    ).select_related('faculty', 'faculty__department', 'phase', 'original_faculty')
    if phase_id:
        qs = qs.filter(phase_id=phase_id)
    if exam_semester_ids:
        qs = qs.filter(q_supervision_duty_phase_in_semesters(exam_semester_ids))
    return qs


def _apply_supervision_completed_date_filter(qs, date_from_str: str, date_to_str: str):
    if date_from_str:
        try:
            d0 = datetime.strptime(date_from_str.strip(), '%Y-%m-%d').date()
            qs = qs.filter(completed_at__date__gte=d0)
        except ValueError:
            pass
    if date_to_str:
        try:
            d1 = datetime.strptime(date_to_str.strip(), '%Y-%m-%d').date()
            qs = qs.filter(completed_at__date__lte=d1)
        except ValueError:
            pass
    return qs.order_by('faculty_id', 'completed_at', 'pk')


def _supervision_summaries_and_detail_map(fac_ordered: list, duties: list):
    by_f: dict[int, list] = {}
    for d in duties:
        by_f.setdefault(d.faculty_id, []).append(d)
    summaries = []
    grand = 0
    detail_map: dict[int, list] = {}
    for fac in fac_ordered:
        rows = by_f.get(fac.pk, [])
        j = k = l_n = 0
        for duty in rows:
            b = _phase_supervision_bucket(duty.phase.name if duty.phase else '')
            if b == 'J':
                j += 1
            elif b == 'K':
                k += 1
            else:
                l_n += 1
        n = len(rows)
        summaries.append(
            {
                'faculty': fac,
                'total_completed': n,
                'bucket_j': j,
                'bucket_k': k,
                'bucket_l': l_n,
            }
        )
        grand += n
        detail_map[fac.pk] = sorted(rows, key=lambda x: (x.completed_at or x.supervision_date, x.pk))
    return summaries, grand, detail_map


def _build_paper_credit_excel_workbook(
    *,
    scope_line: str,
    summaries: list,
    grand_credit: Decimal,
    detail_map: dict[int, list],
    fac_ordered: list,
    single_faculty_id: int | None,
):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    bold = Font(bold=True)

    if single_faculty_id:
        fac = next((x['faculty'] for x in summaries if x['faculty'].pk == single_faculty_id), None)
        items = detail_map.get(single_faculty_id, [])
        ws = wb.active
        ws.title = 'Paper credits'
        sub = fac.full_name if fac else ''
        ws['A1'] = f'Approved paper-check credits — {sub}'
        ws['A2'] = scope_line
        hr = 4
        for c, h in enumerate(
            ['Phase', 'Exam date', 'Subject', 'Papers', 'Credit', 'Running', 'Decided', 'By'],
            1,
        ):
            cell = ws.cell(hr, c)
            cell.value = h
            cell.font = bold
        row = hr + 1
        for item in items:
            r = item['req']
            ws.cell(row, 1).value = r.duty.phase.name if r.duty.phase else ''
            ws.cell(row, 2).value = r.duty.exam_date.isoformat() if r.duty.exam_date else ''
            ws.cell(row, 3).value = r.duty.subject_name or ''
            ws.cell(row, 4).value = r.papers_for_faculty_display()
            ws.cell(row, 5).value = float(item['credit'])
            ws.cell(row, 6).value = float(item['running'])
            ws.cell(row, 7).value = r.decided_at.isoformat() if r.decided_at else ''
            ws.cell(row, 8).value = r.decided_by.username if r.decided_by_id else ''
            row += 1
        return wb

    ws_sum = wb.active
    ws_sum.title = 'Summary'
    ws_sum['A1'] = 'Paper-check credits — all faculty in scope'
    ws_sum['A2'] = scope_line
    hr = 4
    for c, h in enumerate(
        ['Faculty', 'Short', 'Department', 'Total credit', 'Completions'],
        1,
    ):
        cell = ws_sum.cell(hr, c)
        cell.value = h
        cell.font = bold
    row = hr + 1
    for s in summaries:
        f = s['faculty']
        ws_sum.cell(row, 1).value = f.full_name
        ws_sum.cell(row, 2).value = f.short_name or ''
        ws_sum.cell(row, 3).value = f.department.name if f.department_id else ''
        ws_sum.cell(row, 4).value = float(s['total_credit'])
        ws_sum.cell(row, 5).value = s['n_rows']
        row += 1
    cell = ws_sum.cell(row, 1)
    cell.value = 'GRAND TOTAL'
    cell.font = bold
    ws_sum.cell(row, 4).value = float(grand_credit)
    ws_sum.cell(row, 5).value = sum(s['n_rows'] for s in summaries)

    ws_det = wb.create_sheet('All completions')
    dhr = 1
    for c, h in enumerate(
        [
            'Faculty',
            'Short',
            'Dept',
            'Phase',
            'Exam date',
            'Subject',
            'Papers',
            'Credit',
            'Running',
            'Decided',
            'By',
        ],
        1,
    ):
        cell = ws_det.cell(dhr, c)
        cell.value = h
        cell.font = bold
    drow = dhr + 1
    for fac in fac_ordered:
        for item in detail_map.get(fac.pk, []):
            r = item['req']
            ws_det.cell(drow, 1).value = fac.full_name
            ws_det.cell(drow, 2).value = fac.short_name or ''
            ws_det.cell(drow, 3).value = fac.department.name if fac.department_id else ''
            ws_det.cell(drow, 4).value = r.duty.phase.name if r.duty.phase else ''
            ws_det.cell(drow, 5).value = r.duty.exam_date.isoformat() if r.duty.exam_date else ''
            ws_det.cell(drow, 6).value = r.duty.subject_name or ''
            ws_det.cell(drow, 7).value = r.papers_for_faculty_display()
            ws_det.cell(drow, 8).value = float(item['credit'])
            ws_det.cell(drow, 9).value = float(item['running'])
            ws_det.cell(drow, 10).value = r.decided_at.isoformat() if r.decided_at else ''
            ws_det.cell(drow, 11).value = r.decided_by.username if r.decided_by_id else ''
            drow += 1
    return wb


def _build_supervision_credit_excel_workbook(
    *,
    scope_line: str,
    summaries: list,
    grand_n: int,
    detail_map: dict[int, list],
    fac_ordered: list,
    single_faculty_id: int | None,
):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    bold = Font(bold=True)

    def _bucket_label(phase_name: str) -> str:
        b = _phase_supervision_bucket(phase_name)
        if b == 'J':
            return 'T1–T3'
        if b == 'K':
            return 'SEE/T4'
        return 'REM'

    if single_faculty_id:
        fac = next((x['faculty'] for x in summaries if x['faculty'].pk == single_faculty_id), None)
        rows = detail_map.get(single_faculty_id, [])
        ws = wb.active
        ws.title = 'Supervision'
        ws['A1'] = f'Completed supervision — {fac.full_name if fac else ""}'
        ws['A2'] = scope_line
        hr = 4
        for c, h in enumerate(
            [
                'Phase',
                'Bucket',
                'Supervision date',
                'Time',
                'Subject',
                'Division',
                'Room',
                'Block',
                'Proxy',
                'Completed at',
            ],
            1,
        ):
            ws.cell(hr, c).value = h
            ws.cell(hr, c).font = bold
        row = hr + 1
        for duty in rows:
            ws.cell(row, 1).value = duty.phase.name if duty.phase else ''
            ws.cell(row, 2).value = _bucket_label(duty.phase.name if duty.phase else '')
            ws.cell(row, 3).value = duty.supervision_date.isoformat() if duty.supervision_date else ''
            ws.cell(row, 4).value = duty.time_slot or ''
            ws.cell(row, 5).value = duty.subject_name or ''
            ws.cell(row, 6).value = duty.division_code or ''
            ws.cell(row, 7).value = duty.room_no or ''
            ws.cell(row, 8).value = duty.block_no or ''
            ws.cell(row, 9).value = 'Y' if duty.is_proxy else ''
            ws.cell(row, 10).value = duty.completed_at.isoformat() if duty.completed_at else ''
            row += 1
        return wb

    ws_sum = wb.active
    ws_sum.title = 'Summary'
    ws_sum['A1'] = 'Supervision — completed duties (all faculty in scope)'
    ws_sum['A2'] = scope_line
    hr = 4
    for c, h in enumerate(
        [
            'Faculty',
            'Short',
            'Department',
            'Total',
            'T1–T3',
            'SEE/T4',
            'REM',
        ],
        1,
    ):
        ws_sum.cell(hr, c).value = h
        ws_sum.cell(hr, c).font = bold
    row = hr + 1
    for s in summaries:
        f = s['faculty']
        ws_sum.cell(row, 1).value = f.full_name
        ws_sum.cell(row, 2).value = f.short_name or ''
        ws_sum.cell(row, 3).value = f.department.name if f.department_id else ''
        ws_sum.cell(row, 4).value = s['total_completed']
        ws_sum.cell(row, 5).value = s['bucket_j']
        ws_sum.cell(row, 6).value = s['bucket_k']
        ws_sum.cell(row, 7).value = s['bucket_l']
        row += 1
    ws_sum.cell(row, 1).value = 'GRAND TOTAL'
    ws_sum.cell(row, 1).font = bold
    ws_sum.cell(row, 4).value = grand_n

    ws_det = wb.create_sheet('All completions')
    dhr = 1
    for c, h in enumerate(
        [
            'Faculty',
            'Short',
            'Dept',
            'Phase',
            'Bucket',
            'Supervision date',
            'Time',
            'Subject',
            'Division',
            'Room',
            'Block',
            'Proxy',
            'Completed at',
        ],
        1,
    ):
        ws_det.cell(dhr, c).value = h
        ws_det.cell(dhr, c).font = bold
    drow = dhr + 1
    for fac in fac_ordered:
        for duty in detail_map.get(fac.pk, []):
            ws_det.cell(drow, 1).value = fac.full_name
            ws_det.cell(drow, 2).value = fac.short_name or ''
            ws_det.cell(drow, 3).value = fac.department.name if fac.department_id else ''
            ws_det.cell(drow, 4).value = duty.phase.name if duty.phase else ''
            ws_det.cell(drow, 5).value = _bucket_label(duty.phase.name if duty.phase else '')
            ws_det.cell(drow, 6).value = duty.supervision_date.isoformat() if duty.supervision_date else ''
            ws_det.cell(drow, 7).value = duty.time_slot or ''
            ws_det.cell(drow, 8).value = duty.subject_name or ''
            ws_det.cell(drow, 9).value = duty.division_code or ''
            ws_det.cell(drow, 10).value = duty.room_no or ''
            ws_det.cell(drow, 11).value = duty.block_no or ''
            ws_det.cell(drow, 12).value = 'Y' if duty.is_proxy else ''
            ws_det.cell(drow, 13).value = duty.completed_at.isoformat() if duty.completed_at else ''
            drow += 1
    return wb


def _get_phase_for_coordinator(prof, user, phase_id):
    if _is_hub_coordinator(prof):
        return get_object_or_404(
            SupervisionExamPhase,
            pk=phase_id,
            hub_coordinator=user,
            institute_semester_id=prof.institute_semester_id,
        )
    return get_object_or_404(SupervisionExamPhase, pk=phase_id, department=prof.department)


def _generate_faculty_portal_username(faculty: Faculty) -> tuple[str, str]:
    """Return (username, password) for a new faculty login; same pattern as admin credential generator."""
    base_username = (faculty.short_name or 'f').strip().lower()
    base_username = re.sub(r'[^\w]', '', base_username)[:30] or 'f'
    username = base_username
    if User.objects.filter(username=username).exists():
        username = f'{base_username}{faculty.id}'
    while User.objects.filter(username=username).exists():
        username = f'{base_username}{faculty.id}_{random.randint(100, 999)}'
    password = str(random.randint(0, 9999)).zfill(4)
    return username, password


@login_required
def exam_section_dashboard(request):
    if not _exam_section_portal_access(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    coord_profiles = (
        DepartmentExamProfile.objects.filter(
            parent__isnull=True,
            user__role_profile__role='dept_exam_parent',
        )
        .select_related('user', 'department', 'institute_semester')
        .order_by(
            'user__username',
            '-institute_semester__sort_order',
            '-institute_semester_id',
        )
    )
    active_semesters_for_coordinators = InstituteSemester.objects.filter(
        faculty_portal_active=True,
    ).order_by('-sort_order', '-pk')
    all_department_choices = departments_for_exam_coordination_request(request)
    sa = is_super_admin(request)
    exam_operators = []
    if sa:
        exam_operators = (
            UserRole.objects.filter(role='exam_section')
            .select_related('user')
            .order_by('user__username')
        )

    return render(
        request,
        'core/exam_section/dashboard.html',
        {
            'coord_profiles': coord_profiles,
            'active_semesters_for_coordinators': active_semesters_for_coordinators,
            'all_department_choices': all_department_choices,
            'is_super_admin': sa,
            'exam_section_operators': exam_operators,
            'paper_checking_phases': checking_phases_institute_for_request(request),
            'paper_setting_phases': setting_phases_institute_for_request(request),
        },
    )


@login_required
def exam_section_working_semesters(request):
    """
    Exam section operators: choose which institute academic semesters are in scope
    for paper duties, DR exports, credit reports, and DR facilities (multi-select).
    """
    if not _exam_section_portal_access(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    all_semesters = list(InstituteSemester.objects.order_by('-sort_order', '-pk'))
    is_operator = is_exam_section_operator(request)
    if request.method == 'POST':
        if not is_operator:
            messages.error(request, 'Only exam section operator logins can change working semesters.')
            return redirect('core:exam_section_working_semesters')
        raw_ids = request.POST.getlist('semester_ids')
        valid = {s.pk for s in all_semesters}
        ids: list[int] = []
        for x in raw_ids:
            try:
                pk = int(x)
            except (TypeError, ValueError):
                continue
            if pk in valid and pk not in ids:
                ids.append(pk)
        if not ids:
            messages.error(request, 'Select at least one academic semester.')
            return redirect('core:exam_section_working_semesters')
        set_exam_section_working_semester_ids(request, ids)
        messages.success(
            request,
            'Working academic semesters updated. Paper checking, setting, supervision, DR, and reports now use this scope.',
        )
        return redirect('core:exam_section_working_semesters')
    selected_ids: set[int] = set()
    selected_rows: list[InstituteSemester] = []
    selected_semester_detail: list[dict] = []
    if is_operator:
        selected_ids = set(exam_section_working_semester_ids(request))
        selected_rows = [s for s in all_semesters if s.pk in selected_ids]
        selected_semester_detail = [
            {
                'semester': s,
                'department_count': Department.objects.filter(institute_semester=s).count(),
            }
            for s in selected_rows
        ]
    return render(
        request,
        'core/exam_section/working_semesters.html',
        {
            'all_semesters': all_semesters,
            'selected_ids': selected_ids,
            'selected_semester_detail': selected_semester_detail,
            'can_edit_working_semesters': is_operator,
        },
    )


@login_required
def exam_section_daily_dr_excel(request):
    """Exam section operators and institute super admin: daily DR workbook (exam.xlsx layout)."""
    if not _exam_section_portal_access(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    raw = (request.GET.get('date') or request.GET.get('date_from') or '').strip()
    raw_to = (request.GET.get('date_to') or '').strip()
    if not raw:
        messages.error(request, 'Select an exam date (From).')
        return redirect('core:exam_section_dashboard')
    try:
        sem_ids = exam_section_working_semester_ids(request) if is_exam_section_operator(request) else None
        dates = parse_dates_from_request(raw, raw_to, duty_phase_semester_ids=sem_ids)
        return build_exam_daily_dr_workbook(dates, duty_phase_semester_ids=sem_ids)
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect('core:exam_section_dashboard')
    except FileNotFoundError as exc:
        messages.error(request, str(exc))
        return redirect('core:exam_section_dashboard')
    except Exception as exc:
        messages.error(request, f'Could not build workbook: {exc}')
        return redirect('core:exam_section_dashboard')


@login_required
@require_http_methods(['POST'])
def exam_section_create_coordinator(request):
    if not _exam_section_portal_access(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    username = (request.POST.get('username') or '').strip()
    password = request.POST.get('password') or ''
    if not username:
        messages.error(request, 'Username is required.')
        return redirect('core:exam_section_dashboard')
    dept = None
    is_hub = True
    raw_dept = (request.POST.get('department_id') or '').strip()
    if raw_dept:
        try:
            did = int(raw_dept)
        except (TypeError, ValueError):
            messages.error(request, 'Invalid department.')
            return redirect('core:exam_section_dashboard')
        dept = departments_for_exam_coordination_request(request).filter(pk=did).first()
        if not dept:
            messages.error(request, 'Invalid department for the current academic semester.')
            return redirect('core:exam_section_dashboard')
        is_hub = False
    if is_hub:
        raw_sid = (request.POST.get('institute_semester_id') or '').strip()
        try:
            sid = int(raw_sid)
        except (TypeError, ValueError):
            messages.error(request, 'Choose an academic semester for this hub coordinator.')
            return redirect('core:exam_section_dashboard')
        sem_for_profile = InstituteSemester.objects.filter(
            pk=sid,
            faculty_portal_active=True,
        ).first()
        if not sem_for_profile:
            messages.error(
                request,
                'Invalid or inactive academic semester (super admin must turn Faculty portal active on for that period).',
            )
            return redirect('core:exam_section_dashboard')
    else:
        sem_for_profile = dept.institute_semester
        if not sem_for_profile.faculty_portal_active:
            messages.error(
                request,
                'That department’s academic semester is not faculty-portal active — pick another department or ask super admin to enable the period.',
            )
            return redirect('core:exam_section_dashboard')
    invited_by_user = None
    if not is_hub:
        raw_hub_u = (request.POST.get('reporting_hub_username') or '').strip()
        if raw_hub_u:
            cand = User.objects.filter(username__iexact=raw_hub_u).first()
            if cand and DepartmentExamProfile.objects.filter(
                user=cand,
                is_hub_coordinator=True,
                parent__isnull=True,
                institute_semester=sem_for_profile,
            ).exists():
                invited_by_user = cand
            elif cand:
                messages.warning(
                    request,
                    f'"{raw_hub_u}" is not a hub coordinator for {sem_for_profile.label}; sub-units may not see that hub supervision '
                    'until you set Reporting hub (edit coordinator or use a valid hub username).',
                )
            else:
                messages.warning(request, f'No user named "{raw_hub_u}"; reporting hub not linked.')
    existing = User.objects.filter(username=username).first()
    if existing:
        try:
            if existing.role_profile.role != 'dept_exam_parent':
                messages.error(
                    request,
                    'That username exists for another role. Pick a different username.',
                )
                return redirect('core:exam_section_dashboard')
        except Exception:
            messages.error(request, 'That username is not usable for a coordinator.')
            return redirect('core:exam_section_dashboard')
        if DepartmentExamProfile.objects.filter(
            user=existing,
            parent__isnull=True,
            institute_semester=sem_for_profile,
        ).exists():
            messages.error(
                request,
                f'This coordinator already has a profile for {sem_for_profile.label}.',
            )
            return redirect('core:exam_section_dashboard')
        with transaction.atomic():
            if password:
                existing.set_password(password)
                existing.save(update_fields=['password'])
            if dept:
                existing.role_profile.department = dept
                existing.role_profile.save(update_fields=['department'])
            DepartmentExamProfile.objects.create(
                user=existing,
                department=dept,
                parent=None,
                subunit_code='',
                is_hub_coordinator=is_hub,
                institute_semester=sem_for_profile,
                invited_by=invited_by_user,
            )
        if is_hub:
            msg = (
                f'Added hub context {sem_for_profile.label} for "{username}". They will pick this semester when more than one is linked.'
            )
        else:
            msg = f'Added coordinator context for {dept.name} ({sem_for_profile.label}) on "{username}".'
        messages.success(request, msg)
        return redirect('core:exam_section_dashboard')
    if not password:
        messages.error(request, 'Password is required for a new coordinator login.')
        return redirect('core:exam_section_dashboard')
    with transaction.atomic():
        user = User.objects.create_user(username=username, password=password)
        UserRole.objects.create(user=user, role='dept_exam_parent', department=dept)
        DepartmentExamProfile.objects.create(
            user=user,
            department=dept,
            parent=None,
            subunit_code='',
            is_hub_coordinator=is_hub,
            institute_semester=sem_for_profile,
            invited_by=invited_by_user,
        )
    if is_hub:
        msg = (
            f'Hub coordinator "{username}" created for {sem_for_profile.label}. They can manage sub-units and institute-wide '
            'supervision / paper workflows for that academic period.'
        )
    else:
        msg = f'Coordinator "{username}" created for {dept.name} ({sem_for_profile.label}).'
    messages.success(request, msg)
    return redirect('core:exam_section_dashboard')


@login_required
@require_http_methods(['POST'])
def exam_section_create_operator(request):
    """Institute super admin only: create login with role exam_section."""
    if not is_super_admin(request):
        messages.error(request, 'Only institute super admin can create exam section accounts.')
        return redirect('core:exam_section_dashboard')
    username = (request.POST.get('username') or '').strip()
    password = request.POST.get('password') or ''
    if not username or not password:
        messages.error(request, 'Username and password are required.')
        return redirect('core:exam_section_dashboard')
    if User.objects.filter(username=username).exists():
        messages.error(request, 'That username is already taken.')
        return redirect('core:exam_section_dashboard')
    with transaction.atomic():
        user = User.objects.create_user(username=username, password=password)
        UserRole.objects.create(user=user, role='exam_section', department=None)
    messages.success(request, f'Exam section operator "{username}" created.')
    return redirect('core:exam_section_dashboard')


@login_required
@require_http_methods(['POST'])
def exam_section_delete_coordinator(request):
    """Remove one parent coordinator profile (one academic-semester context) and its sub-units; same login may keep other semesters."""
    if not _exam_section_portal_access(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    try:
        pid = int(request.POST.get('profile_id', ''))
    except (TypeError, ValueError):
        messages.error(request, 'Invalid coordinator profile.')
        return redirect('core:exam_section_dashboard')
    prof = (
        DepartmentExamProfile.objects.filter(pk=pid, parent__isnull=True)
        .select_related('user', 'institute_semester')
        .first()
    )
    if not prof:
        messages.error(request, 'Coordinator profile not found.')
        return redirect('core:exam_section_dashboard')
    try:
        if prof.user.role_profile.role != 'dept_exam_parent':
            messages.error(request, 'Not a department exam coordinator.')
            return redirect('core:exam_section_dashboard')
    except Exception:
        messages.error(request, 'Not a department exam coordinator.')
        return redirect('core:exam_section_dashboard')
    hub_user = prof.user
    sem_label = prof.institute_semester.label
    semester_id = prof.institute_semester_id
    uname = hub_user.username
    with transaction.atomic():
        for ch in DepartmentExamProfile.objects.filter(parent=prof).select_related('user'):
            ch.user.delete()
        if prof.is_hub_coordinator:
            invited_delegates = list(
                DepartmentExamProfile.objects.filter(
                    invited_by=hub_user,
                    parent__isnull=True,
                    institute_semester_id=semester_id,
                ).select_related('user')
            )
            for inv_prof in invited_delegates:
                inv_u = inv_prof.user
                for ch in DepartmentExamProfile.objects.filter(parent=inv_prof).select_related('user'):
                    ch.user.delete()
                inv_prof.delete()
                if not DepartmentExamProfile.objects.filter(user=inv_u).exists():
                    inv_u.delete()
        prof.delete()
        if not DepartmentExamProfile.objects.filter(user=hub_user).exists():
            hub_user.delete()
    messages.success(
        request,
        f'Removed coordinator "{uname}" for {sem_label} (sub-units and linked delegates for that period).',
    )
    return redirect('core:exam_section_dashboard')


@login_required
@require_http_methods(['POST'])
def exam_section_delete_operator(request):
    """Remove an exam_section operator login."""
    if not is_super_admin(request):
        messages.error(request, 'Only institute super admin can remove exam section accounts.')
        return redirect('core:exam_section_dashboard')
    try:
        uid = int(request.POST.get('user_id', ''))
    except (TypeError, ValueError):
        messages.error(request, 'Invalid user.')
        return redirect('core:exam_section_dashboard')
    if uid == request.user.id:
        messages.error(request, 'You cannot remove your own account from this screen.')
        return redirect('core:exam_section_dashboard')
    role = UserRole.objects.filter(user_id=uid, role='exam_section').select_related('user').first()
    if not role:
        messages.error(request, 'Not an exam section operator.')
        return redirect('core:exam_section_dashboard')
    uname = role.user.username
    role.user.delete()
    messages.success(request, f'Exam section operator "{uname}" was removed.')
    return redirect('core:exam_section_dashboard')


@login_required
def exam_section_edit_coordinator(request, profile_id):
    if not _exam_section_portal_access(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    prof = (
        DepartmentExamProfile.objects.filter(pk=profile_id, parent__isnull=True)
        .select_related('user', 'department', 'institute_semester', 'invited_by', 'user__role_profile')
        .first()
    )
    if not prof:
        messages.error(request, 'Coordinator not found.')
        return redirect('core:exam_section_dashboard')
    try:
        if prof.user.role_profile.role != 'dept_exam_parent':
            messages.error(request, 'Coordinator not found.')
            return redirect('core:exam_section_dashboard')
    except Exception:
        messages.error(request, 'Coordinator not found.')
        return redirect('core:exam_section_dashboard')
    role = prof.user.role_profile
    u = role.user
    if request.method == 'POST' and request.POST.get('add_coordinator_semester'):
        raw_dept = (request.POST.get('department_id') or '').strip()
        dept = None
        is_hub = True
        sem_for_profile = None
        if raw_dept:
            try:
                did = int(raw_dept)
            except (TypeError, ValueError):
                messages.error(request, 'Invalid department.')
                return redirect('core:exam_section_edit_coordinator', profile_id=profile_id)
            dept = departments_for_exam_coordination_request(request).filter(pk=did).first()
            if not dept:
                messages.error(request, 'Invalid department.')
                return redirect('core:exam_section_edit_coordinator', profile_id=profile_id)
            sem_for_profile = dept.institute_semester
            is_hub = False
            if not sem_for_profile.faculty_portal_active:
                messages.error(request, 'That department’s academic semester is not faculty-portal active.')
                return redirect('core:exam_section_edit_coordinator', profile_id=profile_id)
        else:
            raw_sid = (request.POST.get('institute_semester_id') or '').strip()
            try:
                sid = int(raw_sid)
            except (TypeError, ValueError):
                messages.error(request, 'Choose an academic semester for hub mode.')
                return redirect('core:exam_section_edit_coordinator', profile_id=profile_id)
            sem_for_profile = InstituteSemester.objects.filter(
                pk=sid,
                faculty_portal_active=True,
            ).first()
            if not sem_for_profile:
                messages.error(request, 'Invalid or inactive academic semester.')
                return redirect('core:exam_section_edit_coordinator', profile_id=profile_id)
        if DepartmentExamProfile.objects.filter(
            user=u,
            parent__isnull=True,
            institute_semester=sem_for_profile,
        ).exists():
            messages.error(
                request,
                f'This login already has a profile for {sem_for_profile.label}.',
            )
            return redirect('core:exam_section_edit_coordinator', profile_id=profile_id)
        invited_by_user = None
        if not is_hub:
            raw_hub_u = (request.POST.get('reporting_hub_username') or '').strip()
            if raw_hub_u:
                cand = User.objects.filter(username__iexact=raw_hub_u).first()
                if cand and DepartmentExamProfile.objects.filter(
                    user=cand,
                    is_hub_coordinator=True,
                    parent__isnull=True,
                    institute_semester=sem_for_profile,
                ).exists():
                    invited_by_user = cand
                elif cand:
                    messages.warning(
                        request,
                        f'"{raw_hub_u}" is not a hub coordinator for {sem_for_profile.label}; reporting hub not linked.',
                    )
                else:
                    messages.warning(request, f'No user named "{raw_hub_u}"; reporting hub not linked.')
        with transaction.atomic():
            new_prof = DepartmentExamProfile.objects.create(
                user=u,
                department=dept,
                parent=None,
                subunit_code='',
                is_hub_coordinator=is_hub,
                institute_semester=sem_for_profile,
                invited_by=invited_by_user,
            )
            if dept:
                role.department = dept
                role.save(update_fields=['department'])
        messages.success(
            request,
            f'Added coordinator context for {sem_for_profile.label}'
            + (f' ({dept.name})' if dept else ' (hub)'),
        )
        return redirect('core:exam_section_edit_coordinator', profile_id=new_prof.pk)
    if request.method == 'POST':
        new_username = (request.POST.get('username') or '').strip()
        new_password = request.POST.get('password') or ''
        username_changed = password_changed = False
        if new_username and new_username != u.username:
            if User.objects.filter(username=new_username).exclude(pk=u.pk).exists():
                messages.error(request, 'That username is already taken.')
                return redirect('core:exam_section_edit_coordinator', profile_id=profile_id)
            u.username = new_username
            u.save(update_fields=['username'])
            username_changed = True
        if new_password:
            u.set_password(new_password)
            u.save(update_fields=['password'])
            password_changed = True
        hub_changed = False
        if prof and not prof.is_hub_coordinator and prof.department_id:
            raw_hub = (request.POST.get('reporting_hub_username') or '').strip()
            new_invited = None
            if raw_hub:
                cand = User.objects.filter(username__iexact=raw_hub).first()
                if cand and DepartmentExamProfile.objects.filter(
                    user=cand,
                    is_hub_coordinator=True,
                    parent__isnull=True,
                    institute_semester_id=prof.institute_semester_id,
                ).exists():
                    new_invited = cand
                elif cand:
                    messages.warning(
                        request,
                        f'"{raw_hub}" is not a hub coordinator for this academic semester; reporting hub not updated.',
                    )
                else:
                    messages.warning(request, f'No user named "{raw_hub}"; reporting hub not updated.')
            if raw_hub == '' and prof.invited_by_id is not None:
                prof.invited_by = None
                prof.save(update_fields=['invited_by'])
                hub_changed = True
            elif new_invited is not None and prof.invited_by_id != new_invited.id:
                prof.invited_by = new_invited
                prof.save(update_fields=['invited_by'])
                hub_changed = True
        account_changed = username_changed or password_changed
        if hub_changed and not account_changed:
            messages.success(request, 'Reporting hub link updated.')
        elif hub_changed and account_changed:
            messages.success(request, 'Coordinator updated (including reporting hub).')
        elif account_changed:
            messages.success(request, 'Coordinator account updated.')
        else:
            messages.info(request, 'No changes submitted.')
        return redirect('core:exam_section_dashboard')
    all_parent_profiles = list(
        DepartmentExamProfile.objects.filter(user=u, parent__isnull=True)
        .select_related('institute_semester', 'department')
        .order_by('-institute_semester__sort_order', 'institute_semester_id', 'pk')
    )
    linked_sem_ids = {p.institute_semester_id for p in all_parent_profiles}
    semesters_to_add = list(
        InstituteSemester.objects.filter(faculty_portal_active=True)
        .exclude(pk__in=linked_sem_ids)
        .order_by('-sort_order', '-pk')
    )
    dept_sem_ids = [s.pk for s in semesters_to_add]
    department_choices_for_add = (
        departments_for_exam_coordination_request(request)
        .filter(institute_semester_id__in=dept_sem_ids)
        .select_related('institute_semester')
        .order_by('-institute_semester__sort_order', 'institute_semester_id', 'name')
    )
    return render(
        request,
        'core/exam_section/edit_coordinator.html',
        {
            'role_obj': role,
            'coord_profile': prof,
            'all_parent_profiles': all_parent_profiles,
            'semesters_to_add': semesters_to_add,
            'department_choices_for_add': department_choices_for_add,
        },
    )


@login_required
def exam_section_edit_operator(request, user_id):
    if not is_super_admin(request):
        messages.error(request, 'Only institute super admin can edit exam section operators.')
        return redirect('core:exam_section_dashboard')
    role = UserRole.objects.filter(user_id=user_id, role='exam_section').select_related('user').first()
    if not role:
        messages.error(request, 'Operator not found.')
        return redirect('core:exam_section_dashboard')
    if request.method == 'POST':
        new_username = (request.POST.get('username') or '').strip()
        new_password = request.POST.get('password') or ''
        u = role.user
        if new_username and new_username != u.username:
            if User.objects.filter(username=new_username).exclude(pk=u.pk).exists():
                messages.error(request, 'That username is already taken.')
                return redirect('core:exam_section_edit_operator', user_id=user_id)
            u.username = new_username
            u.save(update_fields=['username'])
        if new_password:
            u.set_password(new_password)
            u.save(update_fields=['password'])
        if new_username or new_password:
            messages.success(request, 'Operator account updated.')
        else:
            messages.info(request, 'No changes submitted.')
        return redirect('core:exam_section_dashboard')
    return render(request, 'core/exam_section/edit_operator.html', {'role_obj': role})


@login_required
def dept_exam_child_edit(request, child_profile_id):
    if not _dept_parent_only(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    parent_prof = _parent_profile(request)
    if not parent_prof or not parent_prof.department_id:
        return redirect('core:dept_exam_link_department')
    child_prof = get_object_or_404(
        DepartmentExamProfile.objects.select_related('user'),
        pk=child_profile_id,
        parent=parent_prof,
    )
    all_departments = departments_for_exam_coordination_request(request)
    if request.method == 'POST':
        new_username = (request.POST.get('username') or '').strip()
        new_password = request.POST.get('password') or ''
        subunit = (request.POST.get('subunit_code') or '').strip().upper().replace(' ', '_')
        try:
            chosen_dept_id = int(request.POST.get('department_id', ''))
        except (TypeError, ValueError):
            chosen_dept_id = None
        if not subunit or not chosen_dept_id:
            messages.error(request, 'Sub-unit code and department are required.')
            return redirect('core:dept_exam_child_edit', child_profile_id=child_prof.pk)
        if not _is_hub_coordinator(parent_prof):
            if chosen_dept_id != parent_prof.department_id:
                messages.error(request, 'Department must match your linked attendance department.')
                return redirect('core:dept_exam_child_edit', child_profile_id=child_prof.pk)
        if DepartmentExamProfile.objects.filter(parent=parent_prof, subunit_code__iexact=subunit).exclude(pk=child_prof.pk).exists():
            messages.error(request, f'Sub-unit code "{subunit}" is already in use.')
            return redirect('core:dept_exam_child_edit', child_profile_id=child_prof.pk)
        u = child_prof.user
        if new_username and new_username != u.username:
            if User.objects.filter(username=new_username).exclude(pk=u.pk).exists():
                messages.error(request, 'That username is already taken.')
                return redirect('core:dept_exam_child_edit', child_profile_id=child_prof.pk)
            u.username = new_username
            u.save(update_fields=['username'])
        if new_password:
            u.set_password(new_password)
            u.save(update_fields=['password'])
        child_prof.subunit_code = subunit
        if _is_hub_coordinator(parent_prof):
            ch_dept = departments_for_exam_coordination_request(request).filter(pk=chosen_dept_id).first()
        else:
            ch_dept = parent_prof.department
        if not ch_dept:
            messages.error(request, 'Invalid department.')
            return redirect('core:dept_exam_child_edit', child_profile_id=child_prof.pk)
        child_prof.department = ch_dept
        child_prof.institute_semester = ch_dept.institute_semester
        child_prof.save(update_fields=['subunit_code', 'department', 'institute_semester'])
        rp = UserRole.objects.filter(user=u, role='dept_exam_child').first()
        if rp:
            rp.department_id = chosen_dept_id
            rp.save(update_fields=['department'])
        messages.success(request, 'Sub-unit account updated.')
        return redirect('core:dept_exam_dashboard')
    return render(
        request,
        'core/exam_dept/child_edit.html',
        {'parent_profile': parent_prof, 'child': child_prof, 'all_departments': all_departments},
    )


@login_required
@require_http_methods(['POST'])
def dept_exam_child_delete(request, child_profile_id):
    if not _dept_parent_only(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    parent_prof = _parent_profile(request)
    if not parent_prof:
        return redirect('accounts:role_redirect')
    child_prof = get_object_or_404(DepartmentExamProfile, pk=child_profile_id, parent=parent_prof)
    uname = child_prof.user.username
    child_prof.user.delete()
    messages.success(request, f'Sub-unit login "{uname}" removed.')
    return redirect('core:dept_exam_dashboard')


@login_required
@require_http_methods(['POST'])
def dept_exam_phase_clear_duties(request, phase_id):
    """Remove all imported supervision rows for this phase (Excel data only; phase remains)."""
    if not _dept_parent_only(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    prof = _parent_profile(request)
    if not prof:
        return redirect('accounts:role_redirect')
    rds = _parent_must_have_department(request, prof)
    if rds:
        return rds
    phase = _get_phase_for_coordinator(prof, request.user, phase_id)
    clear_staging(request, 'supervision', phase.id)
    n, _ = phase.duties.all().delete()
    messages.success(request, f'Removed {n} imported duty row(s) for phase {phase.name}. The phase is still available for a new upload.')
    return redirect('core:dept_exam_phase_detail', phase_id=phase.id)


def _parent_must_have_department(request, prof):
    """Single-dept coordinators must link a department; hub coordinators (examsy) skip this."""
    if prof and prof.department_id is None and not _is_hub_coordinator(prof):
        return redirect('core:dept_exam_link_department')
    return None


@login_required
def dept_exam_link_department(request):
    """First-time (or pending) parent coordinator: attach an attendance department from the master list."""
    if not _dept_parent_only(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    if coordinator_must_select_exam_context(request):
        return redirect('core:dept_exam_select_context')
    prof = _parent_profile(request)
    if not prof:
        messages.error(request, 'Your account has no department exam profile. Contact exam section.')
        return redirect('accounts:role_redirect')
    if prof.department_id:
        return redirect('core:dept_exam_dashboard')
    if prof.is_hub_coordinator:
        messages.info(request, 'You use hub mode: add department logins from your supervision dashboard instead of linking one department here.')
        return redirect('core:dept_exam_dashboard')
    departments = departments_for_exam_coordination_request(request)
    if request.method == 'POST':
        try:
            did = int(request.POST.get('department_id', ''))
        except (TypeError, ValueError):
            messages.error(request, 'Choose a valid department.')
            return redirect('core:dept_exam_link_department')
        dept = departments.filter(pk=did).first()
        if not dept:
            messages.error(request, 'Invalid department for the current academic semester.')
            return redirect('core:dept_exam_link_department')
        with transaction.atomic():
            prof.department = dept
            prof.save(update_fields=['department'])
            rp = request.user.role_profile
            rp.department = dept
            rp.save(update_fields=['department'])
        messages.success(request, f'Department set to {dept.name}. You can add supervision phases and sub-unit logins.')
        return redirect('core:dept_exam_dashboard')
    return render(
        request,
        'core/exam_dept/link_department.html',
        {'profile': prof, 'departments': departments},
    )


@login_required
def dept_exam_select_context(request):
    """dept_exam_parent with multiple academic-semester profiles: pick working context."""
    if not _dept_parent_only(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    profiles = list(
        DepartmentExamProfile.objects.filter(user=request.user, parent__isnull=True)
        .select_related('institute_semester', 'department')
        .order_by('-institute_semester__sort_order', 'institute_semester_id', 'pk')
    )
    if len(profiles) <= 1:
        return redirect('core:dept_exam_dashboard')
    if request.method == 'POST':
        try:
            pid = int(request.POST.get('profile_id', ''))
        except (TypeError, ValueError):
            messages.error(request, 'Invalid selection.')
            return redirect('core:dept_exam_select_context')
        chosen = next((p for p in profiles if p.pk == pid), None)
        if not chosen:
            messages.error(request, 'That context is not available for your account.')
            return redirect('core:dept_exam_select_context')
        request.session[SESSION_KEY_COORD_PARENT_PROFILE_ID] = chosen.pk
        messages.success(
            request,
            f'Working context: {chosen.institute_semester.label}'
            + (f' ({chosen.department.name})' if chosen.department_id else ' — hub'),
        )
        nxt = (request.POST.get('next') or '').strip()
        if nxt.startswith('/') and not nxt.startswith('//'):
            return redirect(nxt)
        return redirect('core:dept_exam_dashboard')
    return render(
        request,
        'core/exam_dept/select_context.html',
        {'profiles': profiles},
    )


@login_required
def dept_exam_child_select_context(request):
    """dept_exam_child with profiles in more than one academic semester: pick working context."""
    if not _dept_child_only(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    profiles = list(
        DepartmentExamProfile.objects.filter(user=request.user, parent__isnull=False)
        .select_related('institute_semester', 'department', 'parent')
        .order_by('-institute_semester__sort_order', 'institute_semester_id', 'pk')
    )
    if len(profiles) <= 1:
        return redirect('core:dept_exam_dashboard')
    if request.method == 'POST':
        try:
            pid = int(request.POST.get('profile_id', ''))
        except (TypeError, ValueError):
            messages.error(request, 'Invalid selection.')
            return redirect('core:dept_exam_child_select_context')
        chosen = next((p for p in profiles if p.pk == pid), None)
        if not chosen:
            messages.error(request, 'That context is not available for your account.')
            return redirect('core:dept_exam_child_select_context')
        request.session[SESSION_KEY_COORD_CHILD_PROFILE_ID] = chosen.pk
        messages.success(
            request,
            f'Working context: {chosen.institute_semester.label} — {chosen.subunit_code or "sub-unit"}',
        )
        nxt = (request.POST.get('next') or '').strip()
        if nxt.startswith('/') and not nxt.startswith('//'):
            return redirect(nxt)
        return redirect('core:dept_exam_dashboard')
    return render(
        request,
        'core/exam_dept/select_child_context.html',
        {'profiles': profiles},
    )


@login_required
def dept_exam_dashboard(request):
    if _dept_parent_only(request):
        if coordinator_must_select_exam_context(request):
            return redirect('core:dept_exam_select_context')
        prof = _parent_profile(request)
        if not prof:
            messages.error(request, 'Your account has no department exam profile. Contact exam section.')
            return redirect('accounts:role_redirect')
        rds = _parent_must_have_department(request, prof)
        if rds:
            return rds
        if _is_hub_coordinator(prof):
            phases = SupervisionExamPhase.objects.filter(
                hub_coordinator=request.user,
                institute_semester_id=prof.institute_semester_id,
            ).order_by('name')
            children = DepartmentExamProfile.objects.filter(parent=prof).select_related('user').order_by('subunit_code')
            delegates = (
                DepartmentExamProfile.objects.filter(
                    invited_by=request.user,
                    parent__isnull=True,
                    institute_semester_id=prof.institute_semester_id,
                )
                .select_related('user', 'department')
                .order_by('department__name', 'user__username')
            )
            ps_pending = list(pending_paper_setting_completion_requests_for_parent(prof))
            ps_pending_enriched = [
                {'r': r, 'credit': credit_for_paper_setting_request(r)} for r in ps_pending
            ]
            ps_hist = history_paper_setting_completion_requests_for_parent(prof)
            ps_history_enriched = []
            for r in ps_hist:
                cr = (
                    credit_for_paper_setting_request(r)
                    if r.status == PaperSettingCompletionRequest.APPROVED
                    else None
                )
                ps_history_enriched.append({'r': r, 'credit': cr})
            return render(
                request,
                'core/exam_dept/hub_dashboard.html',
                {
                    'profile': prof,
                    'phases': phases,
                    'children': children,
                    'delegates': delegates,
                    'all_departments': departments_for_exam_coordination_request(request),
                    'paper_checking_phases': checking_phases_hub_user(
                        request.user, prof.institute_semester_id
                    ),
                    'paper_setting_phases': setting_phases_hub_user(
                        request.user, prof.institute_semester_id
                    ),
                    'coord_parent_profiles': list(
                        DepartmentExamProfile.objects.filter(user=request.user, parent__isnull=True)
                        .select_related('institute_semester', 'department')
                        .order_by('-institute_semester__sort_order', 'institute_semester_id', 'pk')
                    ),
                    'paper_credit_by_faculty': [],
                    'paper_setting_pending_enriched': ps_pending_enriched,
                    'paper_setting_history_enriched': ps_history_enriched,
                    'paper_setting_credit_rows': [],
                },
            )
        phases = SupervisionExamPhase.objects.filter(department=prof.department).order_by('name')
        children = DepartmentExamProfile.objects.filter(parent=prof).select_related('user').order_by('subunit_code')
        pcredit = (
            department_approved_paper_credit_rows(prof.department_id) if prof.department_id else []
        )
        ps_pending = list(pending_paper_setting_completion_requests_for_parent(prof))
        ps_pending_enriched = [
            {'r': r, 'credit': credit_for_paper_setting_request(r)} for r in ps_pending
        ]
        ps_hist = history_paper_setting_completion_requests_for_parent(prof)
        ps_history_enriched = []
        for r in ps_hist:
            cr = (
                credit_for_paper_setting_request(r)
                if r.status == PaperSettingCompletionRequest.APPROVED
                else None
            )
            ps_history_enriched.append({'r': r, 'credit': cr})
        pset_credit = (
            department_paper_setting_credit_rows(prof.department_id) if prof.department_id else []
        )
        return render(
            request,
            'core/exam_dept/parent_dashboard.html',
            {
                'profile': prof,
                'phases': phases,
                'children': children,
                'all_departments': departments_for_exam_coordination_request(request),
                'coord_parent_profiles': list(
                    DepartmentExamProfile.objects.filter(user=request.user, parent__isnull=True)
                    .select_related('institute_semester', 'department')
                    .order_by('-institute_semester__sort_order', 'institute_semester_id', 'pk')
                ),
                'paper_checking_phases': checking_phases_department(prof.department),
                'paper_setting_phases': setting_phases_department(prof.department),
                'paper_credit_by_faculty': pcredit,
                'paper_setting_pending_enriched': ps_pending_enriched,
                'paper_setting_history_enriched': ps_history_enriched,
                'paper_setting_credit_rows': pset_credit,
            },
        )
    if _dept_child_only(request):
        if child_must_select_exam_context(request):
            return redirect('core:dept_exam_child_select_context')
        prof = _child_profile(request)
        if not prof:
            messages.error(request, 'Your account has no department exam profile.')
            return redirect('accounts:role_redirect')
        code = (prof.subunit_code or '').strip().upper()
        duties_base = SupervisionDuty.objects.filter(subunit_supervision_duty_filter_q(prof))
        sup_phases_list = list(phases_for_subunit_prof(prof))
        allowed_sup_ids = {p.pk for p in sup_phases_list}
        sel_sup = _parse_optional_int_param(request, 'sup_phase_id')
        if sel_sup and sel_sup not in allowed_sup_ids:
            sel_sup = None
        duties_qs = duties_base.filter(phase_id=sel_sup) if sel_sup else duties_base
        duties = list(
            duties_qs.select_related('faculty', 'original_faculty', 'phase').order_by(
                'supervision_date', 'faculty__full_name'
            )
        )
        by_faculty = {}
        for d in duties:
            if not d.faculty_id:
                continue
            by_faculty.setdefault(
                d.faculty_id,
                {'faculty': d.faculty, 'n': 0, 'completed': 0, 'open': 0, 'proxy_n': 0},
            )
            ent = by_faculty[d.faculty_id]
            ent['n'] += 1
            if d.completion_status == SupervisionDuty.COMPLETED:
                ent['completed'] += 1
            else:
                ent['open'] += 1
            if d.is_proxy:
                ent['proxy_n'] += 1
        faculty_list = sorted(
            (
                {
                    'faculty': v['faculty'],
                    'duty_count': v['n'],
                    'completed': v['completed'],
                    'open': v['open'],
                    'proxy_n': v['proxy_n'],
                }
                for v in by_faculty.values()
            ),
            key=lambda x: x['faculty'].full_name,
        )
        global_faculty = Faculty.objects.select_related('department').order_by(
            'department__name', 'full_name'
        )
        pcredit = (
            department_approved_paper_credit_rows(prof.department_id) if prof.department_id else []
        )
        pc_base = paper_checking_duties_for_child_prof(prof)
        pc_phase_ids = set(pc_base.values_list('phase_id', flat=True))
        pc_phases_list = PaperCheckingPhase.objects.filter(pk__in=pc_phase_ids).order_by('name')
        sel_pc = _parse_optional_int_param(request, 'pc_phase_id')
        if sel_pc and sel_pc not in pc_phase_ids:
            sel_pc = None
        paper_checking_duties = list(
            (pc_base.filter(phase_id=sel_pc) if sel_pc else pc_base).select_related('phase', 'faculty')
        )

        ps_base = paper_setting_duties_for_child_prof(prof)
        ps_phase_ids = set(ps_base.values_list('phase_id', flat=True))
        ps_phases_list = PaperSettingPhase.objects.filter(pk__in=ps_phase_ids).order_by('name')
        sel_ps = _parse_optional_int_param(request, 'ps_phase_id')
        if sel_ps and sel_ps not in ps_phase_ids:
            sel_ps = None
        paper_setting_duties = list(
            (ps_base.filter(phase_id=sel_ps) if sel_ps else ps_base).select_related('phase', 'faculty')
        )

        return render(
            request,
            'core/exam_dept/child_dashboard.html',
            {
                'profile': prof,
                'coord_child_profiles': list(coordinator_child_profiles_qs(request)),
                'faculty_list': faculty_list,
                'duties': duties,
                'subunit': code,
                'global_faculty': global_faculty,
                'paper_checking_duties': paper_checking_duties,
                'paper_setting_duties': paper_setting_duties,
                'paper_credit_by_faculty': pcredit,
                'sup_phases': sup_phases_list,
                'selected_sup_phase_id': sel_sup,
                'pc_phases': pc_phases_list,
                'selected_pc_phase_id': sel_pc,
                'ps_phases': ps_phases_list,
                'selected_ps_phase_id': sel_ps,
                'show_exam_bulk_complete': getattr(settings, 'EXAM_PORTAL_BULK_COMPLETE', False),
            },
        )
    messages.error(request, 'Access denied.')
    return redirect('accounts:role_redirect')


@login_required
@require_http_methods(['POST'])
def dept_exam_hub_invite_coordinator(request):
    """Deprecated: department-scoped coordinators are created from Exam section only."""
    messages.error(
        request,
        'Department coordinator accounts can only be created from the Exam section portal '
        '(New department coordinator → select the attendance department).',
    )
    if _exam_section_portal_access(request):
        return redirect('core:exam_section_dashboard')
    return redirect('core:dept_exam_dashboard')


@login_required
@require_http_methods(['POST'])
def dept_exam_phase_add(request):
    if not _dept_parent_only(request):
        messages.error(request, 'Only department coordinators can create supervision phases.')
        return redirect('accounts:role_redirect')
    prof = _parent_profile(request)
    if not prof:
        return redirect('accounts:role_redirect')
    rds = _parent_must_have_department(request, prof)
    if rds:
        return rds
    name = (request.POST.get('name') or '').strip()
    if not name:
        messages.error(request, 'Phase name is required (e.g. T1, SEE).')
        return redirect('core:dept_exam_dashboard')
    if _is_hub_coordinator(prof):
        obj, created = SupervisionExamPhase.objects.get_or_create(
            hub_coordinator=request.user,
            name=name,
            institute_semester=prof.institute_semester,
            department=None,
            defaults={'created_by': request.user},
        )
    else:
        obj, created = SupervisionExamPhase.objects.get_or_create(
            department=prof.department,
            name=name,
            defaults={
                'created_by': request.user,
                'hub_coordinator': None,
                'institute_semester': prof.department.institute_semester,
            },
        )
    if created:
        messages.success(request, f'Phase "{name}" created.')
    else:
        messages.info(request, f'Phase "{name}" already exists.')
    return redirect('core:dept_exam_dashboard')


@login_required
@require_http_methods(['POST'])
def dept_exam_phase_rename(request, phase_id):
    if not _dept_parent_only(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    prof = _parent_profile(request)
    if not prof:
        return redirect('accounts:role_redirect')
    rds = _parent_must_have_department(request, prof)
    if rds:
        return rds
    phase = _get_phase_for_coordinator(prof, request.user, phase_id)
    name = (request.POST.get('name') or '').strip()
    if not name:
        messages.error(request, 'Phase name is required.')
        return redirect('core:dept_exam_dashboard')
    phase.name = name
    try:
        phase.save(update_fields=['name'])
    except IntegrityError:
        messages.error(request, 'That phase name already exists in your scope.')
        return redirect('core:dept_exam_dashboard')
    messages.success(request, f'Phase renamed to "{name}".')
    return redirect('core:dept_exam_dashboard')


@login_required
@require_http_methods(['POST'])
def dept_exam_phase_delete(request, phase_id):
    if not _dept_parent_only(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    prof = _parent_profile(request)
    if not prof:
        return redirect('accounts:role_redirect')
    rds = _parent_must_have_department(request, prof)
    if rds:
        return rds
    phase = _get_phase_for_coordinator(prof, request.user, phase_id)
    if phase.duties.exists():
        messages.error(
            request,
            'This phase still has supervision rows. Clear duties on the phase page first.',
        )
        return redirect('core:dept_exam_dashboard')
    label = phase.name
    phase.delete()
    messages.success(request, f'Supervision phase "{label}" deleted.')
    return redirect('core:dept_exam_dashboard')


@login_required
@require_http_methods(['POST'])
def dept_exam_child_create(request):
    if not _dept_parent_only(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    prof = _parent_profile(request)
    if not prof:
        return redirect('accounts:role_redirect')
    rds = _parent_must_have_department(request, prof)
    if rds:
        return rds
    username = (request.POST.get('username') or '').strip()
    password = request.POST.get('password') or ''
    subunit = (request.POST.get('subunit_code') or '').strip().upper().replace(' ', '_')
    try:
        chosen_dept_id = int(request.POST.get('department_id', ''))
    except (TypeError, ValueError):
        chosen_dept_id = None
    if not username or not password or not subunit or not chosen_dept_id:
        messages.error(request, 'Username, password, sub-unit code, and department are required.')
        return redirect('core:dept_exam_dashboard')
    if _is_hub_coordinator(prof):
        dept = departments_for_exam_coordination_request(request).filter(pk=chosen_dept_id).first()
        if not dept:
            messages.error(request, 'Invalid department for the current academic semester.')
            return redirect('core:dept_exam_dashboard')
        if dept.institute_semester_id != prof.institute_semester_id:
            messages.error(
                request,
                'Choose a department that belongs to the same academic semester as this hub context.',
            )
            return redirect('core:dept_exam_dashboard')
    else:
        if chosen_dept_id != prof.department_id:
            messages.error(request, 'Sub-unit department must match your linked attendance department.')
            return redirect('core:dept_exam_dashboard')
        dept = prof.department
    if DepartmentExamProfile.objects.filter(parent=prof, subunit_code__iexact=subunit).exists():
        messages.error(request, f'Sub-unit "{subunit}" already exists under your account for this period.')
        return redirect('core:dept_exam_dashboard')
    child_sem_id = prof.institute_semester_id
    existing_child = User.objects.filter(username=username).first()
    if existing_child:
        try:
            if existing_child.role_profile.role != 'dept_exam_child':
                messages.error(request, 'That username exists for another role.')
                return redirect('core:dept_exam_dashboard')
        except Exception:
            messages.error(request, 'That username is not usable as a sub-unit login.')
            return redirect('core:dept_exam_dashboard')
        if DepartmentExamProfile.objects.filter(user=existing_child, parent=prof).exists():
            messages.error(request, f'This user is already a sub-unit under you ({subunit}).')
            return redirect('core:dept_exam_dashboard')
        with transaction.atomic():
            DepartmentExamProfile.objects.create(
                user=existing_child,
                department=dept,
                parent=prof,
                subunit_code=subunit,
                institute_semester_id=child_sem_id,
            )
            if password:
                existing_child.set_password(password)
                existing_child.save(update_fields=['password'])
            existing_child.role_profile.department = dept
            existing_child.role_profile.save(update_fields=['department'])
        messages.success(
            request,
            f'Linked existing login "{username}" ({subunit}) to this hub for {prof.institute_semester.label}.',
        )
        return redirect('core:dept_exam_dashboard')
    with transaction.atomic():
        user = User.objects.create_user(username=username, password=password)
        UserRole.objects.create(
            user=user,
            role='dept_exam_child',
            department=dept,
        )
        DepartmentExamProfile.objects.create(
            user=user,
            department=dept,
            parent=prof,
            subunit_code=subunit,
            institute_semester_id=child_sem_id,
        )
    messages.success(request, f'Sub-unit login "{username}" ({subunit}) created.')
    return redirect('core:dept_exam_dashboard')


@login_required
def dept_exam_phase_detail(request, phase_id):
    if not _dept_parent_only(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    prof = _parent_profile(request)
    if not prof:
        return redirect('accounts:role_redirect')
    rds = _parent_must_have_department(request, prof)
    if rds:
        return rds
    phase = _get_phase_for_coordinator(prof, request.user, phase_id)
    staging_blob = supervision_stage_get(request, phase.id)
    staging_preview = None
    if staging_blob:
        srows = supervision_stage_deserialize_rows(staging_blob)
        staging_preview = []
        for row in srows:
            if phase.hub_coordinator_id:
                fac = match_faculty_global(row['faculty_name'], row['faculty_initial'])
            else:
                fac = match_faculty_for_department(
                    prof.department,
                    row['faculty_name'],
                    row['faculty_initial'],
                )
            staging_preview.append({**row, 'faculty': fac})
    duties = phase.duties.select_related('faculty', 'original_faculty').order_by(
        'supervision_date', 'division_code', 'faculty__full_name'
    )
    unmatched = duties.filter(faculty__isnull=True)
    groups = {}
    for d in unmatched:
        key = (d.faculty_name_raw or '', d.faculty_short_raw or '')
        if key not in groups:
            groups[key] = {'name_raw': key[0], 'short_raw': key[1], 'count': 0}
        groups[key]['count'] += 1
    unmatched_groups = sorted(groups.values(), key=lambda g: (g['name_raw'] or '').upper())
    return render(
        request,
        'core/exam_dept/phase_detail.html',
        {
            'profile': prof,
            'phase': phase,
            'duties': duties,
            'has_unmatched_duties': unmatched.exists(),
            'unmatched_groups': unmatched_groups,
            'is_hub_phase': bool(phase.hub_coordinator_id),
            'all_departments': departments_for_exam_coordination_request(request),
            'staging_preview': staging_preview,
            'staging_unmatched': staging_blob.get('n_unmatched') if staging_blob else None,
        },
    )


@login_required
@require_http_methods(['POST'])
def dept_exam_phase_create_faculty_assign(request, phase_id):
    """Create Faculty + optional portal User for sheet names that did not match; link all matching duty rows in this phase."""
    if not _dept_parent_only(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    prof = _parent_profile(request)
    if not prof:
        return redirect('accounts:role_redirect')
    rds = _parent_must_have_department(request, prof)
    if rds:
        return rds
    phase = _get_phase_for_coordinator(prof, request.user, phase_id)

    full_name = (request.POST.get('full_name') or '').strip()
    short_name = (request.POST.get('short_name') or '').strip()
    email = (request.POST.get('email') or '').strip()
    name_raw = (request.POST.get('match_faculty_name_raw') or '').strip()
    short_raw = (request.POST.get('match_faculty_short_raw') or '').strip()

    if not full_name or not short_name:
        messages.error(request, 'Full name and short name (initial) are required.')
        return redirect('core:dept_exam_phase_detail', phase_id=phase.id)

    fdept = prof.department
    if phase.hub_coordinator_id:
        try:
            fdept_id = int(request.POST.get('new_faculty_department_id', ''))
        except (TypeError, ValueError):
            fdept_id = None
        fdept = (
            departments_for_exam_coordination_request(request).filter(pk=fdept_id).first()
            if fdept_id
            else None
        )
        if not fdept:
            messages.error(request, 'Choose which attendance department this faculty belongs to.')
            return redirect('core:dept_exam_phase_detail', phase_id=phase.id)

    to_link = phase.duties.filter(faculty__isnull=True, faculty_name_raw=name_raw, faculty_short_raw=short_raw)
    if not to_link.exists():
        messages.error(request, 'No unmatched duties found for that sheet name/initial in this phase. Re-upload or refresh.')
        return redirect('core:dept_exam_phase_detail', phase_id=phase.id)

    existing = Faculty.objects.filter(department=fdept, short_name__iexact=short_name).first()
    if existing and existing.full_name.strip().upper() != full_name.strip().upper():
        messages.error(
            request,
            f'Short name "{short_name}" is already used by {existing.full_name}. '
            'Use a different initial or edit the existing faculty in Admin.',
        )
        return redirect('core:dept_exam_phase_detail', phase_id=phase.id)

    cred_line = ''
    with transaction.atomic():
        if existing:
            fac = existing
            if email and not fac.email:
                fac.email = email
                fac.save(update_fields=['email'])
        else:
            fac = Faculty.objects.create(
                department=fdept,
                full_name=full_name,
                short_name=short_name,
                email=email,
            )

        if not fac.user_id:
            username, password = _generate_faculty_portal_username(fac)
            user = User.objects.create_user(username=username, password=password)
            fac.user = user
            fac.save(update_fields=['user'])
            UserRole.objects.get_or_create(user=user, defaults={'role': 'faculty'})
            cred_line = f' Portal login created — username: {username}  password: {password} (copy now; not shown again).'
        else:
            cred_line = ' This faculty already had portal login; only supervision duties were linked.'

        n = to_link.update(faculty=fac, original_faculty=fac)

    messages.success(
        request,
        f'Linked {n} duty row(s) to {fac.full_name} ({fac.short_name}).{cred_line}',
    )
    return redirect('core:dept_exam_phase_detail', phase_id=phase.id)


@login_required
@require_http_methods(['POST'])
def dept_exam_phase_upload(request, phase_id):
    if not _dept_parent_only(request):
        messages.error(request, 'Only department coordinators can upload supervision sheets.')
        return redirect('accounts:role_redirect')
    prof = _parent_profile(request)
    if not prof:
        return redirect('accounts:role_redirect')
    rds = _parent_must_have_department(request, prof)
    if rds:
        return rds
    phase = _get_phase_for_coordinator(prof, request.user, phase_id)
    f = request.FILES.get('supervision_file')
    if not f:
        messages.error(request, 'Choose an Excel file to upload.')
        return redirect('core:dept_exam_phase_detail', phase_id=phase.id)
    try:
        rows = parse_combined_supervision_workbook(f)
    except Exception as e:
        messages.error(request, str(e))
        return redirect('core:dept_exam_phase_detail', phase_id=phase.id)

    n_unmatched = 0
    for row in rows:
        if phase.hub_coordinator_id:
            fac = match_faculty_global(row['faculty_name'], row['faculty_initial'])
        else:
            fac = match_faculty_for_department(
                prof.department,
                row['faculty_name'],
                row['faculty_initial'],
            )
        if not fac:
            n_unmatched += 1
    supervision_stage_put(request, phase.id, rows, n_unmatched)
    messages.info(
        request,
        f'Loaded {len(rows)} row(s) for review (not saved yet). '
        f'{"No" if n_unmatched == 0 else n_unmatched} row(s) could not be linked to a faculty record. '
        'Click Save to database to replace current duties for this phase.',
    )
    return redirect('core:dept_exam_phase_detail', phase_id=phase.id)


@login_required
@require_http_methods(['POST'])
def dept_exam_phase_commit_import(request, phase_id):
    if not _dept_parent_only(request):
        messages.error(request, 'Only department coordinators can save supervision imports.')
        return redirect('accounts:role_redirect')
    prof = _parent_profile(request)
    if not prof:
        return redirect('accounts:role_redirect')
    rds = _parent_must_have_department(request, prof)
    if rds:
        return rds
    phase = _get_phase_for_coordinator(prof, request.user, phase_id)
    blob = supervision_stage_get(request, phase.id)
    if not blob:
        messages.error(request, 'No imported data in this session. Upload an Excel file first.')
        return redirect('core:dept_exam_phase_detail', phase_id=phase.id)
    rows = supervision_stage_deserialize_rows(blob)
    with transaction.atomic():
        phase.duties.all().delete()
        created = 0
        n_unmatched = 0
        for row in rows:
            if phase.hub_coordinator_id:
                fac = match_faculty_global(row['faculty_name'], row['faculty_initial'])
            else:
                fac = match_faculty_for_department(
                    prof.department,
                    row['faculty_name'],
                    row['faculty_initial'],
                )
            if not fac:
                n_unmatched += 1
            SupervisionDuty.objects.create(
                phase=phase,
                faculty=fac,
                original_faculty=fac,
                faculty_name_raw=row['faculty_name'],
                faculty_short_raw=row['faculty_initial'],
                supervision_date=row['supervision_date'],
                time_slot=row['time_slot'],
                subject_name=row['subject_name'],
                division_code=row['division_code'],
            )
            created += 1
    clear_staging(request, 'supervision', phase.id)
    messages.success(
        request,
        f'Saved {created} supervision duty row(s). '
        f'{"No" if n_unmatched == 0 else n_unmatched} row(s) could not be linked to a faculty record (check names/initials).',
    )
    return redirect('core:dept_exam_phase_detail', phase_id=phase.id)


@login_required
@require_http_methods(['POST'])
def dept_exam_phase_discard_import(request, phase_id):
    if not _dept_parent_only(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    prof = _parent_profile(request)
    if not prof:
        return redirect('accounts:role_redirect')
    rds = _parent_must_have_department(request, prof)
    if rds:
        return rds
    phase = _get_phase_for_coordinator(prof, request.user, phase_id)
    if supervision_stage_get(request, phase.id):
        clear_staging(request, 'supervision', phase.id)
        messages.info(request, 'Discarded the draft import.')
    else:
        messages.info(request, 'No draft import to discard.')
    return redirect('core:dept_exam_phase_detail', phase_id=phase.id)


def _faculty_supervision_duties_page(request, *, history_only: bool):
    from core.views import _faculty_portal_guard_redirect, get_faculty_user, user_can_faculty

    guard_name = 'faculty_exam_history_duties' if history_only else 'faculty_supervision_duties'
    if not user_can_faculty(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    blocked = _faculty_portal_guard_redirect(request, guard_name)
    if blocked:
        return blocked
    fac = get_faculty_user(request)
    if not fac:
        messages.error(request, 'Faculty profile not found.')
        return redirect('accounts:role_redirect')
    duties_all = list(
        SupervisionDuty.objects.filter(faculty=fac)
        .select_related('phase', 'phase__institute_semester')
        .order_by('-phase__name', 'supervision_date', 'time_slot')
    )
    duties = [d for d in duties_all if not supervision_duty_in_faculty_exam_history(d)]
    duties_history = [d for d in duties_all if supervision_duty_in_faculty_exam_history(d)]
    pcr = build_faculty_paper_checking_rows(fac)
    paper_checking_rows = [r for r in pcr if not paper_checking_duty_in_faculty_exam_history(r['duty'], fac)]
    paper_checking_rows_history = [
        r for r in pcr if paper_checking_duty_in_faculty_exam_history(r['duty'], fac)
    ]
    psr = build_faculty_paper_setting_rows(fac)
    paper_setting_rows = []
    paper_setting_rows_history = []
    for r in psr:
        d = r['duty']
        if paper_setting_duty_in_faculty_exam_history(d, fac):
            paper_setting_rows_history.append(r)
        else:
            paper_setting_rows.append(r)
    if history_only:
        return render(
            request,
            'core/faculty/exam_duties.html',
            {
                'faculty': fac,
                'faculty_exam_duties_history_page': True,
                'exam_portal_semester_history': False,
                'duties': duties_history,
                'duties_history': [],
                'paper_checking_rows': paper_checking_rows_history,
                'paper_checking_rows_history': [],
                'paper_setting_rows': paper_setting_rows_history,
                'paper_setting_rows_history': [],
                'has_exam_history_elsewhere': False,
            },
        )
    return render(
        request,
        'core/faculty/exam_duties.html',
        {
            'faculty': fac,
            'faculty_exam_duties_history_page': False,
            'exam_portal_semester_history': bool(
                duties_history or paper_checking_rows_history or paper_setting_rows_history
            ),
            'duties': duties,
            'duties_history': [],
            'paper_checking_rows': paper_checking_rows,
            'paper_checking_rows_history': [],
            'paper_setting_rows': paper_setting_rows,
            'paper_setting_rows_history': [],
            'has_exam_history_elsewhere': bool(
                duties_history or paper_checking_rows_history or paper_setting_rows_history
            ),
        },
    )


@login_required
def faculty_supervision_duties(request):
    return _faculty_supervision_duties_page(request, history_only=False)


@login_required
def faculty_exam_history_duties(request):
    return _faculty_supervision_duties_page(request, history_only=True)


@login_required
@require_http_methods(['POST'])
def faculty_supervision_duty_complete(request):
    from core.views import _faculty_portal_guard_redirect, get_faculty_user, user_can_faculty

    if not user_can_faculty(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    blocked = _faculty_portal_guard_redirect(request, 'faculty_supervision_duty_complete')
    if blocked:
        return blocked
    fac = get_faculty_user(request)
    if not fac:
        return redirect('accounts:role_redirect')
    try:
        duty_id = int(request.POST.get('duty_id', ''))
    except (TypeError, ValueError):
        messages.error(request, 'Invalid duty.')
        return redirect('core:faculty_supervision_duties')
    duty = get_object_or_404(
        SupervisionDuty.objects.select_related('phase', 'phase__institute_semester'),
        pk=duty_id,
        faculty=fac,
    )
    if supervision_duty_in_faculty_exam_history(duty):
        messages.error(
            request,
            'This supervision slot belongs to an academic semester with Faculty portal off — open History (read-only).',
        )
        return redirect('core:faculty_supervision_duties')
    if duty.completion_status == SupervisionDuty.COMPLETED:
        messages.info(request, 'This duty is already marked complete.')
        return redirect('core:faculty_supervision_duties')
    room_no = (request.POST.get('room_no') or '').strip()
    block_no = (request.POST.get('block_no') or '').strip()
    if not room_no or not block_no:
        messages.error(request, 'Room number and block number are required.')
        return redirect('core:faculty_supervision_duties')
    duty.completion_status = SupervisionDuty.COMPLETED
    duty.room_no = room_no
    duty.block_no = block_no
    duty.completed_at = timezone.now()
    duty.save(update_fields=['completion_status', 'room_no', 'block_no', 'completed_at'])
    messages.success(request, 'Supervision duty marked complete with room and block.')
    return redirect('core:faculty_supervision_duties')


@login_required
@require_http_methods(['POST'])
def faculty_paper_checking_completion_request(request):
    from django.urls import reverse
    from core.views import _faculty_portal_guard_redirect, get_faculty_user, user_can_faculty

    if not user_can_faculty(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    blocked = _faculty_portal_guard_redirect(request, 'faculty_paper_checking_completion_request')
    if blocked:
        return blocked
    fac = get_faculty_user(request)
    if not fac:
        return redirect('accounts:role_redirect')
    try:
        duty_id = int(request.POST.get('duty_id', ''))
    except (TypeError, ValueError):
        messages.error(request, 'Invalid duty.')
        return redirect(reverse('core:faculty_supervision_duties') + '#checking')
    duty = get_object_or_404(
        PaperCheckingDuty.objects.select_related('phase', 'phase__institute_semester'),
        pk=duty_id,
    )
    if paper_checking_duty_in_faculty_exam_history(duty, fac):
        messages.error(
            request,
            'This row belongs to an academic semester with Faculty portal off — use History (read-only).',
        )
        return redirect(reverse('core:faculty_supervision_duties') + '#checking')
    if not (
        duty.faculty_id == fac.id
        or PaperCheckingAdjustedShare.objects.filter(duty=duty, faculty=fac).exists()
    ):
        messages.error(request, 'This paper checking duty is not assigned to you.')
        return redirect(reverse('core:faculty_supervision_duties') + '#checking')
    if PaperCheckingCompletionRequest.objects.filter(
        duty=duty, faculty=fac, status=PaperCheckingCompletionRequest.APPROVED
    ).exists():
        messages.info(request, 'This paper checking duty is already approved.')
        return redirect(reverse('core:faculty_supervision_duties') + '#checking')
    if PaperCheckingCompletionRequest.objects.filter(
        duty=duty, faculty=fac, status=PaperCheckingCompletionRequest.PENDING
    ).exists():
        messages.info(request, 'A completion request is already pending with your coordinator.')
        return redirect(reverse('core:faculty_supervision_duties') + '#checking')
    try:
        PaperCheckingCompletionRequest.objects.create(
            duty=duty,
            faculty=fac,
            status=PaperCheckingCompletionRequest.PENDING,
        )
    except IntegrityError:
        messages.info(request, 'A completion request is already pending.')
        return redirect(reverse('core:faculty_supervision_duties') + '#checking')
    messages.success(
        request,
        'Completion submitted to your department exam coordinator for approval.',
    )
    return redirect(reverse('core:faculty_supervision_duties') + '#checking')


@login_required
@require_http_methods(['POST'])
def faculty_paper_setting_completion_request(request):
    from django.db import IntegrityError
    from django.urls import reverse
    from core.views import _faculty_portal_guard_redirect, get_faculty_user, user_can_faculty

    if not user_can_faculty(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    blocked = _faculty_portal_guard_redirect(request, 'faculty_paper_setting_completion_request')
    if blocked:
        return blocked
    fac = get_faculty_user(request)
    if not fac:
        return redirect('accounts:role_redirect')
    try:
        duty_id = int(request.POST.get('duty_id', ''))
    except (TypeError, ValueError):
        messages.error(request, 'Invalid duty.')
        return redirect(reverse('core:faculty_supervision_duties') + '#setting')
    duty = get_object_or_404(
        PaperSettingDuty.objects.select_related('phase', 'phase__institute_semester'),
        pk=duty_id,
    )
    if paper_setting_duty_in_faculty_exam_history(duty, fac):
        messages.error(
            request,
            'This row belongs to an academic semester with Faculty portal off — use History (read-only).',
        )
        return redirect(reverse('core:faculty_supervision_duties') + '#setting')
    if duty.faculty_id != fac.id:
        messages.error(request, 'This paper setting duty is not assigned to you.')
        return redirect(reverse('core:faculty_supervision_duties') + '#setting')
    if PaperSettingCompletionRequest.objects.filter(
        duty=duty, faculty=fac, status=PaperSettingCompletionRequest.APPROVED
    ).exists():
        messages.info(request, 'This paper setting duty is already approved.')
        return redirect(reverse('core:faculty_supervision_duties') + '#setting')
    if PaperSettingCompletionRequest.objects.filter(
        duty=duty, faculty=fac, status=PaperSettingCompletionRequest.PENDING
    ).exists():
        messages.info(request, 'A completion request is already pending with your coordinator.')
        return redirect(reverse('core:faculty_supervision_duties') + '#setting')
    try:
        PaperSettingCompletionRequest.objects.create(
            duty=duty,
            faculty=fac,
            status=PaperSettingCompletionRequest.PENDING,
        )
    except IntegrityError:
        messages.info(request, 'A completion request is already pending.')
        return redirect(reverse('core:faculty_supervision_duties') + '#setting')
    messages.success(
        request,
        'Paper setting completion sent to your department exam coordinator for approval.',
    )
    return redirect(reverse('core:faculty_supervision_duties') + '#setting')


@login_required
@require_http_methods(['POST'])
def dept_exam_proxy_supervision(request):
    """Sub-unit login: reassign a duty to any institute faculty and mark as proxy."""
    if not _dept_child_only(request):
        messages.error(request, 'Only sub-unit exam coordinators can assign proxy supervision.')
        return redirect('accounts:role_redirect')
    prof = _child_profile(request)
    if not prof:
        return redirect('accounts:role_redirect')
    try:
        duty_id = int(request.POST.get('duty_id', ''))
    except (TypeError, ValueError):
        messages.error(request, 'Invalid duty.')
        return redirect('core:dept_exam_dashboard')
    duty = get_object_or_404(
        SupervisionDuty.objects.select_related('faculty'),
        pk=duty_id,
    )
    if not _child_can_access_duty(prof, duty):
        messages.error(request, 'This duty is outside your sub-unit.')
        return redirect('core:dept_exam_dashboard')
    if not duty.faculty_id:
        messages.error(request, 'Link this row to a faculty first (parent phase screen).')
        return redirect('core:dept_exam_dashboard')
    try:
        to_faculty_id = int(request.POST.get('to_faculty_id', ''))
    except (TypeError, ValueError):
        messages.error(request, 'Choose a faculty to assign as proxy.')
        return redirect('core:dept_exam_dashboard')
    new_fac = Faculty.objects.filter(pk=to_faculty_id).select_related('department').first()
    if not new_fac:
        messages.error(request, 'Invalid faculty selected.')
        return redirect('core:dept_exam_dashboard')
    if new_fac.id == duty.faculty_id:
        messages.error(request, 'Choose a different faculty to assign as proxy.')
        return redirect('core:dept_exam_dashboard')
    old = duty.faculty
    duty.faculty = new_fac
    duty.is_proxy = True
    if not duty.original_faculty_id and old:
        duty.original_faculty = old
    duty.completion_status = SupervisionDuty.OPEN
    duty.room_no = ''
    duty.block_no = ''
    duty.completed_at = None
    duty.save()
    messages.success(
        request,
        f'Proxy set: duty moved from {old.full_name} to {new_fac.full_name}. They must mark complete with room/block.',
    )
    return redirect('core:dept_exam_dashboard')


def _faculties_for_bulk_paper_check_child(duty: PaperCheckingDuty, dept_id: int) -> list[Faculty]:
    """Who should get an approved completion row for this duty in a sub-unit bulk action.

    Prefer adjusted shares assigned to faculty in this attendance department; otherwise
    the duty's sheet evaluator (faculty) — the row is visible to this sub-unit because of
    block allocations, even when the evaluator sits in another department.
    """
    facs: list[Faculty] = []
    from core.faculty_scope import faculty_has_department_access

    dept = Department.objects.filter(pk=dept_id).first()
    for sh in duty.adjusted_shares.all():
        if sh.faculty_id and dept and faculty_has_department_access(sh.faculty, dept):
            facs.append(sh.faculty)
    if facs:
        return facs
    if duty.faculty_id:
        return [duty.faculty]
    return []


def _bulk_approve_paper_check_row(duty: PaperCheckingDuty, faculty: Faculty, user: User) -> int:
    now = timezone.now()
    pending = PaperCheckingCompletionRequest.objects.filter(
        duty=duty, faculty=faculty, status=PaperCheckingCompletionRequest.PENDING
    ).first()
    if pending:
        pending.status = PaperCheckingCompletionRequest.APPROVED
        pending.decided_at = now
        pending.decided_by = user
        pending.save(update_fields=['status', 'decided_at', 'decided_by'])
        return 1
    if PaperCheckingCompletionRequest.objects.filter(
        duty=duty, faculty=faculty, status=PaperCheckingCompletionRequest.APPROVED
    ).exists():
        return 0
    PaperCheckingCompletionRequest.objects.create(
        duty=duty,
        faculty=faculty,
        status=PaperCheckingCompletionRequest.APPROVED,
        decided_at=now,
        decided_by=user,
    )
    return 1


def _bulk_approve_paper_setting_row(duty: PaperSettingDuty, faculty: Faculty, user: User) -> int:
    now = timezone.now()
    pending = PaperSettingCompletionRequest.objects.filter(
        duty=duty, faculty=faculty, status=PaperSettingCompletionRequest.PENDING
    ).first()
    if pending:
        pending.status = PaperSettingCompletionRequest.APPROVED
        pending.decided_at = now
        pending.decided_by = user
        pending.save(update_fields=['status', 'decided_at', 'decided_by'])
        return 1
    if PaperSettingCompletionRequest.objects.filter(
        duty=duty, faculty=faculty, status=PaperSettingCompletionRequest.APPROVED
    ).exists():
        return 0
    PaperSettingCompletionRequest.objects.create(
        duty=duty,
        faculty=faculty,
        status=PaperSettingCompletionRequest.APPROVED,
        decided_at=now,
        decided_by=user,
    )
    return 1


@login_required
@require_http_methods(['POST'])
def dept_exam_child_bulk_complete_all(request):
    """Temporary: sub-unit marks supervision complete and approves paper check/setting without faculty workflow."""
    if not getattr(settings, 'EXAM_PORTAL_BULK_COMPLETE', False):
        messages.error(request, 'This action is not enabled (set EXAM_PORTAL_BULK_COMPLETE=1).')
        return redirect('core:dept_exam_dashboard')
    if not _dept_child_only(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    prof = _child_profile(request)
    if not prof or not prof.department_id:
        return redirect('accounts:role_redirect')
    sup_base = SupervisionDuty.objects.filter(subunit_supervision_duty_filter_q(prof))
    dept_id = prof.department_id
    n_sup = n_pc = n_ps = 0
    now = timezone.now()
    with transaction.atomic():
        for d in (
            sup_base.filter(faculty_id__isnull=False)
            .exclude(completion_status=SupervisionDuty.COMPLETED)
            .select_related('faculty')
        ):
            d.completion_status = SupervisionDuty.COMPLETED
            if not (d.room_no or '').strip():
                d.room_no = '—'
            if not (d.block_no or '').strip():
                d.block_no = '—'
            d.completed_at = now
            d.save(update_fields=['completion_status', 'room_no', 'block_no', 'completed_at'])
            n_sup += 1

        pc_qs = paper_checking_duties_for_child_prof(prof).prefetch_related(
            Prefetch(
                'adjusted_shares',
                queryset=PaperCheckingAdjustedShare.objects.select_related('faculty'),
            ),
            'faculty',
        )
        seen_pc: set[tuple[int, int]] = set()
        for duty in pc_qs:
            for fac in _faculties_for_bulk_paper_check_child(duty, dept_id):
                key = (duty.pk, fac.pk)
                if key in seen_pc:
                    continue
                seen_pc.add(key)
                n_pc += _bulk_approve_paper_check_row(duty, fac, request.user)

        for duty in paper_setting_duties_for_child_prof(prof).select_related('faculty'):
            if not duty.faculty_id:
                continue
            n_ps += _bulk_approve_paper_setting_row(duty, duty.faculty, request.user)

    messages.success(
        request,
        f'Bulk update: {n_sup} supervision slot(s) marked complete; '
        f'{n_pc} paper-check completion(s) approved or created; '
        f'{n_ps} paper-setting completion(s) approved or created.',
    )
    return redirect('core:dept_exam_dashboard')


@login_required
def dept_exam_dr_report_excel(request):
    """Supervision DR workbook: sub-unit, department, hub phases, or exam section (all phases)."""
    if _exam_section_portal_access(request):
        duties = list(
            SupervisionDuty.objects.select_related('faculty', 'original_faculty', 'phase')
            .order_by('supervision_date', 'phase__name', 'division_code', 'pk')
        )
        title = 'Institute-wide supervision — exam section (all phases)'
        prefix = 'exam_section_institute'
        return build_supervision_dr_excel(duties, title_line=title, sheet_prefix=prefix)
    if _dept_child_only(request):
        prof = _child_profile(request)
        if not prof:
            messages.error(request, 'No profile.')
            return redirect('accounts:role_redirect')
        code = (prof.subunit_code or '').strip().upper()
        qs = SupervisionDuty.objects.filter(subunit_supervision_duty_filter_q(prof))
        title = f'{prof.department.name} — sub-unit {code}'
        duties = list(qs.select_related('faculty', 'original_faculty', 'phase'))
        prefix = f'subunit_{code}'
    elif _dept_parent_only(request):
        prof = _parent_profile(request)
        if not prof:
            messages.error(request, 'No profile.')
            return redirect('accounts:role_redirect')
        rds = _parent_must_have_department(request, prof)
        if rds:
            return rds
        if _is_hub_coordinator(prof):
            qs = SupervisionDuty.objects.filter(
                phase__hub_coordinator=request.user,
                phase__institute_semester_id=prof.institute_semester_id,
            )
            title = f'Institute-wide supervision — hub {request.user.username} ({prof.institute_semester.label})'
            prefix = f'hub_{request.user.id}_sem{prof.institute_semester_id}'
        else:
            qs = SupervisionDuty.objects.filter(phase__department=prof.department)
            title = f'{prof.department.name} — all divisions (supervision)'
            prefix = f'dept_{prof.department_id}'
        duties = list(qs.select_related('faculty', 'original_faculty', 'phase'))
    else:
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    return build_supervision_dr_excel(duties, title_line=title, sheet_prefix=prefix)


@login_required
def dept_exam_daily_dr_excel(request):
    """Official exam.xlsx daily layout: department child = one dept; hub parent = that hub's phases only."""
    if not (_dept_parent_only(request) or _dept_child_only(request)):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    raw = (request.GET.get('date') or request.GET.get('date_from') or '').strip()
    raw_to = (request.GET.get('date_to') or '').strip()
    if not raw:
        messages.error(request, 'Select an exam date (From).')
        return redirect('core:dept_exam_dashboard')
    dept_ids = None
    hub_uid = None
    hub_sem_id = None
    if _dept_child_only(request):
        prof = _child_profile(request)
        if not prof or not prof.department_id:
            messages.error(request, 'No department on profile.')
            return redirect('core:dept_exam_dashboard')
        dept_ids = [prof.department_id]
    else:
        prof = _parent_profile(request)
        if not prof:
            messages.error(request, 'No profile.')
            return redirect('core:dept_exam_dashboard')
        if _is_hub_coordinator(prof):
            dept_ids = None
            hub_uid = request.user.id
            hub_sem_id = prof.institute_semester_id
        else:
            rds = _parent_must_have_department(request, prof)
            if rds:
                return rds
            dept_ids = [prof.department_id]
    try:
        dates = parse_dates_from_request(
            raw,
            raw_to,
            hub_coordinator_id=hub_uid,
            hub_institute_semester_id=hub_sem_id,
        )
        return build_exam_daily_dr_workbook(
            dates,
            department_ids=dept_ids,
            hub_coordinator_id=hub_uid,
            hub_institute_semester_id=hub_sem_id,
        )
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect('core:dept_exam_dashboard')
    except FileNotFoundError as exc:
        messages.error(request, str(exc))
        return redirect('core:dept_exam_dashboard')
    except Exception as exc:
        messages.error(request, f'Could not build workbook: {exc}')
        return redirect('core:dept_exam_dashboard')


def _faculty_exam_credits_analytics_page(request, *, history_only: bool):
    from core.faculty_scope import get_faculty_exam_context_department
    from core.views import _faculty_portal_guard_redirect, get_faculty_user, user_can_faculty

    guard_name = (
        'faculty_exam_history_credits_analytics' if history_only else 'faculty_exam_credits_analytics'
    )
    if not user_can_faculty(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    blocked = _faculty_portal_guard_redirect(request, guard_name)
    if blocked:
        return blocked
    fac = get_faculty_user(request)
    if not fac:
        return redirect('accounts:role_redirect')
    exam_dept = get_faculty_exam_context_department(request)
    dept_id_for_credits = (exam_dept.pk if exam_dept else None) or fac.department_id

    paper_phase_ids = set(
        PaperCheckingDuty.objects.filter(Q(faculty=fac) | Q(adjusted_shares__faculty=fac))
        .values_list('phase_id', flat=True)
        .distinct()
    )
    paper_phases = PaperCheckingPhase.objects.filter(pk__in=paper_phase_ids).order_by('name')

    raw_pp = (request.GET.get('paper_phase_id') or '').strip()
    sel_paper_phase = None
    try:
        if raw_pp:
            sel_paper_phase = int(raw_pp)
    except (TypeError, ValueError):
        sel_paper_phase = None
    if sel_paper_phase not in paper_phase_ids:
        sel_paper_phase = None

    paper_total = Decimal('0')
    paper_approved_rows: list[dict] = []
    paper_approved_rows_history: list[dict] = []
    if sel_paper_phase:
        reqs = list(
            PaperCheckingCompletionRequest.objects.filter(
                faculty=fac,
                status=PaperCheckingCompletionRequest.APPROVED,
                duty__phase_id=sel_paper_phase,
            )
            .select_related('duty', 'duty__phase')
            .prefetch_related(
                Prefetch(
                    'duty__allocations',
                    queryset=PaperCheckingAllocation.objects.select_related('department'),
                ),
            )
            .order_by('-decided_at', '-id')
        )
        for r in reqs:
            cr = credit_for_completion_request(r)
            rowd = {
                'req': r,
                'credit': cr,
                'papers': paper_count_for_completion(r),
                'rem': remuneration_for_completion_request(r),
            }
            if paper_checking_completion_in_faculty_exam_history(r, fac):
                paper_approved_rows_history.append(rowd)
            else:
                paper_total += cr
                paper_approved_rows.append(rowd)

    sup_phase_ids = list(
        SupervisionDuty.objects.filter(faculty=fac).values_list('phase_id', flat=True).distinct()
    )
    sup_phases = SupervisionExamPhase.objects.filter(pk__in=sup_phase_ids).order_by('name')

    raw_sp = (request.GET.get('sup_phase_id') or '').strip()
    sel_sup = None
    try:
        if raw_sp:
            sel_sup = int(raw_sp)
    except (TypeError, ValueError):
        sel_sup = None
    if sel_sup not in sup_phase_ids:
        sel_sup = None

    sup_done = sup_open = 0
    sup_credit_total = Decimal('0')
    sup_rm_per = Decimal('0')
    sup_completed_rows: list[dict] = []
    sup_completed_rows_history: list[dict] = []
    if sel_sup:
        sup_phase_name = ''
        _sp = SupervisionExamPhase.objects.filter(pk=sel_sup).only('name').first()
        if _sp:
            sup_phase_name = _sp.name or ''
        sup_cr_per = supervision_credit_for_phase(dept_id_for_credits, sup_phase_name)
        sup_rm_per = supervision_remuneration_for_phase(dept_id_for_credits, sup_phase_name)
        sup_open_qs = (
            SupervisionDuty.objects.filter(faculty=fac, phase_id=sel_sup)
            .exclude(completion_status=SupervisionDuty.COMPLETED)
            .select_related('phase', 'phase__institute_semester')
        )
        sup_open = sum(1 for d in sup_open_qs if not supervision_duty_in_faculty_exam_history(d))
        for d in (
            SupervisionDuty.objects.filter(
                faculty=fac, phase_id=sel_sup, completion_status=SupervisionDuty.COMPLETED
            )
            .select_related('phase', 'phase__institute_semester')
            .order_by('supervision_date', 'time_slot', 'pk')
        ):
            rowd = {'duty': d, 'credit': sup_cr_per, 'rem': sup_rm_per}
            if supervision_duty_in_faculty_exam_history(d):
                sup_completed_rows_history.append(rowd)
            else:
                sup_credit_total += sup_cr_per
                sup_completed_rows.append(rowd)
        sup_done = len(sup_completed_rows)

    setting_phase_ids = set(
        PaperSettingDuty.objects.filter(faculty=fac).values_list('phase_id', flat=True)
    )
    setting_phases = PaperSettingPhase.objects.filter(pk__in=setting_phase_ids).order_by('name')

    raw_set = (request.GET.get('setting_phase_id') or '').strip()
    sel_setting_phase = None
    try:
        if raw_set:
            sel_setting_phase = int(raw_set)
    except (TypeError, ValueError):
        sel_setting_phase = None
    if sel_setting_phase not in setting_phase_ids:
        sel_setting_phase = None

    setting_total = Decimal('0')
    setting_approved_rows: list[dict] = []
    setting_approved_rows_history: list[dict] = []
    if sel_setting_phase:
        for r in (
            PaperSettingCompletionRequest.objects.filter(
                faculty=fac,
                status=PaperSettingCompletionRequest.APPROVED,
                duty__phase_id=sel_setting_phase,
            )
            .select_related('duty', 'duty__phase')
            .order_by('-decided_at', '-id')
        ):
            cr = credit_for_paper_setting_request(r)
            rowd = {
                'req': r,
                'credit': cr,
                'rem': remuneration_for_paper_setting_request(r),
            }
            if paper_setting_completion_in_faculty_exam_history(r, fac):
                setting_approved_rows_history.append(rowd)
            else:
                setting_total += cr
                setting_approved_rows.append(rowd)

    def _paper_summarize(qs_list):
        tot = rem = Decimal('0')
        papers = 0
        for r in qs_list:
            tot += credit_for_completion_request(r)
            rem += remuneration_for_completion_request(r)
            papers += int(paper_count_for_completion(r))
        return tot, rem, papers

    _paper_approved_all = list(
        PaperCheckingCompletionRequest.objects.filter(
            faculty=fac,
            status=PaperCheckingCompletionRequest.APPROVED,
        )
        .select_related('duty', 'duty__phase')
        .prefetch_related(
            Prefetch(
                'duty__allocations',
                queryset=PaperCheckingAllocation.objects.select_related('department'),
            ),
        )
    )
    paper_cur_list = [r for r in _paper_approved_all if not paper_checking_completion_in_faculty_exam_history(r, fac)]
    paper_hist_list = [r for r in _paper_approved_all if paper_checking_completion_in_faculty_exam_history(r, fac)]
    paper_summary_total, paper_summary_rem, paper_summary_papers = _paper_summarize(paper_cur_list)
    hist_paper_total, hist_paper_rem, hist_paper_papers = _paper_summarize(paper_hist_list)

    sup_all_completed = list(
        SupervisionDuty.objects.filter(
            faculty=fac,
            completion_status=SupervisionDuty.COMPLETED,
        ).select_related('phase', 'phase__institute_semester')
    )
    sup_cur_duties = [d for d in sup_all_completed if not supervision_duty_in_faculty_exam_history(d)]
    sup_hist_duties = [d for d in sup_all_completed if supervision_duty_in_faculty_exam_history(d)]

    sup_completed_by_phase = Counter(d.phase_id for d in sup_cur_duties)
    sup_hist_by_phase = Counter(d.phase_id for d in sup_hist_duties)

    def _sup_sum(counter: Counter) -> tuple[Decimal, Decimal]:
        tot = rem = Decimal('0')
        if not counter:
            return tot, rem
        phase_names = dict(
            SupervisionExamPhase.objects.filter(pk__in=counter.keys()).values_list('pk', 'name')
        )
        for phase_id, n in counter.items():
            pname = phase_names.get(phase_id) or ''
            tot += supervision_credit_for_phase(dept_id_for_credits, pname) * n
            rem += supervision_remuneration_for_phase(dept_id_for_credits, pname) * n
        return tot, rem

    sup_summary_total, sup_summary_rem = _sup_sum(sup_completed_by_phase)
    hist_sup_total, hist_sup_rem = _sup_sum(sup_hist_by_phase)

    _setting_all = list(
        PaperSettingCompletionRequest.objects.filter(
            faculty=fac,
            status=PaperSettingCompletionRequest.APPROVED,
        ).select_related('duty', 'duty__phase')
    )
    setting_cur = [r for r in _setting_all if not paper_setting_completion_in_faculty_exam_history(r, fac)]
    setting_hist = [r for r in _setting_all if paper_setting_completion_in_faculty_exam_history(r, fac)]
    setting_summary_total = sum((credit_for_paper_setting_request(r) for r in setting_cur), Decimal('0'))
    setting_summary_rem = sum(
        (remuneration_for_paper_setting_request(r) for r in setting_cur), Decimal('0')
    )
    hist_setting_total = sum(
        (credit_for_paper_setting_request(r) for r in setting_hist), Decimal('0')
    )
    hist_setting_rem = sum(
        (remuneration_for_paper_setting_request(r) for r in setting_hist), Decimal('0')
    )

    summary_grand_total = paper_summary_total + sup_summary_total + setting_summary_total
    summary_grand_rem = paper_summary_rem + sup_summary_rem + setting_summary_rem
    history_grand_total = hist_paper_total + hist_sup_total + hist_setting_total
    history_grand_rem = hist_paper_rem + hist_sup_rem + hist_setting_rem
    show_faculty_exam_history_totals = history_grand_total > 0

    if history_only:
        paper_phases = PaperCheckingPhase.objects.filter(
            pk__in={r.duty.phase_id for r in paper_hist_list}
        ).order_by('name')
        paper_phase_ids = set(paper_phases.values_list('pk', flat=True))
        if sel_paper_phase not in paper_phase_ids:
            sel_paper_phase = None
        paper_total = Decimal('0')
        paper_approved_rows = []
        paper_approved_rows_history = []
        if sel_paper_phase:
            sel_ids = {sel_paper_phase}
            for r in paper_hist_list:
                if r.duty.phase_id not in sel_ids:
                    continue
                cr = credit_for_completion_request(r)
                paper_total += cr
                paper_approved_rows.append(
                    {
                        'req': r,
                        'credit': cr,
                        'papers': paper_count_for_completion(r),
                        'rem': remuneration_for_completion_request(r),
                    }
                )
        sup_phases = SupervisionExamPhase.objects.filter(
            pk__in={d.phase_id for d in sup_hist_duties}
        ).order_by('name')
        sup_phase_ids_hist = list(sup_phases.values_list('pk', flat=True))
        if sel_sup not in sup_phase_ids_hist:
            sel_sup = None
        sup_done = sup_open = 0
        sup_credit_total = Decimal('0')
        sup_rm_per = Decimal('0')
        sup_completed_rows = []
        sup_completed_rows_history = []
        if sel_sup:
            sup_phase_name = ''
            _sp = SupervisionExamPhase.objects.filter(pk=sel_sup).only('name').first()
            if _sp:
                sup_phase_name = _sp.name or ''
            sup_cr_per = supervision_credit_for_phase(dept_id_for_credits, sup_phase_name)
            sup_rm_per = supervision_remuneration_for_phase(dept_id_for_credits, sup_phase_name)
            sup_open_qs = (
                SupervisionDuty.objects.filter(faculty=fac, phase_id=sel_sup)
                .exclude(completion_status=SupervisionDuty.COMPLETED)
                .select_related('phase', 'phase__institute_semester')
            )
            sup_open = sum(1 for d in sup_open_qs if supervision_duty_in_faculty_exam_history(d))
            for d in (
                SupervisionDuty.objects.filter(
                    faculty=fac, phase_id=sel_sup, completion_status=SupervisionDuty.COMPLETED
                )
                .select_related('phase', 'phase__institute_semester')
                .order_by('supervision_date', 'time_slot', 'pk')
            ):
                if not supervision_duty_in_faculty_exam_history(d):
                    continue
                sup_completed_rows.append({'duty': d, 'credit': sup_cr_per, 'rem': sup_rm_per})
            sup_credit_total = sup_cr_per * len(sup_completed_rows)
            sup_done = len(sup_completed_rows)
        setting_phases = PaperSettingPhase.objects.filter(
            pk__in={r.duty.phase_id for r in setting_hist}
        ).order_by('name')
        setting_phase_ids_hist = set(setting_phases.values_list('pk', flat=True))
        if sel_setting_phase not in setting_phase_ids_hist:
            sel_setting_phase = None
        setting_total = Decimal('0')
        setting_approved_rows = []
        setting_approved_rows_history = []
        if sel_setting_phase:
            for r in setting_hist:
                if r.duty.phase_id != sel_setting_phase:
                    continue
                cr = credit_for_paper_setting_request(r)
                setting_total += cr
                setting_approved_rows.append(
                    {
                        'req': r,
                        'credit': cr,
                        'rem': remuneration_for_paper_setting_request(r),
                    }
                )
        summary_grand_total = history_grand_total
        summary_grand_rem = history_grand_rem
        summary_paper_credits = hist_paper_total
        summary_paper_rem = hist_paper_rem
        summary_paper_papers = hist_paper_papers
        summary_sup_credits = hist_sup_total
        summary_sup_rem = hist_sup_rem
        summary_setting_credits = hist_setting_total
        summary_setting_rem = hist_setting_rem
        show_faculty_exam_history_totals = False
    else:
        paper_approved_rows_history = []
        sup_completed_rows_history = []
        setting_approved_rows_history = []
        show_faculty_exam_history_totals = False

    paper_total_rem = sum((r['rem'] for r in paper_approved_rows), Decimal('0'))
    paper_phase_papers = sum((int(r['papers']) for r in paper_approved_rows), 0)
    sup_total_rem = (sup_rm_per * sup_done) if sel_sup else Decimal('0')
    setting_total_rem = sum((r['rem'] for r in setting_approved_rows), Decimal('0'))

    return render(
        request,
        'core/faculty/exam_credits_analytics.html',
        {
            'faculty': fac,
            'faculty_exam_credits_history_page': history_only,
            'paper_phases': paper_phases,
            'selected_paper_phase_id': sel_paper_phase,
            'paper_approved_rows': paper_approved_rows,
            'paper_approved_rows_history': paper_approved_rows_history,
            'paper_total_credits': paper_total,
            'paper_total_rem': paper_total_rem,
            'paper_phase_papers': paper_phase_papers,
            'sup_phases': sup_phases,
            'selected_sup_phase_id': sel_sup,
            'sup_completed_count': sup_done,
            'sup_open_count': sup_open,
            'sup_credit_total': sup_credit_total,
            'sup_total_rem': sup_total_rem,
            'sup_completed_rows': sup_completed_rows,
            'sup_completed_rows_history': sup_completed_rows_history,
            'setting_phases': setting_phases,
            'selected_setting_phase_id': sel_setting_phase,
            'setting_total_credits': setting_total,
            'setting_total_rem': setting_total_rem,
            'setting_approved_rows': setting_approved_rows,
            'setting_approved_rows_history': setting_approved_rows_history,
            'summary_paper_credits': paper_summary_total,
            'summary_sup_credits': sup_summary_total,
            'summary_setting_credits': setting_summary_total,
            'summary_grand_total': summary_grand_total,
            'summary_paper_rem': paper_summary_rem,
            'summary_sup_rem': sup_summary_rem,
            'summary_setting_rem': setting_summary_rem,
            'summary_grand_rem': summary_grand_rem,
            'summary_paper_papers': paper_summary_papers,
            'history_paper_credits': hist_paper_total,
            'history_sup_credits': hist_sup_total,
            'history_setting_credits': hist_setting_total,
            'history_grand_total': history_grand_total,
            'history_paper_rem': hist_paper_rem,
            'history_sup_rem': hist_sup_rem,
            'history_setting_rem': hist_setting_rem,
            'history_grand_rem': history_grand_rem,
            'history_paper_papers': hist_paper_papers,
            'show_faculty_exam_history_totals': show_faculty_exam_history_totals,
        },
    )


@login_required
def faculty_exam_credits_analytics(request):
    return _faculty_exam_credits_analytics_page(request, history_only=False)


@login_required
def faculty_exam_history_credits_analytics(request):
    return _faculty_exam_credits_analytics_page(request, history_only=True)


@login_required
def dept_exam_credit_analytics(request):
    if not (_dept_parent_only(request) or _dept_child_only(request)):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    fac_qs = _credit_analytics_faculty_qs(request)
    fac_list = list(fac_qs)
    selected = None
    detail_rows = []
    total_credits = Decimal('0')
    summaries = []
    grand_total_credit = Decimal('0')
    raw_id = (request.GET.get('faculty_id') or '').strip()
    df = (request.GET.get('date_from') or '').strip()
    dt = (request.GET.get('date_to') or '').strip()
    sel_paper_phase = _parse_optional_int_param(request, 'paper_phase_id')
    scope_label = 'your department'
    if _dept_parent_only(request):
        prof = _parent_profile(request)
        if prof and _is_hub_coordinator(prof):
            scope_label = 'invited departments & sub-units under your hub'
        elif prof and prof.department_id:
            scope_label = prof.department.name
    elif _dept_child_only(request):
        ch = _child_profile(request)
        if ch and ch.department_id:
            scope_label = ch.department.name
    allowed_paper_phase_ids: set[int] = set()
    paper_phase_choices: list = []
    if fac_list:
        allowed_paper_phase_ids = set(
            _bulk_approved_paper_completions_qs(fac_qs, df, dt, None).values_list(
                'duty__phase_id', flat=True
            )
        )
        paper_phase_choices = list(
            PaperCheckingPhase.objects.filter(pk__in=allowed_paper_phase_ids).order_by('name')
        )
        if sel_paper_phase and sel_paper_phase not in allowed_paper_phase_ids:
            sel_paper_phase = None
        completions = list(_bulk_approved_paper_completions_qs(fac_qs, df, dt, sel_paper_phase))
        summaries, grand_total_credit, detail_map = _paper_summaries_and_detail_map(
            fac_list, completions
        )
        if raw_id:
            try:
                fid = int(raw_id)
            except (TypeError, ValueError):
                fid = None
            if fid:
                selected = get_object_or_404(fac_qs, pk=fid)
                detail_rows = detail_map.get(selected.pk, [])
                total_credits = next(
                    (s['total_credit'] for s in summaries if s['faculty'].pk == selected.pk),
                    Decimal('0'),
                )
    return render(
        request,
        'core/exam_dept/credit_analytics.html',
        {
            'institute_scope': False,
            'scope_label': scope_label,
            'analytics_mode': 'paper',
            'faculty_list': fac_qs,
            'selected_faculty': selected,
            'detail_rows': detail_rows,
            'total_credits': total_credits,
            'paper_summaries': summaries,
            'grand_total_credit': grand_total_credit,
            'filter_date_from': df,
            'filter_date_to': dt,
            'paper_phase_choices': paper_phase_choices,
            'selected_paper_phase_id': sel_paper_phase,
            'list_url': 'core:dept_exam_credit_analytics',
            'excel_url': 'core:dept_exam_credit_analytics_excel',
            'excel_query_all': _analytics_excel_query(
                df, dt, paper_phase_id=sel_paper_phase
            ),
            'excel_query_selected': (
                _analytics_excel_query(
                    df, dt, selected.pk, paper_phase_id=sel_paper_phase
                )
                if selected
                else None
            ),
        },
    )


@login_required
def dept_exam_credit_analytics_excel(request):
    if not (_dept_parent_only(request) or _dept_child_only(request)):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    fac_qs = _credit_analytics_faculty_qs(request)
    fac_list = list(fac_qs)
    raw_id = (request.GET.get('faculty_id') or '').strip()
    df = (request.GET.get('date_from') or '').strip()
    dt = (request.GET.get('date_to') or '').strip()
    sel_paper_phase = _parse_optional_int_param(request, 'paper_phase_id')
    if not fac_list:
        messages.error(request, 'No faculty in scope.')
        return redirect('core:dept_exam_credit_analytics')
    allowed = set(
        _bulk_approved_paper_completions_qs(fac_qs, df, dt, None).values_list(
            'duty__phase_id', flat=True
        )
    )
    if sel_paper_phase and sel_paper_phase not in allowed:
        sel_paper_phase = None
    completions = list(_bulk_approved_paper_completions_qs(fac_qs, df, dt, sel_paper_phase))
    summaries, grand_credit, detail_map = _paper_summaries_and_detail_map(fac_list, completions)
    single_id = None
    if raw_id:
        try:
            single_id = int(raw_id)
        except (TypeError, ValueError):
            single_id = None
        if single_id:
            get_object_or_404(fac_qs, pk=single_id)
    dept_name = ''
    if fac_list and fac_list[0].department_id:
        dept_name = fac_list[0].department.name
    scope_line = f'Department scope — {dept_name or "see portal"} · decided-at filter: {df or "—"} → {dt or "—"}'
    wb = _build_paper_credit_excel_workbook(
        scope_line=scope_line,
        summaries=summaries,
        grand_credit=grand_credit,
        detail_map=detail_map,
        fac_ordered=fac_list,
        single_faculty_id=single_id,
    )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    if single_id:
        sf = get_object_or_404(fac_qs, pk=single_id)
        fname = f'paper_check_credits_{sf.short_name or sf.pk}_{sf.pk}.xlsx'
    else:
        fname = 'paper_check_credits_all.xlsx'
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@login_required
def exam_section_credit_analytics(request):
    if not _exam_section_portal_access(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    fac_qs = Faculty.objects.select_related('department').order_by(
        'department__name', 'full_name'
    )
    exam_sem_ids = exam_section_working_semester_ids(request) if is_exam_section_operator(request) else None
    if exam_sem_ids:
        fac_qs = fac_qs.filter(department__institute_semester_id__in=exam_sem_ids)
    fac_list = list(fac_qs)
    selected = None
    detail_rows = []
    total_credits = Decimal('0')
    summaries = []
    grand_total_credit = Decimal('0')
    raw_id = (request.GET.get('faculty_id') or '').strip()
    df = (request.GET.get('date_from') or '').strip()
    dt = (request.GET.get('date_to') or '').strip()
    sel_paper_phase = _parse_optional_int_param(request, 'paper_phase_id')
    allowed_paper_phase_ids: set[int] = set()
    paper_phase_choices: list = []
    if fac_list:
        allowed_paper_phase_ids = set(
            _bulk_approved_paper_completions_qs(
                fac_qs, df, dt, None, exam_semester_ids=exam_sem_ids
            ).values_list('duty__phase_id', flat=True)
        )
        paper_phase_choices = list(
            PaperCheckingPhase.objects.filter(pk__in=allowed_paper_phase_ids).order_by('name')
        )
        if sel_paper_phase and sel_paper_phase not in allowed_paper_phase_ids:
            sel_paper_phase = None
        completions = list(
            _bulk_approved_paper_completions_qs(
                fac_qs, df, dt, sel_paper_phase, exam_semester_ids=exam_sem_ids
            )
        )
        summaries, grand_total_credit, detail_map = _paper_summaries_and_detail_map(
            fac_list, completions
        )
        if raw_id:
            try:
                fid = int(raw_id)
            except (TypeError, ValueError):
                fid = None
            if fid:
                selected = get_object_or_404(fac_qs, pk=fid)
                detail_rows = detail_map.get(selected.pk, [])
                total_credits = next(
                    (s['total_credit'] for s in summaries if s['faculty'].pk == selected.pk),
                    Decimal('0'),
                )
    scope_label_paper = 'All departments'
    if exam_sem_ids:
        scope_label_paper = 'Selected working academic semester(s) — faculty in matching departments'
    return render(
        request,
        'core/exam_dept/credit_analytics.html',
        {
            'institute_scope': True,
            'scope_label': scope_label_paper,
            'analytics_mode': 'paper',
            'faculty_list': fac_qs,
            'selected_faculty': selected,
            'detail_rows': detail_rows,
            'total_credits': total_credits,
            'paper_summaries': summaries,
            'grand_total_credit': grand_total_credit,
            'filter_date_from': df,
            'filter_date_to': dt,
            'paper_phase_choices': paper_phase_choices,
            'selected_paper_phase_id': sel_paper_phase,
            'list_url': 'core:exam_section_credit_analytics',
            'excel_url': 'core:exam_section_credit_analytics_excel',
            'excel_query_all': _analytics_excel_query(
                df, dt, paper_phase_id=sel_paper_phase
            ),
            'excel_query_selected': (
                _analytics_excel_query(
                    df, dt, selected.pk, paper_phase_id=sel_paper_phase
                )
                if selected
                else None
            ),
        },
    )


@login_required
def exam_section_credit_analytics_excel(request):
    if not _exam_section_portal_access(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    fac_qs = Faculty.objects.select_related('department').order_by(
        'department__name', 'full_name'
    )
    exam_sem_ids = exam_section_working_semester_ids(request) if is_exam_section_operator(request) else None
    if exam_sem_ids:
        fac_qs = fac_qs.filter(department__institute_semester_id__in=exam_sem_ids)
    fac_list = list(fac_qs)
    raw_id = (request.GET.get('faculty_id') or '').strip()
    df = (request.GET.get('date_from') or '').strip()
    dt = (request.GET.get('date_to') or '').strip()
    sel_paper_phase = _parse_optional_int_param(request, 'paper_phase_id')
    if not fac_list:
        messages.error(request, 'No faculty records.')
        return redirect('core:exam_section_credit_analytics')
    allowed = set(
        _bulk_approved_paper_completions_qs(
            fac_qs, df, dt, None, exam_semester_ids=exam_sem_ids
        ).values_list('duty__phase_id', flat=True)
    )
    if sel_paper_phase and sel_paper_phase not in allowed:
        sel_paper_phase = None
    completions = list(
        _bulk_approved_paper_completions_qs(
            fac_qs, df, dt, sel_paper_phase, exam_semester_ids=exam_sem_ids
        )
    )
    summaries, grand_credit, detail_map = _paper_summaries_and_detail_map(fac_list, completions)
    single_id = None
    if raw_id:
        try:
            single_id = int(raw_id)
        except (TypeError, ValueError):
            single_id = None
        if single_id:
            get_object_or_404(fac_qs, pk=single_id)
    scope_extra = ''
    if exam_sem_ids:
        scope_extra = ' · working semester(s) filter (paper phases)'
    scope_line = f'Institute scope (exam section / DVP){scope_extra} · decided-at: {df or "—"} → {dt or "—"}'
    wb = _build_paper_credit_excel_workbook(
        scope_line=scope_line,
        summaries=summaries,
        grand_credit=grand_credit,
        detail_map=detail_map,
        fac_ordered=fac_list,
        single_faculty_id=single_id,
    )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    if single_id:
        sf = get_object_or_404(fac_qs, pk=single_id)
        fname = f'paper_check_credits_{sf.short_name or sf.pk}_{sf.pk}.xlsx'
    else:
        fname = 'paper_check_credits_all_institute.xlsx'
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@login_required
def dept_exam_supervision_credit_analytics(request):
    if not (_dept_parent_only(request) or _dept_child_only(request)):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    fac_qs = _credit_analytics_faculty_qs(request)
    fac_list = list(fac_qs)
    selected = None
    supervision_detail_rows = []
    total_completed = 0
    summaries = []
    grand_total_sup = 0
    raw_id = (request.GET.get('faculty_id') or '').strip()
    df = (request.GET.get('completed_from') or request.GET.get('date_from') or '').strip()
    dt = (request.GET.get('completed_to') or request.GET.get('date_to') or '').strip()
    scope_label = 'your department'
    if _dept_parent_only(request):
        prof = _parent_profile(request)
        if prof and _is_hub_coordinator(prof):
            scope_label = 'invited departments & sub-units under your hub'
        elif prof and prof.department_id:
            scope_label = prof.department.name
    elif _dept_child_only(request):
        ch = _child_profile(request)
        if ch and ch.department_id:
            scope_label = ch.department.name
    sel_sup_phase = _parse_optional_int_param(request, 'supervision_phase_id')
    supervision_phase_choices: list = []
    allowed_sup_phase_ids: set[int] = set()
    if fac_list:
        phase_discovery = _apply_supervision_completed_date_filter(
            _supervision_completed_scope_qs(request), df, dt
        )
        allowed_sup_phase_ids = set(phase_discovery.values_list('phase_id', flat=True))
        supervision_phase_choices = list(
            SupervisionExamPhase.objects.filter(pk__in=allowed_sup_phase_ids).order_by('name')
        )
        if sel_sup_phase and sel_sup_phase not in allowed_sup_phase_ids:
            sel_sup_phase = None
        qs = _apply_supervision_completed_date_filter(
            _supervision_completed_scope_qs(request, sel_sup_phase), df, dt
        )
        duties = list(qs)
        summaries, grand_total_sup, detail_map = _supervision_summaries_and_detail_map(
            fac_list, duties
        )
        if raw_id:
            try:
                fid = int(raw_id)
            except (TypeError, ValueError):
                fid = None
            if fid:
                selected = get_object_or_404(fac_qs, pk=fid)
                supervision_detail_rows = detail_map.get(selected.pk, [])
                total_completed = next(
                    (s['total_completed'] for s in summaries if s['faculty'].pk == selected.pk),
                    0,
                )
    return render(
        request,
        'core/exam_dept/credit_analytics.html',
        {
            'institute_scope': False,
            'scope_label': scope_label,
            'analytics_mode': 'supervision',
            'faculty_list': fac_qs,
            'selected_faculty': selected,
            'supervision_detail_rows': supervision_detail_rows,
            'total_completed': total_completed,
            'supervision_summaries': summaries,
            'grand_total_supervision': grand_total_sup,
            'filter_date_from': df,
            'filter_date_to': dt,
            'supervision_phase_choices': supervision_phase_choices,
            'selected_supervision_phase_id': sel_sup_phase,
            'list_url': 'core:dept_exam_supervision_credit_analytics',
            'excel_url': 'core:dept_exam_supervision_credit_analytics_excel',
            'excel_query_all': _analytics_excel_query(
                df, dt, supervision_phase_id=sel_sup_phase
            ),
            'excel_query_selected': (
                _analytics_excel_query(
                    df, dt, selected.pk, supervision_phase_id=sel_sup_phase
                )
                if selected
                else None
            ),
        },
    )


@login_required
def dept_exam_supervision_credit_analytics_excel(request):
    if not (_dept_parent_only(request) or _dept_child_only(request)):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    fac_qs = _credit_analytics_faculty_qs(request)
    fac_list = list(fac_qs)
    raw_id = (request.GET.get('faculty_id') or '').strip()
    df = (request.GET.get('completed_from') or request.GET.get('date_from') or '').strip()
    dt = (request.GET.get('completed_to') or request.GET.get('date_to') or '').strip()
    if not fac_list:
        messages.error(request, 'No faculty in scope.')
        return redirect('core:dept_exam_supervision_credit_analytics')
    sel_sup_phase = _parse_optional_int_param(request, 'supervision_phase_id')
    allowed = set(
        _apply_supervision_completed_date_filter(
            _supervision_completed_scope_qs(request), df, dt
        ).values_list('phase_id', flat=True)
    )
    if sel_sup_phase and sel_sup_phase not in allowed:
        sel_sup_phase = None
    qs = _apply_supervision_completed_date_filter(
        _supervision_completed_scope_qs(request, sel_sup_phase), df, dt
    )
    duties = list(qs)
    summaries, grand_n, detail_map = _supervision_summaries_and_detail_map(fac_list, duties)
    single_id = None
    if raw_id:
        try:
            single_id = int(raw_id)
        except (TypeError, ValueError):
            single_id = None
        if single_id:
            get_object_or_404(fac_qs, pk=single_id)
    scope_line = f'Department supervision scope · completed-at: {df or "—"} → {dt or "—"}'
    wb = _build_supervision_credit_excel_workbook(
        scope_line=scope_line,
        summaries=summaries,
        grand_n=grand_n,
        detail_map=detail_map,
        fac_ordered=fac_list,
        single_faculty_id=single_id,
    )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    if single_id:
        sf = get_object_or_404(fac_qs, pk=single_id)
        fname = f'supervision_credits_{sf.short_name or sf.pk}_{sf.pk}.xlsx'
    else:
        fname = 'supervision_credits_all.xlsx'
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@login_required
def exam_section_supervision_credit_analytics(request):
    if not _exam_section_portal_access(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    fac_qs = Faculty.objects.select_related('department').order_by(
        'department__name', 'full_name'
    )
    exam_sem_ids = exam_section_working_semester_ids(request) if is_exam_section_operator(request) else None
    if exam_sem_ids:
        fac_qs = fac_qs.filter(department__institute_semester_id__in=exam_sem_ids)
    fac_list = list(fac_qs)
    selected = None
    supervision_detail_rows = []
    total_completed = 0
    summaries = []
    grand_total_sup = 0
    raw_id = (request.GET.get('faculty_id') or '').strip()
    df = (request.GET.get('completed_from') or request.GET.get('date_from') or '').strip()
    dt = (request.GET.get('completed_to') or request.GET.get('date_to') or '').strip()
    sel_sup_phase = _parse_optional_int_param(request, 'supervision_phase_id')
    supervision_phase_choices: list = []
    allowed_sup_phase_ids: set[int] = set()
    if fac_list:
        phase_discovery = _apply_supervision_completed_date_filter(
            _supervision_completed_institute_qs(
                exam_semester_ids=exam_sem_ids,
            ),
            df,
            dt,
        )
        allowed_sup_phase_ids = set(phase_discovery.values_list('phase_id', flat=True))
        supervision_phase_choices = list(
            SupervisionExamPhase.objects.filter(pk__in=allowed_sup_phase_ids).order_by('name')
        )
        if sel_sup_phase and sel_sup_phase not in allowed_sup_phase_ids:
            sel_sup_phase = None
        qs = _apply_supervision_completed_date_filter(
            _supervision_completed_institute_qs(
                sel_sup_phase,
                exam_semester_ids=exam_sem_ids,
            ),
            df,
            dt,
        )
        duties = list(qs)
        summaries, grand_total_sup, detail_map = _supervision_summaries_and_detail_map(
            fac_list, duties
        )
        if raw_id:
            try:
                fid = int(raw_id)
            except (TypeError, ValueError):
                fid = None
            if fid:
                selected = get_object_or_404(fac_qs, pk=fid)
                supervision_detail_rows = detail_map.get(selected.pk, [])
                total_completed = next(
                    (s['total_completed'] for s in summaries if s['faculty'].pk == selected.pk),
                    0,
                )
    scope_label_sup = 'All departments'
    if exam_sem_ids:
        scope_label_sup = 'Selected working academic semester(s) — faculty in matching departments'
    return render(
        request,
        'core/exam_dept/credit_analytics.html',
        {
            'institute_scope': True,
            'scope_label': scope_label_sup,
            'analytics_mode': 'supervision',
            'faculty_list': fac_qs,
            'selected_faculty': selected,
            'supervision_detail_rows': supervision_detail_rows,
            'total_completed': total_completed,
            'supervision_summaries': summaries,
            'grand_total_supervision': grand_total_sup,
            'filter_date_from': df,
            'filter_date_to': dt,
            'supervision_phase_choices': supervision_phase_choices,
            'selected_supervision_phase_id': sel_sup_phase,
            'list_url': 'core:exam_section_supervision_credit_analytics',
            'excel_url': 'core:exam_section_supervision_credit_analytics_excel',
            'excel_query_all': _analytics_excel_query(
                df, dt, supervision_phase_id=sel_sup_phase
            ),
            'excel_query_selected': (
                _analytics_excel_query(
                    df, dt, selected.pk, supervision_phase_id=sel_sup_phase
                )
                if selected
                else None
            ),
        },
    )


@login_required
def exam_section_supervision_credit_analytics_excel(request):
    if not _exam_section_portal_access(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    fac_qs = Faculty.objects.select_related('department').order_by(
        'department__name', 'full_name'
    )
    exam_sem_ids = exam_section_working_semester_ids(request) if is_exam_section_operator(request) else None
    if exam_sem_ids:
        fac_qs = fac_qs.filter(department__institute_semester_id__in=exam_sem_ids)
    fac_list = list(fac_qs)
    raw_id = (request.GET.get('faculty_id') or '').strip()
    df = (request.GET.get('completed_from') or request.GET.get('date_from') or '').strip()
    dt = (request.GET.get('completed_to') or request.GET.get('date_to') or '').strip()
    if not fac_list:
        messages.error(request, 'No faculty records.')
        return redirect('core:exam_section_supervision_credit_analytics')
    sel_sup_phase = _parse_optional_int_param(request, 'supervision_phase_id')
    allowed = set(
        _apply_supervision_completed_date_filter(
            _supervision_completed_institute_qs(exam_semester_ids=exam_sem_ids),
            df,
            dt,
        ).values_list('phase_id', flat=True)
    )
    if sel_sup_phase and sel_sup_phase not in allowed:
        sel_sup_phase = None
    qs = _apply_supervision_completed_date_filter(
        _supervision_completed_institute_qs(
            sel_sup_phase,
            exam_semester_ids=exam_sem_ids,
        ),
        df,
        dt,
    )
    duties = list(qs)
    summaries, grand_n, detail_map = _supervision_summaries_and_detail_map(fac_list, duties)
    single_id = None
    if raw_id:
        try:
            single_id = int(raw_id)
        except (TypeError, ValueError):
            single_id = None
        if single_id:
            get_object_or_404(fac_qs, pk=single_id)
    scope_extra = ''
    if exam_sem_ids:
        scope_extra = ' · working semester(s) filter (supervision phases)'
    scope_line = f'Institute supervision{scope_extra} · completed-at: {df or "—"} → {dt or "—"}'
    wb = _build_supervision_credit_excel_workbook(
        scope_line=scope_line,
        summaries=summaries,
        grand_n=grand_n,
        detail_map=detail_map,
        fac_ordered=fac_list,
        single_faculty_id=single_id,
    )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    if single_id:
        sf = get_object_or_404(fac_qs, pk=single_id)
        fname = f'supervision_credits_{sf.short_name or sf.pk}_{sf.pk}.xlsx'
    else:
        fname = 'supervision_credits_all_institute.xlsx'
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp
