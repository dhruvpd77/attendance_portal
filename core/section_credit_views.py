"""Section credit & remuneration report (institute + department coordinators)."""
from __future__ import annotations

from collections import defaultdict
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Prefetch, Q
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.exam_subunit_scope import division_code_match_q, division_codes_equivalent
from core.exam_coordination_views import (
    _child_profile,
    _dept_child_only,
    _dept_parent_only,
    _exam_section_portal_access,
    _is_hub_coordinator,
    _parent_profile,
)
from core.models import (
    Department,
    DepartmentExamProfile,
    Faculty,
    PaperCheckingAllocation,
    PaperCheckingCompletionRequest,
    PaperCheckingPhase,
    PaperSettingCompletionRequest,
    PaperSettingPhase,
    SupervisionDuty,
    SupervisionExamPhase,
)
from core.paper_duty_scope import (
    checking_phases_exam_section_scoped,
    setting_phases_exam_section_scoped,
)
from core.paper_duty_views import _checking_phases_for_user, _setting_phases_for_user
from core.semester_scope import (
    departments_for_exam_coordination_request,
    exam_section_working_semester_ids,
    is_exam_section_operator,
    q_completion_duty_phase_in_semesters,
    q_supervision_duty_phase_in_semesters,
)
from core.paper_checking_credits import (
    credit_for_completion_request,
    paper_count_for_completion,
    remuneration_for_completion_request,
)
from core.paper_duty_scope import hub_managed_department_ids
from core.paper_setting_credits import (
    credit_for_paper_setting_request,
    remuneration_for_paper_setting_request,
    supervision_credit_for_phase,
    supervision_remuneration_for_phase,
)


def _institute_paper_access(request):
    return _exam_section_portal_access(request)


def section_credit_report_access(request) -> bool:
    return (
        _dept_parent_only(request)
        or _dept_child_only(request)
        or _institute_paper_access(request)
    )


def _child_phase_filters(request):
    """Paper / setting / supervision phase dropdowns for sub-unit (dept_exam_child) logins."""
    prof = _child_profile(request)
    if not prof or not prof.department_id:
        return (
            PaperCheckingPhase.objects.none(),
            PaperSettingPhase.objects.none(),
            SupervisionExamPhase.objects.none(),
        )
    dept_id = prof.department_id
    parent = prof.parent
    hub_uid = parent.user_id if parent and not parent.department_id else None
    hub_sem = parent.institute_semester_id if parent and hub_uid else None
    q_phase = Q(department_id=dept_id) | Q(institute_scope=True)
    if hub_uid and hub_sem:
        q_phase |= Q(hub_coordinator_id=hub_uid, institute_semester_id=hub_sem)
    elif hub_uid:
        q_phase |= Q(hub_coordinator_id=hub_uid)
    paper_q = PaperCheckingPhase.objects.filter(q_phase).order_by('name')
    setting_q = PaperSettingPhase.objects.filter(q_phase).order_by('name')
    if hub_uid and hub_sem:
        sup_q = SupervisionExamPhase.objects.filter(
            hub_coordinator_id=hub_uid, institute_semester_id=hub_sem
        ).order_by('name')
    elif hub_uid:
        sup_q = SupervisionExamPhase.objects.filter(hub_coordinator_id=hub_uid).order_by('name')
    else:
        sup_q = SupervisionExamPhase.objects.filter(department_id=dept_id).order_by('name')
    return paper_q, setting_q, sup_q


def _child_supervision_extra_q(request):
    """Limit supervision rows to this sub-unit’s division (same scope as DR / supervision analytics)."""
    if not _dept_child_only(request):
        return None
    prof = _child_profile(request)
    if not prof:
        return None
    code = (prof.subunit_code or '').strip().upper()
    return division_code_match_q(code)


def _faculty_queryset(request, department_id: int | None):
    if _institute_paper_access(request):
        qs = Faculty.objects.select_related('department').order_by('full_name')
        if is_exam_section_operator(request):
            sem_ids = exam_section_working_semester_ids(request)
            if sem_ids:
                qs = qs.filter(department__institute_semester_id__in=sem_ids)
        if department_id:
            qs = qs.filter(department_id=department_id)
        return qs
    if _dept_child_only(request):
        prof = _child_profile(request)
        if not prof or not prof.department_id:
            return Faculty.objects.none()
        return (
            Faculty.objects.filter(department_id=prof.department_id)
            .select_related('department')
            .order_by('full_name')
        )
    prof = _parent_profile(request)
    if not prof:
        return Faculty.objects.none()
    if _is_hub_coordinator(prof):
        ids = hub_managed_department_ids(prof)
        qs = Faculty.objects.filter(department_id__in=ids).select_related('department').order_by(
            'full_name'
        )
        if department_id and department_id in ids:
            qs = qs.filter(department_id=department_id)
        return qs
    qs = Faculty.objects.filter(department_id=prof.department_id).select_related(
        'department'
    ).order_by('full_name')
    return qs


def _managed_scope_for_coordination(
    request,
) -> tuple[set[int] | None, list[Department]]:
    """Institute: (None, []). Coordinators: managed department id set + dropdown choices."""
    if _institute_paper_access(request):
        return None, []
    if _dept_child_only(request):
        prof = _child_profile(request)
        if not prof or not prof.department_id:
            return set(), []
        ids = {prof.department_id}
        return ids, list(Department.objects.filter(pk__in=ids).order_by('name'))
    prof = _parent_profile(request)
    if not prof:
        return set(), []
    if _is_hub_coordinator(prof):
        ids = hub_managed_department_ids(prof)
        return ids, list(Department.objects.filter(pk__in=ids).order_by('name'))
    if prof.department_id:
        ids = {prof.department_id}
        return ids, list(Department.objects.filter(pk__in=ids).order_by('name'))
    return set(), []


def _parse_duty_department_id(request, managed: set[int] | None) -> int | None:
    if managed is None:
        return None
    raw = (request.GET.get('duty_department_id') or '').strip()
    if not raw:
        return None
    try:
        v = int(raw)
    except (TypeError, ValueError):
        return None
    if v not in managed:
        return None
    return v


def _section_credit_excel_banner_lines(
    request, d0: date, d1: date, dept_f: int | None
) -> tuple[str, str]:
    """Main title (institution) and subtitle (scope + period), DR-style."""
    college = getattr(
        settings,
        'COLLEGE_DISPLAY_NAME',
        'L.J. Institute of Engineering and Technology',
    )
    if _institute_paper_access(request):
        if dept_f:
            dn = (
                Department.objects.filter(pk=dept_f).values_list('name', flat=True).first()
                or 'Department'
            )
            scope = f'Section credit report — {dn}'
        else:
            scope = 'Section credit report — All departments'
    elif _dept_child_only(request):
        cp = _child_profile(request)
        sub = (cp.subunit_code or '').strip() if cp else ''
        dname = cp.department.name if cp and cp.department else ''
        scope = f'Section credit report — Sub-unit {sub or "—"} — {dname}'
    else:
        prof = _parent_profile(request)
        if prof and _is_hub_coordinator(prof):
            scope = 'Section credit report — Hub coordinator scope'
        elif prof and prof.department:
            scope = f'Section credit report — {prof.department.name}'
        else:
            scope = 'Section credit report'
    sub = (
        f'{scope}  |  Period: {d0.strftime("%d-%b-%Y")} to {d1.strftime("%d-%b-%Y")}'
    )
    return college, sub


def _write_section_credit_workbook(ws, rows: list[dict], line1: str, line2: str) -> None:
    """Merged banners, grouped header row, sub-headers, borders — aligned with DR / supervision exports."""
    ncols = 13
    thin = Side(style='thin', color='555555')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ctr_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    title_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    title_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    sub_fill = PatternFill(start_color='2D4F76', end_color='2D4F76', fill_type='solid')
    sub_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

    grp_id_fill = PatternFill(start_color='D6DCE5', end_color='D6DCE5', fill_type='solid')
    grp_paper_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    grp_sup_fill = PatternFill(start_color='E2F0D9', end_color='E2F0D9', fill_type='solid')
    grp_set_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    grp_total_fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')
    grp_font = Font(name='Calibri', size=10, bold=True, color='000000')
    grp_total_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')

    hdr_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    hdr_font = Font(name='Calibri', size=10, bold=True)
    data_font = Font(name='Calibri', size=10)
    num_align = Alignment(horizontal='right', vertical='center')

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c1 = ws.cell(row=1, column=1, value=line1)
    c1.font = title_font
    c1.fill = title_fill
    c1.alignment = ctr
    for c in range(1, ncols + 1):
        ws.cell(row=1, column=c).border = border

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c2 = ws.cell(row=2, column=1, value=line2)
    c2.font = sub_font
    c2.fill = sub_fill
    c2.alignment = ctr
    for c in range(1, ncols + 1):
        ws.cell(row=2, column=c).border = border

    gr = 3
    ws.merge_cells(start_row=gr, start_column=1, end_row=gr, end_column=2)
    g1 = ws.cell(row=gr, column=1, value='Faculty & department')
    g1.font = grp_font
    g1.fill = grp_id_fill
    g1.alignment = ctr
    for c in (1, 2):
        ws.cell(row=gr, column=c).border = border

    for start_c, end_c, label, fill in (
        (3, 5, 'Paper checking', grp_paper_fill),
        (6, 8, 'Supervision', grp_sup_fill),
        (9, 11, 'Paper setting', grp_set_fill),
    ):
        ws.merge_cells(start_row=gr, start_column=start_c, end_row=gr, end_column=end_c)
        gc = ws.cell(row=gr, column=start_c, value=label)
        gc.font = grp_font
        gc.fill = fill
        gc.alignment = ctr
        for c in range(start_c, end_c + 1):
            ws.cell(row=gr, column=c).border = border

    for col_g, gtitle in (
        (12, 'Total credit'),
        (13, 'Total rupees ₹'),
    ):
        gt = ws.cell(row=gr, column=col_g, value=gtitle)
        gt.font = grp_total_font
        gt.fill = grp_total_fill
        gt.alignment = ctr
        ws.cell(row=gr, column=col_g).border = border

    hdr = [
        'Faculty',
        'Department',
        'Paper checking count',
        'Paper checking credit',
        'Paper checking rupees ₹',
        'No. of supervision',
        'Supervision credit',
        'Supervision rupees ₹',
        'No. paper setting',
        'Paper setting credit',
        'Paper setting rupees ₹',
        'Total credit',
        'Total rupees ₹',
    ]
    hr = 4
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = border
        cell.alignment = ctr if c >= 3 else Alignment(horizontal='center', vertical='center', wrap_text=True)

    band_a = PatternFill(start_color='FAFBFC', end_color='FAFBFC', fill_type='solid')
    band_b = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    rno = hr + 1
    for i, row in enumerate(rows):
        f = row['faculty']
        fill = band_a if i % 2 == 0 else band_b
        ws.cell(row=rno, column=1, value=f.full_name).font = data_font
        ws.cell(row=rno, column=2, value=f.department.name if f.department_id else '').font = data_font
        ws.cell(row=rno, column=3, value=row['paper_n']).font = data_font
        ws.cell(row=rno, column=4, value=float(row['paper_credits'])).font = data_font
        ws.cell(row=rno, column=5, value=float(row['paper_rem'])).font = data_font
        ws.cell(row=rno, column=6, value=row['sup_count']).font = data_font
        ws.cell(row=rno, column=7, value=float(row['sup_credits'])).font = data_font
        ws.cell(row=rno, column=8, value=float(row['sup_rem'])).font = data_font
        ws.cell(row=rno, column=9, value=row['setting_n']).font = data_font
        ws.cell(row=rno, column=10, value=float(row['setting_credits'])).font = data_font
        ws.cell(row=rno, column=11, value=float(row['setting_rem'])).font = data_font
        ws.cell(row=rno, column=12, value=float(row['total_credits'])).font = data_font
        ws.cell(row=rno, column=13, value=float(row['total_rem'])).font = data_font
        for c in range(1, 14):
            cell = ws.cell(row=rno, column=c)
            cell.border = border
            cell.fill = fill
            if c >= 3:
                cell.alignment = num_align
            else:
                cell.alignment = ctr_left
        rno += 1

    tot_row = rno
    sum_papers = sum(row['paper_n'] for row in rows)
    sum_pc = sum((row['paper_credits'] for row in rows), Decimal('0'))
    sum_pr = sum((row['paper_rem'] for row in rows), Decimal('0'))
    sum_sc = sum(row['sup_count'] for row in rows)
    sum_suc = sum((row['sup_credits'] for row in rows), Decimal('0'))
    sum_sur = sum((row['sup_rem'] for row in rows), Decimal('0'))
    sum_stn = sum(row['setting_n'] for row in rows)
    sum_stc = sum((row['setting_credits'] for row in rows), Decimal('0'))
    sum_str = sum((row['setting_rem'] for row in rows), Decimal('0'))
    sum_tcr = sum((row['total_credits'] for row in rows), Decimal('0'))
    sum_total = sum((row['total_rem'] for row in rows), Decimal('0'))

    total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    total_font = Font(name='Calibri', size=10, bold=True)
    lbl = ws.cell(row=tot_row, column=1, value='Totals')
    lbl.font = total_font
    lbl.fill = total_fill
    lbl.border = border
    ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=2)
    ws.cell(row=tot_row, column=2).border = border
    ws.cell(row=tot_row, column=2).fill = total_fill

    vals = [
        None,
        None,
        sum_papers,
        float(sum_pc),
        float(sum_pr),
        sum_sc,
        float(sum_suc),
        float(sum_sur),
        sum_stn,
        float(sum_stc),
        float(sum_str),
        float(sum_tcr),
        float(sum_total),
    ]
    for c in range(3, 14):
        cell = ws.cell(row=tot_row, column=c, value=vals[c - 1])
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        cell.alignment = num_align

    widths = (22, 14, 10, 12, 12, 10, 12, 12, 10, 12, 12, 12, 14)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 28


def _write_overall_line_detail_sheet(ws, rows: list[dict]) -> None:
    """Second sheet: every line item with pay department (overall faculty export)."""
    thin = Side(style='thin', color='555555')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    hdr_font = Font(name='Calibri', size=10, bold=True)
    data_font = Font(name='Calibri', size=10)
    num_align = Alignment(horizontal='right', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    hdr = [
        'Faculty',
        'Home department',
        'Type',
        'Subject',
        'Phase',
        'Date',
        'Slot',
        'Papers',
        'Credit',
        '₹',
        'Work for (department)',
    ]
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    rno = 2
    for block in rows:
        f = block['faculty']
        home = f.department.name if f.department_id else ''
        for ln in block.get('line_paper') or []:
            ws.cell(row=rno, column=1, value=f.full_name).font = data_font
            ws.cell(row=rno, column=2, value=home).font = data_font
            ws.cell(row=rno, column=3, value='Paper checking').font = data_font
            ws.cell(row=rno, column=4, value=ln.get('subject', '')).font = data_font
            ws.cell(row=rno, column=5, value=ln.get('phase', '')).font = data_font
            ddt = ln.get('decided')
            ws.cell(row=rno, column=6, value=ddt.strftime('%d-%b-%Y') if ddt else '').font = data_font
            ws.cell(row=rno, column=7, value='').font = data_font
            ws.cell(row=rno, column=8, value=ln.get('papers', '')).font = data_font
            ws.cell(row=rno, column=9, value=float(ln.get('credit') or 0)).font = data_font
            ws.cell(row=rno, column=10, value=float(ln.get('rem') or 0)).font = data_font
            ws.cell(row=rno, column=11, value=ln.get('work_for', '')).font = data_font
            for c in range(1, 12):
                cell = ws.cell(row=rno, column=c)
                cell.border = border
                cell.alignment = num_align if c >= 8 else left_align
            rno += 1
        for ln in block.get('line_sup') or []:
            ws.cell(row=rno, column=1, value=f.full_name).font = data_font
            ws.cell(row=rno, column=2, value=home).font = data_font
            ws.cell(row=rno, column=3, value='Supervision').font = data_font
            ws.cell(row=rno, column=4, value=ln.get('subject', '')).font = data_font
            ws.cell(row=rno, column=5, value=ln.get('phase', '')).font = data_font
            sdt = ln.get('date')
            ws.cell(row=rno, column=6, value=sdt.strftime('%d-%b-%Y') if sdt else '').font = data_font
            ws.cell(row=rno, column=7, value=str(ln.get('slot') or '')).font = data_font
            ws.cell(row=rno, column=8, value='').font = data_font
            ws.cell(row=rno, column=9, value=float(ln.get('credit') or 0)).font = data_font
            ws.cell(row=rno, column=10, value=float(ln.get('rem') or 0)).font = data_font
            ws.cell(row=rno, column=11, value=ln.get('work_for', '')).font = data_font
            for c in range(1, 12):
                cell = ws.cell(row=rno, column=c)
                cell.border = border
                cell.alignment = num_align if c >= 8 else left_align
            rno += 1
        for ln in block.get('line_set') or []:
            ws.cell(row=rno, column=1, value=f.full_name).font = data_font
            ws.cell(row=rno, column=2, value=home).font = data_font
            ws.cell(row=rno, column=3, value='Paper setting').font = data_font
            ws.cell(row=rno, column=4, value=ln.get('subject', '')).font = data_font
            ws.cell(row=rno, column=5, value=ln.get('phase', '')).font = data_font
            ddt = ln.get('decided')
            ws.cell(row=rno, column=6, value=ddt.strftime('%d-%b-%Y') if ddt else '').font = data_font
            ws.cell(row=rno, column=7, value='').font = data_font
            ws.cell(row=rno, column=8, value='').font = data_font
            ws.cell(row=rno, column=9, value=float(ln.get('credit') or 0)).font = data_font
            ws.cell(row=rno, column=10, value=float(ln.get('rem') or 0)).font = data_font
            ws.cell(row=rno, column=11, value=ln.get('work_for', '')).font = data_font
            for c in range(1, 12):
                cell = ws.cell(row=rno, column=c)
                cell.border = border
                cell.alignment = num_align if c >= 8 else left_align
            rno += 1

    widths = (22, 18, 14, 24, 10, 12, 10, 8, 10, 10, 22)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _parse_date(s: str | None) -> date | None:
    if not (s or '').strip():
        return None
    try:
        y, m, d = (int(x) for x in s.strip().split('-')[:3])
        return date(y, m, d)
    except (TypeError, ValueError):
        return None


def _work_dept_id_paper(r: PaperCheckingCompletionRequest) -> int | None:
    duty = r.duty
    ph = duty.phase
    if ph and ph.department_id:
        return ph.department_id
    for a in duty.allocations.all():
        if a.department_id:
            return a.department_id
    return None


def _work_dept_id_sup(d: SupervisionDuty, managed_ids: set[int]) -> int | None:
    ph = d.phase
    if ph.department_id:
        return ph.department_id
    code = (d.division_code or '').strip()
    if not code:
        return None
    for did in managed_ids:
        dn = Department.objects.filter(pk=did).values_list('name', flat=True).first()
        if dn and division_codes_equivalent(str(dn), code):
            return did
    for p in DepartmentExamProfile.objects.filter(department_id__in=managed_ids).exclude(
        subunit_code=''
    ):
        if division_codes_equivalent((p.subunit_code or '').strip(), code):
            return p.department_id
    return None


def _work_dept_id_sup_any(d: SupervisionDuty) -> int | None:
    """Resolve supervision duty stream to a department (any division in the institute)."""
    ph = d.phase
    if ph.department_id:
        return ph.department_id
    code = (d.division_code or '').strip()
    if not code:
        return None
    for did, dn in Department.objects.values_list('id', 'name'):
        if dn and division_codes_equivalent(str(dn), code):
            return int(did)
    for p in DepartmentExamProfile.objects.exclude(subunit_code='').only(
        'department_id', 'subunit_code'
    ):
        if division_codes_equivalent((p.subunit_code or '').strip(), code):
            return p.department_id
    return None


def _work_dept_id_setting(r: PaperSettingCompletionRequest) -> int | None:
    ph = r.duty.phase
    return ph.department_id if ph and ph.department_id else None


def _dept_name(wid: int | None) -> str:
    if not wid:
        return '—'
    return Department.objects.filter(pk=wid).values_list('name', flat=True).first() or '—'


def _collect_duty_scoped_buckets(
    managed_ids: set[int],
    duty_dept_filter: int | None,
    d0: date,
    d1: date,
    *,
    paper_phase_id: int | None,
    sup_phase_id: int | None,
    setting_phase_id: int | None,
    supervision_extra_q: Q | None,
    setting_hub_coordinator_id: int | None = None,
    hub_scope_user_id: int | None = None,
    hub_scope_semester_id: int | None = None,
) -> tuple[dict[int, list], dict[int, list], dict[int, list]]:
    pc_by_f: dict[int, list] = defaultdict(list)
    pc_q = PaperCheckingCompletionRequest.objects.filter(
        status=PaperCheckingCompletionRequest.APPROVED,
        decided_at__date__gte=d0,
        decided_at__date__lte=d1,
    )
    if paper_phase_id:
        pc_q = pc_q.filter(duty__phase_id=paper_phase_id)
    if hub_scope_user_id and hub_scope_semester_id:
        pc_q = pc_q.filter(
            ~Q(duty__phase__hub_coordinator_id=hub_scope_user_id)
            | Q(duty__phase__institute_semester_id=hub_scope_semester_id)
        )
    pc_q = pc_q.select_related('duty', 'duty__phase', 'faculty', 'faculty__department').prefetch_related(
        Prefetch(
            'duty__allocations',
            queryset=PaperCheckingAllocation.objects.select_related('department'),
        ),
    )
    for r in pc_q:
        if not r.faculty_id:
            continue
        wid = _work_dept_id_paper(r)
        if not wid or wid not in managed_ids:
            continue
        if duty_dept_filter and wid != duty_dept_filter:
            continue
        pc_by_f[r.faculty_id].append(r)

    sup_by_f: dict[int, list] = defaultdict(list)
    sup_q = SupervisionDuty.objects.filter(
        completion_status=SupervisionDuty.COMPLETED,
    ).select_related('phase', 'faculty', 'faculty__department')
    sup_q = sup_q.filter(
        Q(completed_at__date__gte=d0, completed_at__date__lte=d1)
        | Q(completed_at__isnull=True, supervision_date__gte=d0, supervision_date__lte=d1)
    )
    if sup_phase_id:
        sup_q = sup_q.filter(phase_id=sup_phase_id)
    if supervision_extra_q:
        sup_q = sup_q.filter(supervision_extra_q)
    if hub_scope_user_id and hub_scope_semester_id:
        sup_q = sup_q.filter(
            ~Q(phase__hub_coordinator_id=hub_scope_user_id)
            | Q(phase__institute_semester_id=hub_scope_semester_id)
        )
    for d in sup_q:
        if not d.faculty_id:
            continue
        wid = _work_dept_id_sup(d, managed_ids)
        if not wid or wid not in managed_ids:
            continue
        if duty_dept_filter and wid != duty_dept_filter:
            continue
        sup_by_f[d.faculty_id].append(d)

    st_by_f: dict[int, list] = defaultdict(list)
    st_q = PaperSettingCompletionRequest.objects.filter(
        status=PaperSettingCompletionRequest.APPROVED,
        decided_at__date__gte=d0,
        decided_at__date__lte=d1,
    ).select_related('duty', 'duty__phase', 'faculty', 'faculty__department')
    if setting_phase_id:
        st_q = st_q.filter(duty__phase_id=setting_phase_id)
    for r in st_q:
        if not r.faculty_id:
            continue
        ph = r.duty.phase
        if not ph:
            continue
        wid = ph.department_id

        in_scope = False
        if wid and wid in managed_ids:
            in_scope = True
        elif setting_hub_coordinator_id is not None and ph.hub_coordinator_id == setting_hub_coordinator_id:
            if (
                hub_scope_semester_id is not None
                and ph.institute_semester_id != hub_scope_semester_id
            ):
                in_scope = False
            else:
                in_scope = True
        elif ph.institute_scope and r.faculty.department_id in managed_ids:
            # Institute-wide phase: count rows for faculty in managed departments (e.g. sub-units).
            in_scope = True

        if not in_scope:
            continue

        if duty_dept_filter:
            if wid:
                if wid != duty_dept_filter:
                    continue
            elif setting_hub_coordinator_id and ph.hub_coordinator_id == setting_hub_coordinator_id:
                if r.faculty.department_id != duty_dept_filter:
                    continue
            elif ph.institute_scope:
                if r.faculty.department_id != duty_dept_filter:
                    continue
            else:
                continue

        st_by_f[r.faculty_id].append(r)

    return pc_by_f, sup_by_f, st_by_f


def _aggregate_from_buckets(
    faculty_list: list[Faculty],
    pc_by_f: dict[int, list],
    sup_by_f: dict[int, list],
    st_by_f: dict[int, list],
    *,
    managed_ids: set[int] | None,
    include_work_for: bool,
) -> tuple[list[dict], list[str], list[str]]:
    sup_phases_set: set[str] = set()
    set_phases_set: set[str] = set()
    for rows in sup_by_f.values():
        for d in rows:
            if d.phase:
                sup_phases_set.add(d.phase.name or f'#{d.phase_id}')
    for rows in st_by_f.values():
        for r in rows:
            if r.duty.phase:
                set_phases_set.add(r.duty.phase.name or f'#{r.duty.phase_id}')
    sup_phases_list = sorted(sup_phases_set)
    set_phases_list = sorted(set_phases_set)

    out: list[dict] = []
    for fac in faculty_list:
        fid = fac.pk
        pc_rows = pc_by_f.get(fid, [])
        sup_rows = sup_by_f.get(fid, [])
        st_rows = st_by_f.get(fid, [])

        p_cr = sum((credit_for_completion_request(r) for r in pc_rows), Decimal('0'))
        p_rm = sum((remuneration_for_completion_request(r) for r in pc_rows), Decimal('0'))

        sup_by_phase: dict[str, list] = defaultdict(list)
        for d in sup_rows:
            pn = d.phase.name if d.phase else ''
            sup_by_phase[pn or '?'].append(d)
        s_cr = Decimal('0')
        s_rm = Decimal('0')
        sup_phase_detail = []
        for pname, duties in sorted(sup_by_phase.items()):
            cr_tot = Decimal('0')
            rm_tot = Decimal('0')
            for d in duties:
                wd_id = fac.department_id
                if managed_ids is not None:
                    wd_id = _work_dept_id_sup(d, managed_ids) or wd_id
                else:
                    wd_id = _work_dept_id_sup_any(d) or wd_id
                cr_tot += supervision_credit_for_phase(wd_id, pname)
                rm_tot += supervision_remuneration_for_phase(wd_id, pname)
            n = len(duties)
            ce = cr_tot / n if n else Decimal('0')
            re = rm_tot / n if n else Decimal('0')
            s_cr += cr_tot
            s_rm += rm_tot
            sup_phase_detail.append(
                {
                    'phase': pname,
                    'count': n,
                    'credit_each': ce,
                    'credit_total': cr_tot,
                    'rem_each': re,
                    'rem_total': rm_tot,
                }
            )

        st_by_phase: dict[str, list] = defaultdict(list)
        for r in st_rows:
            pn = r.duty.phase.name if r.duty.phase else ''
            st_by_phase[pn or '?'].append(r)
        t_cr = Decimal('0')
        t_rm = Decimal('0')
        set_phase_detail = []
        for pname, reqs in sorted(st_by_phase.items()):
            cr_sum = sum((credit_for_paper_setting_request(r) for r in reqs), Decimal('0'))
            rm_sum = sum((remuneration_for_paper_setting_request(r) for r in reqs), Decimal('0'))
            t_cr += cr_sum
            t_rm += rm_sum
            set_phase_detail.append(
                {'phase': pname, 'count': len(reqs), 'credit_total': cr_sum, 'rem_total': rm_sum}
            )

        line_paper = []
        for r in pc_rows:
            duty = r.duty
            ln = {
                'subject': duty.subject_name,
                'phase': duty.phase.name if duty.phase else '',
                'papers': paper_count_for_completion(r),
                'credit': credit_for_completion_request(r),
                'rem': remuneration_for_completion_request(r),
                'decided': r.decided_at,
            }
            if include_work_for:
                ln['work_for'] = _dept_name(_work_dept_id_paper(r))
            line_paper.append(ln)
        line_sup = []
        for d in sup_rows:
            pn = d.phase.name if d.phase else ''
            wd_id = fac.department_id
            if managed_ids is not None:
                wd_id = _work_dept_id_sup(d, managed_ids) or wd_id
            else:
                wd_id = _work_dept_id_sup_any(d) or wd_id
            cr_o = supervision_credit_for_phase(wd_id, pn)
            rm_o = supervision_remuneration_for_phase(wd_id, pn)
            ln = {
                'date': d.supervision_date,
                'subject': d.subject_name,
                'phase': pn,
                'slot': d.time_slot,
                'credit': cr_o,
                'rem': rm_o,
            }
            if include_work_for:
                ln['work_for'] = _dept_name(wd_id)
            line_sup.append(ln)
        line_set = []
        for r in st_rows:
            ln = {
                'subject': r.duty.subject_name,
                'phase': r.duty.phase.name if r.duty.phase else '',
                'credit': credit_for_paper_setting_request(r),
                'rem': remuneration_for_paper_setting_request(r),
                'decided': r.decided_at,
            }
            if include_work_for:
                ln['work_for'] = _dept_name(_work_dept_id_setting(r))
            line_set.append(ln)

        total_rm = p_rm + s_rm + t_rm
        total_cr = p_cr + s_cr + t_cr
        paper_papers = sum((paper_count_for_completion(r) for r in pc_rows), 0)
        out.append(
            {
                'faculty': fac,
                'paper_credits': p_cr,
                'paper_rem': p_rm,
                'paper_n': paper_papers,
                'sup_count': len(sup_rows),
                'sup_credits': s_cr,
                'sup_rem': s_rm,
                'setting_credits': t_cr,
                'setting_rem': t_rm,
                'setting_n': len(st_rows),
                'total_credits': total_cr,
                'total_rem': total_rm,
                'sup_phase_detail': sup_phase_detail,
                'set_phase_detail': set_phase_detail,
                'line_paper': line_paper,
                'line_sup': line_sup,
                'line_set': line_set,
            }
        )

    return out, sup_phases_list, set_phases_list


def _fetch_faculty_duty_buckets(
    fac_ids: list[int],
    d0: date,
    d1: date,
    *,
    paper_phase_id: int | None,
    sup_phase_id: int | None,
    setting_phase_id: int | None,
    supervision_extra_q: Q | None = None,
    exam_semester_ids: list[int] | None = None,
) -> tuple[dict[int, list], dict[int, list], dict[int, list]]:
    if not fac_ids:
        return {}, {}, {}

    pc_base = PaperCheckingCompletionRequest.objects.filter(
        faculty_id__in=fac_ids,
        status=PaperCheckingCompletionRequest.APPROVED,
        decided_at__date__gte=d0,
        decided_at__date__lte=d1,
    ).select_related('duty', 'duty__phase', 'faculty', 'faculty__department')
    if exam_semester_ids:
        pc_base = pc_base.filter(q_completion_duty_phase_in_semesters(exam_semester_ids))
    if paper_phase_id:
        pc_base = pc_base.filter(duty__phase_id=paper_phase_id)
    pc_base = pc_base.prefetch_related(
        Prefetch(
            'duty__allocations',
            queryset=PaperCheckingAllocation.objects.select_related('department'),
        ),
    )
    pc_by_f: dict[int, list] = defaultdict(list)
    for r in pc_base:
        pc_by_f[r.faculty_id].append(r)

    sup_base = SupervisionDuty.objects.filter(
        faculty_id__in=fac_ids,
        completion_status=SupervisionDuty.COMPLETED,
    ).select_related('phase', 'faculty', 'faculty__department')
    sup_base = sup_base.filter(
        Q(completed_at__date__gte=d0, completed_at__date__lte=d1)
        | Q(completed_at__isnull=True, supervision_date__gte=d0, supervision_date__lte=d1)
    )
    if sup_phase_id:
        sup_base = sup_base.filter(phase_id=sup_phase_id)
    if supervision_extra_q:
        sup_base = sup_base.filter(supervision_extra_q)
    if exam_semester_ids:
        sup_base = sup_base.filter(q_supervision_duty_phase_in_semesters(exam_semester_ids))
    sup_by_f: dict[int, list] = defaultdict(list)
    for d in sup_base:
        sup_by_f[d.faculty_id].append(d)

    st_base = PaperSettingCompletionRequest.objects.filter(
        faculty_id__in=fac_ids,
        status=PaperSettingCompletionRequest.APPROVED,
        decided_at__date__gte=d0,
        decided_at__date__lte=d1,
    ).select_related('duty', 'duty__phase', 'faculty', 'faculty__department')
    if exam_semester_ids:
        st_base = st_base.filter(q_completion_duty_phase_in_semesters(exam_semester_ids))
    if setting_phase_id:
        st_base = st_base.filter(duty__phase_id=setting_phase_id)
    st_by_f: dict[int, list] = defaultdict(list)
    for r in st_base:
        st_by_f[r.faculty_id].append(r)

    return pc_by_f, sup_by_f, st_by_f


def _build_faculty_report(
    faculty_list: list[Faculty],
    d0: date,
    d1: date,
    *,
    paper_phase_id: int | None,
    sup_phase_id: int | None,
    setting_phase_id: int | None,
    supervision_extra_q: Q | None = None,
    exam_semester_ids: list[int] | None = None,
) -> tuple[list[dict], list[str], list[str]]:
    fac_ids = [f.pk for f in faculty_list]
    if not fac_ids:
        return [], [], []
    pc_by_f, sup_by_f, st_by_f = _fetch_faculty_duty_buckets(
        fac_ids,
        d0,
        d1,
        paper_phase_id=paper_phase_id,
        sup_phase_id=sup_phase_id,
        setting_phase_id=setting_phase_id,
        supervision_extra_q=supervision_extra_q,
        exam_semester_ids=exam_semester_ids,
    )
    return _aggregate_from_buckets(
        faculty_list,
        pc_by_f,
        sup_by_f,
        st_by_f,
        managed_ids=None,
        include_work_for=False,
    )


def _daily_dr_block_for_section_credit(request) -> dict:
    """URLs and copy for the institute-style daily DR download (exam.xlsx) on the section credit page."""
    if _institute_paper_access(request):
        return {
            'daily_dr_url': 'core:exam_section_daily_dr_excel',
            'daily_dr_blurb': (
                'Official institute daily DR (<strong>exam.xlsx</strong>). Same export as the exam section dashboard: '
                'choose <strong>From</strong> (and optional <strong>To</strong>) — sheets include days with supervision or approved paper-check / paper-setting fills.'
            ),
        }
    if _dept_child_only(request):
        return {
            'daily_dr_url': 'core:dept_exam_daily_dr_excel',
            'daily_dr_blurb': (
                'Daily DR in the official layout, scoped to your <strong>sub-unit</strong> (same as the department dashboard download).'
            ),
        }
    prof = _parent_profile(request)
    if prof and _is_hub_coordinator(prof):
        return {
            'daily_dr_url': 'core:dept_exam_daily_dr_excel',
            'daily_dr_blurb': (
                'Workbook lists all faculty like the main template, but only <strong>this hub’s</strong> completed supervision and approved '
                'paper-check / paper-setting credits are filled in. The exam section uses the same layout with every hub combined.'
            ),
        }
    if prof and prof.department:
        return {
            'daily_dr_url': 'core:dept_exam_daily_dr_excel',
            'daily_dr_blurb': (
                f'Same <strong>exam.xlsx</strong> layout, restricted to <strong>{prof.department.name}</strong>.'
            ),
        }
    return {
        'daily_dr_url': 'core:dept_exam_daily_dr_excel',
        'daily_dr_blurb': 'Official daily DR export for your coordinator scope.',
    }


def _section_credit_rows(
    request,
    d0: date,
    d1: date,
    dept_f: int | None,
    paper_phase_id: int | None,
    sup_phase_id: int | None,
    setting_phase_id: int | None,
) -> tuple[list[dict], list[str], list[str]]:
    """Institute: home-department faculty + all their duties. Coordinators: every assignee for managed exam streams."""
    sup_xq = _child_supervision_extra_q(request)
    managed, _duty_dept_choices = _managed_scope_for_coordination(request)
    duty_department_id = _parse_duty_department_id(request, managed)

    if managed is None:
        fac_qs = _faculty_queryset(request, dept_f)
        faculty_list = list(fac_qs)
        exam_sids = (
            exam_section_working_semester_ids(request) if is_exam_section_operator(request) else None
        )
        return _build_faculty_report(
            faculty_list,
            d0,
            d1,
            paper_phase_id=paper_phase_id,
            sup_phase_id=sup_phase_id,
            setting_phase_id=setting_phase_id,
            supervision_extra_q=sup_xq,
            exam_semester_ids=exam_sids,
        )

    if not managed:
        return [], [], []

    hub_uid = None
    hub_sem = None
    if _dept_parent_only(request):
        pr = _parent_profile(request)
        if pr and _is_hub_coordinator(pr):
            hub_uid = request.user.pk
            hub_sem = pr.institute_semester_id

    pc_by_f, sup_by_f, st_by_f = _collect_duty_scoped_buckets(
        managed,
        duty_department_id,
        d0,
        d1,
        paper_phase_id=paper_phase_id,
        sup_phase_id=sup_phase_id,
        setting_phase_id=setting_phase_id,
        supervision_extra_q=sup_xq,
        setting_hub_coordinator_id=hub_uid,
        hub_scope_user_id=hub_uid,
        hub_scope_semester_id=hub_sem,
    )
    fac_ids = set(pc_by_f) | set(sup_by_f) | set(st_by_f)
    faculty_list = list(
        Faculty.objects.filter(pk__in=fac_ids)
        .select_related('department')
        .order_by('full_name')
    )
    return _aggregate_from_buckets(
        faculty_list,
        pc_by_f,
        sup_by_f,
        st_by_f,
        managed_ids=managed,
        include_work_for=False,
    )


def _overall_faculty_rows(
    request,
    d0: date,
    d1: date,
    dept_f: int | None,
    paper_phase_id: int | None,
    sup_phase_id: int | None,
    setting_phase_id: int | None,
) -> tuple[list[dict], list[str], list[str]]:
    """Coordinator hub / child / parent: home faculties only; all duties with work_for labels. Institute: optional home-dept filter."""
    sup_xq = _child_supervision_extra_q(request)
    fac_qs = _faculty_queryset(request, dept_f if _institute_paper_access(request) else None)
    faculty_list = list(fac_qs)
    fac_ids = [f.pk for f in faculty_list]
    if not fac_ids:
        return [], [], []
    exam_sids = (
        exam_section_working_semester_ids(request) if is_exam_section_operator(request) else None
    )
    pc_by_f, sup_by_f, st_by_f = _fetch_faculty_duty_buckets(
        fac_ids,
        d0,
        d1,
        paper_phase_id=paper_phase_id,
        sup_phase_id=sup_phase_id,
        setting_phase_id=setting_phase_id,
        supervision_extra_q=sup_xq,
        exam_semester_ids=exam_sids,
    )
    return _aggregate_from_buckets(
        faculty_list,
        pc_by_f,
        sup_by_f,
        st_by_f,
        managed_ids=None,
        include_work_for=True,
    )


@login_required
def section_credit_report(request):
    if not section_credit_report_access(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')

    today = timezone.localdate()
    d1 = _parse_date(request.GET.get('date_to')) or today
    d0 = _parse_date(request.GET.get('date_from')) or (d1 - timedelta(days=30))
    if d0 > d1:
        d0, d1 = d1, d0

    raw_dept = (request.GET.get('department_id') or '').strip()
    dept_f = None
    if raw_dept and _institute_paper_access(request):
        try:
            dept_f = int(raw_dept)
        except (TypeError, ValueError):
            dept_f = None

    def _opt_int(name: str) -> int | None:
        raw = (request.GET.get(name) or '').strip()
        if not raw:
            return None
        try:
            return int(raw)
        except (TypeError, ValueError):
            return None

    paper_phase_id = _opt_int('paper_phase_id')
    sup_phase_id = _opt_int('sup_phase_id')
    setting_phase_id = _opt_int('setting_phase_id')

    managed_scope, duty_dept_choices = _managed_scope_for_coordination(request)
    duty_department_id = _parse_duty_department_id(request, managed_scope)

    rows, sup_ph_list, set_ph_list = _section_credit_rows(
        request,
        d0,
        d1,
        dept_f,
        paper_phase_id,
        sup_phase_id,
        setting_phase_id,
    )

    dept_choices = []
    if _institute_paper_access(request):
        if is_exam_section_operator(request):
            dept_choices = list(departments_for_exam_coordination_request(request).order_by('name'))
        else:
            dept_choices = list(Department.objects.order_by('name'))

    if _institute_paper_access(request):
        if is_exam_section_operator(request):
            sids = exam_section_working_semester_ids(request)
            paper_phases_q = checking_phases_exam_section_scoped(request)
            setting_phases_q = setting_phases_exam_section_scoped(request)
            sup_phases_q = SupervisionExamPhase.objects.filter(
                institute_semester_id__in=sids,
            ).order_by('name')
        else:
            paper_phases_q = PaperCheckingPhase.objects.order_by('name')
            setting_phases_q = PaperSettingPhase.objects.order_by('name')
            sup_phases_q = SupervisionExamPhase.objects.order_by('name')
    elif _dept_child_only(request):
        paper_phases_q, setting_phases_q, sup_phases_q = _child_phase_filters(request)
    else:
        prof = _parent_profile(request)
        paper_phases_q = _checking_phases_for_user(request)
        setting_phases_q = _setting_phases_for_user(request)
        if prof and _is_hub_coordinator(prof):
            sup_phases_q = SupervisionExamPhase.objects.filter(
                hub_coordinator=request.user,
                institute_semester_id=prof.institute_semester_id,
            ).order_by('name')
        elif prof and prof.department_id:
            sup_phases_q = SupervisionExamPhase.objects.filter(
                department_id=prof.department_id
            ).order_by('name')
        else:
            sup_phases_q = SupervisionExamPhase.objects.none()

    prof = _parent_profile(request) if _dept_parent_only(request) else None
    child_prof = _child_profile(request) if _dept_child_only(request) else None
    if _institute_paper_access(request):
        if is_exam_section_operator(request):
            report_scope_hint = (
                'Exam section view: data is limited to your selected working academic semesters '
                '(sidebar → Working academic semesters). Use Department to narrow further; export uses the same scope.'
            )
        else:
            report_scope_hint = (
                'Institute view: all faculties are listed. Use Department to narrow the table; export respects the same filters.'
            )
    elif child_prof and child_prof.department:
        sub = (child_prof.subunit_code or '').strip() or '—'
        report_scope_hint = (
            f'Sub-unit {sub}: lists everyone who completed paper checking, paper setting, or supervision '
            f'for {child_prof.department.name} in this period — including faculty from other departments. '
            'Supervision rows follow your division scope where applicable.'
        )
    elif prof and _is_hub_coordinator(prof):
        report_scope_hint = (
            'Lists every faculty who earned credits on duties for your hub’s departments (including assignees from outside the hub). '
            'Optional “Work for department” limits rows to one sub-department. Pay uses each stream’s configured rates.'
        )
    elif prof and prof.department_id:
        report_scope_hint = (
            f'Lists every faculty with duties recorded against {prof.department.name} exam activity in this range, '
            'including people assigned from other departments.'
        )
    else:
        report_scope_hint = ''

    show_duty_dept_filter = managed_scope is not None and len(duty_dept_choices) > 1
    dr_block = _daily_dr_block_for_section_credit(request)

    return render(
        request,
        'core/exam_paper/section_credit_report.html',
        {
            'date_from': d0,
            'date_to': d1,
            'department_id': dept_f,
            'dept_choices': dept_choices,
            'institute_scope': _institute_paper_access(request),
            'report_scope_hint': report_scope_hint,
            'rows': rows,
            'paper_phase_id': paper_phase_id,
            'sup_phase_id': sup_phase_id,
            'setting_phase_id': setting_phase_id,
            'sup_phases_labels': sup_ph_list,
            'set_phases_labels': set_ph_list,
            'paper_phases_filter': paper_phases_q,
            'setting_phases_filter': setting_phases_q,
            'sup_phases_filter': sup_phases_q,
            'duty_department_id': duty_department_id,
            'duty_dept_choices': duty_dept_choices,
            'show_duty_dept_filter': show_duty_dept_filter,
            **dr_block,
        },
    )


@login_required
def section_credit_report_excel(request):
    if not section_credit_report_access(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')

    today = timezone.localdate()
    d1 = _parse_date(request.GET.get('date_to')) or today
    d0 = _parse_date(request.GET.get('date_from')) or (d1 - timedelta(days=30))
    if d0 > d1:
        d0, d1 = d1, d0

    raw_dept = (request.GET.get('department_id') or '').strip()
    dept_f = None
    if raw_dept and _institute_paper_access(request):
        try:
            dept_f = int(raw_dept)
        except (TypeError, ValueError):
            dept_f = None

    def _opt_int(name: str) -> int | None:
        raw = (request.GET.get(name) or '').strip()
        if not raw:
            return None
        try:
            return int(raw)
        except (TypeError, ValueError):
            return None

    paper_phase_id = _opt_int('paper_phase_id')
    sup_phase_id = _opt_int('sup_phase_id')
    setting_phase_id = _opt_int('setting_phase_id')

    rows, _, _ = _section_credit_rows(
        request,
        d0,
        d1,
        dept_f,
        paper_phase_id,
        sup_phase_id,
        setting_phase_id,
    )

    line1, line2 = _section_credit_excel_banner_lines(request, d0, d1, dept_f)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Section credits'
    _write_section_credit_workbook(ws, rows, line1, line2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    fname = f'section_credit_report_{d0}_{d1}.xlsx'
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@login_required
def overall_faculty_report(request):
    if not section_credit_report_access(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')

    today = timezone.localdate()
    d1 = _parse_date(request.GET.get('date_to')) or today
    d0 = _parse_date(request.GET.get('date_from')) or (d1 - timedelta(days=30))
    if d0 > d1:
        d0, d1 = d1, d0

    raw_dept = (request.GET.get('department_id') or '').strip()
    dept_f = None
    if raw_dept and _institute_paper_access(request):
        try:
            dept_f = int(raw_dept)
        except (TypeError, ValueError):
            dept_f = None

    def _opt_int(name: str) -> int | None:
        raw = (request.GET.get(name) or '').strip()
        if not raw:
            return None
        try:
            return int(raw)
        except (TypeError, ValueError):
            return None

    paper_phase_id = _opt_int('paper_phase_id')
    sup_phase_id = _opt_int('sup_phase_id')
    setting_phase_id = _opt_int('setting_phase_id')

    rows, sup_ph_list, set_ph_list = _overall_faculty_rows(
        request,
        d0,
        d1,
        dept_f,
        paper_phase_id,
        sup_phase_id,
        setting_phase_id,
    )

    dept_choices = []
    if _institute_paper_access(request):
        if is_exam_section_operator(request):
            dept_choices = list(departments_for_exam_coordination_request(request).order_by('name'))
        else:
            dept_choices = list(Department.objects.order_by('name'))

    if _institute_paper_access(request):
        if is_exam_section_operator(request):
            sids = exam_section_working_semester_ids(request)
            paper_phases_q = checking_phases_exam_section_scoped(request)
            setting_phases_q = setting_phases_exam_section_scoped(request)
            sup_phases_q = SupervisionExamPhase.objects.filter(
                institute_semester_id__in=sids,
            ).order_by('name')
        else:
            paper_phases_q = PaperCheckingPhase.objects.order_by('name')
            setting_phases_q = PaperSettingPhase.objects.order_by('name')
            sup_phases_q = SupervisionExamPhase.objects.order_by('name')
    elif _dept_child_only(request):
        paper_phases_q, setting_phases_q, sup_phases_q = _child_phase_filters(request)
    else:
        prof = _parent_profile(request)
        paper_phases_q = _checking_phases_for_user(request)
        setting_phases_q = _setting_phases_for_user(request)
        if prof and _is_hub_coordinator(prof):
            sup_phases_q = SupervisionExamPhase.objects.filter(
                hub_coordinator=request.user,
                institute_semester_id=prof.institute_semester_id,
            ).order_by('name')
        elif prof and prof.department_id:
            sup_phases_q = SupervisionExamPhase.objects.filter(
                department_id=prof.department_id
            ).order_by('name')
        else:
            sup_phases_q = SupervisionExamPhase.objects.none()

    prof = _parent_profile(request) if _dept_parent_only(request) else None
    child_prof = _child_profile(request) if _dept_child_only(request) else None
    if _institute_paper_access(request):
        if is_exam_section_operator(request):
            report_scope_hint = (
                'Exam section: faculties and duties are limited to your selected working academic semesters.'
            )
        else:
            report_scope_hint = (
                'Institute view: all faculties in scope; expandable rows show which department each duty belonged to.'
            )
    elif child_prof and child_prof.department:
        sub = (child_prof.subunit_code or '').strip() or '—'
        report_scope_hint = (
            f'Sub-unit {sub}: only faculty linked to {child_prof.department.name} appear, '
            'but every duty they performed anywhere is listed with the department that owns the exam stream.'
        )
    elif prof and _is_hub_coordinator(prof):
        report_scope_hint = (
            'Hub scope: only faculty from your hub departments are listed; detail shows where each task was performed across the institute.'
        )
    elif prof and prof.department_id:
        report_scope_hint = (
            f'Only faculty from {prof.department.name} appear; line items name the department that hosted each duty.'
        )
    else:
        report_scope_hint = ''

    return render(
        request,
        'core/exam_paper/overall_faculty_report.html',
        {
            'date_from': d0,
            'date_to': d1,
            'department_id': dept_f,
            'dept_choices': dept_choices,
            'institute_scope': _institute_paper_access(request),
            'report_scope_hint': report_scope_hint,
            'rows': rows,
            'paper_phase_id': paper_phase_id,
            'sup_phase_id': sup_phase_id,
            'setting_phase_id': setting_phase_id,
            'sup_phases_labels': sup_ph_list,
            'set_phases_labels': set_ph_list,
            'paper_phases_filter': paper_phases_q,
            'setting_phases_filter': setting_phases_q,
            'sup_phases_filter': sup_phases_q,
        },
    )


@login_required
def overall_faculty_report_excel(request):
    if not section_credit_report_access(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')

    today = timezone.localdate()
    d1 = _parse_date(request.GET.get('date_to')) or today
    d0 = _parse_date(request.GET.get('date_from')) or (d1 - timedelta(days=30))
    if d0 > d1:
        d0, d1 = d1, d0

    raw_dept = (request.GET.get('department_id') or '').strip()
    dept_f = None
    if raw_dept and _institute_paper_access(request):
        try:
            dept_f = int(raw_dept)
        except (TypeError, ValueError):
            dept_f = None

    def _opt_int(name: str) -> int | None:
        raw = (request.GET.get(name) or '').strip()
        if not raw:
            return None
        try:
            return int(raw)
        except (TypeError, ValueError):
            return None

    paper_phase_id = _opt_int('paper_phase_id')
    sup_phase_id = _opt_int('sup_phase_id')
    setting_phase_id = _opt_int('setting_phase_id')

    rows, _, _ = _overall_faculty_rows(
        request,
        d0,
        d1,
        dept_f,
        paper_phase_id,
        sup_phase_id,
        setting_phase_id,
    )

    line1, line2 = _section_credit_excel_banner_lines(request, d0, d1, dept_f)
    line2 = line2.replace('Section credit report', 'Overall faculty report')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Summary'
    _write_section_credit_workbook(ws, rows, line1, line2)
    detail = wb.create_sheet('Line detail')
    _write_overall_line_detail_sheet(detail, rows)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    fname = f'overall_faculty_report_{d0}_{d1}.xlsx'
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp
