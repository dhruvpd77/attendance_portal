"""
Excel export for exam supervision DR-style reporting (openpyxl).
"""
from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.models import SupervisionDuty


def _session_label(time_slot: str) -> str:
    s = (time_slot or '').upper()
    if 'MORNING' in s or ' AM' in f' {s}' or s.startswith('08') or s.startswith('09') or s.startswith('10'):
        return 'Morning'
    if 'EVENING' in s or ' PM' in f' {s}':
        return 'Evening'
    return '—'


def build_supervision_dr_excel(
    duties: list[SupervisionDuty],
    *,
    title_line: str,
    sheet_prefix: str,
) -> HttpResponse:
    """Build a workbook: Detail sheet + faculty summary (DR-style columns)."""
    wb = Workbook()

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='1e3a5f')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    sub_fill = PatternFill('solid', fgColor='FFF2CC')
    sub_font = Font(bold=True, size=9)

    # --- Sheet 1: every duty row ---
    ws1 = wb.active
    ws1.title = 'Supervision_detail'[:31]

    h1 = [
        'Phase',
        'Division',
        'Assigned faculty',
        'Original from sheet',
        'Proxy?',
        'Status',
        'Date',
        'Session',
        'Time slot',
        'Subject',
        'Block',
        'Room',
        'Completed at (IST)',
    ]
    ws1.append(h1)
    for c in range(1, len(h1) + 1):
        cell = ws1.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for d in sorted(duties, key=lambda x: (x.supervision_date, x.phase.name, x.time_slot or '')):
        orig = d.original_faculty.full_name if d.original_faculty_id else (d.faculty_name_raw or '—')
        cur = d.faculty.full_name if d.faculty_id else (d.faculty_name_raw or '—')
        ws1.append(
            [
                d.phase.name,
                d.division_code,
                cur,
                orig,
                'Yes' if d.is_proxy else 'No',
                d.get_completion_status_display(),
                d.supervision_date.isoformat() if d.supervision_date else '',
                _session_label(d.time_slot),
                d.time_slot,
                d.subject_name,
                d.block_no,
                d.room_no,
                timezone.localtime(d.completed_at).strftime('%Y-%m-%d %H:%M') if d.completed_at else '',
            ]
        )
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=len(h1)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    for i in range(1, len(h1) + 1):
        ws1.column_dimensions[get_column_letter(i)].width = 14

    # --- Sheet 2: one row per current faculty ---
    ws2 = wb.create_sheet('DR_supervision_summary'[:31])

    by_faculty: dict = {}
    phase_names = sorted({d.phase.name for d in duties if d.faculty_id})

    for d in duties:
        if not d.faculty_id:
            continue
        fid = d.faculty_id
        if fid not in by_faculty:
            by_faculty[fid] = {'faculty': d.faculty, 'duties': [], 'any_proxy': False}
        by_faculty[fid]['duties'].append(d)
        if d.is_proxy:
            by_faculty[fid]['any_proxy'] = True

    hdr = [
        'Sr',
        'Name — contact',
        'Initials',
        'Assigned / Proxy',
        'Open slots',
        'Completed slots',
        'Block / Room (completed)',
    ]
    hdr.extend([f'Duty count · {pn}' for pn in phase_names])
    ws2.append([title_line])
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(hdr))
    ws2.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws2.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    ws2.append(hdr)
    for c in range(1, len(hdr) + 1):
        cell = ws2.cell(row=2, column=c)
        cell.fill = sub_fill
        cell.font = sub_font
        cell.border = border

    sr = 0
    for fid in sorted(by_faculty.keys(), key=lambda i: by_faculty[i]['faculty'].full_name.upper()):
        rowd = by_faculty[fid]
        fac = rowd['faculty']
        ds = rowd['duties']
        sr += 1
        contact = fac.email or '—'
        name_line = f'{fac.full_name}  |  {contact}'
        open_n = sum(1 for x in ds if x.completion_status == SupervisionDuty.OPEN)
        done_n = sum(1 for x in ds if x.completion_status == SupervisionDuty.COMPLETED)
        br_parts = []
        for x in ds:
            if x.completion_status == SupervisionDuty.COMPLETED and (x.block_no or x.room_no):
                br_parts.append(
                    f"{x.supervision_date}: B{x.block_no or '—'} R{x.room_no or '—'} ({x.phase.name})"
                )
        br_cell = '; '.join(br_parts) if br_parts else ''
        assign_proxy = 'Proxy' if rowd['any_proxy'] else 'Assigned'
        per_phase = defaultdict(int)
        for x in ds:
            per_phase[x.phase.name] += 1
        out = [
            sr,
            name_line,
            fac.short_name,
            assign_proxy,
            open_n,
            done_n,
            br_cell,
        ]
        for pn in phase_names:
            out.append(per_phase.get(pn, 0))
        ws2.append(out)

    for row in ws2.iter_rows(min_row=3, max_row=ws2.max_row, min_col=1, max_col=len(hdr)):
        for cell in row:
            cell.border = border

    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 42
    ws2.column_dimensions['C'].width = 10
    ws2.column_dimensions['D'].width = 14
    ws2.column_dimensions['G'].width = 48

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"supervision_dr_{sheet_prefix}_{datetime.now():%Y%m%d_%H%M}.xlsx".replace(' ', '_')
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp
