"""
Core views: Admin, Faculty, Student dashboards and features.
"""
import csv
import json
import os
import re
import random
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from io import BytesIO

import openpyxl
from django.conf import settings
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.http import Http404, HttpResponse, JsonResponse

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db import OperationalError
from django.db.models import Count, Q, IntegerField
from django.db.models.functions import Cast
from django.utils import timezone
import zoneinfo
from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from .models import (
    Department, Batch, Subject, Faculty, Student,
    ScheduleSlot, TermPhase, FacultyAttendance, LectureAdjustment, LectureCancellation, ExtraLecture, PhaseHoliday,
    AttendanceNotificationLog, AttendanceLockSetting, HODWeekLock,
    ExamPhase, ExamPhaseSubject, StudentMark, FacultyDoubtSession,
    FacultyDoubtRequest, FacultyDoubtRequestStudent, FacultyCombineDrCache,
)
from accounts.models import UserRole


# ---------- Error handlers ----------

def handler404(request, exception):
    return render(request, '404.html', status=404)


def handler500(request):
    return render(request, '500.html', status=500)


# ---------- Helpers ----------

def _roll_sort_key(s):
    """Sort key for students: ascending numeric roll_no (75 before 109). Non-numeric sorts last."""
    r = str(getattr(s, 'roll_no', '') or '').strip()
    return (int(r) if r.isdigit() else 999999, r)


def get_phase_holidays(dept, phase):
    """Return set of holiday dates for this department and phase (T1, T2, T3, T4)."""
    if not dept or not phase:
        return set()
    return set(
        PhaseHoliday.objects.filter(department=dept, phase=phase.upper()).values_list('date', flat=True)
    )


def get_all_holiday_dates(dept):
    """Return set of all holiday dates for this department (all phases). Used to exclude from valid/available dates."""
    if not dept:
        return set()
    return set(
        PhaseHoliday.objects.filter(department=dept).values_list('date', flat=True)
    )


def get_cancelled_lectures_set(dept):
    """Return set of (date, batch_id, time_slot) for cancelled lectures in this department."""
    if not dept:
        return set()
    try:
        return set(
            LectureCancellation.objects.filter(batch__department=dept)
            .values_list('date', 'batch_id', 'time_slot')
        )
    except OperationalError:
        return set()  # Table may not exist if migrations not run yet


def _is_admin_manual_locked_by_hod(dept, target_date):
    """Return True if this date falls in a week locked for the department (HOD or super admin). Departmental admins cannot edit manual attendance that week; HOD and super admin may still edit."""
    if not dept:
        return False
    try:
        week_map = {p: _compile_phase_weeks_date_objects(dept, p) for p in ['T1', 'T2', 'T3', 'T4']}
        for phase, weeks in week_map.items():
            for week_idx, week_dates in enumerate(weeks):
                if target_date in week_dates:
                    if HODWeekLock.objects.filter(department=dept, phase=phase, week_index=week_idx).exists():
                        return True
                    return False
    except Exception:
        pass
    return False


def _is_attendance_locked_for_date(target_date):
    """Return True if faculty cannot edit attendance for this date. Uses IST. Admin manual attendance never calls this."""
    try:
        lock = AttendanceLockSetting.objects.filter(pk=1).first()
    except OperationalError:
        return False  # Table may not exist if migrations not run yet
    if not lock or not lock.enabled:
        return False
    ist = zoneinfo.ZoneInfo('Asia/Kolkata')
    now_ist = timezone.now().astimezone(ist)
    today_ist = now_ist.date()
    if target_date < today_ist:
        return True  # past date always locked
    if target_date > today_ist:
        return False  # future date not locked
    lock_minutes = lock.lock_hour * 60 + lock.lock_minute
    now_minutes = now_ist.hour * 60 + now_ist.minute
    return now_minutes >= lock_minutes


# ----------

def _effective_slots_for_date(dept, date, extra_filters=None):
    """Return list of ScheduleSlot objects effective on this date."""
    from core.schedule_utils import get_effective_slots_for_date
    return get_effective_slots_for_date(dept, date, extra_filters)


def _effective_slots_for_faculty_on_date(faculty, date):
    """Return list of ScheduleSlot objects for this faculty effective on date."""
    return _effective_slots_for_date(faculty.department, date, extra_filters={'faculty': faculty})


def _effective_day_set_for_dept(dept, date):
    """Return set of weekday names that have schedule effective on this date."""
    from core.schedule_utils import get_effective_day_set
    return get_effective_day_set(dept, date)


def _effective_day_set_for_batch(batch, date):
    """Return set of weekday names for this batch effective on this date."""
    slots = _effective_slots_for_date(batch.department, date, extra_filters={'batch': batch})
    return {s.day for s in slots if s.day}


def get_faculty_subject_for_slot(date, batch, time_slot):
    """Return (faculty, subject) for this date/batch/slot; ExtraLecture > LectureAdjustment > ScheduleSlot."""
    from datetime import date as date_type
    if not isinstance(date, date_type):
        date = date
    weekday = date.strftime('%A')
    extra = ExtraLecture.objects.filter(
        date=date, batch=batch, time_slot=time_slot
    ).select_related('faculty', 'subject').first()
    if extra:
        return extra.faculty, extra.subject
    adj = LectureAdjustment.objects.filter(
        date=date, batch=batch, time_slot=time_slot
    ).select_related('new_faculty', 'new_subject').first()
    if adj:
        return adj.new_faculty, adj.new_subject
    slot = ScheduleSlot.objects.filter(
        batch=batch, day=weekday, time_slot=time_slot,
        effective_from__lte=date
    ).select_related('faculty', 'subject').order_by('-effective_from').first()
    if slot:
        return slot.faculty, slot.subject
    return None, None


def _is_extra_lecture_slot(dept, d, batch, time_slot):
    if not batch or not time_slot:
        return False
    return ExtraLecture.objects.filter(date=d, batch=batch, time_slot=time_slot).exists()


def _dr_slot_effective_load(is_extra_lecture, attendance_filled, present_count):
    """DR effective load: 0 if attendance not marked; extra (ETL) with present < 24 → 0.5 else 0.75."""
    if not attendance_filled:
        return 0
    try:
        p = int(present_count)
    except (TypeError, ValueError):
        p = 0
    if is_extra_lecture and p < 24:
        return 0.5
    return 0.75


def _add_batch_schedule_pairs_for_attendance(dept, batch, dates_iter, target_set, cancelled_set=None):
    """Add (date, time_slot) to target_set for held/percent: timetable rows + extra lectures if dept flag on."""
    if cancelled_set is None:
        cancelled_set = get_cancelled_lectures_set(dept)
    include_extra = bool(getattr(dept, 'include_extra_lectures_in_attendance', False))
    for d in dates_iter:
        weekday = d.strftime('%A')
        slots = [s for s in _effective_slots_for_date(dept, d, extra_filters={'batch': batch}) if s.day == weekday]
        for slot in set(s.time_slot for s in slots if s.time_slot):
            if (d, batch.id, slot) not in cancelled_set:
                target_set.add((d, slot))
        if include_extra:
            for ts in ExtraLecture.objects.filter(date=d, batch=batch).values_list('time_slot', flat=True):
                if ts and (d, batch.id, ts) not in cancelled_set:
                    target_set.add((d, ts))


# ----------

def get_admin_department(request):
    """Department for admin/HOD: departmental admin or HOD has fixed dept; super admin uses session or first."""
    try:
        if request.user.is_authenticated and hasattr(request.user, 'role_profile'):
            rp = request.user.role_profile
            if rp.role in ('admin', 'hod') and rp.department_id:
                return rp.department
    except Exception:
        pass
    dept_id = request.session.get('admin_department_id')
    if dept_id:
        return Department.objects.filter(pk=dept_id).first()
    return Department.objects.first()


def is_super_admin(request):
    """True if current user is admin with no department (can create depts, admins, and HODs)."""
    if not user_can_admin(request):
        return False
    if request.user.is_superuser or request.user.is_staff:
        return True
    try:
        rp = request.user.role_profile
        return rp.role == 'admin' and not rp.department_id
    except (UserRole.DoesNotExist, AttributeError):
        return True


def get_faculty_user(request):
    """Current user's Faculty or None."""
    if not request.user.is_authenticated:
        return None
    return getattr(request.user, 'faculty_profile', None)


def get_student_user(request):
    """Current user's Student or None."""
    if not request.user.is_authenticated:
        return None
    return getattr(request.user, 'student_profile', None)


def user_can_admin(request):
    """True if user can access admin features (admin, HOD, or superuser/staff)."""
    try:
        role = request.user.role_profile.role
        return role in ('admin', 'hod') or request.user.is_superuser or request.user.is_staff
    except (UserRole.DoesNotExist, AttributeError):
        return request.user.is_superuser or request.user.is_staff


def is_hod(request):
    """True if current user is HOD (can manage week locks for their department)."""
    try:
        return request.user.role_profile.role == 'hod'
    except (UserRole.DoesNotExist, AttributeError):
        return False


def user_can_faculty(request):
    try:
        return request.user.role_profile.role == 'faculty'
    except (UserRole.DoesNotExist, AttributeError):
        return False


def user_can_student(request):
    try:
        return request.user.role_profile.role == 'student'
    except (UserRole.DoesNotExist, AttributeError):
        return False


def faculty_portal_feature_allowed(faculty, url_name):
    """Per-department toggles for optional faculty menu items (HOD / super admin). Default allow."""
    if not faculty or not faculty.department:
        return True
    d = faculty.department
    mapping = {
        'faculty_doubt_solving': d.faculty_show_doubt_solving,
        'faculty_doubt_students_data': d.faculty_show_doubt_solving,
        'faculty_dr_load': d.faculty_show_dr_weekly_load,
        'faculty_mark_analytics': d.faculty_show_mark_analytics,
        'faculty_mark_analytics_risk_excel': d.faculty_show_mark_analytics,
        'faculty_mark_analytics_risk_all_excel': d.faculty_show_mark_analytics,
        'faculty_mark_analytics_report_excel': d.faculty_show_mark_analytics,
        'faculty_marks_report': d.faculty_show_marks_report,
        'faculty_student_marksheet': d.faculty_show_student_marksheet,
    }
    return mapping.get(url_name, True)


def _faculty_portal_guard_redirect(request, url_name):
    """If optional faculty feature is off, redirect to dashboard with message."""
    faculty = get_faculty_user(request)
    if not faculty_portal_feature_allowed(faculty, url_name):
        messages.error(
            request,
            'This section is turned off for your department. Contact your HOD if you need access.',
        )
        return redirect('core:faculty_dashboard')
    return None


# ---------- Home & Dashboards ----------

def portal_root(request):
    """Redirect /portal/ to main app (role redirect or login)."""
    return redirect('accounts:role_redirect')


def home(request):
    if request.user.is_authenticated:
        return redirect('accounts:role_redirect')
    return redirect('accounts:login')


@login_required
def admin_dashboard(request):
    if not user_can_admin(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    super_admin = is_super_admin(request)
    ctx = {
        'department': dept,
        'departments': Department.objects.all() if super_admin else ([dept] if dept else []),
        'batch_count': Batch.objects.filter(department=dept).count() if dept else 0,
        'faculty_count': Faculty.objects.filter(department=dept).count() if dept else 0,
        'student_count': Student.objects.filter(department=dept).count() if dept else 0,
        'is_super_admin': super_admin,
    }
    return render(request, 'core/admin/dashboard.html', ctx)


@login_required
def attendance_lock_setting(request):
    """Admin: set lock time after which faculty cannot edit attendance. Uses IST."""
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    try:
        lock_setting, _ = AttendanceLockSetting.objects.get_or_create(
            pk=1, defaults={'lock_hour': 17, 'lock_minute': 0, 'enabled': False}
        )
    except OperationalError:
        messages.error(request, 'Run migrations first: python manage.py migrate')
        return redirect('core:admin_dashboard')
    if request.method == 'POST':
        lock_setting.enabled = request.POST.get('enabled') == 'on'
        try:
            lock_setting.lock_hour = int(request.POST.get('lock_hour', 17))
            lock_setting.lock_minute = int(request.POST.get('lock_minute', 0))
        except (TypeError, ValueError):
            pass
        lock_setting.lock_hour = max(0, min(23, lock_setting.lock_hour))
        lock_setting.lock_minute = max(0, min(59, lock_setting.lock_minute))
        lock_setting.save()
        messages.success(request, 'Lock time saved.')
        return redirect('core:attendance_lock_setting')
    minute_options = list(range(0, 60, 5))  # 0, 5, 10, ..., 55
    ctx = {'lock_setting': lock_setting, 'minute_options': minute_options}
    return render(request, 'core/admin/attendance_lock_setting.html', ctx)


def _dates_for_lecture_cancellation(dept):
    """Dates for lecture cancellation dropdown. Uses term phases if set; otherwise a reasonable range."""
    dates = _dates_for_department(dept)
    if dates:
        return dates
    # Fallback when no term phases: wide range (past 6 months + next 18 months), weekdays with schedule
    day_set = _effective_day_set_for_dept(dept, datetime.now().date())
    day_set = {d.lower() for d in day_set if d}
    if not day_set:
        return []
    today = datetime.now().date()
    out = []
    for offset in range(-180, 550):
        d = today + timedelta(days=offset)
        if d.strftime('%A').lower() in day_set:
            out.append(d)
    return sorted(out)


@login_required
def lecture_cancellation(request):
    """Admin: select date, list all lectures for that date, allow delete (cancel) each lecture."""
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first from Dashboard.')
        return redirect('core:admin_dashboard')

    available_dates = _dates_for_lecture_cancellation(dept)
    date_str = request.GET.get('date')
    selected_date = None
    lectures = []

    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            if selected_date not in available_dates:
                available_dates = sorted(set(available_dates) | {selected_date})
        except Exception:
            pass

    if selected_date:
        cancelled_set = get_cancelled_lectures_set(dept)
        weekday = selected_date.strftime('%A')
        effective_slots = [s for s in _effective_slots_for_date(dept, selected_date) if s.day == weekday]
        seen = set()
        for slot in sorted(effective_slots, key=lambda s: (s.batch.name if s.batch else '', s.time_slot or '')):
            if (selected_date, slot.batch_id, slot.time_slot) in cancelled_set:
                continue
            seen.add((slot.batch_id, slot.time_slot))
            fac, subj = get_faculty_subject_for_slot(selected_date, slot.batch, slot.time_slot)
            lectures.append({
                'slot': slot,
                'faculty': fac or slot.faculty,
                'subject': subj or slot.subject,
                'batch': slot.batch,
            })
        for ex in ExtraLecture.objects.filter(date=selected_date, batch__department=dept).select_related('batch', 'faculty', 'subject'):
            if (ex.batch_id, ex.time_slot) in seen or (selected_date, ex.batch_id, ex.time_slot) in cancelled_set:
                continue
            seen.add((ex.batch_id, ex.time_slot))
            virtual_slot = type('Slot', (), {'batch': ex.batch, 'time_slot': ex.time_slot})()
            lectures.append({
                'slot': virtual_slot,
                'faculty': ex.faculty,
                'subject': ex.subject,
                'batch': ex.batch,
            })
        lectures.sort(key=lambda lec: (lec['batch'].name if lec.get('batch') else '', lec['slot'].time_slot or ''))

    cancellation_history = list(
        LectureCancellation.objects.filter(batch__department=dept)
        .select_related('batch')
        .order_by('-date', 'batch', 'time_slot')[:100]
    )

    ctx = {
        'department': dept,
        'available_dates': available_dates,
        'selected_date': selected_date,
        'lectures': lectures,
        'cancellation_history': cancellation_history,
    }
    return render(request, 'core/admin/lecture_cancellation.html', ctx)


@login_required
def lecture_cancellation_delete(request):
    """Cancel a lecture: create LectureCancellation, delete FacultyAttendance. Removes from all counts."""
    if not request.method == 'POST' or not user_can_admin(request):
        return redirect('core:lecture_cancellation')
    dept = get_admin_department(request)
    if not dept:
        return redirect('core:admin_dashboard')
    date_str = request.POST.get('date')
    batch_id = request.POST.get('batch_id')
    time_slot = request.POST.get('time_slot', '').strip()
    if not date_str or not batch_id or not time_slot:
        messages.error(request, 'Missing data.')
        return redirect('core:lecture_cancellation')
    try:
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        messages.error(request, 'Invalid date.')
        return redirect('core:lecture_cancellation')
    batch = Batch.objects.filter(pk=batch_id, department=dept).first()
    if not batch:
        messages.error(request, 'Invalid batch.')
        return redirect('core:lecture_cancellation')

    try:
        LectureCancellation.objects.get_or_create(date=selected_date, batch=batch, time_slot=time_slot)
        FacultyAttendance.objects.filter(date=selected_date, batch=batch, lecture_slot=time_slot).delete()
        ExtraLecture.objects.filter(date=selected_date, batch=batch, time_slot=time_slot).delete()
    except OperationalError:
        messages.error(request, 'Run migrations first: python manage.py migrate')
        return redirect('core:lecture_cancellation')
    messages.success(request, f'Lecture cancelled: {batch.name} {time_slot}. Removed from all records and counts.')
    url = reverse('core:lecture_cancellation') + f'?date={date_str}'
    return redirect(url)


def _time_slots_for_department(dept):
    """Distinct time_slot values from ScheduleSlot for this department."""
    if not dept:
        return []
    return list(
        ScheduleSlot.objects.filter(department=dept)
        .values_list('time_slot', flat=True).distinct().order_by('time_slot')
    )


@login_required
def extra_lecture(request):
    """Admin: Add extra lecture — select batch, date, time slot, subject, faculty, room."""
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first from Dashboard.')
        return redirect('core:admin_dashboard')

    batches = list(Batch.objects.filter(department=dept).order_by('name'))
    subjects = list(Subject.objects.filter(department=dept).order_by('name'))
    faculties = list(Faculty.objects.filter(department=dept).order_by('full_name'))
    time_slots = _time_slots_for_department(dept)
    if not time_slots:
        time_slots = ['Lec 1', 'Lec 2', 'Lec 3', 'Lec 4', 'Lec 5', 'Lec 6', 'Lec 7', 'Lec 8']

    if request.method == 'POST':
        date_str = request.POST.get('date', '').strip()
        batch_id = request.POST.get('batch_id', '').strip()
        time_slot = request.POST.get('time_slot', '').strip()
        subject_id = request.POST.get('subject_id', '').strip()
        faculty_id = request.POST.get('faculty_id', '').strip()
        room_number = request.POST.get('room_number', '').strip()
        if not all([date_str, batch_id, time_slot, subject_id, faculty_id]):
            messages.error(request, 'Please fill Batch, Date, Time Slot, Subject, and Faculty.')
            return redirect('core:extra_lecture')
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except Exception:
            messages.error(request, 'Invalid date.')
            return redirect('core:extra_lecture')
        batch = Batch.objects.filter(pk=batch_id, department=dept).first()
        subject = Subject.objects.filter(pk=subject_id, department=dept).first()
        faculty = Faculty.objects.filter(pk=faculty_id, department=dept).first()
        if not batch or not subject or not faculty:
            messages.error(request, 'Invalid batch, subject, or faculty.')
            return redirect('core:extra_lecture')
        obj, created = ExtraLecture.objects.update_or_create(
            date=selected_date, batch=batch, time_slot=time_slot,
            defaults={'subject': subject, 'faculty': faculty, 'room_number': room_number}
        )
        if created:
            messages.success(request, f'Extra lecture added: {batch.name} {time_slot} — {subject.name} ({faculty.short_name})')
        else:
            messages.success(request, f'Extra lecture updated: {batch.name} {time_slot}')
        return redirect('core:extra_lecture')

    extra_list = list(
        ExtraLecture.objects.filter(batch__department=dept)
        .select_related('batch', 'subject', 'faculty')
        .order_by('-date', 'batch', 'time_slot')[:100]
    )

    ctx = {
        'department': dept,
        'batches': batches,
        'subjects': subjects,
        'faculties': faculties,
        'time_slots': time_slots,
        'extra_list': extra_list,
    }
    return render(request, 'core/admin/extra_lecture.html', ctx)


@login_required
def extra_lecture_delete(request):
    """Delete an extra lecture."""
    if request.method != 'POST' or not user_can_admin(request):
        return redirect('core:extra_lecture')
    dept = get_admin_department(request)
    if not dept:
        return redirect('core:admin_dashboard')
    pk = request.POST.get('id')
    obj = ExtraLecture.objects.filter(pk=pk, batch__department=dept).first()
    if obj:
        batch_name, time_slot = obj.batch.name, obj.time_slot
        d, bid = obj.date, obj.batch_id
        obj.delete()
        FacultyAttendance.objects.filter(date=d, batch_id=bid, lecture_slot=time_slot).delete()
        messages.success(request, f'Extra lecture removed: {batch_name} {time_slot}')
    return redirect('core:extra_lecture')


# ---------- Admin: Result / Exam Phases ----------

@login_required
def exam_phases_list(request):
    """Admin: List exam phases, add new phase."""
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first from Dashboard.')
        return redirect('core:admin_dashboard')

    if request.method == 'POST' and request.POST.get('action') == 'add_phase':
        name = (request.POST.get('name', '') or '').strip().upper()
        if name:
            obj, created = ExamPhase.objects.get_or_create(department=dept, name=name)
            if created:
                messages.success(request, f'Exam phase "{name}" created.')
            else:
                messages.info(request, f'Exam phase "{name}" already exists.')
        else:
            messages.error(request, 'Enter a phase name.')
        return redirect('core:exam_phases_list')

    if request.method == 'POST' and request.POST.get('action') == 'delete_phase':
        pk = request.POST.get('phase_id')
        obj = ExamPhase.objects.filter(pk=pk, department=dept).first()
        if obj:
            obj.delete()
            messages.success(request, f'Exam phase "{obj.name}" deleted.')
        return redirect('core:exam_phases_list')

    phases = list(ExamPhase.objects.filter(department=dept).order_by('name'))
    ctx = {'department': dept, 'phases': phases}
    return render(request, 'core/admin/exam_phases_list.html', ctx)


@login_required
def admin_performance_students(request):
    """Hub: links to result, marks, and attendance-performance tools (HOD / departmental admin / super admin)."""
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    return render(request, 'core/admin/performance_students.html', {'department': dept})


@login_required
def admin_faculty_portal_management(request):
    """HOD / super admin: toggle optional faculty sidebar features for the current department."""
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    if not (is_hod(request) or is_super_admin(request)):
        messages.error(request, 'Only HOD or super admin can open Management.')
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    toggle_fields = (
        'faculty_show_doubt_solving',
        'faculty_show_dr_weekly_load',
        'faculty_show_mark_analytics',
        'faculty_show_marks_report',
        'faculty_show_student_marksheet',
    )
    if request.method == 'POST':
        for fn in toggle_fields:
            setattr(dept, fn, request.POST.get(fn) == 'on')
        dept.save(update_fields=list(toggle_fields))
        messages.success(request, 'Faculty portal options saved.')
        return redirect('core:admin_faculty_portal_management')
    return render(request, 'core/admin/faculty_portal_management.html', {'department': dept})


@login_required
def exam_phase_detail(request, phase_id):
    """Admin: Phase detail — manage subjects, upload marksheet per subject."""
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        return redirect('core:admin_dashboard')
    phase = ExamPhase.objects.filter(pk=phase_id, department=dept).first()
    if not phase:
        messages.error(request, 'Exam phase not found.')
        return redirect('core:exam_phases_list')

    dept_subjects = list(Subject.objects.filter(department=dept).order_by('name'))
    phase_subject_ids = set(ExamPhaseSubject.objects.filter(exam_phase=phase).values_list('subject_id', flat=True))
    phase_subjects = [s for s in dept_subjects if s.id in phase_subject_ids]
    available_to_add = [s for s in dept_subjects if s.id not in phase_subject_ids]

    if request.method == 'POST' and request.POST.get('action') == 'add_subject':
        sub_id = request.POST.get('subject_id')
        sub = Subject.objects.filter(pk=sub_id, department=dept).first()
        if sub:
            ExamPhaseSubject.objects.get_or_create(exam_phase=phase, subject=sub)
            messages.success(request, f'Added subject "{sub.name}" to phase.')
        return redirect('core:exam_phase_detail', phase_id=phase.id)

    if request.method == 'POST' and request.POST.get('action') == 'remove_subject':
        sub_id = request.POST.get('subject_id')
        ExamPhaseSubject.objects.filter(exam_phase=phase, subject_id=sub_id).delete()
        StudentMark.objects.filter(exam_phase=phase, subject_id=sub_id).delete()
        messages.success(request, 'Subject removed from phase.')
        return redirect('core:exam_phase_detail', phase_id=phase.id)

    if request.method == 'POST' and request.POST.get('action') == 'add_all_subjects':
        added = 0
        for s in dept_subjects:
            _, c = ExamPhaseSubject.objects.get_or_create(exam_phase=phase, subject=s)
            if c:
                added += 1
        messages.success(request, f'Added {added} subject(s) to phase.')
        return redirect('core:exam_phase_detail', phase_id=phase.id)

    batches = list(Batch.objects.filter(department=dept).order_by('name'))
    selected_subject_id = request.GET.get('subject_id')
    selected_batch_id = request.GET.get('batch_id')
    marks_list = []
    selected_subject = None
    if selected_subject_id and phase_subjects:
        sub = next((s for s in phase_subjects if str(s.id) == str(selected_subject_id)), None)
        if sub:
            selected_subject = sub
            qs = StudentMark.objects.filter(exam_phase=phase, subject=sub).select_related('student', 'student__batch')
            if selected_batch_id:
                qs = qs.filter(student__batch_id=selected_batch_id)
            marks_list = list(qs)
            marks_list.sort(key=lambda m: (m.student.batch.name if m.student.batch else '', _roll_sort_key(m.student)))

    ctx = {
        'department': dept,
        'phase': phase,
        'phase_subjects': phase_subjects,
        'available_to_add': available_to_add,
        'batches': batches,
        'selected_subject_id': selected_subject_id,
        'selected_batch_id': selected_batch_id,
        'marks_list': marks_list,
        'selected_subject': selected_subject,
    }
    return render(request, 'core/admin/exam_phase_detail.html', ctx)


@login_required
def exam_phase_upload_marks(request):
    """Admin: Upload marksheet Excel for a phase+subject. Map enrollment_no, save marks."""
    if request.method != 'POST' or not user_can_admin(request):
        return redirect('core:exam_phases_list')
    dept = get_admin_department(request)
    if not dept:
        return redirect('core:admin_dashboard')
    phase_id = request.POST.get('phase_id')
    subject_id = request.POST.get('subject_id')
    phase = ExamPhase.objects.filter(pk=phase_id, department=dept).first()
    subject = Subject.objects.filter(pk=subject_id, department=dept).first()
    if not phase or not subject:
        messages.error(request, 'Invalid phase or subject.')
        return redirect('core:exam_phases_list')
    if not ExamPhaseSubject.objects.filter(exam_phase=phase, subject=subject).exists():
        messages.error(request, 'Subject not in this phase.')
        return redirect('core:exam_phase_detail', phase_id=phase.id)

    excel_file = request.FILES.get('marksheet_file')
    if not excel_file:
        messages.error(request, 'Select an Excel file.')
        return redirect('core:exam_phase_detail', phase_id=phase.id)

    try:
        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        messages.error(request, f'Invalid Excel file: {e}')
        return redirect('core:exam_phase_detail', phase_id=phase.id)

    # Find header row and columns: Enrollment (Enrolllment/Enrollment), Marks
    enroll_col = marks_col = None
    data_start_row = 0
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx > 30:
            break
        row_str = [str(c).lower() if c is not None else '' for c in (row or [])]
        if 'enroll' in ' '.join(row_str) or 'enrolllment' in ' '.join(row_str):
            for c_idx, cell in enumerate(row_str):
                if 'enroll' in cell or 'enrolllment' in cell:
                    enroll_col = c_idx
                if 'mark' in cell:
                    marks_col = c_idx
            if enroll_col is not None:
                data_start_row = row_idx
                break
    if enroll_col is None:
        enroll_col = 3  # Fallback: column D
    if marks_col is None:
        marks_col = 6   # Fallback: column G

    students_by_enrollment = {
        str(s.enrollment_no).strip(): s
        for s in Student.objects.filter(department=dept).exclude(enrollment_no='')
        if s.enrollment_no
    }

    created = updated = skipped = 0
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx <= data_start_row or not row:  # Skip header and empty
            continue
        enroll_val = row[enroll_col] if enroll_col < len(row) else None
        if enroll_val is None:
            continue
        enroll_str = str(enroll_val).strip()
        if not enroll_str or not enroll_str.isdigit():
            continue
        marks_val = row[marks_col] if marks_col < len(row) else None
        try:
            marks_decimal = float(marks_val) if marks_val is not None else None
        except (TypeError, ValueError):
            marks_decimal = None

        student = students_by_enrollment.get(enroll_str)
        if not student:
            skipped += 1
            continue
        obj, created_flag = StudentMark.objects.update_or_create(
            student=student, exam_phase=phase, subject=subject,
            defaults={'marks_obtained': marks_decimal}
        )
        if created_flag:
            created += 1
        else:
            updated += 1

    wb.close()
    messages.success(request, f'Marks uploaded: {created} new, {updated} updated. {skipped} rows skipped (enrollment not found).')
    return redirect('core:exam_phase_detail', phase_id=phase.id)


@login_required
def faculty_student_marksheet(request):
    """Faculty: View mentorship students' marks — phase-wise, subject-wise (same format as attendance)."""
    if not user_can_faculty(request):
        return redirect('accounts:role_redirect')
    blocked = _faculty_portal_guard_redirect(request, 'faculty_student_marksheet')
    if blocked:
        return blocked
    faculty = get_faculty_user(request)
    if not faculty:
        return redirect('accounts:logout')
    dept = faculty.department
    phases = list(ExamPhase.objects.filter(department=dept).order_by('name'))
    phase_subjects_map = {}
    for p in phases:
        subs = list(ExamPhaseSubject.objects.filter(exam_phase=p).select_related('subject').order_by('subject__name'))
        phase_subjects_map[p.id] = subs

    mentorship_students = list(
        Student.objects.filter(mentor=faculty, department=dept)
        .select_related('batch', 'mentor')
    )
    mentorship_students.sort(key=lambda s: (s.batch.name if s.batch else '', _roll_sort_key(s)))

    marks_data = []
    if mentorship_students:
        student_ids = [s.id for s in mentorship_students]
        marks_qs = StudentMark.objects.filter(
            student_id__in=student_ids
        ).select_related('exam_phase', 'subject', 'student')
        marks_map = defaultdict(lambda: defaultdict(dict))  # student_id -> phase_id -> {subject_id: marks}
        for m in marks_qs:
            marks_map[m.student_id][m.exam_phase_id][m.subject_id] = m.marks_obtained

    for student in mentorship_students:
        row = {
            'student': student,
            'phases': [],
        }
        for phase in phases:
            phase_subs = phase_subjects_map.get(phase.id, [])
            subj_marks = []
            for eps in phase_subs:
                m = marks_map.get(student.id, {}).get(phase.id, {}).get(eps.subject_id)
                subj_marks.append({'subject': eps.subject, 'marks': m})
            row['phases'].append({'phase': phase, 'subjects': subj_marks})
        marks_data.append(row)

    phase_with_subjects = [(p, phase_subjects_map.get(p.id, [])) for p in phases]

    ctx = {
        'faculty': faculty,
        'phases': phases,
        'phase_with_subjects': phase_with_subjects,
        'marks_data': marks_data,
    }
    return render(request, 'core/faculty/student_marksheet.html', ctx)


@login_required
def faculty_dashboard(request):
    if not user_can_faculty(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    faculty = get_faculty_user(request)
    if not faculty:
        messages.error(request, 'No faculty profile linked.')
        return redirect('accounts:logout')
    from datetime import date
    today = date.today()
    weekday = today.strftime('%A')
    today_slots = sorted(
        [s for s in _effective_slots_for_faculty_on_date(faculty, today) if s.day == weekday],
        key=lambda s: s.time_slot or ''
    )
    extra_today = [
        {'time_slot': ex.time_slot, 'batch': ex.batch, 'subject': ex.subject}
        for ex in ExtraLecture.objects.filter(date=today, faculty=faculty).select_related('batch', 'subject')
    ]
    for ex in extra_today:
        today_slots.append(type('Slot', (), {'time_slot': ex['time_slot'], 'batch': ex['batch'], 'subject': ex['subject']})())
    today_slots.sort(key=lambda s: s.time_slot or '')
    # Batches where this faculty teaches (from schedule)
    faculty_batches = list(
        Batch.objects.filter(scheduleslot__faculty=faculty)
        .distinct().order_by('name')
    )
    ctx = {
        'faculty': faculty,
        'today_slots': today_slots,
        'today': today,
        'faculty_batches': faculty_batches,
    }
    return render(request, 'core/faculty/dashboard.html', ctx)


@login_required
def student_dashboard(request):
    if not user_can_student(request):
        messages.error(request, 'Access denied.')
        return redirect('accounts:role_redirect')
    student = get_student_user(request)
    if not student:
        messages.error(request, 'No student profile linked.')
        return redirect('accounts:logout')
    # Today's schedule for this batch (use adjusted faculty/subject if any)
    from datetime import date as date_type
    today = date_type.today()
    weekday = today.strftime('%A')
    slots = [s for s in _effective_slots_for_date(student.batch.department, today, extra_filters={'batch': student.batch}) if s.day == weekday]
    slots = sorted(slots, key=lambda s: s.time_slot or '')
    schedule = []
    for slot in slots:
        fac, subj = get_faculty_subject_for_slot(today, student.batch, slot.time_slot)
        schedule.append({
            'time_slot': slot.time_slot,
            'subject': subj.name if subj else (slot.subject.name if slot.subject else 'N/A'),
            'faculty': fac.short_name if fac else (slot.faculty.short_name if slot.faculty else '—'),
        })
    ctx = {'student': student, 'schedule': schedule}
    return render(request, 'core/student/dashboard.html', ctx)


# ---------- Admin: Department ----------

@login_required
def department_list(request):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    # Only super admin can switch department via form
    if is_super_admin(request) and request.method == 'POST' and request.POST.get('set_department'):
        request.session['admin_department_id'] = request.POST.get('department_id')
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    departments = Department.objects.all() if is_super_admin(request) else ([dept] if dept else [])
    ctx = {
        'departments': departments,
        'department': dept,
        'is_super_admin': is_super_admin(request),
    }
    return render(request, 'core/admin/department_list.html', ctx)


@login_required
def department_add(request):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    if not is_super_admin(request):
        messages.error(request, 'Only super admin can create departments.')
        return redirect('core:admin_dashboard')
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        semester = request.POST.get('semester', '').strip()
        include_x = request.POST.get('include_extra_lectures_in_attendance') == 'on'
        if name:
            Department.objects.create(
                name=name, semester=semester,
                include_extra_lectures_in_attendance=include_x,
            )
            messages.success(request, 'Department added.')
            return redirect('core:department_list')
    return render(request, 'core/admin/department_form.html', {'form_type': 'add'})


@login_required
def department_edit(request, pk):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    obj = get_object_or_404(Department, pk=pk)
    if not is_super_admin(request) and get_admin_department(request) != obj:
        messages.error(request, 'You can only edit your own department.')
        return redirect('core:department_list')
    if request.method == 'POST':
        obj.name = request.POST.get('name', '').strip() or obj.name
        obj.semester = request.POST.get('semester', '').strip()
        obj.include_extra_lectures_in_attendance = request.POST.get('include_extra_lectures_in_attendance') == 'on'
        obj.save()
        messages.success(request, 'Department updated.')
        return redirect('core:department_list')
    return render(request, 'core/admin/department_form.html', {'obj': obj, 'form_type': 'edit'})


@login_required
def department_delete(request, pk):
    if not user_can_admin(request) or not is_super_admin(request):
        messages.error(request, 'Only super admin can delete departments.')
        return redirect('core:department_list')
    obj = get_object_or_404(Department, pk=pk)
    name = obj.name
    obj.delete()
    messages.success(request, f'Department "{name}" deleted.')
    return redirect('core:department_list')


@login_required
def departmental_admin_list(request):
    """List departmental admins (admin users with a department). Super admin only."""
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    if not is_super_admin(request):
        messages.error(request, 'Only super admin can manage departmental admins.')
        return redirect('core:admin_dashboard')
    admins = UserRole.objects.filter(role='admin', department__isnull=False).select_related('user', 'department').order_by('department__name', 'user__username')
    ctx = {'admins': admins}
    return render(request, 'core/admin/departmental_admin_list.html', ctx)


@login_required
def departmental_admin_create(request):
    """Create a departmental admin: username, password, department. Super admin only."""
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    if not is_super_admin(request):
        messages.error(request, 'Only super admin can create departmental admins.')
        return redirect('core:admin_dashboard')
    if request.method == 'POST':
        username = (request.POST.get('username') or '').strip()
        password = request.POST.get('password') or ''
        password2 = request.POST.get('password2') or ''
        department_id = request.POST.get('department_id')
        if not username:
            messages.error(request, 'Username is required.')
        elif User.objects.filter(username=username).exists():
            messages.error(request, 'That username is already taken.')
        elif not password or len(password) < 6:
            messages.error(request, 'Password must be at least 6 characters.')
        elif password != password2:
            messages.error(request, 'Passwords do not match.')
        elif not department_id:
            messages.error(request, 'Please select a department.')
        else:
            dept = Department.objects.filter(pk=department_id).first()
            if not dept:
                messages.error(request, 'Invalid department.')
            else:
                user = User.objects.create_user(username=username, password=password)
                UserRole.objects.create(user=user, role='admin', department=dept)
                messages.success(request, f'Departmental admin "{username}" created for {dept.name}.')
                return redirect('core:departmental_admin_list')
    ctx = {'departments': Department.objects.order_by('name')}
    return render(request, 'core/admin/departmental_admin_form.html', ctx)


@login_required
def departmental_hod_list(request):
    """List departmental HODs. Super admin only."""
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    if not is_super_admin(request):
        messages.error(request, 'Only super admin can manage departmental HODs.')
        return redirect('core:admin_dashboard')
    if request.method == 'POST' and request.POST.get('action') == 'unlock_all_departments':
        deleted, _ = HODWeekLock.objects.all().delete()
        messages.success(request, f'All {deleted} week lock(s) removed. All weeks are now unlocked by default.')
        return redirect('core:departmental_hod_list')
    hods = UserRole.objects.filter(role='hod', department__isnull=False).select_related('user', 'department').order_by('department__name', 'user__username')
    lock_count = HODWeekLock.objects.count()
    ctx = {'hods': hods, 'lock_count': lock_count}
    return render(request, 'core/admin/departmental_hod_list.html', ctx)


@login_required
def departmental_hod_create(request):
    """Create a departmental HOD. Super admin only."""
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    if not is_super_admin(request):
        messages.error(request, 'Only super admin can create departmental HODs.')
        return redirect('core:admin_dashboard')
    if request.method == 'POST':
        username = (request.POST.get('username') or '').strip()
        password = request.POST.get('password') or ''
        password2 = request.POST.get('password2') or ''
        department_id = request.POST.get('department_id')
        if not username:
            messages.error(request, 'Username is required.')
        elif User.objects.filter(username=username).exists():
            messages.error(request, 'That username is already taken.')
        elif not password or len(password) < 6:
            messages.error(request, 'Password must be at least 6 characters.')
        elif password != password2:
            messages.error(request, 'Passwords do not match.')
        elif not department_id:
            messages.error(request, 'Please select a department.')
        else:
            dept = Department.objects.filter(pk=department_id).first()
            if not dept:
                messages.error(request, 'Invalid department.')
            else:
                user = User.objects.create_user(username=username, password=password)
                UserRole.objects.create(user=user, role='hod', department=dept)
                messages.success(request, f'HOD "{username}" created for {dept.name}.')
                return redirect('core:departmental_hod_list')
    ctx = {'departments': Department.objects.order_by('name')}
    return render(request, 'core/admin/departmental_hod_form.html', ctx)


@login_required
def hod_lock_admin_weeks(request):
    """HOD or super admin: lock/unlock weeks for the current department only. When locked, departmental admins cannot edit manual attendance for that week (daily faculty time lock is unchanged)."""
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    if not (is_hod(request) or is_super_admin(request)):
        messages.error(request, 'Only HOD or super admin can manage week locks.')
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first from the Dashboard.')
        return redirect('core:admin_dashboard')
    tp = TermPhase.objects.filter(department=dept).first()
    phases = ['T1', 'T2', 'T3', 'T4']
    week_map = {}
    for p in phases:
        weeks = _compile_phase_weeks_date_objects(dept, p)
        week_map[p] = weeks
    locked_set = set(
        HODWeekLock.objects.filter(department=dept).values_list('phase', 'week_index')
    )
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'unlock_all':
            deleted, _ = HODWeekLock.objects.filter(department=dept).delete()
            messages.success(request, f'All {deleted} week(s) unlocked. Admin can now edit manual attendance for all weeks.')
            return redirect('core:hod_lock_admin_weeks')
        phase = request.POST.get('phase', '').strip()
        week_index_str = request.POST.get('week_index', '')
        if action and phase in phases and week_index_str != '':
            try:
                week_index = int(week_index_str)
                weeks = week_map.get(phase, [])
                if 0 <= week_index < len(weeks):
                    if action == 'lock':
                        HODWeekLock.objects.get_or_create(
                            department=dept, phase=phase, week_index=week_index
                        )
                        messages.success(request, f'{phase} Week {week_index + 1} locked. Admin cannot edit manual attendance for that week.')
                    elif action == 'unlock':
                        HODWeekLock.objects.filter(
                            department=dept, phase=phase, week_index=week_index
                        ).delete()
                        messages.success(request, f'{phase} Week {week_index + 1} unlocked.')
            except (ValueError, TypeError):
                pass
        return redirect('core:hod_lock_admin_weeks')
    phase_week_offsets = _get_phase_week_offsets(week_map)
    phase_weeks_list = []
    for p in phases:
        weeks = week_map.get(p, [])
        opts = [(i, phase_week_offsets.get(p, 0) + i + 1, (p, i) in locked_set, len(weeks[i]) if i < len(weeks) else 0) for i in range(len(weeks))]
        phase_weeks_list.append((p, opts, weeks))
    ctx = {
        'department': dept,
        'phases': phases,
        'phase_weeks_list': phase_weeks_list,
        'locked_set': locked_set,
        'is_super_admin': is_super_admin(request),
    }
    return render(request, 'core/admin/hod_lock_admin_weeks.html', ctx)


# ---------- Admin: Batch ----------

@login_required
def batch_list(request):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    batches = Batch.objects.filter(department=dept) if dept else []
    ctx = {'batches': batches, 'department': dept}
    return render(request, 'core/admin/batch_list.html', ctx)


@login_required
def batch_add(request):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            Batch.objects.get_or_create(department=dept, name=name)
            messages.success(request, 'Batch added.')
            return redirect('core:batch_list')
    return render(request, 'core/admin/batch_form.html', {'form_type': 'add', 'department': dept})


@login_required
def batch_edit(request, pk):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    obj = get_object_or_404(Batch, pk=pk)
    dept = get_admin_department(request)
    if dept and obj.department != dept:
        messages.error(request, 'You can only edit batches in your department.')
        return redirect('core:batch_list')
    if request.method == 'POST':
        obj.name = request.POST.get('name', '').strip() or obj.name
        obj.save()
        messages.success(request, 'Batch updated.')
        return redirect('core:batch_list')
    return render(request, 'core/admin/batch_form.html', {'obj': obj, 'form_type': 'edit'})


@login_required
def batch_delete(request, pk):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    obj = get_object_or_404(Batch, pk=pk)
    dept = get_admin_department(request)
    if dept and obj.department != dept:
        messages.error(request, 'You can only manage batches in your department.')
        return redirect('core:batch_list')
    name = obj.name
    obj.delete()
    messages.success(request, f'Batch "{name}" deleted.')
    return redirect('core:batch_list')


# ---------- Admin: Subject ----------

@login_required
def subject_list(request):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    subjects = Subject.objects.filter(department=dept) if dept else []
    ctx = {'subjects': subjects, 'department': dept}
    return render(request, 'core/admin/subject_list.html', ctx)


@login_required
def subject_add(request):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip()
        if name:
            Subject.objects.get_or_create(department=dept, name=name, defaults={'code': code})
            messages.success(request, 'Subject added.')
            return redirect('core:subject_list')
    return render(request, 'core/admin/subject_form.html', {'form_type': 'add', 'department': dept})


@login_required
def subject_edit(request, pk):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    obj = get_object_or_404(Subject, pk=pk)
    dept = get_admin_department(request)
    if dept and obj.department != dept:
        messages.error(request, 'You can only edit subjects in your department.')
        return redirect('core:subject_list')
    if request.method == 'POST':
        obj.name = request.POST.get('name', '').strip() or obj.name
        obj.code = request.POST.get('code', '').strip()
        obj.save()
        messages.success(request, 'Subject updated.')
        return redirect('core:subject_list')
    return render(request, 'core/admin/subject_form.html', {'obj': obj, 'form_type': 'edit'})


@login_required
def subject_delete(request, pk):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    obj = get_object_or_404(Subject, pk=pk)
    dept = get_admin_department(request)
    if dept and obj.department != dept:
        messages.error(request, 'You can only manage subjects in your department.')
        return redirect('core:subject_list')
    name = obj.name
    obj.delete()
    messages.success(request, f'Subject "{name}" deleted.')
    return redirect('core:subject_list')


# ---------- Admin: Faculty ----------

@login_required
def faculty_list(request):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    faculties = Faculty.objects.filter(department=dept) if dept else []
    ctx = {'faculties': faculties, 'department': dept}
    return render(request, 'core/admin/faculty_list.html', ctx)


@login_required
def faculty_add(request):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    if request.method == 'POST':
        full_name = request.POST.get('full_name', '').strip()
        short_name = request.POST.get('short_name', '').strip()
        email = request.POST.get('email', '').strip()
        if full_name and short_name:
            Faculty.objects.get_or_create(
                department=dept, short_name=short_name,
                defaults={'full_name': full_name, 'email': email}
            )
            messages.success(request, 'Faculty added.')
            return redirect('core:faculty_list')
    return render(request, 'core/admin/faculty_form.html', {'form_type': 'add', 'department': dept})


@login_required
def faculty_edit(request, pk):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    obj = get_object_or_404(Faculty, pk=pk)
    dept = get_admin_department(request)
    if dept and obj.department != dept:
        messages.error(request, 'You can only edit faculty in your department.')
        return redirect('core:faculty_list')
    if request.method == 'POST':
        obj.full_name = request.POST.get('full_name', '').strip() or obj.full_name
        obj.short_name = request.POST.get('short_name', '').strip() or obj.short_name
        obj.email = request.POST.get('email', '').strip()
        obj.save()
        messages.success(request, 'Faculty updated.')
        return redirect('core:faculty_list')
    return render(request, 'core/admin/faculty_form.html', {'obj': obj, 'form_type': 'edit'})


@login_required
def faculty_delete(request, pk):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    obj = get_object_or_404(Faculty, pk=pk)
    dept = get_admin_department(request)
    if dept and obj.department != dept:
        messages.error(request, 'You can only manage faculties in your department.')
        return redirect('core:faculty_list')
    name = obj.full_name
    obj.delete()
    messages.success(request, f'Faculty "{name}" deleted.')
    return redirect('core:faculty_list')


@login_required
def generate_credentials_choice(request):
    """Choose whether to generate credentials for Students or Faculty."""
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first from Dashboard.')
        return redirect('core:admin_dashboard')
    ctx = {'department': dept}
    return render(request, 'core/admin/generate_credentials_choice.html', ctx)


@login_required
def faculty_generate_credentials(request):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first from Dashboard.')
        return redirect('core:admin_dashboard')

    if request.method == 'POST':
        if request.POST.get('action') in ('reset_passwords', 'reset_and_send_email'):
            faculties_with_user = Faculty.objects.filter(department=dept, user__isnull=False).select_related('user').order_by('short_name')
            if not faculties_with_user.exists():
                messages.warning(request, 'No faculty with credentials to reset.')
                return redirect('core:faculty_generate_credentials')
            rows = []
            for faculty in faculties_with_user:
                password = str(random.randint(0, 9999)).zfill(4)
                faculty.user.set_password(password)
                faculty.user.save()
                rows.append({
                    'department': dept.name,
                    'full_name': faculty.full_name,
                    'short_name': faculty.short_name,
                    'username': faculty.user.username,
                    'password': password,
                    'email': (faculty.email or getattr(faculty.user, 'email', '') or '').strip(),
                })
            cred_dir = os.path.join(settings.MEDIA_ROOT, 'credentials')
            os.makedirs(cred_dir, exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            safe_dept = re.sub(r'[^\w\-]', '_', dept.name)[:50]
            filename = f'faculty_credentials_reset_{safe_dept}_{timestamp}.csv'
            filepath = os.path.join(cred_dir, filename)
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                w = csv.DictWriter(f, fieldnames=['department', 'full_name', 'short_name', 'username', 'password', 'email'])
                w.writeheader()
                w.writerows(rows)
            if request.POST.get('action') == 'reset_and_send_email':
                login_url = (settings.SITE_URL + reverse('accounts:login')) if settings.SITE_URL else request.build_absolute_uri(reverse('accounts:login'))
                sent = 0
                skipped = []
                for row in rows:
                    email = row.get('email', '').strip()
                    if not email:
                        skipped.append(row.get('full_name', 'Unknown'))
                        continue
                    html = render(request, 'core/admin/email_faculty_credentials.html', {
                        'full_name': row.get('full_name', 'User'),
                        'username': row.get('username', ''),
                        'password': row.get('password', ''),
                        'login_url': login_url,
                    }).content.decode('utf-8')
                    try:
                        plain_msg = f"Username: {row.get('username', '')}\nPassword: {row.get('password', '')}\n\nSign in at: {login_url}\n\nKeep these credentials secure. Use Change Password on the login page to reset."
                        send_mail(
                            subject='Your LJIET Attendance Portal Login Credentials',
                            message=plain_msg,
                            from_email=settings.DEFAULT_FROM_EMAIL,
                            recipient_list=[email],
                            html_message=html,
                            fail_silently=False,
                        )
                        sent += 1
                    except Exception:
                        skipped.append(row.get('full_name', 'Unknown'))
                if sent:
                    messages.success(request, f'Passwords reset. Email sent to {sent} faculty with their new username and password.')
                if skipped:
                    messages.warning(request, f'Could not email {len(skipped)} (no email or failed): {", ".join(skipped[:5])}{"..." if len(skipped) > 5 else ""}')
            else:
                messages.success(request, f'Passwords reset for {len(rows)} faculty. Download the file below — it contains the new passwords.')
            request.session['credentials_filename'] = filename
            request.session['credentials_count'] = len(rows)
            request.session['credentials_type'] = 'faculty'
            return redirect('core:credentials_result')

        faculties_without_user = Faculty.objects.filter(department=dept, user__isnull=True).order_by('short_name')
        if not faculties_without_user.exists():
            messages.warning(request, 'All faculties in this department already have login credentials.')
            return redirect('core:generate_credentials_choice')

        rows = []
        for faculty in faculties_without_user:
            base_username = (faculty.short_name or 'f').strip().lower()
            base_username = re.sub(r'[^\w]', '', base_username)[:30] or 'f'
            username = base_username
            if User.objects.filter(username=username).exists():
                username = f"{base_username}{faculty.id}"
            while User.objects.filter(username=username).exists():
                username = f"{base_username}{faculty.id}_{random.randint(100, 999)}"
            password = str(random.randint(0, 9999)).zfill(4)
            user = User.objects.create_user(username=username, password=password)
            faculty.user = user
            faculty.save()
            UserRole.objects.get_or_create(user=user, defaults={'role': 'faculty'})
            rows.append({
                'department': dept.name,
                'full_name': faculty.full_name,
                'short_name': faculty.short_name,
                'username': username,
                'password': password,
                'email': faculty.email or '',
            })

        cred_dir = os.path.join(settings.MEDIA_ROOT, 'credentials')
        os.makedirs(cred_dir, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_dept = re.sub(r'[^\w\-]', '_', dept.name)[:50]
        filename = f'faculty_credentials_{safe_dept}_{timestamp}.csv'
        filepath = os.path.join(cred_dir, filename)

        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            w = csv.DictWriter(f, fieldnames=['department', 'full_name', 'short_name', 'username', 'password', 'email'])
            w.writeheader()
            w.writerows(rows)

        request.session['credentials_filename'] = filename
        request.session['credentials_count'] = len(rows)
        request.session['credentials_type'] = 'faculty'
        messages.success(request, f'Credentials generated for {len(rows)} faculty. Download and store the file securely.')
        return redirect('core:credentials_result')

    faculties_without = Faculty.objects.filter(department=dept, user__isnull=True).count()
    faculties_total = Faculty.objects.filter(department=dept).count()
    ctx = {'department': dept, 'faculties_without': faculties_without, 'faculties_total': faculties_total}
    return render(request, 'core/admin/faculty_generate_credentials.html', ctx)


@login_required
def student_generate_credentials(request):
    """Generate login credentials for students: username = enrollment number, password = 4-digit random."""
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first from Dashboard.')
        return redirect('core:admin_dashboard')

    if request.method == 'POST':
        students_without_user = Student.objects.filter(department=dept, user__isnull=True).select_related('batch').annotate(roll_no_int=Cast('roll_no', IntegerField())).order_by('batch__name', 'roll_no_int', 'roll_no')
        if not students_without_user.exists():
            messages.warning(request, 'All students in this department already have login credentials.')
            return redirect('core:generate_credentials_choice')

        rows = []
        for student in students_without_user:
            base_username = (student.enrollment_no or '').strip()
            if not base_username:
                base_username = f'stu{student.id}'
            base_username = re.sub(r'[^\w]', '', base_username)[:30] or f'stu{student.id}'
            username = base_username
            if User.objects.filter(username=username).exists():
                username = f"{base_username}_{student.id}"
            while User.objects.filter(username=username).exists():
                username = f"{base_username}_{student.id}_{random.randint(100, 999)}"
            password = str(random.randint(0, 9999)).zfill(4)
            user = User.objects.create_user(username=username, password=password)
            student.user = user
            student.save()
            UserRole.objects.get_or_create(user=user, defaults={'role': 'student'})
            rows.append({
                'department': dept.name,
                'batch': student.batch.name,
                'roll_no': student.roll_no,
                'enrollment_no': student.enrollment_no or '',
                'name': student.name,
                'username': username,
                'password': password,
                'email': student.email or '',
            })

        cred_dir = os.path.join(settings.MEDIA_ROOT, 'credentials')
        os.makedirs(cred_dir, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_dept = re.sub(r'[^\w\-]', '_', dept.name)[:50]
        filename = f'student_credentials_{safe_dept}_{timestamp}.csv'
        filepath = os.path.join(cred_dir, filename)

        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            w = csv.DictWriter(f, fieldnames=['department', 'batch', 'roll_no', 'enrollment_no', 'name', 'username', 'password', 'email'])
            w.writeheader()
            w.writerows(rows)

        request.session['credentials_filename'] = filename
        request.session['credentials_count'] = len(rows)
        request.session['credentials_type'] = 'student'
        messages.success(request, f'Credentials generated for {len(rows)} students. Download and store the file securely.')
        return redirect('core:credentials_result')

    students_without = Student.objects.filter(department=dept, user__isnull=True).count()
    students_total = Student.objects.filter(department=dept).count()
    ctx = {'department': dept, 'students_without': students_without, 'students_total': students_total}
    return render(request, 'core/admin/student_generate_credentials.html', ctx)


@login_required
def credentials_result(request):
    """Shared result page after generating faculty or student credentials."""
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    filename = request.session.pop('credentials_filename', None)
    count = request.session.pop('credentials_count', 0)
    cred_type = request.session.pop('credentials_type', 'faculty')
    if not filename:
        messages.info(request, 'No credentials file from this session.')
        return redirect('core:generate_credentials_choice')
    ctx = {'filename': filename, 'count': count, 'credentials_type': cred_type}
    return render(request, 'core/admin/credentials_result.html', ctx)


@login_required
def faculty_credentials_result(request):
    """Legacy URL: redirect to shared credentials result (session may already be consumed)."""
    return redirect('core:credentials_result')


@login_required
def download_credentials_file(request, filename):
    if not user_can_admin(request):
        raise Http404()
    if not filename or not re.match(r'^[a-zA-Z0-9_.\-]+\.csv$', filename):
        raise Http404()
    filepath = os.path.join(settings.MEDIA_ROOT, 'credentials', filename)
    if not os.path.isfile(filepath):
        raise Http404()
    with open(filepath, 'rb') as f:
        response = HttpResponse(f.read(), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def download_credentials_excel(request, filename):
    """Download Excel from the credentials CSV (includes passwords - only available right after generation)."""
    if not user_can_admin(request):
        raise Http404()
    if not filename or not re.match(r'^[a-zA-Z0-9_.\-]+\.csv$', filename):
        raise Http404()
    filepath = os.path.join(settings.MEDIA_ROOT, 'credentials', filename)
    if not os.path.isfile(filepath):
        raise Http404()
    cred_type = 'Faculty' if 'faculty' in filename.lower() else 'Student'
    wb = Workbook()
    ws = wb.active
    ws.title = f'{cred_type} Credentials'
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    if not rows:
        raise Http404()
    fieldnames = list(rows[0].keys())
    for c, h in enumerate(fieldnames, 1):
        cell = ws.cell(1, c, h.replace('_', ' ').title())
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(fieldnames, 1):
            ws.cell(row_idx, col_idx, row.get(key, '')).border = thin_border
    for c in range(1, len(fieldnames) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    xlsx_name = filename.replace('.csv', '.xlsx')
    resp = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{xlsx_name}"'
    return resp


@login_required
def send_credentials_emails(request):
    """Send credentials email to each faculty/student from the CSV (after generation)."""
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    if request.method != 'POST':
        return redirect('core:generate_credentials_choice')
    filename = request.POST.get('filename', '').strip()
    if not filename or not re.match(r'^[a-zA-Z0-9_.\-]+\.csv$', filename):
        messages.error(request, 'Invalid file.')
        return redirect('core:generate_credentials_choice')
    filepath = os.path.join(settings.MEDIA_ROOT, 'credentials', filename)
    if not os.path.isfile(filepath):
        messages.error(request, 'Credentials file not found.')
        return redirect('core:generate_credentials_choice')
    cred_type = 'faculty' if 'faculty' in filename.lower() else 'student'
    login_url = (settings.SITE_URL + reverse('accounts:login')) if settings.SITE_URL else request.build_absolute_uri(reverse('accounts:login'))
    sent = 0
    skipped = []
    try:
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        for row in rows:
            email = row.get('email', '').strip()
            if not email and cred_type == 'faculty':
                try:
                    user = User.objects.get(username=row.get('username', ''))
                    faculty = getattr(user, 'faculty_profile', None)
                    if faculty:
                        email = (faculty.email or getattr(user, 'email', '') or '').strip()
                except User.DoesNotExist:
                    pass
            if not email:
                skipped.append(row.get('full_name') or row.get('name', 'Unknown'))
                continue
            full_name = row.get('full_name') or row.get('name', 'User')
            username = row.get('username', '')
            password = row.get('password', '')
            if cred_type == 'faculty':
                html = render(request, 'core/admin/email_faculty_credentials.html', {
                    'full_name': full_name,
                    'username': username,
                    'password': password,
                    'login_url': login_url,
                }).content.decode('utf-8')
                subject = 'Your LJIET Attendance Portal Login Credentials'
            else:
                html = render(request, 'core/admin/email_student_credentials.html', {
                    'full_name': full_name,
                    'username': username,
                    'password': password,
                    'login_url': login_url,
                }).content.decode('utf-8')
                subject = 'Your LJIET Attendance Portal Login Credentials'
            try:
                plain_msg = f"Username: {username}\nPassword: {password}\n\nSign in at: {login_url}\n\nKeep these credentials secure. Use Change Password on the login page to reset."
                send_mail(
                    subject=subject,
                    message=plain_msg,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[email],
                    html_message=html,
                    fail_silently=False,
                )
                sent += 1
            except Exception:
                skipped.append(full_name)
    except Exception as e:
        messages.error(request, f'Error: {e}')
        return redirect('core:generate_credentials_choice')
    if sent:
        messages.success(request, f'Email sent to {sent} {cred_type}(s).')
    if skipped:
        messages.warning(request, f'Skipped (no email or failed): {", ".join(skipped[:5])}{"..." if len(skipped) > 5 else ""}')
    return redirect('core:faculty_generate_credentials' if cred_type == 'faculty' else 'core:student_generate_credentials')


@login_required
def send_faculty_existing_emails(request):
    """Send username-only email to all faculty with credentials in the department."""
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:generate_credentials_choice')
    faculties = Faculty.objects.filter(department=dept, user__isnull=False).select_related('user')
    login_url = (settings.SITE_URL + reverse('accounts:login')) if settings.SITE_URL else request.build_absolute_uri(reverse('accounts:login'))
    sent = 0
    skipped = []
    for faculty in faculties:
        email = (faculty.email or getattr(faculty.user, 'email', '') or '').strip()
        if not email:
            skipped.append(faculty.full_name)
            continue
        html = render(request, 'core/admin/email_faculty_username_only.html', {
            'full_name': faculty.full_name,
            'username': faculty.user.username,
            'login_url': login_url,
        }).content.decode('utf-8')
        try:
            send_mail(
                subject='Your LJIET Attendance Portal Login',
                message='Please view this email in HTML format.',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                html_message=html,
                fail_silently=False,
            )
            sent += 1
        except Exception:
            skipped.append(faculty.full_name)
    if sent:
        messages.success(request, f'Email sent to {sent} faculty.')
    if skipped:
        messages.warning(request, f'Skipped (no email or failed): {", ".join(skipped[:5])}{"..." if len(skipped) > 5 else ""}')
    return redirect('core:faculty_generate_credentials')


@login_required
def download_faculty_credentials_excel(request):
    """Download Excel of all faculty credentials (username, name, etc.) in the department. Passwords cannot be exported."""
    if not user_can_admin(request):
        raise Http404()
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first from Dashboard.')
        return redirect('core:generate_credentials_choice')
    faculties = Faculty.objects.filter(department=dept, user__isnull=False).select_related('user').order_by('short_name')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Faculty Credentials'
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    headers = ['Department', 'Full Name', 'Short Name', 'Username', 'Password', 'Note']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    for row_idx, f in enumerate(faculties, 2):
        ws.cell(row_idx, 1, f.department.name).border = thin_border
        ws.cell(row_idx, 2, f.full_name).border = thin_border
        ws.cell(row_idx, 3, f.short_name or '').border = thin_border
        ws.cell(row_idx, 4, f.user.username).border = thin_border
        ws.cell(row_idx, 5, '—').border = thin_border  # Passwords are hashed, not retrievable
        ws.cell(row_idx, 6, 'Use Change Password on login page to reset.').border = thin_border
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 38
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    safe_dept = re.sub(r'[^\w\-]', '_', dept.name)[:50]
    fname = f'Faculty_Credentials_{safe_dept}.xlsx'
    resp = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@login_required
def download_student_credentials_excel(request):
    """Download Excel of all student credentials (username, roll_no, name, etc.) in the department. Passwords cannot be exported."""
    if not user_can_admin(request):
        raise Http404()
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first from Dashboard.')
        return redirect('core:generate_credentials_choice')
    students = Student.objects.filter(department=dept, user__isnull=False).select_related('user', 'batch').annotate(roll_no_int=Cast('roll_no', IntegerField())).order_by('batch__name', 'roll_no_int', 'roll_no')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Student Credentials'
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    headers = ['Department', 'Batch', 'Roll No', 'Enrollment No', 'Name', 'Username', 'Note']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    for row_idx, s in enumerate(students, 2):
        ws.cell(row_idx, 1, s.department.name).border = thin_border
        ws.cell(row_idx, 2, s.batch.name).border = thin_border
        ws.cell(row_idx, 3, s.roll_no).border = thin_border
        ws.cell(row_idx, 4, s.enrollment_no or '').border = thin_border
        ws.cell(row_idx, 5, s.name).border = thin_border
        ws.cell(row_idx, 6, s.user.username).border = thin_border
        ws.cell(row_idx, 7, 'Password set at creation. Use Change Password to reset.').border = thin_border
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 42
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    safe_dept = re.sub(r'[^\w\-]', '_', dept.name)[:50]
    fname = f'Student_Credentials_{safe_dept}.xlsx'
    resp = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


# ---------- Admin: Students ----------

@login_required
def student_list(request):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    batch_id = request.GET.get('batch')
    q = request.GET.get('q', '').strip()
    students = Student.objects.filter(department=dept) if dept else Student.objects.none()
    if batch_id and dept:
        students = students.filter(batch_id=batch_id)
    if q:
        from django.db.models import Q
        students = students.filter(
            Q(roll_no__icontains=q) | Q(name__icontains=q) | Q(enrollment_no__icontains=q)
        )
    students = students.select_related('batch', 'mentor').annotate(
        roll_no_int=Cast('roll_no', IntegerField())
    ).order_by('batch__name', 'roll_no_int', 'roll_no')
    from django.core.paginator import Paginator
    paginator = Paginator(students, 25)
    page = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(int(page))
    except (ValueError, Paginator.EmptyPage):
        page_obj = paginator.page(1)
    batches = Batch.objects.filter(department=dept) if dept else []
    ctx = {
        'students': page_obj,
        'page_obj': page_obj,
        'batches': batches,
        'department': dept,
        'selected_batch_id': batch_id,
        'search_q': q,
    }
    return render(request, 'core/admin/student_list.html', ctx)


@login_required
def student_upload(request):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    batches = Batch.objects.filter(department=dept)
    if request.method == 'POST':
        batch_id = request.POST.get('batch')
        file = request.FILES.get('file')
        if not batch_id or not file:
            messages.error(request, 'Select batch and upload CSV. See format info for required columns.')
            return redirect('core:student_upload')
        batch = Batch.objects.filter(pk=batch_id, department=dept).first()
        if not batch:
            messages.error(request, 'Invalid batch.')
            return redirect('core:student_upload')
        try:
            decoded = file.read().decode('utf-8-sig')
            reader = csv.DictReader(decoded.splitlines())
            rows = []
            faculties = list(Faculty.objects.filter(department=dept))

            def get_col(row, *names):
                for n in names:
                    for k in row:
                        if k.strip().lower().replace(' ', '_') == n.lower().replace(' ', '_'):
                            return (row.get(k) or '').strip()
                return ''

            for row in reader:
                rn = get_col(row, 'roll_no', 'roll no')
                nm = get_col(row, 'name')
                en = get_col(row, 'enrollment_no', 'enrollment no')
                mentor_val = get_col(row, 'mentor')
                student_phone = get_col(row, 'student_phone_number', 'student phone number', 'student_phone')
                parents_contact = get_col(row, 'parents_contact_number', 'parents contact number', 'parents_contact')
                mentor = None
                if mentor_val:
                    for f in faculties:
                        if mentor_val.lower() in (f.short_name.lower(), f.full_name.lower()):
                            mentor = f
                            break
                if rn and nm:
                    rows.append({
                        'roll_no': rn, 'name': nm, 'enrollment_no': en, 'mentor': mentor,
                        'student_phone_number': student_phone, 'parents_contact_number': parents_contact,
                    })
            if not rows:
                messages.error(request, 'No valid rows (need roll_no, name).')
                return redirect('core:student_upload')
            Student.objects.filter(department=dept, batch=batch).delete()
            for r in rows:
                Student.objects.create(
                    department=dept, batch=batch,
                    roll_no=r['roll_no'], name=r['name'], enrollment_no=r.get('enrollment_no', ''),
                    mentor=r.get('mentor'),
                    student_phone_number=r.get('student_phone_number', ''),
                    parents_contact_number=r.get('parents_contact_number', ''),
                )
            messages.success(request, f'{len(rows)} students uploaded for {batch.name}.')
            return redirect('core:student_list')
        except Exception as e:
            messages.error(request, str(e))
    ctx = {'batches': batches, 'department': dept}
    return render(request, 'core/admin/student_upload.html', ctx)


@login_required
def student_add(request):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    batches = Batch.objects.filter(department=dept)
    faculties = Faculty.objects.filter(department=dept).order_by('short_name')
    if request.method == 'POST':
        roll_no = request.POST.get('roll_no', '').strip()
        name = request.POST.get('name', '').strip()
        batch_id = request.POST.get('batch_id', '').strip()
        enrollment_no = request.POST.get('enrollment_no', '').strip()
        mentor_id = request.POST.get('mentor_id', '').strip()
        student_phone = request.POST.get('student_phone_number', '').strip()
        parents_contact = request.POST.get('parents_contact_number', '').strip()
        if not roll_no or not name or not batch_id:
            messages.error(request, 'Roll No, Name, and Batch are required.')
            return redirect('core:student_add')
        batch = Batch.objects.filter(pk=batch_id, department=dept).first()
        if not batch:
            messages.error(request, 'Invalid batch.')
            return redirect('core:student_add')
        if Student.objects.filter(department=dept, batch=batch, roll_no=roll_no).exists():
            messages.error(request, f'Student with roll no {roll_no} already exists in batch {batch.name}.')
            return redirect('core:student_add')
        mentor = Faculty.objects.filter(pk=mentor_id, department=dept).first() if mentor_id else None
        Student.objects.create(
            department=dept, batch=batch,
            roll_no=roll_no, name=name, enrollment_no=enrollment_no,
            mentor=mentor,
            student_phone_number=student_phone, parents_contact_number=parents_contact,
        )
        messages.success(request, f'Student {name} added.')
        return redirect('core:student_list')
    ctx = {'batches': batches, 'faculties': faculties, 'department': dept}
    return render(request, 'core/admin/student_form.html', {'form_type': 'add', **ctx})


@login_required
def student_edit(request, pk):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    obj = get_object_or_404(Student, pk=pk)
    dept = get_admin_department(request)
    if dept and obj.department != dept:
        messages.error(request, 'You can only edit students in your department.')
        return redirect('core:student_list')
    batches = Batch.objects.filter(department=obj.department)
    faculties = Faculty.objects.filter(department=obj.department).order_by('short_name')
    if request.method == 'POST':
        roll_no = request.POST.get('roll_no', '').strip()
        name = request.POST.get('name', '').strip()
        batch_id = request.POST.get('batch_id', '').strip()
        enrollment_no = request.POST.get('enrollment_no', '').strip()
        mentor_id = request.POST.get('mentor_id', '').strip()
        student_phone = request.POST.get('student_phone_number', '').strip()
        parents_contact = request.POST.get('parents_contact_number', '').strip()
        if not roll_no or not name or not batch_id:
            messages.error(request, 'Roll No, Name, and Batch are required.')
            return redirect('core:student_edit', pk=pk)
        batch = Batch.objects.filter(pk=batch_id, department=obj.department).first()
        if not batch:
            messages.error(request, 'Invalid batch.')
            return redirect('core:student_edit', pk=pk)
        dup = Student.objects.filter(department=obj.department, batch=batch, roll_no=roll_no).exclude(pk=pk)
        if dup.exists():
            messages.error(request, f'Student with roll no {roll_no} already exists in batch {batch.name}.')
            return redirect('core:student_edit', pk=pk)
        mentor = Faculty.objects.filter(pk=mentor_id, department=obj.department).first() if mentor_id else None
        obj.roll_no = roll_no
        obj.name = name
        obj.batch = batch
        obj.enrollment_no = enrollment_no
        obj.mentor = mentor
        obj.student_phone_number = student_phone
        obj.parents_contact_number = parents_contact
        obj.save()
        messages.success(request, 'Student updated.')
        return redirect('core:student_list')
    ctx = {'obj': obj, 'batches': batches, 'faculties': faculties, 'department': obj.department}
    return render(request, 'core/admin/student_form.html', {'form_type': 'edit', **ctx})


@login_required
def student_delete(request, pk):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    obj = get_object_or_404(Student, pk=pk)
    dept = get_admin_department(request)
    if dept and obj.department != dept:
        messages.error(request, 'You can only manage students in your department.')
        return redirect('core:student_list')
    name = obj.name
    obj.delete()
    messages.success(request, f'Student "{name}" deleted.')
    return redirect('core:student_list')


# ---------- Admin: Schedule ----------

@login_required
def schedule_list(request):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    today = datetime.now().date()
    slots = sorted(
        _effective_slots_for_date(dept, today),
        key=lambda s: (s.day or '', s.time_slot or '')
    ) if dept else []
    ctx = {'slots': slots, 'department': dept}
    return render(request, 'core/admin/schedule_list.html', ctx)


@login_required
def schedule_add(request):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    if request.method == 'POST':
        faculty_id = request.POST.get('faculty')
        subject_id = request.POST.get('subject')
        batch_id = request.POST.get('batch')
        day = request.POST.get('day', '').strip()
        time_slot = request.POST.get('time_slot', '').strip()
        if faculty_id and subject_id and batch_id and day and time_slot:
            faculty = Faculty.objects.filter(pk=faculty_id, department=dept).first()
            subject = Subject.objects.filter(pk=subject_id, department=dept).first()
            batch = Batch.objects.filter(pk=batch_id, department=dept).first()
            if faculty and subject and batch:
                ScheduleSlot.objects.get_or_create(
                    department=dept, batch=batch, day=day, time_slot=time_slot,
                    effective_from=datetime.now().date(),
                    defaults={'faculty': faculty, 'subject': subject}
                )
                messages.success(request, 'Schedule slot added.')
                return redirect('core:schedule_list')
    faculties = Faculty.objects.filter(department=dept)
    subjects = Subject.objects.filter(department=dept)
    batches = Batch.objects.filter(department=dept)
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    ctx = {
        'department': dept, 'faculties': faculties, 'subjects': subjects,
        'batches': batches, 'days': days, 'slot': None,
    }
    return render(request, 'core/admin/schedule_form.html', ctx)


@login_required
def schedule_delete(request, pk):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    slot = get_object_or_404(ScheduleSlot, pk=pk)
    if slot.department != dept:
        messages.error(request, 'You can only manage schedule for your selected department.')
        return redirect('core:schedule_list')
    slot.delete()
    messages.success(request, 'Slot removed.')
    return redirect('core:schedule_list')


@login_required
def schedule_edit(request, pk):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    slot = get_object_or_404(ScheduleSlot, pk=pk)
    if slot.department != dept:
        messages.error(request, 'You can only edit schedule for your selected department.')
        return redirect('core:schedule_list')
    if request.method == 'POST':
        faculty_id = request.POST.get('faculty')
        subject_id = request.POST.get('subject')
        batch_id = request.POST.get('batch')
        day = request.POST.get('day', '').strip()
        time_slot = request.POST.get('time_slot', '').strip()
        if faculty_id and subject_id and batch_id and day and time_slot:
            faculty = Faculty.objects.filter(pk=faculty_id, department=dept).first()
            subject = Subject.objects.filter(pk=subject_id, department=dept).first()
            batch = Batch.objects.filter(pk=batch_id, department=dept).first()
            if faculty and subject and batch:
                slot.faculty = faculty
                slot.subject = subject
                slot.batch = batch
                slot.day = day
                slot.time_slot = time_slot
                slot.save()
                messages.success(request, 'Schedule slot updated.')
                return redirect('core:schedule_list')
    faculties = Faculty.objects.filter(department=dept)
    subjects = Subject.objects.filter(department=dept)
    batches = Batch.objects.filter(department=dept)
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    ctx = {
        'slot': slot,
        'department': dept,
        'faculties': faculties,
        'subjects': subjects,
        'batches': batches,
        'days': days,
    }
    return render(request, 'core/admin/schedule_form.html', ctx)


@login_required
def schedule_clear_all(request):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    if request.method != 'POST':
        return redirect('core:schedule_list')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    deleted, _ = ScheduleSlot.objects.filter(department=dept).delete()
    messages.success(request, f'All schedule entries for {dept.name} have been deleted ({deleted} slot(s)).')
    return redirect('core:schedule_list')


# ---------- Admin: Lecture Adjustment ----------

@login_required
def lecture_adjustment(request):
    """Adjust lectures for a date (e.g. substitute faculty): select date, batch, then add changes and apply. History shown below."""
    if not user_can_admin(request):
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')

    batches = Batch.objects.filter(department=dept).order_by('name')
    faculties = Faculty.objects.filter(department=dept).order_by('full_name')
    subjects = Subject.objects.filter(department=dept).order_by('name')

    # Pending changes in session: list of {time_slot, new_subject_id, new_faculty_id, ...}
    pending_key = 'lecture_adj_pending'
    pending = request.session.get(pending_key, [])

    if request.method == 'POST':
        action = request.POST.get('action')
        date_str = request.POST.get('date')
        batch_id = request.POST.get('batch')
        if action == 'add':
            time_slot = request.POST.get('time_slot', '').strip()
            new_subject_id = request.POST.get('new_subject')
            new_faculty_id = request.POST.get('new_faculty')
            if date_str and batch_id and time_slot and new_subject_id and new_faculty_id:
                # Remove any existing pending for same slot
                pending = [p for p in pending if p.get('time_slot') != time_slot or p.get('_date') != date_str or p.get('_batch_id') != batch_id]
                pending.append({
                    '_date': date_str, '_batch_id': batch_id,
                    'time_slot': time_slot,
                    'new_subject_id': new_subject_id,
                    'new_faculty_id': new_faculty_id,
                })
                request.session[pending_key] = pending
                request.session.modified = True
                messages.success(request, 'Change added. Add more or click "Apply changes" below.')
            return redirect(reverse('core:lecture_adjustment') + f'?date={date_str}&batch={batch_id}')
        if action == 'apply' and date_str and batch_id:
            try:
                selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except Exception:
                messages.error(request, 'Invalid date.')
                return redirect('core:lecture_adjustment')
            batch = Batch.objects.filter(pk=batch_id, department=dept).first()
            if not batch:
                messages.error(request, 'Invalid batch.')
                return redirect('core:lecture_adjustment')
            weekday = selected_date.strftime('%A')
            to_apply = [p for p in pending if p.get('_date') == date_str and str(p.get('_batch_id')) == str(batch_id)]
            for p in to_apply:
                slots = [s for s in _effective_slots_for_date(dept, selected_date, extra_filters={'batch': batch}) if s.day == weekday and s.time_slot == p.get('time_slot')]
                slot = slots[0] if slots else None
                if not slot:
                    continue
                new_faculty = Faculty.objects.filter(pk=p.get('new_faculty_id'), department=dept).first()
                new_subject = Subject.objects.filter(pk=p.get('new_subject_id'), department=dept).first()
                if not new_faculty or not new_subject:
                    continue
                LectureAdjustment.objects.update_or_create(
                    date=selected_date,
                    batch=batch,
                    time_slot=slot.time_slot,
                    defaults={
                        'original_faculty': slot.faculty,
                        'original_subject': slot.subject,
                        'new_faculty': new_faculty,
                        'new_subject': new_subject,
                    },
                )
            pending = [p for p in pending if not (p.get('_date') == date_str and str(p.get('_batch_id')) == str(batch_id))]
            request.session[pending_key] = pending
            request.session.modified = True
            messages.success(request, f'Applied {len(to_apply)} adjustment(s). They will reflect in all Excel and displays.')
            return redirect(reverse('core:lecture_adjustment') + f'?date={date_str}&batch={batch_id}')
        if action == 'clear_pending':
            request.session[pending_key] = []
            request.session.modified = True
            return redirect('core:lecture_adjustment')

    selected_date = None
    selected_batch = None
    day_slots = []
    date_str = request.GET.get('date')
    batch_id = request.GET.get('batch')
    if date_str and batch_id:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except Exception:
            pass
        selected_batch = Batch.objects.filter(pk=batch_id, department=dept).first()
        if selected_date and selected_batch:
            weekday = selected_date.strftime('%A')
            slots = [s for s in _effective_slots_for_date(dept, selected_date, extra_filters={'batch': selected_batch}) if s.day == weekday]
            slots = sorted(slots, key=lambda s: s.time_slot or '')
            # Check existing adjustments for this date/batch
            existing_adj = {a.time_slot: (a.new_faculty, a.new_subject) for a in
                           LectureAdjustment.objects.filter(date=selected_date, batch=selected_batch).select_related('new_faculty', 'new_subject')}
            for i, slot in enumerate(slots, 1):
                fac, subj = existing_adj.get(slot.time_slot, (slot.faculty, slot.subject))
                day_slots.append({
                    'lecture_num': i,
                    'time_slot': slot.time_slot,
                    'subject': subj.name,
                    'faculty': fac.short_name,
                    'original_subject': slot.subject.name,
                    'original_faculty': slot.faculty.short_name,
                    'is_adjusted': slot.time_slot in existing_adj,
                })

    # Pending for this page's date/batch
    pending_for_page = [p for p in pending if p.get('_date') == date_str and str(p.get('_batch_id')) == str(batch_id)] if date_str and batch_id else []
    for p in pending_for_page:
        new_f = Faculty.objects.filter(pk=p.get('new_faculty_id')).first()
        new_s = Subject.objects.filter(pk=p.get('new_subject_id')).first()
        p['new_faculty_name'] = new_f.short_name if new_f else '—'
        p['new_subject_name'] = new_s.name if new_s else '—'

    # History: recent adjustments (for this dept)
    batch_ids = list(batches.values_list('pk', flat=True))
    history = LectureAdjustment.objects.filter(
        batch_id__in=batch_ids
    ).select_related('batch', 'new_faculty', 'new_subject', 'original_faculty', 'original_subject').order_by('-date', '-created_at')[:50]

    ctx = {
        'department': dept,
        'batches': batches,
        'faculties': faculties,
        'subjects': subjects,
        'selected_date': selected_date,
        'date_str': date_str or '',
        'selected_batch': selected_batch,
        'batch_id': batch_id or '',
        'day_slots': day_slots,
        'pending': pending_for_page,
        'history': history,
    }
    return render(request, 'core/admin/lecture_adjustment.html', ctx)


# ---------- Admin: Upload Timetable (Excel) ----------

def _looks_like_timing_header(val):
    """Return True if header cell looks like timing/time/slot, not a batch name."""
    if not val:
        return True
    s = str(val).strip().lower()
    if not s:
        return True
    if s in ('timing', 'time', 'slot', 'slots', 'lecture', 'lec', 'period', 'sr no', 's.no', 'no.'):
        return True
    if re.match(r'^lec\s*\d+$', s):
        return True
    if re.match(r'^\d{1,2}:\d{2}\s*[-–—to]+\s*\d{1,2}:\d{2}', s):
        return True
    if re.match(r'^\d{1,2}-\d{1,2}$', s):
        return True
    return False


def _normalize_day(excel_day):
    """Map Mon, Tue, ... to Monday, Tuesday, ..."""
    if not excel_day or not str(excel_day).strip():
        return None
    s = str(excel_day).strip().lower()[:3]
    day_map = {
        'mon': 'Monday', 'tue': 'Tuesday', 'wed': 'Wednesday',
        'thu': 'Thursday', 'fri': 'Friday', 'sat': 'Saturday', 'sun': 'Sunday',
    }
    return day_map.get(s, str(excel_day).strip().capitalize())


def _normalize_time_slot(val):
    """Normalize time slot string (e.g. 08:45-09:45, 8:45 to 9:45, Lec 1)."""
    if not val:
        return None
    t = str(val).strip().replace('to', '-').replace('–', '-').replace('—', '-').replace(' ', '')
    if t:
        return t
    return str(val).strip()


def _batch_date_lecture_labels(dept, batch, d):
    """Map time_slot -> 'Lec N' — same ordering as Lecture Adjustment (that weekday's slots for batch, sorted by time_slot), then ExtraLecture times."""
    weekday = d.strftime('%A')
    base_slots = sorted(
        [s for s in _effective_slots_for_date(dept, d, extra_filters={'batch': batch}) if s.day == weekday],
        key=lambda s: s.time_slot or ''
    )
    labels = {}
    for i, s in enumerate(base_slots, start=1):
        ts = (s.time_slot or '').strip()
        if ts:
            labels[ts] = f'Lec {i}'
    scheduled = set(labels.keys())
    extra_times = sorted(
        set(ExtraLecture.objects.filter(date=d, batch=batch).values_list('time_slot', flat=True))
        - scheduled
    )
    n = len(labels)
    for j, ts in enumerate(extra_times):
        ts = (ts or '').strip()
        if ts:
            labels[ts] = f'Lec {n + j + 1}'
    return labels


def _lecture_label_for_slot(dept, batch, d, time_slot_str, fallback_index=1):
    """Excel column header label: Lec 1, Lec 2, … matching Lecture Adjustment; extra slots continue numbering."""
    labels = _batch_date_lecture_labels(dept, batch, d)
    ts = (time_slot_str or '').strip()
    if ts in labels:
        return labels[ts]
    if ts:
        return ts
    return f'Lec {fallback_index}'


def _user_can_daily_report(request):
    """HOD or Super Admin only (not departmental admin)."""
    if not request.user.is_authenticated:
        return False
    return is_hod(request) or is_super_admin(request)


# Roman numerals for Daily Report semester (user-selected on export form).
DR_SEMESTER_ROMAN_OPTIONS = ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X', 'XI', 'XII']
_DR_SEMESTER_ROMAN_CHOICES = frozenset(DR_SEMESTER_ROMAN_OPTIONS)


def _dr_normalize_semester(value):
    """Return trimmed semester label or '' if invalid."""
    if value is None:
        return ''
    s = re.sub(r'\s+', '', str(value).strip().upper())
    if not s:
        return ''
    romans = {
        '1': 'I', '2': 'II', '3': 'III', '4': 'IV', '5': 'V', '6': 'VI',
        '7': 'VII', '8': 'VIII', '9': 'IX', '10': 'X', '11': 'XI', '12': 'XII',
    }
    if s.isdigit() and s in romans:
        s = romans[s]
    return s if s in _DR_SEMESTER_ROMAN_CHOICES else ''


def _dr_semester_for_department(dept):
    """Display semester for DR exports (from Department.semester)."""
    if not dept:
        return '—'
    raw = (getattr(dept, 'semester', None) or '').strip()
    if not raw:
        return '—'
    norm = _dr_normalize_semester(raw)
    return norm if norm else raw[:20]


def _dr_faculty_display_name(faculty):
    return (faculty.full_name or '').strip()


def _dr_slots_for_batch_on_date(dept, batch, d):
    """Schedule + extra lectures for this batch/date, sorted by time_slot; excludes cancellations."""
    cancelled = get_cancelled_lectures_set(dept)
    weekday = d.strftime('%A')
    slots = [s for s in _effective_slots_for_date(dept, d, extra_filters={'batch': batch}) if s.day == weekday]
    seen = {(s.time_slot or '').strip() for s in slots if s.time_slot}
    for ex in ExtraLecture.objects.filter(date=d, batch=batch):
        ts = (ex.time_slot or '').strip()
        if not ts or ts in seen:
            continue
        if (d, batch.id, ex.time_slot) in cancelled:
            continue
        seen.add(ts)
        virtual = type('S', (), {'time_slot': ex.time_slot, 'subject': ex.subject, 'faculty': ex.faculty})()
        slots.append(virtual)
    slots.sort(key=lambda s: (s.time_slot or ''))
    return [s for s in slots if (d, batch.id, s.time_slot) not in cancelled]


def _dr_lec_no_from_label(lec_label):
    s = (lec_label or '').strip()
    low = s.lower()
    if low.startswith('lec'):
        parts = s.split()
        if len(parts) >= 2:
            try:
                return int(parts[1])
            except ValueError:
                pass
    return 0


DR_FIXED_LECTURE_ROWS = 5


def _dr_expand_fixed_rows_per_faculty(sorted_rows):
    """
    Exactly DR_FIXED_LECTURE_ROWS rows per faculty (Lec. 1–5). Places each real lecture in the
    row matching its timetable lecture number when possible; conflicts use the next free row.
    """
    from itertools import groupby
    if not sorted_rows:
        return []
    out = []
    dr = DR_FIXED_LECTURE_ROWS
    for _fac_id, group in groupby(sorted_rows, key=lambda r: r['faculty'].id):
        group = list(group)
        fac = group[0]['faculty']
        sem = group[0].get('sem') or '—'
        dept_label = group[0].get('dept') or ''
        by_slot = []
        no_num = []
        for r in group:
            ln = int(r.get('lec_no') or 0)
            if 1 <= ln <= dr:
                by_slot.append((ln, r))
            elif ln > dr:
                by_slot.append((dr, r))
            else:
                no_num.append(r)
        by_slot.sort(key=lambda x: (x[0], x[1]['batch'].name))
        placed = {}
        for ln, r in by_slot:
            slot = ln
            while slot <= dr and slot in placed:
                slot += 1
            if slot > dr:
                free = next((i for i in range(1, dr + 1) if i not in placed), None)
                if free is not None:
                    placed[free] = r
                else:
                    placed[dr] = r
            else:
                placed[slot] = r
        for r in sorted(no_num, key=lambda x: x['batch'].name):
            free = next((i for i in range(1, dr + 1) if i not in placed), None)
            if free is not None:
                placed[free] = r
            else:
                placed[dr] = r
        for i in range(1, dr + 1):
            if i in placed:
                row = dict(placed[i])
                row['lec_no'] = i
                row['lec_label'] = f'Lec {i}'
                row['is_blank'] = False
                out.append(row)
            else:
                out.append({
                    'faculty': fac,
                    'batch': None,
                    'subject': None,
                    'lec_no': i,
                    'lec_label': f'Lec {i}',
                    'initials': fac.short_name,
                    'course': '',
                    'dept': dept_label,
                    'sem': sem,
                    'div': '',
                    'sub_name': '',
                    'lecture_type': '',
                    'proxy': '',
                    'present': '',
                    'total': '',
                    'eff': 0,
                    'attendance_filled': False,
                    'is_extra_lecture': False,
                    'is_blank': True,
                })
    return out


def _dr_collect_rows_for_date(dept, d):
    """Teaching rows for date, then expanded to exactly 5 rows per faculty (Lec. 1–5); blanks fill gaps."""
    sem = _dr_semester_for_department(dept)
    dept_label = (dept.name or '')[:80]
    batch_ids = list(Batch.objects.filter(department=dept).values_list('id', flat=True))
    strength = {}
    for bid in batch_ids:
        strength[bid] = Student.objects.filter(batch_id=bid).count()
    rows = []
    for batch in Batch.objects.filter(department=dept).order_by('name'):
        labels = _batch_date_lecture_labels(dept, batch, d)
        for slot in _dr_slots_for_batch_on_date(dept, batch, d):
            ts = (slot.time_slot or '').strip()
            if not ts:
                continue
            fac, subj = get_faculty_subject_for_slot(d, batch, ts)
            if not fac or not subj:
                continue
            lec_label = labels.get(ts) or _lecture_label_for_slot(dept, batch, d, ts)
            lec_no = _dr_lec_no_from_label(lec_label)
            is_extra = _is_extra_lecture_slot(dept, d, batch, ts)
            adj = None
            if is_extra:
                proxy = ''
                lecture_type = 'ETL'
            else:
                adj = LectureAdjustment.objects.filter(
                    date=d, batch=batch, time_slot=slot.time_slot
                ).select_related('original_faculty', 'new_faculty').first()
                proxy = adj.original_faculty.short_name if adj else ''
                lecture_type = 'PTL' if adj else 'TL'
            tot = strength.get(batch.id, 0)
            att = FacultyAttendance.objects.filter(date=d, batch=batch, lecture_slot=ts, faculty=fac).first()
            if not att:
                att = FacultyAttendance.objects.filter(date=d, batch=batch, lecture_slot=ts).first()
            attendance_filled = att is not None
            if attendance_filled:
                absent_n = len([x for x in (att.absent_roll_numbers or '').split(',') if x.strip()])
                present = tot - absent_n
                eff = _dr_slot_effective_load(is_extra, True, present)
            else:
                present = '—'
                eff = 0
            rows.append({
                'faculty': fac,
                'batch': batch,
                'subject': subj,
                'lec_no': lec_no,
                'lec_label': lec_label,
                'initials': fac.short_name,
                'course': 'UG',
                'dept': dept_label,
                'sem': sem,
                'div': batch.name,
                'sub_name': subj.name,
                'lecture_type': lecture_type,
                'proxy': proxy,
                'present': present,
                'total': tot,
                'eff': eff,
                'attendance_filled': attendance_filled,
                'is_extra_lecture': is_extra,
                'is_blank': False,
            })
    rows.sort(key=lambda r: (r['faculty'].full_name.lower(), r['lec_no'], r['batch'].name))
    return _dr_expand_fixed_rows_per_faculty(rows)


def _dr_write_daily_sheet(ws, dept, d, college_name):
    """One date tab: DR layout with merged faculty blocks."""
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    block_side = Side(style='medium', color='5C6B7A')
    thick_sep = Side(style='thick', color='1F2937')
    hdr_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
    present_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    eff_fill = PatternFill(start_color='D9D2E9', end_color='D9D2E9', fill_type='solid')
    band_fills = [
        PatternFill(start_color='F7F9FC', end_color='F7F9FC', fill_type='solid'),
        PatternFill(start_color='E8EDF5', end_color='E8EDF5', fill_type='solid'),
    ]
    etl_row_fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
    etl_data_font = Font(name='Calibri', size=10, color='FFFFFF')
    etl_num_font = Font(name='Calibri', size=10, color='FFFFFF', bold=True)
    ptl_row_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    ptl_data_font = Font(name='Calibri', size=10, color='006100', bold=True)
    ptl_num_font = Font(name='Calibri', size=10, color='006100', bold=True)
    title_font = Font(name='Calibri', size=14, bold=True)
    subtitle_font = Font(name='Calibri', size=11, bold=True)
    hdr_font = Font(name='Calibri', size=10, bold=True)
    data_font = Font(name='Calibri', size=10)
    num_font = Font(name='Calibri', size=10)
    present_missing_font = Font(name='Calibri', size=10, bold=True, color='CC0000')
    rows = _dr_collect_rows_for_date(dept, d)
    ncols = 16
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, value=f'COLLEGE NAME : {college_name}')
    c.font = title_font
    c.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c2 = ws.cell(2, 1, value=f'DAILY REPORT — {d.strftime("%A")}, {d.strftime("%d-%b-%Y")} — Sem {_dr_semester_for_department(dept)}')
    c2.font = subtitle_font
    c2.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
    ws.cell(3, 1, value='FACULTY DETAILS').font = hdr_font
    ws.cell(3, 1).fill = hdr_fill
    ws.cell(3, 1).alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=ncols)
    ws.cell(3, 3, value='LECTURE DETAILS').font = hdr_font
    ws.cell(3, 3).fill = hdr_fill
    ws.cell(3, 3).alignment = Alignment(horizontal='center', vertical='center')

    hdrs = [
        'Sr No', 'Name of Faculty', 'Lec. No', 'Faculty Initials',
        'Course (UG/PG)', 'Dept', 'SEM (Roman)', 'Div', 'Batch', 'Sub.', 'Lecture type', 'Faculty Initials for Proxy',
        'No. of Students Present', 'Total Students', 'Effective Load', "Today's Eff. Load Alloted"
    ]
    for col, h in enumerate(hdrs, start=1):
        cell = ws.cell(4, col, value=h)
        cell.font = hdr_font
        cell.border = thin
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if col >= 12 and col <= 13:
            cell.fill = present_fill
        elif col == 15:
            cell.fill = eff_fill
        elif col == 16:
            cell.fill = eff_fill

    if not rows:
        msg = ws.cell(5, 1, value='No scheduled lectures for this date.')
        msg.font = data_font
        for col_letter, w in zip(
            'ABCDEFGHIJKLMNOP',
            (6, 32, 8, 10, 8, 12, 8, 10, 8, 22, 8, 12, 12, 10, 10, 12),
        ):
            ws.column_dimensions[col_letter].width = w
        return

    from itertools import groupby
    data_row = 5
    sr_fac = 0
    group_idx = 0
    for fac_id, group in groupby(rows, key=lambda r: r['faculty'].id):
        group = list(group)
        sr_fac += 1
        band_fill = band_fills[group_idx % len(band_fills)]
        group_idx += 1
        day_total = sum(r['eff'] for r in group if not r.get('is_blank'))
        first_r = data_row
        for r in group:
            is_blank = r.get('is_blank')
            lt_raw = r.get('lecture_type')
            if is_blank:
                lt_disp = ''
                is_ex = is_ptl = False
                row_fill = band_fill
            else:
                lt_disp = lt_raw or 'TL'
                is_ex = lt_disp == 'ETL'
                is_ptl = lt_disp == 'PTL'
                if is_ex:
                    row_fill = etl_row_fill
                elif is_ptl:
                    row_fill = ptl_row_fill
                else:
                    row_fill = band_fill
            ws.cell(data_row, 1, value=sr_fac)
            ws.cell(data_row, 2, value=_dr_faculty_display_name(r['faculty']))
            ws.cell(data_row, 3, value=r['lec_no'] if r['lec_no'] else r['lec_label'])
            ws.cell(data_row, 4, value=r['initials'])
            ws.cell(data_row, 5, value='' if is_blank else r['course'])
            ws.cell(data_row, 6, value=r['dept'])
            ws.cell(data_row, 7, value=r['sem'])
            ws.cell(data_row, 8, value='' if is_blank else r['div'])
            ws.cell(data_row, 9, value='')
            ws.cell(data_row, 10, value='' if is_blank else r['sub_name'])
            ws.cell(data_row, 11, value=lt_disp)
            ws.cell(data_row, 12, value='' if is_blank else r['proxy'])
            ws.cell(data_row, 13, value='' if is_blank else r['present'])
            ws.cell(data_row, 14, value='' if is_blank else r['total'])
            ws.cell(data_row, 15, value='' if is_blank else r['eff'])
            for col in range(1, ncols + 1):
                cell = ws.cell(data_row, col)
                cell.border = thin
                if is_blank:
                    cell.font = data_font
                elif col == 13:
                    if r.get('attendance_filled'):
                        if is_ex:
                            cell.font = etl_num_font
                        elif is_ptl:
                            cell.font = ptl_num_font
                        else:
                            cell.font = num_font
                    else:
                        cell.font = present_missing_font
                elif col in (1, 3, 14, 15, 16):
                    if is_ex:
                        cell.font = etl_num_font
                    elif is_ptl:
                        cell.font = ptl_num_font
                    else:
                        cell.font = num_font
                else:
                    if is_ex:
                        cell.font = etl_data_font
                    elif is_ptl:
                        cell.font = ptl_data_font
                    else:
                        cell.font = data_font
                cell.fill = row_fill
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            data_row += 1
        last_r = data_row - 1
        if last_r > first_r:
            ws.merge_cells(start_row=first_r, start_column=1, end_row=last_r, end_column=1)
            ws.merge_cells(start_row=first_r, start_column=2, end_row=last_r, end_column=2)
            ws.merge_cells(start_row=first_r, start_column=16, end_row=last_r, end_column=16)
        tot_cell = ws.cell(first_r, 16, value=day_total)
        tot_cell.fill = eff_fill
        tot_cell.font = Font(name='Calibri', size=10, bold=True)
        tot_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(first_r, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(first_r, 2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for rr in range(first_r, last_r + 1):
            for cc in range(1, ncols + 1):
                b_kw = {'left': thin.left, 'right': thin.right, 'top': thin.top, 'bottom': thin.bottom}
                if rr == first_r:
                    b_kw['top'] = block_side
                if rr == last_r:
                    b_kw['bottom'] = thick_sep
                elif rr < last_r:
                    b_kw['bottom'] = thin.bottom
                if cc == 1:
                    b_kw['left'] = block_side
                if cc == ncols:
                    b_kw['right'] = block_side
                ws.cell(rr, cc).border = Border(**b_kw)

    for col_letter, w in zip(
        'ABCDEFGHIJKLMNOP',
        (6, 32, 8, 10, 8, 12, 8, 10, 8, 22, 8, 12, 12, 10, 10, 12),
    ):
        ws.column_dimensions[col_letter].width = w


def _dr_aggregate_week(dept, week_dates):
    """(fac_id, batch_id, subj_id) -> {'n': lecture count, 'eff': sum effective load} (attendance marked only)."""
    agg = defaultdict(lambda: {'n': 0, 'eff': 0.0})
    for d in week_dates:
        for r in _dr_collect_rows_for_date(dept, d):
            if r.get('is_blank') or not r.get('attendance_filled'):
                continue
            b, s = r.get('batch'), r.get('subject')
            if b is None or s is None:
                continue
            key = (r['faculty'].id, b.id, s.id)
            agg[key]['n'] += 1
            agg[key]['eff'] += float(r.get('eff') or 0)
    return agg


def _dr_write_combine_sheet(
    ws,
    dept,
    week_dates,
    phase,
    global_week_num,
    college_name,
    *,
    counts_override=None,
    title_line1=None,
    subtitle_line3=None,
    hdr_lectures='Total lectures (week)',
    hdr_overall='Weekly overall effective',
):
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    block_side = Side(style='medium', color='5C6B7A')
    thick_sep = Side(style='thick', color='1F2937')
    hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    hdr_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    title_font = Font(name='Calibri', size=14, bold=True)
    subtitle_font = Font(name='Calibri', size=11, bold=True)
    data_font = Font(name='Calibri', size=10)
    num_font = Font(name='Calibri', size=10)
    band_fills = [
        PatternFill(start_color='F7F9FC', end_color='F7F9FC', fill_type='solid'),
        PatternFill(start_color='E8EDF5', end_color='E8EDF5', fill_type='solid'),
    ]
    agg = counts_override if counts_override is not None else _dr_aggregate_week(dept, week_dates)
    fac_objs = {f.id: f for f in Faculty.objects.filter(department=dept)}
    batch_objs = {b.id: b for b in Batch.objects.filter(department=dept)}
    sub_objs = {s.id: s for s in Subject.objects.filter(department=dept)}
    dept_label = (dept.name or '')[:80]
    sem = _dr_semester_for_department(dept)

    keys_sorted = sorted(
        agg.keys(),
        key=lambda k: (
            fac_objs[k[0]].full_name.lower() if k[0] in fac_objs else '',
            batch_objs[k[1]].name if k[1] in batch_objs else '',
            sub_objs[k[2]].name.lower() if k[2] in sub_objs else '',
        ),
    )
    if title_line1 is None:
        title_line1 = 'WEEKLY REPORT — COMBINE'
    if subtitle_line3 is None:
        subtitle_line3 = f'{phase} — Week {global_week_num} — Sem {sem} ({len(week_dates)} lecture day(s))'

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    ws.cell(1, 1, value=title_line1).font = title_font
    ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
    ws.cell(2, 1, value=f'COLLEGE NAME : {college_name}').font = subtitle_font
    ws.cell(2, 1).alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=11)
    ws.cell(3, 1, value=subtitle_line3).font = subtitle_font
    ws.cell(3, 1).alignment = Alignment(horizontal='center', vertical='center')

    hdrs = [
        'Sr No', 'Name of Faculty', 'Short Form', 'Course', 'DEPT', 'SEM', 'DIV', 'SUBJECT',
        hdr_lectures, 'Total effective load (0.75×)', hdr_overall,
    ]
    for col, h in enumerate(hdrs, start=1):
        cell = ws.cell(4, col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = thin
        cell.alignment = Alignment(wrap_text=True, vertical='center')

    if not keys_sorted:
        ws.cell(5, 1, value='No data for this week.').font = data_font
        return

    from itertools import groupby
    row = 5
    sr = 0
    group_idx = 0
    for fac_id, fac_keys_iter in groupby(keys_sorted, key=lambda k: k[0]):
        fac_keys = list(fac_keys_iter)
        sr += 1
        band_fill = band_fills[group_idx % len(band_fills)]
        group_idx += 1
        fac = fac_objs.get(fac_id)
        week_sum = sum(agg[k]['eff'] for k in fac_keys)
        first_r = row
        for (fid, bid, sid) in fac_keys:
            cell_agg = agg[(fid, bid, sid)]
            nlec = cell_agg['n']
            row_eff = cell_agg['eff']
            b = batch_objs.get(bid)
            su = sub_objs.get(sid)
            ws.cell(row, 1, value=sr)
            ws.cell(row, 2, value=_dr_faculty_display_name(fac) if fac else '')
            ws.cell(row, 3, value=fac.short_name if fac else '')
            ws.cell(row, 4, value='UG')
            ws.cell(row, 5, value=dept_label)
            ws.cell(row, 6, value=sem)
            ws.cell(row, 7, value=b.name if b else '')
            ws.cell(row, 8, value=su.name if su else '')
            ws.cell(row, 9, value=nlec)
            ws.cell(row, 10, value=round(row_eff, 2))
            for col in range(1, 12):
                cell = ws.cell(row, col)
                cell.border = thin
                cell.fill = band_fill
                cell.font = num_font if col in (1, 9, 10, 11) else data_font
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            row += 1
        last_r = row - 1
        if last_r > first_r:
            ws.merge_cells(start_row=first_r, start_column=1, end_row=last_r, end_column=1)
            ws.merge_cells(start_row=first_r, start_column=2, end_row=last_r, end_column=2)
            ws.merge_cells(start_row=first_r, start_column=3, end_row=last_r, end_column=3)
            ws.merge_cells(start_row=first_r, start_column=11, end_row=last_r, end_column=11)
        tot = ws.cell(first_r, 11, value=round(week_sum, 2))
        tot.font = Font(name='Calibri', size=10, bold=True)
        tot.alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(first_r, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(first_r, 2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.cell(first_r, 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for rr in range(first_r, last_r + 1):
            for cc in range(1, 12):
                b_kw = {'left': thin.left, 'right': thin.right, 'top': thin.top, 'bottom': thin.bottom}
                if rr == first_r:
                    b_kw['top'] = block_side
                if rr == last_r:
                    b_kw['bottom'] = thick_sep
                elif rr < last_r:
                    b_kw['bottom'] = thin.bottom
                if cc == 1:
                    b_kw['left'] = block_side
                if cc == 11:
                    b_kw['right'] = block_side
                ws.cell(rr, cc).border = Border(**b_kw)

    for col_letter, w in zip(
        'ABCDEFGHIJK',
        (6, 32, 10, 8, 12, 8, 10, 26, 12, 14, 14),
    ):
        ws.column_dimensions[col_letter].width = w


def _build_daily_report_phase_workbook(dept, phase):
    """One sheet per week in phase (weekly combine only) + final phase-wide combined sheet."""
    college_name = getattr(settings, 'COLLEGE_DISPLAY_NAME', 'L.J. INSTITUTE OF ENGINEERING AND TECHNOLOGY')
    phases = ['T1', 'T2', 'T3', 'T4']
    week_map = {p: _compile_phase_weeks_date_objects(dept, p) for p in phases}
    phase_week_offsets = _get_phase_week_offsets(week_map)
    weeks_in_phase = week_map.get(phase, [])
    wb = Workbook()
    wb.remove(wb.active)
    all_phase_dates = []
    sem = _dr_semester_for_department(dept)
    for week_idx, week_dates in enumerate(weeks_in_phase):
        global_week_num = phase_week_offsets.get(phase, 0) + week_idx + 1
        ws = wb.create_sheet(title=f'Week-{global_week_num}'[:31])
        _dr_write_combine_sheet(
            ws,
            dept,
            week_dates,
            phase,
            global_week_num,
            college_name,
            title_line1='WEEKLY COMBINE',
            subtitle_line3=(
                f'{phase} — Week {global_week_num} — Sem {sem} '
                f'({len(week_dates)} lecture day(s))'
            ),
        )
        all_phase_dates.extend(week_dates)
    all_phase_dates = sorted(set(all_phase_dates))
    counts_phase = _dr_aggregate_week(dept, all_phase_dates)
    ws_p = wb.create_sheet(title='Combined'[:31])
    _dr_write_combine_sheet(
        ws_p,
        dept,
        all_phase_dates,
        phase,
        0,
        college_name,
        counts_override=counts_phase,
        title_line1='PHASE COMBINE — ALL WEEKS',
        subtitle_line3=(
            f'{phase} — Full phase — Sem {sem} '
            f'({len(all_phase_dates)} lecture day(s), {len(weeks_in_phase)} week(s))'
        ),
        hdr_lectures='Total lectures (phase)',
        hdr_overall='Phase overall effective',
    )
    return wb


def _build_daily_report_phases_range_workbook(dept, phases_included):
    """One combine sheet per week (global Week-1…N) across consecutive phases, then one Combined for the full date range."""
    college_name = getattr(settings, 'COLLEGE_DISPLAY_NAME', 'L.J. INSTITUTE OF ENGINEERING AND TECHNOLOGY')
    all_phases = ['T1', 'T2', 'T3', 'T4']
    week_map = {p: _compile_phase_weeks_date_objects(dept, p) for p in all_phases}
    phase_week_offsets = _get_phase_week_offsets(week_map)
    wb = Workbook()
    wb.remove(wb.active)
    all_range_dates = []
    sem = _dr_semester_for_department(dept)
    label_range = '+'.join(phases_included)
    for phase in phases_included:
        for week_idx, week_dates in enumerate(week_map.get(phase, [])):
            global_week_num = phase_week_offsets.get(phase, 0) + week_idx + 1
            ws = wb.create_sheet(title=f'Week-{global_week_num}'[:31])
            _dr_write_combine_sheet(
                ws,
                dept,
                week_dates,
                phase,
                global_week_num,
                college_name,
                title_line1='WEEKLY COMBINE',
                subtitle_line3=(
                    f'{phase} — Week {global_week_num} — Sem {sem} '
                    f'({len(week_dates)} lecture day(s))'
                ),
            )
            all_range_dates.extend(week_dates)
    all_range_dates = sorted(set(all_range_dates))
    n_weeks = sum(len(week_map.get(p, [])) for p in phases_included)
    counts_range = _dr_aggregate_week(dept, all_range_dates)
    ws_p = wb.create_sheet(title='Combined'[:31])
    _dr_write_combine_sheet(
        ws_p,
        dept,
        all_range_dates,
        phases_included[-1],
        0,
        college_name,
        counts_override=counts_range,
        title_line1='MULTI-PHASE COMBINE',
        subtitle_line3=(
            f'{label_range} — Sem {sem} '
            f'({len(all_range_dates)} lecture day(s), {n_weeks} week(s))'
        ),
        hdr_lectures='Total lectures (range)',
        hdr_overall='Range overall effective',
    )
    return wb


def _build_daily_report_workbook(dept, phase, week_index, week_dates, global_week_num):
    college_name = getattr(settings, 'COLLEGE_DISPLAY_NAME', 'L.J. INSTITUTE OF ENGINEERING AND TECHNOLOGY')
    wb = Workbook()
    wb.remove(wb.active)
    for d in week_dates:
        title = d.strftime('%d-%b-%y')
        ws = wb.create_sheet(title=title[:31])
        _dr_write_daily_sheet(ws, dept, d, college_name)
    combine_title = f'Combine W{global_week_num}'
    ws_c = wb.create_sheet(title=combine_title[:31])
    _dr_write_combine_sheet(ws_c, dept, week_dates, phase, global_week_num, college_name)
    return wb


@login_required
def daily_report_export(request):
    """HOD / Super Admin: page to download phase/week Daily Report Excel."""
    if not _user_can_daily_report(request):
        messages.error(request, 'Only HOD or Super Admin can download the Daily Report.')
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    if is_super_admin(request):
        dept_override = request.GET.get('department_id')
        if dept_override:
            alt = Department.objects.filter(pk=dept_override).first()
            if alt:
                dept = alt
    phases = ['T1', 'T2', 'T3', 'T4']
    week_map = {p: _compile_phase_weeks_date_objects(dept, p) for p in phases}
    phase_week_offsets = _get_phase_week_offsets(week_map)
    week_map_serial = {p: [[d.isoformat() for d in w] for w in week_map[p]] for p in phases}
    phase_weeks_list = []
    for p in phases:
        weeks = week_map.get(p, [])
        opts = [
            (i, phase_week_offsets.get(p, 0) + i + 1, len(weeks[i]) if i < len(weeks) else 0)
            for i in range(len(weeks))
        ]
        phase_weeks_list.append((p, opts))
    ctx = {
        'department': dept,
        'phases': phases,
        'phase_weeks_list': phase_weeks_list,
        'week_map_json': json.dumps(week_map_serial),
        'phase_week_offsets_json': json.dumps(phase_week_offsets),
        'departments': Department.objects.order_by('name') if is_super_admin(request) else [],
        'is_super_admin': is_super_admin(request),
    }
    return render(request, 'core/admin/daily_report_export.html', ctx)


@login_required
def daily_report_excel(request):
    """Download DR workbook: one sheet per date in week + Combine sheet."""
    if not _user_can_daily_report(request):
        messages.error(request, 'Only HOD or Super Admin can download the Daily Report.')
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    if not dept:
        return redirect('core:daily_report_export')
    if is_super_admin(request):
        dept_override = request.GET.get('department_id')
        if dept_override:
            alt = Department.objects.filter(pk=dept_override).first()
            if alt:
                dept = alt

    phases = ['T1', 'T2', 'T3', 'T4']
    safe_dept = re.sub(r'[^\w\-]+', '_', (dept.name or 'dept')[:40])
    sem_tag = re.sub(r'[^\w\-]+', '', _dr_semester_for_department(dept)) or 'Sem'
    report_scope = (request.GET.get('report_scope') or 'week').lower().strip()

    range_scope_phases = {
        'range_t1_t2': ['T1', 'T2'],
        'range_t1_t2_t3': ['T1', 'T2', 'T3'],
        'range_all': ['T1', 'T2', 'T3', 'T4'],
    }
    range_fname_tag = {
        'range_t1_t2': 'T1plusT2',
        'range_t1_t2_t3': 'T1plusT2plusT3',
        'range_all': 'AllT1toT4',
    }

    if report_scope in range_scope_phases:
        phases_subset = range_scope_phases[report_scope]
        week_map_chk = {p: _compile_phase_weeks_date_objects(dept, p) for p in phases}
        if sum(len(week_map_chk.get(p, [])) for p in phases_subset) == 0:
            messages.error(
                request,
                'No weeks in the selected range. Set term phases and schedule.',
            )
            return redirect('core:daily_report_export')
        wb = _build_daily_report_phases_range_workbook(dept, phases_subset)
        fname = f'DR_{safe_dept}_{range_fname_tag[report_scope]}_{sem_tag}.xlsx'
    else:
        phase = (request.GET.get('phase') or '').upper().strip()
        if phase not in phases:
            messages.error(request, 'Invalid phase.')
            return redirect('core:daily_report_export')

        if report_scope == 'phase':
            week_map_chk = {p: _compile_phase_weeks_date_objects(dept, p) for p in phases}
            if not week_map_chk.get(phase):
                messages.error(request, 'No weeks in this phase. Set term phases and schedule.')
                return redirect('core:daily_report_export')
            wb = _build_daily_report_phase_workbook(dept, phase)
            fname = f'DR_{safe_dept}_{phase}_FullPhase_{sem_tag}.xlsx'
        else:
            week_index_str = request.GET.get('week_index', '').strip()
            try:
                week_index = int(week_index_str)
            except (TypeError, ValueError):
                messages.error(request, 'Select a week.')
                return redirect('core:daily_report_export')
            week_map = {p: _compile_phase_weeks_date_objects(dept, p) for p in phases}
            weeks = week_map.get(phase, [])
            if not weeks:
                messages.error(request, 'No weeks in this phase. Set term phases and schedule.')
                return redirect('core:daily_report_export')
            if not (0 <= week_index < len(weeks)):
                messages.error(request, 'Invalid week for this phase.')
                return redirect('core:daily_report_export')
            week_dates = weeks[week_index]
            phase_week_offsets = _get_phase_week_offsets(week_map)
            global_week_num = phase_week_offsets.get(phase, 0) + week_index + 1
            wb = _build_daily_report_workbook(dept, phase, week_index, week_dates, global_week_num)
            fname = f'DR_{safe_dept}_{phase}_Week{global_week_num}_{sem_tag}.xlsx'

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


def _parse_cell_faculty_subject(cell_value):
    """
    Parse timetable cell to get (faculty_short_name, subject_name).
    Formats: "Subject (Faculty) (Room)", "Subject (Faculty) (Lab) (Lab)", "FAC-SUB-310", "FAC-SUB-408(L)"
    """
    if not cell_value or str(cell_value).strip() == '':
        return None, None
    text = str(cell_value).strip()
    # New format: DE (UMS) (301) or FSD-1 (DKU) (408-A) (Lab)
    new_pattern = r'^(.+?)\s*\(([^)]+)\)\s*\(([^)]+)\)'
    match = re.match(new_pattern, text)
    if match:
        subject = match.group(1).strip()
        faculty = match.group(2).strip()
        return faculty, subject
    # Old format: PHA-FSD_2-408-A(L) or PSK-DM-310
    parts = text.split('-')
    if len(parts) >= 2:
        faculty = parts[0].strip()
        subject = parts[1].strip()
        return faculty, subject
    return None, None


@login_required
def upload_timetable(request):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first from Dashboard.')
        return redirect('core:admin_dashboard')

    if request.method == 'POST':
        file = request.FILES.get('excel_file')
        replace_schedule = request.POST.get('replace_schedule') == 'on'
        effective_date_str = request.POST.get('effective_date', '').strip()
        if not file:
            messages.error(request, 'Please select an Excel file.')
            return redirect('core:upload_timetable')
        try:
            effective_date = datetime.strptime(effective_date_str, '%Y-%m-%d').date() if effective_date_str else datetime.now().date()
        except (ValueError, TypeError):
            effective_date = datetime.now().date()
        if not file.name.lower().endswith(('.xlsx', '.xls')):
            messages.error(request, 'Upload a valid Excel file (.xlsx or .xls).')
            return redirect('core:upload_timetable')
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
        except Exception as e:
            messages.error(request, f'Could not read Excel: {e}')
            return redirect('core:upload_timetable')

        sheet = None
        for name in ['TT-CLASSWISE', 'TT CLASSWISE', 'Timetable', 'Sheet1']:
            if name in wb.sheetnames:
                sheet = wb[name]
                break
        if not sheet:
            sheet = wb.active
        if not sheet:
            messages.error(request, 'No sheet found in the workbook.')
            return redirect('core:upload_timetable')

        # Header row: row 3 (1-based) = batch names from column C onwards
        header_row = next(sheet.iter_rows(min_row=3, max_row=3, values_only=True), None)
        if not header_row:
            messages.error(request, 'Could not read header row (row 3). Expected batch names in columns from C.')
            return redirect('core:upload_timetable')

        batches = []
        for col_idx in range(2, len(header_row)):
            val = header_row[col_idx]
            if not val or not str(val).strip():
                continue
            batch_name = str(val).strip()
            if _looks_like_timing_header(batch_name):
                continue
            batches.append((col_idx, batch_name))

        if not batches:
            messages.error(request, 'No batch columns found in row 3 (columns C onwards).')
            return redirect('core:upload_timetable')

        for _, bname in batches:
            Batch.objects.get_or_create(department=dept, name=bname)

        # When replace: delete existing slots with this effective_from, then add new (replaces that version)
        # When add: only add new slots (keep existing)
        if replace_schedule:
            deleted, _ = ScheduleSlot.objects.filter(department=dept, effective_from=effective_date).delete()
            if deleted:
                messages.info(request, f'Replaced {deleted} existing slot(s) with effective date {effective_date}.')
        current_day = None
        created_slots = 0
        for row in sheet.iter_rows(min_row=4, values_only=True):
            first_col = row[0] if len(row) > 0 else None
            time_val = row[1] if len(row) > 1 else None
            if first_col and isinstance(first_col, str) and first_col.strip().upper() not in ('', 'RECESS'):
                current_day = _normalize_day(first_col)
            if not current_day or not time_val:
                continue
            if len(row) > 2 and isinstance(row[2], str) and row[2].strip().upper() == 'RECESS':
                continue
            time_slot = _normalize_time_slot(time_val)
            if not time_slot:
                continue
            for col_idx, batch_name in batches:
                if col_idx >= len(row):
                    continue
                cell_val = row[col_idx]
                fac, subj = _parse_cell_faculty_subject(cell_val)
                if not fac or not subj:
                    continue
                faculty_obj, _ = Faculty.objects.get_or_create(
                    department=dept,
                    short_name=fac,
                    defaults={'full_name': fac}
                )
                subject_obj, _ = Subject.objects.get_or_create(
                    department=dept,
                    name=subj,
                    defaults={}
                )
                batch_obj = Batch.objects.get(department=dept, name=batch_name)
                _, created = ScheduleSlot.objects.get_or_create(
                    department=dept,
                    batch=batch_obj,
                    day=current_day,
                    time_slot=time_slot,
                    effective_from=effective_date,
                    defaults={'faculty': faculty_obj, 'subject': subject_obj}
                )
                if created:
                    created_slots += 1

        # Exact counts after import (from database)
        slots_qs = ScheduleSlot.objects.filter(department=dept)  # All versions for summary count
        total_entries = slots_qs.count()
        total_days = slots_qs.values_list('day', flat=True).distinct().count()
        total_time_slots = slots_qs.values_list('time_slot', flat=True).distinct().count()
        total_batches = Batch.objects.filter(department=dept).count()
        total_subjects = Subject.objects.filter(department=dept).count()
        total_faculties = Faculty.objects.filter(department=dept).count()
        per_batch_raw = slots_qs.values('batch__name').annotate(count=Count('id')).order_by('batch__name')
        per_batch = [{'batch_name': r['batch__name'], 'count': r['count']} for r in per_batch_raw]
        import_summary = {
            'total_batches': total_batches,
            'total_days': total_days,
            'total_time_slots': total_time_slots,
            'total_subjects': total_subjects,
            'total_faculties': total_faculties,
            'total_entries': total_entries,
            'new_slots_added': created_slots,
            'replace_schedule': replace_schedule,
            'effective_date': effective_date,
            'per_batch': per_batch,
        }
        messages.success(request, 'Timetable imported successfully. See summary below.')
        departments = Department.objects.all()
        ctx = {
            'department': dept,
            'departments': departments,
            'import_summary': import_summary,
            'today': datetime.now().date(),
        }
        return render(request, 'core/admin/upload_timetable.html', ctx)

    departments = Department.objects.all()
    ctx = {
        'department': dept,
        'departments': departments,
        'today': datetime.now().date(),
    }
    return render(request, 'core/admin/upload_timetable.html', ctx)


# ---------- Admin: Term Phases ----------

def _parse_holiday_dates(text, phase_start, phase_end):
    """Parse holiday date string (newline or comma separated YYYY-MM-DD). Return list of date objects within phase range."""
    from datetime import datetime
    out = []
    if not text or not phase_start or not phase_end:
        return out
    # Ensure phase_start and phase_end are date objects (POST/model can give strings)
    try:
        if isinstance(phase_start, str):
            phase_start = datetime.strptime(phase_start, '%Y-%m-%d').date()
        if isinstance(phase_end, str):
            phase_end = datetime.strptime(phase_end, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return out
    for part in text.replace(',', '\n').split():
        part = part.strip()
        if not part:
            continue
        try:
            d = datetime.strptime(part, '%Y-%m-%d').date()
            if phase_start <= d <= phase_end:
                out.append(d)
        except ValueError:
            continue
    return sorted(set(out))


@login_required
def term_phases(request):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    tp, _ = TermPhase.objects.get_or_create(department=dept)
    if request.method == 'POST':
        for i in range(1, 5):
            setattr(tp, f't{i}_start', request.POST.get(f't{i}_start') or None)
            setattr(tp, f't{i}_end', request.POST.get(f't{i}_end') or None)
        tp.save()
        # Save holidays per phase (only dates within phase start–end)
        for i in range(1, 5):
            phase = f'T{i}'
            start = getattr(tp, f't{i}_start', None)
            end = getattr(tp, f't{i}_end', None)
            PhaseHoliday.objects.filter(department=dept, phase=phase).delete()
            raw = request.POST.get(f't{i}_holidays', '') or ''
            dates = _parse_holiday_dates(raw, start, end)
            for d in dates:
                PhaseHoliday.objects.get_or_create(department=dept, phase=phase, date=d)
        messages.success(request, 'Term phases and holidays saved.')
        return redirect('core:term_phases')
    # Load existing holidays per phase for the form
    holiday_lists = {}
    for i in range(1, 5):
        phase = f'T{i}'
        dates = list(
            PhaseHoliday.objects.filter(department=dept, phase=phase).order_by('date').values_list('date', flat=True)
        )
        holiday_lists[f't{i}_holidays_list'] = dates
        holiday_lists[f't{i}_holidays'] = '\n'.join(d.strftime('%Y-%m-%d') for d in dates)
    ctx = {'term_phase': tp, 'department': dept, **holiday_lists}
    return render(request, 'core/admin/term_phases.html', ctx)


# ---------- Admin: Daily Absent ----------

def _valid_dates(dept, term_phase):
    """Lecture days across all phases, excluding holidays. Uses get_all_schedule_days for versioned timetables."""
    if not term_phase:
        return []
    from core.schedule_utils import get_all_schedule_days
    days_set = get_all_schedule_days(dept) or _effective_day_set_for_dept(dept, datetime.now().date())
    days_set = {d.lower() for d in days_set if d}
    holidays = get_all_holiday_dates(dept)
    out = []
    for i in range(1, 5):
        start = getattr(term_phase, f't{i}_start', None)
        end = getattr(term_phase, f't{i}_end', None)
        if not start or not end:
            continue
        cur = start
        while cur <= end:
            if cur not in holidays and cur.strftime('%A').lower() in days_set:
                out.append(cur)
            cur += timedelta(days=1)
    return sorted(out)


@login_required
def daily_absent(request):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    tp = TermPhase.objects.filter(department=dept).first()
    valid_dates = _valid_dates(dept, tp)
    date_str = request.GET.get('date')
    selected_date = None
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            if selected_date not in valid_dates:
                selected_date = valid_dates[0] if valid_dates else None
        except Exception:
            selected_date = valid_dates[0] if valid_dates else None
    else:
        selected_date = valid_dates[0] if valid_dates else None

    lectures_by_batch = defaultdict(list)
    if selected_date:
        cancelled_set = get_cancelled_lectures_set(dept)
        weekday = selected_date.strftime('%A')
        slots = [s for s in _effective_slots_for_date(dept, selected_date) if s.day == weekday]
        slots = sorted(slots, key=lambda s: (s.batch.name if s.batch else '', s.time_slot or ''))
        seen = set()
        for s in slots:
            if (selected_date, s.batch_id, s.time_slot) not in cancelled_set:
                seen.add((s.batch_id, s.time_slot))
                lectures_by_batch[s.batch.name].append(s)
        for ex in ExtraLecture.objects.filter(date=selected_date, batch__department=dept).select_related('batch', 'subject', 'faculty'):
            if (selected_date, ex.batch_id, ex.time_slot) in cancelled_set or (ex.batch_id, ex.time_slot) in seen:
                continue
            seen.add((ex.batch_id, ex.time_slot))
            virtual = type('Slot', (), {'batch': ex.batch, 'time_slot': ex.time_slot, 'subject': ex.subject, 'faculty': ex.faculty})()
            lectures_by_batch[ex.batch.name].append(virtual)

    attendance_map = defaultdict(lambda: defaultdict(list))
    if selected_date:
        atts = FacultyAttendance.objects.filter(
            faculty__department=dept, date=selected_date
        )
        for a in atts:
            for r in (a.absent_roll_numbers or '').split(','):
                if r.strip():
                    attendance_map[a.batch.id][a.lecture_slot].append(r.strip())

    missing = []
    for batch_name, lectures in lectures_by_batch.items():
        for slot in lectures:
            batch_id = slot.batch.id
            fac, subj = get_faculty_subject_for_slot(selected_date, slot.batch, slot.time_slot)
            effective_faculty = fac if fac is not None else slot.faculty
            if subj is not None:
                slot.subject = subj
            if fac is not None:
                slot.faculty = fac
            att_qs = FacultyAttendance.objects.filter(
                date=selected_date, batch_id=batch_id, lecture_slot=slot.time_slot
            )
            rec = att_qs.filter(faculty=effective_faculty).first() or att_qs.first()
            slot.absent_list = [x.strip() for x in (rec.absent_roll_numbers or '').split(',') if x.strip()] if rec else []
            # Only "missing" when no attendance record; saved "all present" (empty absent) is valid
            if not rec:
                missing.append((effective_faculty.full_name, batch_name, slot.time_slot))

    can_generate = not missing and bool(lectures_by_batch)
    ctx = {
        'valid_dates': valid_dates, 'selected_date': selected_date,
        'lectures_by_batch': dict(lectures_by_batch),
        'attendance_map': dict(attendance_map),
        'missing_entries': missing, 'can_generate_report': can_generate,
        'department': dept,
    }
    return render(request, 'core/admin/daily_absent.html', ctx)


@login_required
def daily_absent_excel(request):
    """Export daily absent report: title, 2 batches per row, No/Subject/Faculty/Absent Nos (all absent numbers in one cell per lecture, with wrap text)."""
    if not user_can_admin(request):
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    date_str = request.GET.get('date')
    if not date_str or not dept:
        return redirect('core:daily_absent')
    try:
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        return redirect('core:daily_absent')
    if selected_date in get_all_holiday_dates(dept):
        messages.warning(request, 'Selected date is a holiday; no lectures scheduled.')
        return redirect('core:daily_absent')
    weekday = selected_date.strftime('%A')
    cancelled_set = get_cancelled_lectures_set(dept)
    slots = [s for s in _effective_slots_for_date(dept, selected_date) if s.day == weekday]
    slots = sorted(slots, key=lambda s: (s.batch.name if s.batch else '', s.time_slot or ''))
    lectures_by_batch = defaultdict(list)
    seen = set()
    for s in slots:
        if (selected_date, s.batch_id, s.time_slot) not in cancelled_set:
            seen.add((s.batch_id, s.time_slot))
            lectures_by_batch[s.batch.name].append(s)
    for ex in ExtraLecture.objects.filter(date=selected_date, batch__department=dept).select_related('batch', 'subject', 'faculty'):
        if (selected_date, ex.batch_id, ex.time_slot) in cancelled_set or (ex.batch_id, ex.time_slot) in seen:
            continue
        seen.add((ex.batch_id, ex.time_slot))
        virtual = type('Slot', (), {'batch': ex.batch, 'time_slot': ex.time_slot, 'subject': ex.subject, 'faculty': ex.faculty})()
        lectures_by_batch[ex.batch.name].append(virtual)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Daily Absent'

    max_batches_per_row = 2
    headers = ['No', 'Subject', 'Faculty', 'Absent Nos']
    header_fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    bottom_line = Border(bottom=Side(style='thin'))

    batch_names = sorted(lectures_by_batch.keys())
    pairs = [batch_names[i:i + max_batches_per_row] for i in range(0, len(batch_names), max_batches_per_row)]
    n_cols = (len(pairs[0]) * 5) if pairs else 5

    # Header block: institution, dept, legend, date (all bold, large font)
    header_font_style = Font(size=13, bold=True)
    header_rows = [
        'L J Institute of Engineering and Technology',
        dept.name,
        'FONT COLOUR: BLACK (Absent in all Lectures)',
        'RED (Not attended all Lectures)',
        'BLUE (Absent reason: washroom/playing game/others)',
        f"Date-{selected_date.strftime('%d-%m-%Y')}. Day:- {selected_date.strftime('%A').upper()}",
    ]
    current_row = 1
    for text in header_rows:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=n_cols)
        cell = ws.cell(row=current_row, column=1, value=text)
        cell.font = header_font_style
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bottom_line
        current_row += 1
    current_row += 1

    for pair in pairs:
        col = 1
        for batch in pair:
            ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col + 3)
            batch_cell = ws.cell(row=current_row, column=col, value=f'Batch: {batch}')
            batch_cell.font = Font(bold=True, color='FFFFFF')
            batch_cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            batch_cell.alignment = Alignment(horizontal='center')
            col += 5
        current_row += 1
        col = 1
        for batch in pair:
            for cidx, header in enumerate(headers):
                c = ws.cell(row=current_row, column=col + cidx, value=header)
                c.font = header_font
                c.fill = header_fill
                c.border = thin_border
                c.alignment = Alignment(horizontal='center')
            col += 5
        current_row += 1

        max_lects = max(len(lectures_by_batch[b]) for b in pair)
        to_continue = []
        for batch in pair:
            lec_list = lectures_by_batch[batch]
            # Collect absent sets per lecture to find "absent in some, present in others" (red)
            lec_absents = []
            rows = []
            for idx, lec in enumerate(lec_list, 1):
                fac, subj = get_faculty_subject_for_slot(selected_date, lec.batch, lec.time_slot)
                subj_name = subj.name if subj else lec.subject.name
                fac_name = fac.full_name if fac else lec.faculty.full_name
                att = FacultyAttendance.objects.filter(
                    date=selected_date, batch=lec.batch, lecture_slot=lec.time_slot
                ).first()
                if fac and not att:
                    att = FacultyAttendance.objects.filter(
                        faculty=fac, date=selected_date,
                        batch=lec.batch, lecture_slot=lec.time_slot
                    ).first()
                if not att:
                    att = FacultyAttendance.objects.filter(
                        faculty=lec.faculty, date=selected_date,
                        batch=lec.batch, lecture_slot=lec.time_slot
                    ).first()
                absent_nos = (att.absent_roll_numbers or '').strip() if att else ''
                absents = [a.strip() for a in absent_nos.split(',') if a.strip()]
                lec_absents.append(set(absents))
                if not absents:
                    rows.append([idx, subj_name, fac_name, 'NIL', att])
                else:
                    rows.append([idx, subj_name, fac_name, absents, att])
            # Compute: absent_in_all = absent in every lecture (black); absent_in_some = present in at least one (red)
            num_lects = len(lec_absents)
            all_absent = set()
            for s in lec_absents:
                all_absent.update(s)
            absent_in_all = {r for r in all_absent if sum(1 for lec in lec_absents if r in lec) == num_lects}
            absent_in_some = all_absent - absent_in_all
            # Build rich text for each row's absent cell (with reason in brackets, blue for non-general)
            red_font = InlineFont(color='00FF0000')
            black_font = InlineFont(color='00000000')
            blue_font = InlineFont(color='000000FF')
            for i, row in enumerate(rows):
                if row[3] == 'NIL':
                    continue
                absents = row[3]
                att = row[4] if len(row) > 4 else None
                reasons = {}
                if att and att.absent_reasons:
                    try:
                        reasons = json.loads(att.absent_reasons)
                    except Exception:
                        pass
                blocks = []
                for j, r in enumerate(absents):
                    reason = reasons.get(str(r), 'general')
                    display_text = f'{r} ({reason})' if reason != 'general' else r
                    if reason != 'general':
                        blocks.append(TextBlock(blue_font, display_text))
                    elif r in absent_in_some:
                        blocks.append(TextBlock(red_font, r))
                    else:
                        blocks.append(TextBlock(black_font, r))
                    if j < len(absents) - 1:
                        blocks.append(TextBlock(black_font, ', '))
                row[3] = CellRichText(*blocks) if blocks else 'NIL'
            for row in rows:
                if len(row) > 4:
                    row.pop()  # remove att before writing
            to_continue.append(rows)
        block_height = max(len(x) for x in to_continue)
        for data_rows in to_continue:
            while len(data_rows) < block_height:
                data_rows.append(['', '', '', ''])
        for ridx in range(block_height):
            col = 1
            for data_rows in to_continue:
                for cidx, value in enumerate(data_rows[ridx]):
                    c = ws.cell(row=current_row, column=col + cidx, value=value)
                    c.border = thin_border
                    wrap = (cidx == 3)
                    c.alignment = Alignment(horizontal='center' if cidx == 0 else 'left', wrap_text=wrap)
                col += 5
            current_row += 1
        current_row += 1

    for b in range(2):
        offset = b * 5
        ws.column_dimensions[get_column_letter(1 + offset)].width = 5
        ws.column_dimensions[get_column_letter(2 + offset)].width = 28
        ws.column_dimensions[get_column_letter(3 + offset)].width = 22
        ws.column_dimensions[get_column_letter(4 + offset)].width = 35

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename=DailyAbsent_{selected_date:%Y-%m-%d}.xlsx'
    return resp


# ---------- Admin: Attendance Sheet Manager ----------

@login_required
def attendance_sheet_manager(request):
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    tp = TermPhase.objects.filter(department=dept).first()
    batches = Batch.objects.filter(department=dept)
    phases = ['T1', 'T2', 'T3', 'T4']
    week_map = {}
    for p in phases:
        start = getattr(tp, f'{p.lower()}_start', None) if tp else None
        end = getattr(tp, f'{p.lower()}_end', None) if tp else None
        if not start or not end:
            week_map[p] = []
            continue
        from core.schedule_utils import get_all_schedule_days
        days_set = get_all_schedule_days(dept) or _effective_day_set_for_dept(dept, datetime.now().date())
        days_set = {d.lower() for d in days_set if d}
        dates = []
        cur = start
        while cur <= end:
            if cur.strftime('%A').lower() in days_set:
                dates.append(cur)
            cur += timedelta(days=1)
        # chunk by week
        weeks = []
        week = []
        last_w = None
        for d in sorted(dates):
            w = d.isocalendar()[1]
            if last_w is not None and w != last_w and week:
                weeks.append(week)
                week = []
            week.append(d)
            last_w = w
        if week:
            weeks.append(week)
        week_map[p] = [[d.isoformat() for d in w] for w in weeks]
    # All lecture dates for daily picker (any phase)
    all_dates = []
    for p in phases:
        for week_dates in week_map.get(p, []):
            for d_str in week_dates:
                try:
                    all_dates.append(datetime.strptime(d_str, '%Y-%m-%d').date())
                except Exception:
                    pass
    available_dates = sorted(set(all_dates))
    phase_week_offsets = _get_phase_week_offsets(week_map)
    ctx = {
        'department': dept, 'batches': batches, 'phases': phases,
        'week_map': week_map, 'week_map_json': json.dumps(week_map),
        'phase_week_offsets_json': json.dumps(phase_week_offsets),
        'available_dates': available_dates,
    }
    return render(request, 'core/admin/attendance_sheet_manager.html', ctx)


def _build_date_slots_list_for_batch(dept, batch, dates):
    """Return [(date, slots), ...] for this batch and list of dates (excluding holidays, excluding cancelled)."""
    cancelled_set = get_cancelled_lectures_set(dept)
    out = []
    for d in dates:
        weekday = d.strftime('%A')
        slots = [s for s in _effective_slots_for_date(dept, d, extra_filters={'batch': batch}) if s.day == weekday]
        slots = sorted(slots, key=lambda s: s.time_slot or '')
        slots = [s for s in slots if (d, batch.id, s.time_slot) not in cancelled_set]
        seen_slots = {s.time_slot for s in slots if s.time_slot}
        for ex in ExtraLecture.objects.filter(date=d, batch=batch).select_related('subject', 'faculty'):
            if (d, batch.id, ex.time_slot) in cancelled_set or ex.time_slot in seen_slots:
                continue
            seen_slots.add(ex.time_slot)
            virtual = type('Slot', (), {'time_slot': ex.time_slot, 'subject': ex.subject, 'faculty': ex.faculty})()
            slots.append(virtual)
        slots.sort(key=lambda s: s.time_slot or '')
        out.append((d, slots))
    return out


def _filter_date_slots_by_subject(date_slots_list, subject_id):
    """Filter date_slots_list to only slots for the given subject. Returns [(date, slots), ...]."""
    return [(d, [s for s in slots if s.subject_id == subject_id]) for d, slots in date_slots_list]


def _write_all_batches_combined_sheet(ws, batches, all_students, date_slots_union, batch_att_map, batch_date_slots_set, styles, overall_segments_per_batch):
    """Write one sheet with all batches combined. Roll No, Name, Batch, then (date, slot) columns, then Overall.
    date_slots_union: [(d, slots), ...] slots have .time_slot. batch_date_slots_set: batch_id -> set of (d, time_slot).
    overall_segments_per_batch: [(batch, [(label, seg), ...]), ...]
    """
    thin_border = styles['thin_border']
    date_fill = styles['date_fill']
    date_font = styles['date_font']
    date_align = styles['date_align']
    header_font = styles['header_font']
    lect_fill = styles['lect_fill']
    lect_font = styles['lect_font']
    lect_align = styles['lect_align']
    red_font = styles['red_font']
    ws.title = 'All Batches'[:31]
    ws.cell(1, 1, 'Roll No').font = header_font
    ws.cell(1, 2, 'Student Name').font = header_font
    ws.cell(1, 3, 'Batch').font = header_font
    for c in range(1, 4):
        ws.cell(1, c).border = thin_border
        ws.cell(2, c).border = thin_border
    col = 4
    for d, slots in date_slots_union:
        n_lec = max(len(slots), 1)
        if n_lec > 1:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + n_lec - 1)
        for c in range(col, col + n_lec):
            cell = ws.cell(row=1, column=c, value=d.strftime('%d-%b') if c == col else None)
            cell.border = thin_border
            cell.fill = date_fill
            cell.font = date_font
            cell.alignment = date_align
        for i, slot in enumerate(slots, start=1):
            subj_name = getattr(slot, 'subject', None) and getattr(slot.subject, 'name', None) or 'Lec'
            cell = ws.cell(row=2, column=col + i - 1, value=f'Lect {i}\n{subj_name}')
            cell.alignment = lect_align
            cell.fill = lect_fill
            cell.font = lect_font
            cell.border = thin_border
        if not slots:
            ws.cell(2, col, '').border = thin_border
        col += n_lec
    n_overall = len(overall_segments_per_batch[0][1]) * 3 if overall_segments_per_batch else 0
    if n_overall:
        overall_col_start = col
        ws.merge_cells(start_row=1, start_column=overall_col_start, end_row=1, end_column=overall_col_start + n_overall - 1)
        cell = ws.cell(row=1, column=overall_col_start, value='Overall Attendance')
        cell.border = thin_border
        cell.fill = date_fill
        cell.font = date_font
        cell.alignment = date_align
        seg_col = overall_col_start
        for label, _ in overall_segments_per_batch[0][1]:
            ws.merge_cells(start_row=2, start_column=seg_col, end_row=2, end_column=seg_col + 2)
            c = ws.cell(row=2, column=seg_col, value=label)
            c.border = thin_border
            c.fill = lect_fill
            c.font = lect_font
            c.alignment = date_align
            for i, sub in enumerate(('Total Lecture', 'Attended', '%')):
                cc = ws.cell(row=3, column=seg_col + i, value=sub)
                cc.border = thin_border
                cc.fill = lect_fill
                cc.font = lect_font
                cc.alignment = lect_align
            seg_col += 3
        col = overall_col_start + n_overall
    data_start_row = 4 if n_overall else 3
    for idx, s in enumerate(all_students, start=data_start_row):
        ws.cell(idx, 1, s.roll_no).border = thin_border
        ws.cell(idx, 2, s.name).border = thin_border
        ws.cell(idx, 3, s.batch.name).border = thin_border
        str_roll = str(s.roll_no)
        bid = s.batch_id
        att_map = batch_att_map.get(bid, {})
        batch_has = batch_date_slots_set.get(bid, set())
        c = 4
        for d, slots in date_slots_union:
            for slot in slots:
                key = (d, slot.time_slot)
                if key not in batch_has:
                    val = '—'
                elif key not in att_map:
                    val = '—'
                else:
                    val = 'A' if str_roll in att_map[key] else 'P'
                cell = ws.cell(idx, c, value=val)
                cell.border = thin_border
                if val == 'A':
                    cell.font = red_font
                c += 1
            if not slots:
                ws.cell(idx, c, '').border = thin_border
                c += 1
        if n_overall and overall_segments_per_batch:
            segs = next((segs for b, segs in overall_segments_per_batch if b.id == bid), [])
            for label, date_slots_seg in segs:
                held, attended = _student_held_attended_for_segment(date_slots_seg, att_map, str_roll)
                pct = round(attended / held * 100, 2) if held else 0
                ws.cell(idx, c, held).border = thin_border
                ws.cell(idx, c + 1, attended).border = thin_border
                pct_cell = ws.cell(idx, c + 2, f'{pct:.2f}%')
                pct_cell.border = thin_border
                if pct < 75 and held:
                    pct_cell.font = Font(color='FFFFFF', bold=True)
                    pct_cell.fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
                c += 3
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 10
    ws.freeze_panes = f'D{data_start_row}'


def _write_subject_all_batches_sheet(ws, subject_name, batches, all_students, batch_date_slots_subject, batch_att_map, styles, build_overall_fn):
    """Write one sheet per subject with students from all batches. Same format as individual batch subject sheet.
    Roll No, Name, Batch (extra), date+lecture columns (date merged when multiple lectures), Overall.
    Only difference from individual: Batch column + all batches' students in one sheet."""
    # Build date_slots_union: [(date, n_lec), ...] - same structure as individual batch, n_lec = max slots per date
    date_to_nlec = {}
    for batch in batches:
        for d, slots in batch_date_slots_subject.get(batch.id, []):
            if slots:
                date_to_nlec[d] = max(date_to_nlec.get(d, 0), len(slots))
    date_slots_union = [(d, date_to_nlec[d]) for d in sorted(date_to_nlec.keys())]
    if not date_slots_union:
        return False
    # batch_slots_by_date: batch_id -> date -> [slot1, slot2, ...] for attendance lookup
    batch_slots_by_date = {}
    for batch in batches:
        d2slots = {}
        for d, slots in batch_date_slots_subject.get(batch.id, []):
            d2slots[d] = list(slots)
        batch_slots_by_date[batch.id] = d2slots
    thin_border = styles['thin_border']
    date_fill = styles['date_fill']
    date_font = styles['date_font']
    date_align = styles['date_align']
    header_font = styles['header_font']
    lect_fill = styles['lect_fill']
    lect_font = styles['lect_font']
    lect_align = styles['lect_align']
    red_font = styles['red_font']
    ws.title = (subject_name[:31] if subject_name else 'Subject')
    ws.cell(1, 1, 'Roll No').font = header_font
    ws.cell(1, 2, 'Student Name').font = header_font
    ws.cell(1, 3, 'Batch').font = header_font
    for c in range(1, 4):
        ws.cell(1, c).border = thin_border
        ws.cell(2, c).border = thin_border
        ws.cell(3, c).border = thin_border
    col = 4
    for d, n_lec in date_slots_union:
        if n_lec == 1:
            ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
            cell = ws.cell(row=1, column=col, value=f"{d.strftime('%d-%b')}\nLect 1")
            cell.border = thin_border
            cell.fill = date_fill
            cell.font = date_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            if n_lec > 1:
                ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + n_lec - 1)
            for c in range(col, col + n_lec):
                cell = ws.cell(row=1, column=c, value=d.strftime('%d-%b') if c == col else None)
                cell.border = thin_border
                cell.fill = date_fill
                cell.font = date_font
                cell.alignment = date_align
            for i in range(1, n_lec + 1):
                cell = ws.cell(row=2, column=col + i - 1, value=f'Lect {i}\n{subject_name}')
                cell.alignment = lect_align
                cell.fill = lect_fill
                cell.font = lect_font
                cell.border = thin_border
        col += n_lec
    overall_segments_per_batch = []
    for batch in batches:
        segs = build_overall_fn(batch)
        if segs:
            overall_segments_per_batch.append((batch, segs))
    n_overall = len(overall_segments_per_batch[0][1]) * 3 if overall_segments_per_batch else 0
    if n_overall:
        overall_col_start = col
        for c in range(4, overall_col_start):
            ws.cell(3, c, '').border = thin_border
        ws.merge_cells(start_row=1, start_column=overall_col_start, end_row=1, end_column=overall_col_start + n_overall - 1)
        cell = ws.cell(row=1, column=overall_col_start, value='Overall Attendance')
        cell.border = thin_border
        cell.fill = date_fill
        cell.font = date_font
        cell.alignment = date_align
        seg_col = overall_col_start
        for label, _ in overall_segments_per_batch[0][1]:
            ws.merge_cells(start_row=2, start_column=seg_col, end_row=2, end_column=seg_col + 2)
            c = ws.cell(row=2, column=seg_col, value=label)
            c.border = thin_border
            c.fill = lect_fill
            c.font = lect_font
            c.alignment = date_align
            for i, sub in enumerate(('Total Lecture', 'Attended', '%')):
                cc = ws.cell(row=3, column=seg_col + i, value=sub)
                cc.border = thin_border
                cc.fill = lect_fill
                cc.font = lect_font
                cc.alignment = lect_align
            seg_col += 3
        col = overall_col_start + n_overall
    data_start_row = 4 if n_overall else 3
    for idx, s in enumerate(all_students, start=data_start_row):
        ws.cell(idx, 1, s.roll_no).border = thin_border
        ws.cell(idx, 2, s.name).border = thin_border
        ws.cell(idx, 3, s.batch.name).border = thin_border
        str_roll = str(s.roll_no)
        bid = s.batch_id
        att_map = batch_att_map.get(bid, {})
        batch_slots = batch_slots_by_date.get(bid, {})
        c = 4
        for d, n_lec in date_slots_union:
            slots_for_date = batch_slots.get(d, [])
            for col_idx in range(n_lec):
                if col_idx < len(slots_for_date):
                    slot = slots_for_date[col_idx]
                    key = (d, slot.time_slot)
                    val = '—' if key not in att_map else ('A' if str_roll in att_map[key] else 'P')
                else:
                    val = '—'
                cell = ws.cell(idx, c, value=val)
                cell.border = thin_border
                if val == 'A':
                    cell.font = red_font
                c += 1
        if n_overall and overall_segments_per_batch:
            segs = next((segs for b, segs in overall_segments_per_batch if b.id == bid), [])
            for label, date_slots_seg in segs:
                held, attended = _student_held_attended_for_segment(date_slots_seg, att_map, str_roll)
                pct = round(attended / held * 100, 2) if held else 0
                ws.cell(idx, c, held).border = thin_border
                ws.cell(idx, c + 1, attended).border = thin_border
                pct_cell = ws.cell(idx, c + 2, f'{pct:.2f}%')
                pct_cell.border = thin_border
                if pct < 75 and held:
                    pct_cell.font = Font(color='FFFFFF', bold=True)
                    pct_cell.fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
                c += 3
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 10
    ws.freeze_panes = f'D{data_start_row}'
    return True


def _attendance_sheet_dates_for_period(dept, period_type, phase, week_index=None, single_date=None):
    """Return list of dates for the chosen period (excluding holidays). week_index is 0-based index into week_map[phase].
    Uses get_all_schedule_days for versioned timetables."""
    tp = TermPhase.objects.filter(department=dept).first()
    from core.schedule_utils import get_all_schedule_days
    days_set = get_all_schedule_days(dept) or _effective_day_set_for_dept(dept, datetime.now().date())
    days_set = {d.lower() for d in days_set if d}
    if period_type == 'daily' and single_date:
        holidays = get_all_holiday_dates(dept)
        if single_date in holidays:
            return []
        return [single_date] if single_date.strftime('%A').lower() in days_set else []
    if period_type == 'weekly' and phase and week_index is not None:
        start = getattr(tp, f'{phase.lower()}_start', None) if tp else None
        end = getattr(tp, f'{phase.lower()}_end', None) if tp else None
        if not start or not end:
            return []
        holidays = get_phase_holidays(dept, phase)
        dates = []
        cur = start
        while cur <= end:
            if cur not in holidays and cur.strftime('%A').lower() in days_set:
                dates.append(cur)
            cur += timedelta(days=1)
        weeks = []
        week = []
        last_w = None
        for d in sorted(dates):
            w = d.isocalendar()[1]
            if last_w is not None and w != last_w and week:
                weeks.append(week)
                week = []
            week.append(d)
            last_w = w
        if week:
            weeks.append(week)
        if 0 <= week_index < len(weeks):
            return weeks[week_index]
        return []
    if period_type == 'phase' and phase:
        start = getattr(tp, f'{phase.lower()}_start', None) if tp else None
        end = getattr(tp, f'{phase.lower()}_end', None) if tp else None
        if not start or not end:
            return []
        holidays = get_phase_holidays(dept, phase)
        dates = []
        cur = start
        while cur <= end:
            if cur not in holidays and cur.strftime('%A').lower() in days_set:
                dates.append(cur)
            cur += timedelta(days=1)
        return sorted(dates)
    return []


def _student_held_attended_for_segment(date_slots_segment, att_map, str_roll):
    """For one segment (list of (date, slots)), return (held, attended) for one student.
    held = total scheduled (excluding holidays). attended = only where attendance was taken AND student present."""
    held = attended = 0
    for d, slots in date_slots_segment:
        for slot in slots:
            held += 1
            key = (d, slot.time_slot)
            if key in att_map and str_roll not in att_map[key]:
                attended += 1
    return held, attended


def _write_one_batch_attendance_sheet(ws, batch, date_slots_list, students, att_map, styles, overall_segments=None, sheet_title=None):
    """Write one batch's attendance data into worksheet ws. If overall_segments is given, append Overall Attendance block.
    overall_segments: list of (label, date_slots_sub_list) e.g. [('Week 1', w1_list), ('Week 2', w2_list), ('Overall', all_list)].
    sheet_title: optional custom title for the sheet (default: batch.name).
    Only includes dates that have lectures (skips dates with no slots).
    """
    date_slots_list = [(d, s) for d, s in date_slots_list if s]
    if overall_segments:
        overall_segments = [(label, [(d, s) for d, s in seg if s]) for label, seg in overall_segments]
    thin_border = styles['thin_border']
    date_fill = styles['date_fill']
    date_font = styles['date_font']
    date_align = styles['date_align']
    header_font = styles['header_font']
    lect_fill = styles['lect_fill']
    lect_font = styles['lect_font']
    lect_align = styles['lect_align']
    red_font = styles['red_font']

    data_start_row = 3
    header_rows = 2
    if overall_segments:
        header_rows = 3
        data_start_row = 4

    ws.title = ((sheet_title or batch.name or 'Sheet')[:31])
    ws.cell(1, 1, 'Roll No').font = header_font
    ws.cell(1, 2, 'Student Name').font = header_font
    ws.cell(1, 1).border = thin_border
    ws.cell(1, 2).border = thin_border
    ws.cell(2, 1, '').border = thin_border
    ws.cell(2, 2, '').border = thin_border
    if overall_segments:
        ws.cell(3, 1, '').border = thin_border
        ws.cell(3, 2, '').border = thin_border
    col = 3
    is_subject_sheet = bool(sheet_title)
    for d, slots in date_slots_list:
        n_lec = max(len(slots), 1)
        if is_subject_sheet and n_lec == 1:
            ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
            cell = ws.cell(row=1, column=col, value=f"{d.strftime('%d-%b')}\nLect 1")
            cell.border = thin_border
            cell.fill = date_fill
            cell.font = date_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            if n_lec > 1:
                ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + n_lec - 1)
            for c in range(col, col + n_lec):
                cell = ws.cell(row=1, column=c, value=d.strftime('%d-%b') if c == col else None)
                cell.border = thin_border
                cell.fill = date_fill
                cell.font = date_font
                cell.alignment = date_align
            for i, slot in enumerate(slots, start=1):
                fac, subj = get_faculty_subject_for_slot(d, batch, slot.time_slot)
                subj_name = subj.name if subj else (slot.subject.name if slot.subject else 'N/A')
                cell = ws.cell(row=2, column=col + i - 1, value=f'Lect {i}\n{subj_name}')
                cell.alignment = lect_align
                cell.fill = lect_fill
                cell.font = lect_font
                cell.border = thin_border
            if not slots:
                ws.cell(2, col, '').border = thin_border
        col += n_lec

    if overall_segments:
        for c in range(3, col):
            ws.cell(3, c, '').border = thin_border

    # Overall Attendance block: merged header, then per-segment label row and (Total Lecture, Attended, %)
    if overall_segments:
        overall_col_start = col
        n_overall_cols = len(overall_segments) * 3
        ws.merge_cells(start_row=1, start_column=overall_col_start, end_row=1, end_column=overall_col_start + n_overall_cols - 1)
        cell = ws.cell(row=1, column=overall_col_start, value='Overall Attendance')
        cell.border = thin_border
        cell.fill = date_fill
        cell.font = date_font
        cell.alignment = date_align
        seg_col = overall_col_start
        for label, _ in overall_segments:
            ws.merge_cells(start_row=2, start_column=seg_col, end_row=2, end_column=seg_col + 2)
            c = ws.cell(row=2, column=seg_col, value=label)
            c.border = thin_border
            c.fill = lect_fill
            c.font = lect_font
            c.alignment = date_align
            for i, sub in enumerate(('Total Lecture', 'Attended', '%')):
                cc = ws.cell(row=3, column=seg_col + i, value=sub)
                cc.border = thin_border
                cc.fill = lect_fill
                cc.font = lect_font
                cc.alignment = lect_align
            seg_col += 3
        col = overall_col_start + n_overall_cols

    for idx, s in enumerate(students, start=data_start_row):
        ws.cell(idx, 1, s.roll_no).border = thin_border
        ws.cell(idx, 2, s.name).border = thin_border
        str_roll = str(s.roll_no)
        c = 3
        for d, slots in date_slots_list:
            for slot in slots:
                key = (d, slot.time_slot)
                if key not in att_map:
                    val = '—'
                else:
                    val = 'A' if str_roll in att_map[key] else 'P'
                cell = ws.cell(idx, c, value=val)
                cell.border = thin_border
                if val == 'A':
                    cell.font = red_font
                c += 1
            if not slots:
                ws.cell(idx, c, '').border = thin_border
                c += 1
        if overall_segments:
            for label, date_slots_seg in overall_segments:
                held, attended = _student_held_attended_for_segment(date_slots_seg, att_map, str_roll)
                pct = round(attended / held * 100, 2) if held else 0
                ws.cell(idx, c, held).border = thin_border
                ws.cell(idx, c + 1, attended).border = thin_border
                pct_cell = ws.cell(idx, c + 2, f'{pct:.2f}%')
                pct_cell.border = thin_border
                if pct < 75 and held:
                    pct_cell.font = Font(color='FFFFFF', bold=True)
                    pct_cell.fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
                c += 3
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.freeze_panes = f'C{data_start_row}'


@login_required
def attendance_sheet_excel(request):
    if not user_can_admin(request):
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    batch_id = request.GET.get('batch')
    period_type = request.GET.get('period_type', 'phase')
    phase = request.GET.get('phase')
    week_index = request.GET.get('week')
    date_str = request.GET.get('date')
    if not dept:
        return redirect('core:attendance_sheet_manager')
    if not batch_id:
        return redirect('core:attendance_sheet_manager')

    all_batches = batch_id == 'all'
    if all_batches:
        batches = list(Batch.objects.filter(department=dept).order_by('name'))
        if not batches:
            messages.error(request, 'No batches in this department.')
            return redirect('core:attendance_sheet_manager')
    else:
        batch = Batch.objects.filter(pk=batch_id, department=dept).first()
        if not batch:
            return redirect('core:attendance_sheet_manager')
        batches = [batch]

    single_date = None
    if period_type == 'daily' and date_str:
        try:
            single_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except Exception:
            pass
    week_idx = None
    if period_type == 'weekly' and week_index is not None:
        try:
            week_idx = int(week_index)
        except Exception:
            pass
    if period_type in ('weekly', 'phase') and not phase:
        return redirect('core:attendance_sheet_manager')

    # For weekly we want "through week N" (weeks 0..week_idx) so the sheet shows all those weeks and Overall block has W1, W2, ... Overall
    phases_order = ['T1', 'T2', 'T3', 'T4']
    week_map = {p: _compile_phase_weeks_date_objects(dept, p) for p in phases_order} if phase else {}
    phase_week_offsets = _get_phase_week_offsets(week_map) if week_map else {}
    weeks_current_phase = week_map.get(phase, [])
    if period_type == 'weekly' and week_idx is not None and 0 <= week_idx < len(weeks_current_phase):
        dates = []
        for w in weeks_current_phase[: week_idx + 1]:
            dates.extend(w)
        dates = sorted(set(dates))
    else:
        dates = _attendance_sheet_dates_for_period(dept, period_type, phase, week_idx, single_date)
    if not dates:
        messages.error(request, 'No dates in selected period.')
        return redirect('core:attendance_sheet_manager')

    def build_overall_segments(batch, date_slots_list, att_map_batch):
        """Build list of (label, date_slots_sub_list) for Overall Attendance block. att_map_batch must cover all dates in any segment."""
        segments = []
        if period_type == 'daily':
            segments = [('Overall', date_slots_list)]
        elif period_type == 'weekly' and weeks_current_phase:
            # Previous phase overall when phase is T2, T3, T4
            phase_order = ['T1', 'T2', 'T3', 'T4']
            try:
                pi = phase_order.index(phase)
            except ValueError:
                pi = 0
            if pi > 0:
                for prev_i in range(pi):
                    prev_phase = phase_order[prev_i]
                    prev_weeks = _compile_phase_weeks_date_objects(dept, prev_phase)
                    prev_dates = []
                    for w in prev_weeks:
                        prev_dates.extend(w)
                    prev_dates = sorted(set(prev_dates))
                    prev_slots = _build_date_slots_list_for_batch(dept, batch, prev_dates)
                    segments.append((f'{prev_phase} Overall', prev_slots))
            offset = phase_week_offsets.get(phase, 0)
            for i in range(week_idx + 1):
                w_dates = weeks_current_phase[i]
                w_slots = _build_date_slots_list_for_batch(dept, batch, w_dates)
                label = f'Week {offset + i + 1}'
                segments.append((label, w_slots))
            segments.append(('Overall', date_slots_list))
        elif period_type == 'phase' and weeks_current_phase:
            offset = phase_week_offsets.get(phase, 0)
            for i, w in enumerate(weeks_current_phase):
                w_slots = _build_date_slots_list_for_batch(dept, batch, w)
                segments.append((f'Week {offset + i + 1}', w_slots))
            segments.append(('Overall', date_slots_list))
        return segments

    styles = {
        'thin_border': Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        ),
        'date_fill': PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid'),
        'date_font': Font(bold=True, color='FFFFFF'),
        'date_align': Alignment(horizontal='center', vertical='center'),
        'header_font': Font(bold=True),
        'lect_fill': PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid'),
        'lect_font': Font(bold=True, color='FFFFFF'),
        'lect_align': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'red_font': Font(color='FF0000'),
    }

    wb = Workbook()
    first = True
    for batch in batches:
        if first:
            ws = wb.active
            first = False
        else:
            ws = wb.create_sheet(title=(batch.name[:31] if batch.name else 'Sheet'))

        students = list(Student.objects.filter(department=dept, batch=batch))
        students.sort(key=_roll_sort_key)
        date_slots_list = _build_date_slots_list_for_batch(dept, batch, dates)
        overall_segments = build_overall_segments(batch, date_slots_list, None)

        # att_map must cover all dates in main grid and in any segment (e.g. T1 when phase is T2)
        all_dates_for_att = set(d for d, _ in date_slots_list)
        for label, seg_slots in overall_segments:
            for d, _ in seg_slots:
                all_dates_for_att.add(d)
        att_map = {}
        for d in all_dates_for_att:
            for att in FacultyAttendance.objects.filter(batch=batch, date=d):
                key = (d, att.lecture_slot)
                att_map[key] = set(x.strip() for x in (att.absent_roll_numbers or '').split(',') if x.strip())

        _write_one_batch_attendance_sheet(ws, batch, date_slots_list, students, att_map, styles, overall_segments=overall_segments)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    global_week_num = phase_week_offsets.get(phase, 0) + week_idx + 1 if (period_type == 'weekly' and week_idx is not None and weeks_current_phase) else None
    if all_batches:
        if period_type == 'daily':
            fname = f'Attendance_All_{dates[0]:%Y-%m-%d}.xlsx'
        elif period_type == 'weekly' and global_week_num:
            fname = f'Attendance_All_{phase}_week{global_week_num}.xlsx'
        else:
            fname = f'Attendance_All_{phase}.xlsx'
    else:
        batch = batches[0]
        if period_type == 'daily':
            fname = f'Attendance_{batch.name}_{dates[0]:%Y-%m-%d}.xlsx'
        elif period_type == 'weekly' and global_week_num:
            fname = f'Attendance_{batch.name}_{phase}_week{global_week_num}.xlsx'
        else:
            fname = f'Attendance_{batch.name}_{phase}.xlsx'
    resp = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename={fname}'
    return resp


@login_required
def attendance_sheet_subjectwise_manager(request):
    """Admin: Subject-wise Attendance — same form as Attendance Sheet but Excel has one tab per subject."""
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    tp = TermPhase.objects.filter(department=dept).first()
    batches = Batch.objects.filter(department=dept)
    phases = ['T1', 'T2', 'T3', 'T4']
    week_map = {}
    for p in phases:
        start = getattr(tp, f'{p.lower()}_start', None) if tp else None
        end = getattr(tp, f'{p.lower()}_end', None) if tp else None
        if not start or not end:
            week_map[p] = []
            continue
        from core.schedule_utils import get_all_schedule_days
        days_set = get_all_schedule_days(dept) or _effective_day_set_for_dept(dept, datetime.now().date())
        days_set = {d.lower() for d in days_set if d}
        dates = []
        cur = start
        while cur <= end:
            if cur.strftime('%A').lower() in days_set:
                dates.append(cur)
            cur += timedelta(days=1)
        weeks = []
        week = []
        last_w = None
        for d in sorted(dates):
            w = d.isocalendar()[1]
            if last_w is not None and w != last_w and week:
                weeks.append(week)
                week = []
            week.append(d)
            last_w = w
        if week:
            weeks.append(week)
        week_map[p] = [[d.isoformat() for d in w] for w in weeks]
    all_dates = []
    for p in phases:
        for week_dates in week_map.get(p, []):
            for d_str in week_dates:
                try:
                    all_dates.append(datetime.strptime(d_str, '%Y-%m-%d').date())
                except Exception:
                    pass
    available_dates = sorted(set(all_dates))
    phase_week_offsets = _get_phase_week_offsets(week_map)
    ctx = {
        'department': dept, 'batches': batches, 'phases': phases,
        'week_map': week_map, 'week_map_json': json.dumps(week_map),
        'phase_week_offsets_json': json.dumps(phase_week_offsets),
        'available_dates': available_dates,
    }
    return render(request, 'core/admin/attendance_sheet_subjectwise.html', ctx)


@login_required
def attendance_sheet_subjectwise_excel(request):
    """Export subject-wise attendance: one Excel tab per subject for the selected batch."""
    if not user_can_admin(request):
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    batch_id = request.GET.get('batch')
    period_type = request.GET.get('period_type', 'phase')
    phase = request.GET.get('phase')
    week_index = request.GET.get('week')
    date_str = request.GET.get('date')
    if not dept or not batch_id:
        messages.error(request, 'Select a batch.')
        return redirect('core:attendance_sheet_subjectwise_manager')
    all_batches = batch_id == 'all'
    if all_batches:
        batches = list(Batch.objects.filter(department=dept).order_by('name'))
        if not batches:
            messages.error(request, 'No batches in this department.')
            return redirect('core:attendance_sheet_subjectwise_manager')
    else:
        batch = Batch.objects.filter(pk=batch_id, department=dept).first()
        if not batch:
            return redirect('core:attendance_sheet_subjectwise_manager')
        batches = [batch]
    single_date = None
    if period_type == 'daily' and date_str:
        try:
            single_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except Exception:
            pass
    week_idx = None
    if period_type == 'weekly' and week_index is not None:
        try:
            week_idx = int(week_index)
        except Exception:
            pass
    if period_type in ('weekly', 'phase') and not phase:
        return redirect('core:attendance_sheet_subjectwise_manager')
    phases_order = ['T1', 'T2', 'T3', 'T4']
    week_map = {p: _compile_phase_weeks_date_objects(dept, p) for p in phases_order} if phase else {}
    phase_week_offsets = _get_phase_week_offsets(week_map) if week_map else {}
    weeks_current_phase = week_map.get(phase, [])
    if period_type == 'weekly' and week_idx is not None and 0 <= week_idx < len(weeks_current_phase):
        dates = []
        for w in weeks_current_phase[: week_idx + 1]:
            dates.extend(w)
        dates = sorted(set(dates))
    else:
        dates = _attendance_sheet_dates_for_period(dept, period_type, phase, week_idx, single_date)
    if not dates:
        messages.error(request, 'No dates in selected period.')
        return redirect('core:attendance_sheet_subjectwise_manager')
    styles = {
        'thin_border': Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        ),
        'date_fill': PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid'),
        'date_font': Font(bold=True, color='FFFFFF'),
        'date_align': Alignment(horizontal='center', vertical='center'),
        'header_font': Font(bold=True),
        'lect_fill': PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid'),
        'lect_font': Font(bold=True, color='FFFFFF'),
        'lect_align': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'red_font': Font(color='FF0000'),
    }

    def build_overall_segments_subject(batch, date_slots_list, subject_id):
        base_segments = []
        if period_type == 'daily':
            base_segments = [('Overall', date_slots_list)]
        elif period_type == 'weekly' and weeks_current_phase:
            phase_order = ['T1', 'T2', 'T3', 'T4']
            try:
                pi = phase_order.index(phase)
            except ValueError:
                pi = 0
            if pi > 0:
                for prev_i in range(pi):
                    prev_phase = phase_order[prev_i]
                    prev_weeks = _compile_phase_weeks_date_objects(dept, prev_phase)
                    prev_dates = []
                    for w in prev_weeks:
                        prev_dates.extend(w)
                    prev_dates = sorted(set(prev_dates))
                    prev_slots = _build_date_slots_list_for_batch(dept, batch, prev_dates)
                    base_segments.append((f'{prev_phase} Overall', prev_slots))
            offset = phase_week_offsets.get(phase, 0)
            for i in range(week_idx + 1):
                w_dates = weeks_current_phase[i]
                w_slots = _build_date_slots_list_for_batch(dept, batch, w_dates)
                label = f'Week {offset + i + 1}'
                base_segments.append((label, w_slots))
            base_segments.append(('Overall', date_slots_list))
        elif period_type == 'phase' and weeks_current_phase:
            offset = phase_week_offsets.get(phase, 0)
            for i, w in enumerate(weeks_current_phase):
                w_slots = _build_date_slots_list_for_batch(dept, batch, w)
                base_segments.append((f'Week {offset + i + 1}', w_slots))
            base_segments.append(('Overall', date_slots_list))
        return [(label, _filter_date_slots_by_subject(seg, subject_id) if subject_id else seg) for label, seg in base_segments]

    wb = Workbook()
    first = True
    if all_batches:
        all_students = []
        batch_date_slots = {}
        batch_att_map = {}
        batch_date_slots_set = {}
        all_date_slots_pairs = set()
        slot_by_key = {}
        overall_segments_per_batch = []
        subjects_all = set()
        for batch in batches:
            date_slots_list = _build_date_slots_list_for_batch(dept, batch, dates)
            date_slots_list = [(d, s) for d, s in date_slots_list if s]
            batch_date_slots[batch.id] = date_slots_list
            students = list(Student.objects.filter(department=dept, batch=batch))
            students.sort(key=_roll_sort_key)
            all_students.extend(students)
            s_set = set()
            for d, slots in date_slots_list:
                for slot in slots:
                    key = (d, slot.time_slot)
                    all_date_slots_pairs.add(key)
                    if key not in slot_by_key:
                        slot_by_key[key] = slot
                    s_set.add(key)
            batch_date_slots_set[batch.id] = s_set
            all_dates_for_att = set(d for d, _ in date_slots_list)
            att_map = {}
            for d in all_dates_for_att:
                for att in FacultyAttendance.objects.filter(batch=batch, date=d):
                    key = (d, att.lecture_slot)
                    att_map[key] = set(x.strip() for x in (att.absent_roll_numbers or '').split(',') if x.strip())
            batch_att_map[batch.id] = att_map
            segs = build_overall_segments_subject(batch, date_slots_list, None)
            if segs:
                segs = [(label, [(d, s) for d, s in seg if s]) for label, seg in segs]
            overall_segments_per_batch.append((batch, segs or []))
            for d, slots in date_slots_list:
                for slot in slots:
                    if slot.subject_id:
                        subjects_all.add((slot.subject_id, slot.subject.name if slot.subject else 'Subject'))
        all_students.sort(key=lambda s: (s.batch.name, _roll_sort_key(s)))
        date_slots_union = [(d, [slot_by_key[(d, t)]]) for (d, t) in sorted(all_date_slots_pairs)]
        if date_slots_union:
            ws = wb.active
            first = False
            _write_all_batches_combined_sheet(
                ws, batches, all_students, date_slots_union, batch_att_map, batch_date_slots_set,
                styles, overall_segments_per_batch
            )
        for subject_id, subject_name in sorted(subjects_all, key=lambda x: x[1]):
            batch_date_slots_subject = {}
            for batch in batches:
                dsl = batch_date_slots.get(batch.id, [])
                batch_date_slots_subject[batch.id] = [(d, s) for d, s in _filter_date_slots_by_subject(dsl, subject_id) if s]
            if not any(batch_date_slots_subject.values()):
                continue
            def build_fn(b):
                return build_overall_segments_subject(b, batch_date_slots.get(b.id, []), subject_id)
            ws_subj = wb.create_sheet(title=(subject_name[:31] if subject_name else 'Subject'))
            _write_subject_all_batches_sheet(
                ws_subj, subject_name, batches, all_students, batch_date_slots_subject, batch_att_map, styles, build_fn
            )
    else:
        batch = batches[0]
        date_slots_list = _build_date_slots_list_for_batch(dept, batch, dates)
        subjects_in_batch = set()
        for d, slots in date_slots_list:
            for s in slots:
                if s.subject_id:
                    subjects_in_batch.add((s.subject_id, s.subject.name if s.subject else 'Subject'))
        subjects_in_batch = sorted(subjects_in_batch, key=lambda x: x[1])
        students = list(Student.objects.filter(department=dept, batch=batch))
        students.sort(key=_roll_sort_key)
        all_dates_for_att = set(d for d, _ in date_slots_list)
        att_map = {}
        for d in all_dates_for_att:
            for att in FacultyAttendance.objects.filter(batch=batch, date=d):
                key = (d, att.lecture_slot)
                att_map[key] = set(x.strip() for x in (att.absent_roll_numbers or '').split(',') if x.strip())
        for subject_id, subject_name in subjects_in_batch:
            date_slots_subject = _filter_date_slots_by_subject(date_slots_list, subject_id)
            if not any(slots for _, slots in date_slots_subject):
                continue
            overall_segments = build_overall_segments_subject(batch, date_slots_list, subject_id)
            if first:
                ws = wb.active
                first = False
            else:
                ws = wb.create_sheet(title=(subject_name[:31] if subject_name else 'Sheet'))
            _write_one_batch_attendance_sheet(
                ws, batch, date_slots_subject, students, att_map, styles,
                overall_segments=overall_segments, sheet_title=subject_name
            )
    if first:
        messages.warning(request, 'No subject data found for the selected period.')
        return redirect('core:attendance_sheet_subjectwise_manager')
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    batch_label = 'All' if all_batches else batches[0].name
    global_week_num = phase_week_offsets.get(phase, 0) + week_idx + 1 if (period_type == 'weekly' and week_idx is not None and weeks_current_phase) else None
    if period_type == 'daily':
        fname = f'Attendance_Subjectwise_{batch_label}_{dates[0]:%Y-%m-%d}.xlsx'
    elif period_type == 'weekly' and global_week_num:
        fname = f'Attendance_Subjectwise_{batch_label}_{phase}_week{global_week_num}.xlsx'
    else:
        fname = f'Attendance_Subjectwise_{batch_label}_{phase}.xlsx'
    resp = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename={fname}'
    return resp


def _admin_analytics_data(dept, phase=None, week=None):
    """Compute analytics for admin dashboard: at-risk students, batch-wise, subject-wise, weekly trend, heat map.
    week: None or 'all' = all weeks; int (0-based) = cumulative through that week."""
    tp = TermPhase.objects.filter(department=dept).first()
    phases = ['T1', 'T2', 'T3', 'T4']
    if not phase or phase not in phases:
        phase = next((p for p in phases if getattr(tp, f'{p.lower()}_start', None) and getattr(tp, f'{p.lower()}_end', None)), 'T1')
    week_map = {p: _compile_phase_weeks_date_objects(dept, p) for p in phases} if tp else {}
    phase_week_offsets = _get_phase_week_offsets(week_map) if week_map else {}
    weeks = week_map.get(phase, [])
    all_dates = set()
    if week is not None and week != 'all' and isinstance(week, int) and 0 <= week < len(weeks):
        for i in range(week + 1):
            all_dates.update(weeks[i])
    else:
        for w in weeks:
            all_dates.update(w)
    all_dates = sorted(all_dates)
    if not all_dates:
        return {
            'at_risk_students': [], 'batch_wise': [], 'subject_wise': [],
            'weekly_trend': [], 'heat_map': [], 'heat_map_slots': [], 'phase': phase, 'phases': phases,
            'weeks': weeks, 'num_weeks': len(weeks), 'phase_week_offsets': phase_week_offsets,
        }
    batches = list(Batch.objects.filter(department=dept).select_related('department'))
    students = list(Student.objects.filter(department=dept).select_related('batch', 'mentor'))
    students.sort(key=lambda s: (s.batch.name if s.batch else '', _roll_sort_key(s)))
    cancelled_set = get_cancelled_lectures_set(dept)
    batch_scheduled = defaultdict(set)
    for batch in batches:
        _add_batch_schedule_pairs_for_attendance(dept, batch, all_dates, batch_scheduled[batch.id], cancelled_set)
    batch_att_map = defaultdict(lambda: defaultdict(set))
    for batch in batches:
        for att in FacultyAttendance.objects.filter(batch=batch, date__in=all_dates):
            key = (att.date, att.lecture_slot)
            batch_att_map[batch.id][key] = set(x.strip() for x in (att.absent_roll_numbers or '').split(',') if x.strip())
    at_risk = []
    batch_pcts = defaultdict(list)
    subject_totals = defaultdict(lambda: {'held': 0, 'attended': 0})
    heat_map = defaultdict(lambda: defaultdict(lambda: {'held': 0, 'attended': 0}))
    batch_sizes = {b.id: sum(1 for st in students if st.batch_id == b.id) for b in batches}
    batch_held = {}
    for b in batches:
        scheduled = batch_scheduled.get(b.id, set())
        batch_held[b.id] = len(scheduled)
    for s in students:
        str_roll = str(s.roll_no)
        scheduled = batch_scheduled.get(s.batch_id, set())
        held = batch_held.get(s.batch_id, 0)
        attended = sum(1 for (d, slot) in scheduled if (d, slot) in batch_att_map[s.batch_id] and str_roll not in batch_att_map[s.batch_id][(d, slot)])
        pct = round(attended / held * 100, 2) if held else 0
        if held and pct < 75:
            at_risk.append({'student': s, 'held': held, 'attended': attended, 'pct': pct})
        batch_pcts[s.batch_id].append(pct)
    for b in batches:
        sz = batch_sizes.get(b.id, 0)
        for (d, slot) in batch_scheduled.get(b.id, set()):
            day_name = d.strftime('%A')
            heat_map[day_name][slot]['held'] += sz
            attended_count = sum(1 for st in students if st.batch_id == b.id and (d, slot) in batch_att_map[b.id] and str(st.roll_no) not in batch_att_map[b.id][(d, slot)])
            heat_map[day_name][slot]['attended'] += attended_count
            fac, subj = get_faculty_subject_for_slot(d, b, slot)
            subj_name = subj.name if subj else 'N/A'
            subject_totals[subj_name]['held'] += sz
            subject_totals[subj_name]['attended'] += attended_count
    batch_wise = []
    for b in batches:
        pcts = batch_pcts.get(b.id, [])
        avg_pct = round(sum(pcts) / len(pcts), 2) if pcts else 0
        total_held = batch_held.get(b.id, 0) * len(pcts)
        total_attended = sum(
            sum(1 for (d, slot) in batch_scheduled.get(b.id, set())
                if (d, slot) in batch_att_map[b.id] and str(st.roll_no) not in batch_att_map[b.id][(d, slot)])
            for st in students if st.batch_id == b.id
        )
        batch_wise.append({'batch': b, 'held': total_held, 'attended': total_attended, 'pct': avg_pct})
    subject_wise = []
    for name in sorted(subject_totals.keys()):
        t = subject_totals[name]
        pct = round(t['attended'] / t['held'] * 100, 2) if t['held'] else 0
        subject_wise.append({'name': name, 'held': t['held'], 'attended': t['attended'], 'pct': pct})
    offset = phase_week_offsets.get(phase, 0)
    weekly_trend = []
    for i, week_dates in enumerate(weeks):
        if week is not None and week != 'all' and isinstance(week, int) and i > week:
            break
        w_held = w_attended = 0
        for s in students:
            str_roll = str(s.roll_no)
            for (d, slot) in batch_scheduled.get(s.batch_id, set()):
                if d not in week_dates:
                    continue
                w_held += 1
                if (d, slot) in batch_att_map[s.batch_id] and str_roll not in batch_att_map[s.batch_id][(d, slot)]:
                    w_attended += 1
        pct = round(w_attended / w_held * 100, 2) if w_held else 0
        weekly_trend.append({'week': offset + i + 1, 'held': w_held, 'attended': w_attended, 'pct': pct})
    heat_map_list = []
    days_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    slots_order = sorted(set(slot for day_data in heat_map.values() for slot in day_data.keys()))
    for day in days_order:
        if day not in heat_map:
            continue
        row = {'day': day, 'slots': []}
        for slot in slots_order:
            t = heat_map[day].get(slot, {'held': 0, 'attended': 0})
            pct = round(t['attended'] / t['held'] * 100, 2) if t['held'] else None
            row['slots'].append({'slot': slot, 'pct': pct, 'held': t['held']})
        heat_map_list.append(row)
    return {
        'at_risk_students': sorted(at_risk, key=lambda x: x['pct']),
        'batch_wise': batch_wise,
        'subject_wise': subject_wise,
        'weekly_trend': weekly_trend,
        'heat_map': heat_map_list,
        'heat_map_slots': slots_order,
        'phase': phase,
        'phases': phases,
        'weeks': weeks,
        'num_weeks': len(weeks),
        'phase_week_offsets': phase_week_offsets,
    }


@login_required
def admin_analytics_dashboard(request):
    """Admin analytics: charts, at-risk students, heat map."""
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    phase = request.GET.get('phase', 'T1')
    week_param = request.GET.get('week', 'all')
    week = None
    if week_param and week_param != 'all':
        try:
            week = int(week_param)
        except ValueError:
            week = None
    data = _admin_analytics_data(dept, phase, week)
    batch_wise_serial = [{'name': b['batch'].name, 'held': b['held'], 'attended': b['attended'], 'pct': b['pct']} for b in data['batch_wise']]
    phase_week_offsets = data.get('phase_week_offsets', {})
    week_options = [(i, phase_week_offsets.get(phase, 0) + i + 1) for i in range(data.get('num_weeks', 0))]
    selected_week_global_num = (phase_week_offsets.get(phase, 0) + week + 1) if week is not None and 0 <= week < data.get('num_weeks', 0) else None
    ctx = {
        'department': dept,
        'is_super_admin': is_super_admin(request),
        'selected_week': week_param,
        'week_options': week_options,
        'selected_week_global_num': selected_week_global_num,
        **data,
        'weekly_trend_json': json.dumps(data['weekly_trend']),
        'subject_wise_json': json.dumps(data['subject_wise']),
        'batch_wise_json': json.dumps(batch_wise_serial),
    }
    return render(request, 'core/admin/analytics_dashboard.html', ctx)


@login_required
def admin_analytics_at_risk_excel(request):
    """Download at-risk students (below 75%) as Excel."""
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        return redirect('core:admin_dashboard')
    phase = request.GET.get('phase', 'T1')
    week_param = request.GET.get('week', 'all')
    week = None
    if week_param and week_param != 'all':
        try:
            week = int(week_param)
        except ValueError:
            week = None
    data = _admin_analytics_data(dept, phase, week)
    at_risk = data['at_risk_students']
    phase_week_offsets = data.get('phase_week_offsets', {})
    global_week_num = (phase_week_offsets.get(phase, 0) + week + 1) if week is not None and 0 <= week < data.get('num_weeks', 0) else None
    wb = Workbook()
    ws = wb.active
    ws.title = 'At-Risk Students'
    headers = ['Roll No', 'Name', 'Enrollment', 'Mentor Name', 'Batch', 'Lectures Held', 'Attended', 'Attendance %']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    for row_idx, r in enumerate(at_risk, 2):
        s = r['student']
        ws.cell(row_idx, 1, s.roll_no)
        ws.cell(row_idx, 2, s.name)
        ws.cell(row_idx, 3, s.enrollment_no or '')
        ws.cell(row_idx, 4, (s.mentor.short_name if s.mentor else '') or '')
        ws.cell(row_idx, 5, s.batch.name)
        ws.cell(row_idx, 6, r['held'])
        ws.cell(row_idx, 7, r['attended'])
        ws.cell(row_idx, 8, f"{r['pct']:.2f}%")
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 16
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f'At_Risk_Students_{dept.name}_{phase}'
    if global_week_num:
        fname += f'_Week{global_week_num}'
    fname += '.xlsx'
    resp = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


def _get_phase_week_offsets(week_map):
    """Return {phase: offset} so global week num = offset + week_idx + 1. T1=1,2,3; T2=4,5,6...; sequential across phases."""
    phases = ['T1', 'T2', 'T3', 'T4']
    offsets = {}
    cum = 0
    for p in phases:
        offsets[p] = cum
        cum += len(week_map.get(p, []))
    return offsets


def _build_date_to_week_map(dept):
    """Return {date: global_week_num} for all lecture dates in term phases. Used for Week header row in Excel."""
    phases = ['T1', 'T2', 'T3', 'T4']
    week_map = {p: _compile_phase_weeks_date_objects(dept, p) for p in phases}
    phase_week_offsets = _get_phase_week_offsets(week_map)
    date_to_week = {}
    for phase in phases:
        offset = phase_week_offsets.get(phase, 0)
        for i, week_dates in enumerate(week_map.get(phase, [])):
            gw = offset + i + 1
            for d in week_dates:
                date_to_week[d] = gw
    return date_to_week


def _compile_phase_weeks_date_objects(dept, phase):
    """Return list of weeks, each week = list of date objects (lecture days only, excluding holidays)."""
    tp = TermPhase.objects.filter(department=dept).first()
    if not tp:
        return []
    start = getattr(tp, f'{phase.lower()}_start', None)
    end = getattr(tp, f'{phase.lower()}_end', None)
    if not start or not end:
        return []
    from core.schedule_utils import get_all_schedule_days
    days_set = get_all_schedule_days(dept) or _effective_day_set_for_dept(dept, datetime.now().date())
    days_set = {d.lower() for d in days_set if d}
    holidays = get_phase_holidays(dept, phase)
    dates = []
    cur = start
    while cur <= end:
        if cur not in holidays and cur.strftime('%A').lower() in days_set:
            dates.append(cur)
        cur += timedelta(days=1)
    dates = sorted(dates)
    weeks = []
    week = []
    last_w = None
    for d in dates:
        w = d.isocalendar()[1]
        if last_w is not None and w != last_w and week:
            weeks.append(week)
            week = []
        week.append(d)
        last_w = w
    if week:
        weeks.append(week)
    return weeks


@login_required
def compile_attendance(request):
    """Compile attendance page: select phase and week, download single-sheet Excel (all students, week-wise + compile columns)."""
    if not user_can_admin(request):
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    tp = TermPhase.objects.filter(department=dept).first()
    phases = ['T1', 'T2', 'T3', 'T4']
    week_map = {}
    for p in phases:
        start = getattr(tp, f'{p.lower()}_start', None) if tp else None
        end = getattr(tp, f'{p.lower()}_end', None) if tp else None
        if not start or not end:
            week_map[p] = []
            continue
        from core.schedule_utils import get_all_schedule_days
        days_set = get_all_schedule_days(dept) or _effective_day_set_for_dept(dept, datetime.now().date())
        days_set = {d.lower() for d in days_set if d}
        holidays = get_phase_holidays(dept, p)
        dates = []
        cur = start
        while cur <= end:
            if cur not in holidays and cur.strftime('%A').lower() in days_set:
                dates.append(cur)
            cur += timedelta(days=1)
        weeks = []
        week = []
        last_w = None
        for d in sorted(dates):
            w = d.isocalendar()[1]
            if last_w is not None and w != last_w and week:
                weeks.append(week)
                week = []
            week.append(d)
            last_w = w
        if week:
            weeks.append(week)
        week_map[p] = [[d.isoformat() for d in w] for w in weeks]
    phase_week_offsets = _get_phase_week_offsets(week_map)
    ctx = {
        'department': dept,
        'phases': phases,
        'week_map': week_map,
        'week_map_json': json.dumps(week_map),
        'phase_week_offsets_json': json.dumps(phase_week_offsets),
    }
    return render(request, 'core/admin/compile_attendance.html', ctx)


def _admin_notifications_build_mentor_data(dept, phase, week_idx):
    """Build list of mentors with their at-risk mentees (below 75%) and full attendance report.
    Returns: [(mentor, [{'student': s, 'held': n, 'attended': n, 'pct': x, 'week_wise': [...], 'subject_wise': [...]}]), ...]
    """
    weeks = _compile_phase_weeks_date_objects(dept, phase)
    if not weeks:
        return []
    cum_dates = set()
    for i in range(len(weeks)):
        cum_dates.update(weeks[i])
        if i == week_idx:
            break
    students = list(
        Student.objects.filter(department=dept, mentor__isnull=False)
        .select_related('batch', 'mentor')
        .annotate(roll_no_int=Cast('roll_no', IntegerField()))
        .order_by('mentor__full_name', 'batch__name', 'roll_no_int', 'roll_no')
    )
    if not students:
        return []
    cancelled_set = get_cancelled_lectures_set(dept)
    batch_scheduled = defaultdict(set)
    for s in students:
        batch = s.batch
        _add_batch_schedule_pairs_for_attendance(dept, batch, cum_dates, batch_scheduled[batch.id], cancelled_set)
    batch_att_map = defaultdict(lambda: defaultdict(set))
    for batch_id in {s.batch_id for s in students}:
        batch = next(b for b in students if b.batch_id == batch_id).batch
        for att in FacultyAttendance.objects.filter(batch=batch, date__in=cum_dates):
            key = (att.date, att.lecture_slot)
            batch_att_map[batch_id][key] = set(x.strip() for x in (att.absent_roll_numbers or '').split(',') if x.strip())
    mentor_data = defaultdict(list)
    for s in students:
        scheduled = batch_scheduled.get(s.batch_id, set())
        str_roll = str(s.roll_no)
        held = len(scheduled)
        attended = sum(1 for (d, slot) in scheduled if (d, slot) in batch_att_map[s.batch_id] and str_roll not in batch_att_map[s.batch_id][(d, slot)])
        pct = round(attended / held * 100, 2) if held else 0
        if held and pct < 75:
            week_wise = []
            cum_held = cum_attended = 0
            for i, week_dates in enumerate(weeks):
                if i > week_idx:
                    break
                week_set = set(week_dates)
                w_held = sum(1 for (d, slot) in scheduled if d in week_set)
                w_attended = sum(1 for (d, slot) in scheduled if d in week_set and (d, slot) in batch_att_map[s.batch_id] and str_roll not in batch_att_map[s.batch_id][(d, slot)])
                w_pct = round(w_attended / w_held * 100, 2) if w_held else 0
                cum_held += w_held
                cum_attended += w_attended
                cum_pct = round(cum_attended / cum_held * 100, 2) if cum_held else 0
                week_wise.append({'week': i + 1, 'held': w_held, 'attended': w_attended, 'pct': w_pct, 'cum_held': cum_held, 'cum_attended': cum_attended, 'cum_pct': cum_pct})
            subject_wise = defaultdict(lambda: {'held': 0, 'attended': 0})
            for (d, slot) in scheduled:
                fac, subj = get_faculty_subject_for_slot(d, s.batch, slot)
                subj_name = subj.name if subj else 'N/A'
                subject_wise[subj_name]['held'] += 1
                if (d, slot) in batch_att_map[s.batch_id] and str_roll not in batch_att_map[s.batch_id][(d, slot)]:
                    subject_wise[subj_name]['attended'] += 1
            subj_list = [{'name': n, 'held': t['held'], 'attended': t['attended'], 'pct': round(t['attended'] / t['held'] * 100, 2) if t['held'] else 0} for n, t in sorted(subject_wise.items())]
            mentor_data[s.mentor].append({
                'student': s, 'held': held, 'attended': attended, 'pct': pct,
                'week_wise': week_wise, 'subject_wise': subj_list,
            })
    return [(mentor, data) for mentor, data in mentor_data.items() if data]


@login_required
def admin_notifications(request):
    """Admin: Notifications — list students below 75% (phase/week), grouped by mentor. Email to mentor with full report."""
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    tp = TermPhase.objects.filter(department=dept).first()
    phases = ['T1', 'T2', 'T3', 'T4']
    phase = request.GET.get('phase') or request.POST.get('phase', 'T1')
    week_str = request.GET.get('week') or request.POST.get('week')
    week_map = {}
    for p in phases:
        start = getattr(tp, f'{p.lower()}_start', None) if tp else None
        end = getattr(tp, f'{p.lower()}_end', None) if tp else None
        if not start or not end:
            week_map[p] = []
            continue
        from core.schedule_utils import get_all_schedule_days
        days_set = get_all_schedule_days(dept) or _effective_day_set_for_dept(dept, datetime.now().date())
        days_set = {d.lower() for d in days_set if d}
        holidays = get_phase_holidays(dept, p)
        dates = []
        cur = start
        while cur <= end:
            if cur not in holidays and cur.strftime('%A').lower() in days_set:
                dates.append(cur)
            cur += timedelta(days=1)
        weeks = []
        week = []
        last_w = None
        for d in sorted(dates):
            w = d.isocalendar()[1]
            if last_w is not None and w != last_w and week:
                weeks.append(week)
                week = []
            week.append(d)
            last_w = w
        if week:
            weeks.append(week)
        week_map[p] = weeks
    weeks_list = week_map.get(phase, [])
    phase_week_offsets = _get_phase_week_offsets(week_map)
    week_idx = 0
    if week_str is not None:
        try:
            week_idx = int(week_str)
            if week_idx < 0 or week_idx >= len(weeks_list):
                week_idx = max(0, len(weeks_list) - 1)
        except ValueError:
            week_idx = 0
    if weeks_list and week_idx >= len(weeks_list):
        week_idx = len(weeks_list) - 1
    global_week_num = phase_week_offsets.get(phase, 0) + week_idx + 1 if weeks_list else 0
    mentor_data = []
    if weeks_list:
        mentor_data = _admin_notifications_build_mentor_data(dept, phase, week_idx)
    if request.method == 'POST' and request.POST.get('action') == 'email_mentor':
        mentor_id = request.POST.get('mentor_id')
        if mentor_id:
            mentor = Faculty.objects.filter(pk=mentor_id, department=dept).first()
            if mentor:
                mentor_data_full = _admin_notifications_build_mentor_data(dept, phase, week_idx)
                mentor_entry = next((m for m in mentor_data_full if m[0].id == mentor.id), None)
                if mentor_entry:
                    mentor_fac, at_risk_list = mentor_entry
                    email = mentor_fac.email or (mentor_fac.user.email if mentor_fac.user else None)
                    if email:
                        html = render(request, 'core/admin/email_mentor_attendance_report.html', {
                            'mentor': mentor_fac,
                            'phase': phase,
                            'week_num': global_week_num,
                            'at_risk_list': at_risk_list,
                            'department': dept,
                        }).content.decode('utf-8')
                        try:
                            send_mail(
                                subject=f'LJIET Attendance: {len(at_risk_list)} mentee(s) below 75% — Week {global_week_num}',
                                message='Please view this email in HTML format.',
                                from_email=settings.DEFAULT_FROM_EMAIL,
                                recipient_list=[email],
                                html_message=html,
                                fail_silently=False,
                            )
                            messages.success(request, f'Email sent to {mentor_fac.full_name} ({mentor_fac.email or mentor_fac.user.email}).')
                        except Exception as e:
                            err_str = str(e)
                            if '535' in err_str or 'BadCredentials' in err_str or 'Username and Password' in err_str:
                                msg = (
                                    'Email failed: Gmail rejected the login. '
                                    'Use an App Password (not your regular password). '
                                    'Enable 2-Step Verification at myaccount.google.com, then create an App Password at myaccount.google.com/apppasswords. '
                                    'Update EMAIL_HOST_PASSWORD in .env and restart the server.'
                                )
                            else:
                                msg = f'Failed to send email: {e}'
                            messages.error(request, msg)
                    else:
                        messages.error(request, f'No email address for {mentor_fac.full_name}. Add email in Faculty profile.')
                else:
                    messages.error(request, 'Mentor not found in at-risk list.')
            else:
                messages.error(request, 'Invalid mentor.')
        url = reverse('core:admin_notifications') + f'?phase={phase}&week={week_idx}'
        return redirect(url)
    if request.method == 'POST' and request.POST.get('action') == 'email_all':
        mentor_data_full = _admin_notifications_build_mentor_data(dept, phase, week_idx)
        sent = 0
        failed = []
        for mentor_fac, at_risk_list in mentor_data_full:
            email = mentor_fac.email or (mentor_fac.user.email if mentor_fac.user else None)
            if not email:
                failed.append(f'{mentor_fac.full_name} (no email)')
                continue
            html = render(request, 'core/admin/email_mentor_attendance_report.html', {
                'mentor': mentor_fac,
                'phase': phase,
                'week_num': global_week_num,
                'at_risk_list': at_risk_list,
                'department': dept,
            }).content.decode('utf-8')
            try:
                send_mail(
                    subject=f'LJIET Attendance: {len(at_risk_list)} mentee(s) below 75% — Week {global_week_num}',
                    message='Please view this email in HTML format.',
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[email],
                    html_message=html,
                    fail_silently=False,
                )
                sent += 1
            except Exception:
                failed.append(mentor_fac.full_name)
        if sent:
            messages.success(request, f'Email sent to {sent} mentor(s).')
        if failed:
            messages.warning(request, f'Could not email: {", ".join(failed[:5])}{"..." if len(failed) > 5 else ""}')
        url = reverse('core:admin_notifications') + f'?phase={phase}&week={week_idx}'
        return redirect(url)
    week_range = list(range(len(weeks_list)))
    week_offset = phase_week_offsets.get(phase, 0)
    week_options = [(i, week_offset + i + 1) for i in week_range]
    ctx = {
        'department': dept,
        'phases': phases,
        'phase': phase,
        'week_idx': week_idx,
        'global_week_num': global_week_num,
        'week_range': week_range,
        'week_options': week_options,
        'mentor_data': mentor_data,
    }
    return render(request, 'core/admin/notifications.html', ctx)


def _build_slot_subject_cache(batch, cum_dates, batch_scheduled):
    """Pre-build (date, slot) -> subject_name to avoid N queries. Returns dict.
    Must filter by s.day == weekday so slots from other days don't overwrite correct mapping."""
    cache = {}
    if not batch_scheduled:
        return cache
    slot_to_subj = {}
    all_slots = set(slot for (d, slot) in batch_scheduled)
    for d in cum_dates:
        weekday = d.strftime('%A')
        slots = [s for s in _effective_slots_for_date(batch.department, d, extra_filters={'batch': batch}) if s.day == weekday]
        for s in slots:
            if s.time_slot:
                slot_to_subj[(d, s.time_slot)] = (s.subject.name if s.subject else 'N/A')
    adj_list = list(LectureAdjustment.objects.filter(
        batch=batch, date__in=cum_dates, time_slot__in=all_slots
    ).select_related('new_subject').values('date', 'time_slot', 'new_subject__name'))
    adj_map = {(a['date'], a['time_slot']): (a['new_subject__name'] or 'N/A') for a in adj_list}
    extra_list = list(ExtraLecture.objects.filter(batch=batch, date__in=cum_dates).select_related('subject').values('date', 'time_slot', 'subject__name'))
    extra_map = {(a['date'], a['time_slot']): (a['subject__name'] or 'N/A') for a in extra_list}
    for (d, slot) in batch_scheduled:
        key = (d, slot)
        if key in extra_map:
            cache[key] = extra_map[key]
        elif key in adj_map:
            cache[key] = adj_map[key]
        else:
            cache[key] = slot_to_subj.get((d, slot), 'N/A')
    return cache


def _student_analytics_build_data(dept, phase, week_idx, batch_id, roll_search=None):
    """Build student analytics for a batch: list of {student, held, attended, pct, week_wise, subject_wise}.
    Cumulative phases: T2 = T1+T2, T3 = T1+T2+T3, T4 = T1+T2+T3+T4. Optimized for speed."""
    batch = Batch.objects.filter(pk=batch_id, department=dept).first()
    if not batch:
        return [], []
    week_map, _, phase_dates = _student_phase_weeks_and_dates(dept, batch)
    phases = ['T1', 'T2', 'T3', 'T4']
    phase_order_idx = phases.index(phase) if phase in phases else 0
    weeks = week_map.get(phase, [])
    if not weeks:
        return [], []
    if week_idx < 0 or week_idx >= len(weeks):
        week_idx = len(weeks) - 1
    cum_dates = set()
    for i in range(phase_order_idx + 1):
        cum_dates.update(phase_dates.get(phases[i], []))
    if week_idx is not None and weeks:
        cum_dates = set()
        for i in range(phase_order_idx):
            cum_dates.update(phase_dates.get(phases[i], []))
        for i in range(week_idx + 1):
            cum_dates.update(weeks[i])
    students = Student.objects.filter(department=dept, batch=batch).select_related('batch', 'mentor').annotate(roll_no_int=Cast('roll_no', IntegerField())).order_by('roll_no_int', 'roll_no')
    if roll_search:
        q = Q(roll_no__icontains=roll_search) | Q(name__icontains=roll_search) | Q(enrollment_no__icontains=roll_search)
        students = students.filter(q)
    students = list(students)
    if not students:
        return [], []
    cancelled_set = get_cancelled_lectures_set(dept)
    batch_scheduled = set()
    _add_batch_schedule_pairs_for_attendance(dept, batch, cum_dates, batch_scheduled, cancelled_set)
    batch_att_map = {}
    for att in FacultyAttendance.objects.filter(batch=batch, date__in=cum_dates).only('date', 'lecture_slot', 'absent_roll_numbers'):
        key = (att.date, att.lecture_slot)
        batch_att_map[key] = set(x.strip() for x in (att.absent_roll_numbers or '').split(',') if x.strip())
    slot_subj_cache = _build_slot_subject_cache(batch, cum_dates, batch_scheduled)
    prev_dates_list = [set(phase_dates.get(phases[i], [])) for i in range(phase_order_idx)]
    phase_offsets = _get_phase_week_offsets(week_map)
    week_offset = phase_offsets.get(phase, 0)
    result = []
    for s in students:
        str_roll = str(s.roll_no)
        held = len(batch_scheduled)
        attended = sum(1 for (d, slot) in batch_scheduled if (d, slot) in batch_att_map and str_roll not in batch_att_map[(d, slot)])
        pct = round(attended / held * 100, 2) if held else 0
        week_wise = []
        cum_held = cum_attended = 0
        for prev_idx in range(phase_order_idx):
            prev_dates = prev_dates_list[prev_idx]
            prev_held = sum(1 for (d, slot) in batch_scheduled if d in prev_dates)
            prev_attended = sum(1 for (d, slot) in batch_scheduled if d in prev_dates and (d, slot) in batch_att_map and str_roll not in batch_att_map[(d, slot)])
            prev_pct = round(prev_attended / prev_held * 100, 2) if prev_held else 0
            cum_held += prev_held
            cum_attended += prev_attended
            cum_pct = round(cum_attended / cum_held * 100, 2) if cum_held else 0
            week_wise.append({'label': f'{phases[prev_idx]} Overall', 'held': prev_held, 'attended': prev_attended, 'pct': prev_pct, 'cum_held': cum_held, 'cum_attended': cum_attended, 'cum_pct': cum_pct})
        weeks_to_show = range(len(weeks)) if week_idx is None else range(min(week_idx + 1, len(weeks)))
        for i in weeks_to_show:
            week_set = set(weeks[i])
            w_held = sum(1 for (d, slot) in batch_scheduled if d in week_set)
            w_attended = sum(1 for (d, slot) in batch_scheduled if d in week_set and (d, slot) in batch_att_map and str_roll not in batch_att_map[(d, slot)])
            w_pct = round(w_attended / w_held * 100, 2) if w_held else 0
            cum_held += w_held
            cum_attended += w_attended
            cum_pct = round(cum_attended / cum_held * 100, 2) if cum_held else 0
            global_week = week_offset + i + 1
            week_wise.append({'label': f'Week {global_week}', 'week': global_week, 'held': w_held, 'attended': w_attended, 'pct': w_pct, 'cum_held': cum_held, 'cum_attended': cum_attended, 'cum_pct': cum_pct})
        subject_wise = defaultdict(lambda: {'held': 0, 'attended': 0})
        for (d, slot) in batch_scheduled:
            subj_name = slot_subj_cache.get((d, slot), 'N/A')
            subject_wise[subj_name]['held'] += 1
            if (d, slot) in batch_att_map and str_roll not in batch_att_map[(d, slot)]:
                subject_wise[subj_name]['attended'] += 1
        subj_list = [{'name': n, 'held': t['held'], 'attended': t['attended'], 'pct': round(t['attended'] / t['held'] * 100, 2) if t['held'] else 0} for n, t in sorted(subject_wise.items())]
        result.append({'student': s, 'held': held, 'attended': attended, 'pct': pct, 'week_wise': week_wise, 'subject_wise': subj_list})
    batches = list(Batch.objects.filter(department=dept).order_by('name'))
    return result, batches


def _student_analytics_build_data_by_roll_search(dept, phase, week_idx, roll_search):
    """Search students by roll/name/enrollment across ALL batches. Return (result, batches)."""
    batches = list(Batch.objects.filter(department=dept).order_by('name'))
    if not roll_search:
        return [], batches
    weeks = _compile_phase_weeks_date_objects(dept, phase)
    if not weeks or week_idx < 0 or week_idx >= len(weeks):
        return [], batches
    q = Q(roll_no__icontains=roll_search) | Q(name__icontains=roll_search) | Q(enrollment_no__icontains=roll_search)
    students = list(Student.objects.filter(department=dept).filter(q).select_related('batch', 'mentor').annotate(roll_no_int=Cast('roll_no', IntegerField())).order_by('batch__name', 'roll_no_int', 'roll_no'))
    if not students:
        return [], batches
    result = []
    for bid in {s.batch_id for s in students}:
        part, _ = _student_analytics_build_data(dept, phase, week_idx, bid, roll_search)
        result.extend(part)
    result.sort(key=lambda x: (x['student'].batch.name, _roll_sort_key(x['student'])))
    return result, batches


@login_required
def student_analytics(request):
    """Admin & Faculty: Student-wise, batch-wise, phase/week analytics. Filter by batch, search by roll no."""
    dept = None
    is_admin = user_can_admin(request)
    if is_admin:
        # Super admin can set department via POST
        if is_super_admin(request) and request.method == 'POST' and request.POST.get('set_department'):
            request.session['admin_department_id'] = request.POST.get('department_id')
            base = 'core:admin_student_analytics'
            params = request.GET.urlencode()
            url = reverse(base)
            return redirect(f"{url}?{params}" if params else url)
        dept = get_admin_department(request)
        if not dept:
            messages.error(request, 'Select a department first.')
            return redirect('core:admin_dashboard')
    elif user_can_faculty(request):
        faculty = get_faculty_user(request)
        if faculty:
            dept = faculty.department
        if not dept:
            return redirect('accounts:role_redirect')
    else:
        return redirect('accounts:role_redirect')
    phases = ['T1', 'T2', 'T3', 'T4']
    phase = request.GET.get('phase', 'T1')
    week_str = request.GET.get('week')
    batch_id = request.GET.get('batch_id')
    roll_search = request.GET.get('roll_search', '').strip()
    week_map = {p: _compile_phase_weeks_date_objects(dept, p) for p in phases}
    weeks_list = week_map.get(phase, [])
    phase_week_offsets = _get_phase_week_offsets(week_map)
    week_idx = 0
    if week_str is not None:
        try:
            week_idx = int(week_str)
            if week_idx < 0 or week_idx >= len(weeks_list):
                week_idx = max(0, len(weeks_list) - 1)
        except ValueError:
            week_idx = 0
    if weeks_list and week_idx >= len(weeks_list):
        week_idx = len(weeks_list) - 1
    student_data = []
    batches = list(Batch.objects.filter(department=dept).select_related('department').order_by('name'))
    batches_from_all_depts = False
    if not batches:
        # Fallback: show batches from ALL departments so dropdown is never empty
        batches = list(Batch.objects.select_related('department').order_by('department__name', 'name'))
        batches_from_all_depts = bool(batches)
    if weeks_list:
        if roll_search and not batch_id:
            student_data, _batches = _student_analytics_build_data_by_roll_search(dept, phase, week_idx, roll_search)
            if not batches_from_all_depts:
                batches = _batches
        elif batch_id:
            batch = Batch.objects.filter(pk=batch_id).select_related('department').first()
            if batch:
                effective_dept = batch.department
                student_data, _batches = _student_analytics_build_data(effective_dept, phase, week_idx, batch_id, roll_search or None)
                if not batches_from_all_depts:
                    batches = _batches
    week_range = list(range(len(weeks_list)))
    week_offset = phase_week_offsets.get(phase, 0)
    global_week_num = week_offset + week_idx + 1 if weeks_list else 0
    week_options = [(i, week_offset + i + 1) for i in week_range]
    departments = list(Department.objects.all()) if is_admin and is_super_admin(request) else []
    ctx = {
        'department': dept,
        'departments': departments,
        'is_super_admin': is_admin and is_super_admin(request),
        'phases': phases,
        'phase': phase,
        'week_idx': week_idx,
        'global_week_num': global_week_num,
        'week_range': week_range,
        'week_map': week_map,
        'phase_week_offsets': phase_week_offsets,
        'week_options': week_options,
        'batches': batches,
        'batches_from_all_depts': batches_from_all_depts,
        'selected_batch_id': batch_id,
        'roll_search': roll_search,
        'student_data': student_data,
        'is_admin': is_admin,
    }
    return render(request, 'core/student_analytics.html', ctx)


def _mark_analytics_build_data_from_students(students, dept):
    """Build phase_wise marks data for a given student list. Returns list of {student, phase_wise}."""
    if not students:
        return []
    exam_phases = list(ExamPhase.objects.filter(department=dept).order_by('name'))
    phase_subjects = {}
    for ep in exam_phases:
        phase_subjects[ep.id] = list(ExamPhaseSubject.objects.filter(exam_phase=ep).select_related('subject').order_by('subject__name'))
    marks_qs = StudentMark.objects.filter(
        student__in=students,
        exam_phase__department=dept
    ).select_related('subject', 'exam_phase')
    marks_by_student_phase = defaultdict(lambda: defaultdict(dict))
    for m in marks_qs:
        marks_by_student_phase[m.student_id][m.exam_phase_id][m.subject.name] = m.marks_obtained
    result = []
    for s in students:
        phase_wise = []
        for ep in exam_phases:
            subs = phase_subjects.get(ep.id, [])
            subject_marks = []
            for eps in subs:
                marks_val = marks_by_student_phase[s.id][ep.id].get(eps.subject.name)
                try:
                    is_low = marks_val is not None and float(marks_val) < 9
                except (TypeError, ValueError):
                    is_low = False
                subject_marks.append({'name': eps.subject.name, 'marks': marks_val, 'is_low': is_low})
            phase_wise.append({'phase_name': ep.name, 'subjects': subject_marks})
        result.append({'student': s, 'phase_wise': phase_wise})
    return result


def _mark_analytics_build_data(dept, batch_id, roll_search=None):
    """Build mark analytics for students: list of {student, phase_wise: [{phase_name, subjects: [{name, marks}]}]}."""
    batch = Batch.objects.filter(pk=batch_id, department=dept).first()
    if not batch:
        return [], []
    qs = Student.objects.filter(department=dept, batch=batch).select_related('batch')
    if roll_search:
        q = Q(roll_no__icontains=roll_search) | Q(name__icontains=roll_search) | Q(enrollment_no__icontains=roll_search)
        qs = qs.filter(q)
    students = list(qs)
    students.sort(key=lambda s: (s.batch.name if s.batch else '', _roll_sort_key(s)))
    result = _mark_analytics_build_data_from_students(students, dept)
    batches = list(Batch.objects.filter(department=dept).order_by('name'))
    return result, batches


def _mark_analytics_build_risk_data(student_data, threshold=9):
    """From student_data, extract subject-wise at-risk students (marks < threshold)."""
    risk_data = []
    subject_map = {}
    for item in student_data:
        for phase in item['phase_wise']:
            for s in phase['subjects']:
                if s.get('marks') is not None and float(s['marks']) < threshold:
                    subj_name = s['name']
                    if subj_name not in subject_map:
                        subject_map[subj_name] = {'subject_name': subj_name, 'at_risk': []}
                    subject_map[subj_name]['at_risk'].append({
                        'student': item['student'],
                        'phase_name': phase['phase_name'],
                        'marks': s['marks'],
                    })
    for subj_name in sorted(subject_map.keys()):
        entries = subject_map[subj_name]['at_risk']
        entries.sort(key=lambda x: (x['student'].batch.name if x['student'].batch else '', _roll_sort_key(x['student'])))
        risk_data.append(subject_map[subj_name])
    return risk_data


def _mark_analytics_build_data_by_roll_search(dept, roll_search):
    """Search students by roll/name/enrollment across all batches. Return (result, batches)."""
    batches = list(Batch.objects.filter(department=dept).order_by('name'))
    if not roll_search:
        return [], batches
    q = Q(roll_no__icontains=roll_search) | Q(name__icontains=roll_search) | Q(enrollment_no__icontains=roll_search)
    students = list(Student.objects.filter(department=dept).filter(q).select_related('batch'))
    if not students:
        return [], batches
    result = []
    for bid in {s.batch_id for s in students}:
        part, _ = _mark_analytics_build_data(dept, str(bid), roll_search)
        result.extend(part)
    result.sort(key=lambda x: (x['student'].batch.name if x['student'].batch else '', _roll_sort_key(x['student'])))
    return result, batches


@login_required
def mark_analytics(request):
    """Admin: Mark analytics — same design as Student Analytics. Phase-wise: T1, T2, T3, SEE. Each phase shows Subject | Marks."""
    if not user_can_admin(request):
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    batch_id = request.GET.get('batch_id')
    roll_search = request.GET.get('roll_search', '').strip()
    selected_phase = request.GET.get('phase', '').strip()
    student_data = []
    batches = list(Batch.objects.filter(department=dept).select_related('department').order_by('name'))
    if not batches:
        batches = list(Batch.objects.select_related('department').order_by('department__name', 'name'))
    if roll_search and not batch_id:
        student_data, _ = _mark_analytics_build_data_by_roll_search(dept, roll_search)
    elif batch_id == 'all' or (batch_id and batch_id != 'all'):
        if batch_id == 'all':
            qs = Student.objects.filter(department=dept).select_related('batch')
            if roll_search:
                q = Q(roll_no__icontains=roll_search) | Q(name__icontains=roll_search) | Q(enrollment_no__icontains=roll_search)
                qs = qs.filter(q)
            students = list(qs)
            students.sort(key=lambda s: (s.batch.name if s.batch else '', _roll_sort_key(s)))
            student_data = _mark_analytics_build_data_from_students(students, dept)
        else:
            student_data, batches = _mark_analytics_build_data(dept, batch_id, roll_search or None)
    risk_data = _mark_analytics_build_risk_data(student_data)
    if selected_phase:
        for item in student_data:
            item['phase_wise'] = [p for p in item['phase_wise'] if p['phase_name'] == selected_phase]
        _risk_filtered = []
        for r in risk_data:
            filtered_at_risk = [e for e in r['at_risk'] if e['phase_name'] == selected_phase]
            if filtered_at_risk:
                _risk_filtered.append({'subject_name': r['subject_name'], 'at_risk': filtered_at_risk})
        risk_data = _risk_filtered
    exam_phases = list(ExamPhase.objects.filter(department=dept).order_by('name'))
    ctx = {
        'department': dept,
        'batches': batches,
        'exam_phases': exam_phases,
        'selected_batch_id': batch_id,
        'selected_phase': selected_phase,
        'roll_search': roll_search,
        'student_data': student_data,
        'risk_data': risk_data,
        'is_admin': True,
    }
    return render(request, 'core/admin/mark_analytics.html', ctx)


@login_required
def faculty_mark_analytics(request):
    """Faculty: Mark analytics — same design as admin, but only mentorship students."""
    if not user_can_faculty(request):
        return redirect('accounts:role_redirect')
    blocked = _faculty_portal_guard_redirect(request, 'faculty_mark_analytics')
    if blocked:
        return blocked
    faculty = get_faculty_user(request)
    if not faculty:
        return redirect('accounts:logout')
    dept = faculty.department
    batch_id = request.GET.get('batch_id')
    roll_search = request.GET.get('roll_search', '').strip()
    selected_phase = request.GET.get('phase', '').strip()
    qs = Student.objects.filter(mentor=faculty, department=dept).select_related('batch')
    if batch_id and batch_id != 'all':
        qs = qs.filter(batch_id=batch_id)
    if roll_search:
        q = Q(roll_no__icontains=roll_search) | Q(name__icontains=roll_search) | Q(enrollment_no__icontains=roll_search)
        qs = qs.filter(q)
    students = list(qs)
    students.sort(key=lambda s: (s.batch.name if s.batch else '', _roll_sort_key(s)))
    student_data = _mark_analytics_build_data_from_students(students, dept)
    batches = list(Batch.objects.filter(
        id__in=Student.objects.filter(mentor=faculty, department=dept).values_list('batch_id', flat=True).distinct()
    ).order_by('name'))
    risk_data = _mark_analytics_build_risk_data(student_data)
    risk_student_data = []
    if selected_phase:
        for item in student_data:
            item['phase_wise'] = [p for p in item['phase_wise'] if p['phase_name'] == selected_phase]
        _risk_filtered = []
        for r in risk_data:
            filtered_at_risk = [e for e in r['at_risk'] if e['phase_name'] == selected_phase]
            if filtered_at_risk:
                _risk_filtered.append({'subject_name': r['subject_name'], 'at_risk': filtered_at_risk})
        risk_data = _risk_filtered
        for item in student_data:
            has_low = any(
                s.get('marks') is not None and float(s['marks']) < 9
                for p in item['phase_wise'] for s in p['subjects']
            )
            if has_low:
                risk_student_data.append(item)
    else:
        for item in student_data:
            has_low = any(
                s.get('marks') is not None and float(s['marks']) < 9
                for p in item['phase_wise'] for s in p['subjects']
            )
            if has_low:
                risk_student_data.append(item)
    exam_phases = list(ExamPhase.objects.filter(department=dept).order_by('name'))
    ctx = {
        'faculty': faculty,
        'department': dept,
        'batches': batches,
        'exam_phases': exam_phases,
        'selected_batch_id': batch_id,
        'selected_phase': selected_phase,
        'roll_search': roll_search,
        'student_data': student_data,
        'risk_student_data': risk_student_data,
        'risk_data': risk_data,
        'is_admin': False,
    }
    return render(request, 'core/admin/mark_analytics.html', ctx)


def _build_detailed_mark_analytics(dept, phase_id=None, subject_id=None, batch_id=None):
    """Build detailed mark analytics: top 10, highest, avg, batch-wise avg per phase×subject."""
    qs = StudentMark.objects.filter(
        student__department=dept,
        exam_phase__department=dept
    ).select_related('student', 'student__batch', 'exam_phase', 'subject')
    if phase_id:
        qs = qs.filter(exam_phase_id=phase_id)
    if subject_id:
        qs = qs.filter(subject_id=subject_id)
    if batch_id and batch_id != 'all':
        qs = qs.filter(student__batch_id=batch_id)
    marks_list = list(qs)
    phase_subject_data = defaultdict(lambda: defaultdict(list))
    for m in marks_list:
        if m.marks_obtained is not None:
            try:
                val = float(m.marks_obtained)
            except (TypeError, ValueError):
                continue
            phase_subject_data[m.exam_phase.name][m.subject.name].append({
                'student': m.student,
                'marks': val,
                'batch_name': m.student.batch.name if m.student.batch else '',
            })
    result = []
    for phase_name in sorted(phase_subject_data.keys()):
        for subj_name in sorted(phase_subject_data[phase_name].keys()):
            entries = phase_subject_data[phase_name][subj_name]
            entries_sorted = sorted(entries, key=lambda x: (-x['marks'], x['batch_name'], _roll_sort_key(x['student'])))
            top_10 = entries_sorted[:10]
            vals = [e['marks'] for e in entries]
            highest = max(vals) if vals else None
            avg = sum(vals) / len(vals) if vals else None
            batch_totals = defaultdict(list)
            for e in entries:
                batch_totals[e['batch_name']].append(e['marks'])
            batch_avg = {b: sum(v) / len(v) for b, v in batch_totals.items()}
            batch_avg_sorted = sorted(batch_avg.items(), key=lambda x: (-x[1], x[0]))
            result.append({
                'phase_name': phase_name,
                'subject_name': subj_name,
                'top_10': top_10,
                'highest': highest,
                'avg': avg,
                'count': len(entries),
                'batch_avg': batch_avg_sorted,
            })
    return result


@login_required
def detailed_mark_analytics(request):
    """Admin: Detailed Mark Analytics — top 10, highest, avg, batch-wise avg per phase×subject."""
    if not user_can_admin(request):
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    phase_id = request.GET.get('phase_id', '').strip()
    subject_id = request.GET.get('subject_id', '').strip()
    batch_id = request.GET.get('batch_id', 'all')
    if phase_id:
        try:
            phase_id = int(phase_id)
        except ValueError:
            phase_id = None
    if subject_id:
        try:
            subject_id = int(subject_id)
        except ValueError:
            subject_id = None
    analytics = _build_detailed_mark_analytics(dept, phase_id, subject_id, batch_id)
    exam_phases = list(ExamPhase.objects.filter(department=dept).order_by('name'))
    all_subjects = list(Subject.objects.filter(department=dept).order_by('name'))
    batches = list(Batch.objects.filter(department=dept).select_related('department').order_by('name'))
    ctx = {
        'department': dept,
        'analytics': analytics,
        'exam_phases': exam_phases,
        'all_subjects': all_subjects,
        'batches': batches,
        'selected_phase_id': phase_id or '',
        'selected_subject_id': subject_id or '',
        'selected_batch_id': batch_id,
    }
    return render(request, 'core/admin/detailed_mark_analytics.html', ctx)


@login_required
def detailed_mark_analytics_excel(request):
    """Admin: Download Detailed Mark Analytics as Excel."""
    if not user_can_admin(request):
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    if not dept:
        return redirect('core:admin_dashboard')
    phase_id = request.GET.get('phase_id', '').strip()
    subject_id = request.GET.get('subject_id', '').strip()
    batch_id = request.GET.get('batch_id', 'all')
    if phase_id:
        try:
            phase_id = int(phase_id)
        except ValueError:
            phase_id = None
    if subject_id:
        try:
            subject_id = int(subject_id)
        except ValueError:
            subject_id = None
    analytics = _build_detailed_mark_analytics(dept, phase_id, subject_id, batch_id)
    wb = Workbook()
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='1e293b', end_color='1e293b', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    sub_header_fill = PatternFill(start_color='334155', end_color='334155', fill_type='solid')
    for idx, block in enumerate(analytics):
        sheet_title = (block['phase_name'] + '_' + block['subject_name']).replace('/', '-')[:31]
        ws = wb.active if idx == 0 else wb.create_sheet(title=sheet_title)
        if idx == 0:
            ws.title = sheet_title
        row = 1
        for h, v in [('Phase', block['phase_name']), ('Subject', block['subject_name']),
                     ('Highest', block['highest']), ('Avg', f"{block['avg']:.2f}" if block['avg'] is not None else '-'),
                     ('Count', block['count'])]:
            ws.cell(row, 1, h)
            ws.cell(row, 2, str(v) if v is not None else '-')
            row += 1
        row += 1
        headers = ['Rank', 'Roll No', 'Name', 'Batch', 'Marks']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row, c, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1
        for i, e in enumerate(block['top_10'], 1):
            s = e['student']
            for c, v in enumerate([i, s.roll_no, s.name, e['batch_name'], e['marks']], 1):
                cell = ws.cell(row, c, str(v) if v is not None else '')
                cell.border = thin_border
            row += 1
        row += 1
        for c, h in enumerate(['Batch', 'Avg Marks'], 1):
            cell = ws.cell(row, c, h)
            cell.font = header_font
            cell.fill = sub_header_fill
            cell.border = thin_border
        row += 1
        for b, avg in block['batch_avg']:
            for c, v in enumerate([b, f"{avg:.2f}"], 1):
                cell = ws.cell(row, c, str(v) if v is not None else '')
                cell.border = thin_border
            row += 1
    if not analytics:
        ws = wb.active
        ws.title = 'No Data'
        ws.cell(1, 1, 'No marks data for the selected filters.')
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="detailed_mark_analytics.xlsx"'
    wb.save(resp)
    return resp


@login_required
def marks_report(request):
    """Admin: Report Generation — phase, subject, batch selection for marks Excel download."""
    if not user_can_admin(request):
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    exam_phases = list(ExamPhase.objects.filter(department=dept).order_by('name'))
    all_subjects = list(Subject.objects.filter(department=dept).order_by('name'))
    batches = list(Batch.objects.filter(department=dept).select_related('department').order_by('name'))
    ctx = {'department': dept, 'exam_phases': exam_phases, 'all_subjects': all_subjects, 'batches': batches, 'is_admin': True}
    return render(request, 'core/admin/marks_report.html', ctx)


@login_required
def faculty_marks_report(request):
    """Faculty: Report Generation (mentorship students only)."""
    if not user_can_faculty(request):
        return redirect('accounts:role_redirect')
    blocked = _faculty_portal_guard_redirect(request, 'faculty_marks_report')
    if blocked:
        return blocked
    faculty = get_faculty_user(request)
    if not faculty:
        return redirect('accounts:logout')
    dept = faculty.department
    exam_phases = list(ExamPhase.objects.filter(department=dept).order_by('name'))
    all_subjects = list(Subject.objects.filter(department=dept).order_by('name'))
    batches = list(Batch.objects.filter(
        id__in=Student.objects.filter(mentor=faculty, department=dept).values_list('batch_id', flat=True).distinct()
    ).order_by('name'))
    ctx = {'faculty': faculty, 'department': dept, 'exam_phases': exam_phases, 'all_subjects': all_subjects, 'batches': batches, 'is_admin': False}
    return render(request, 'core/admin/marks_report.html', ctx)


@login_required
def mark_analytics_risk_excel(request):
    """Admin: Download at-risk Excel for a specific subject (or all from current filter)."""
    if not user_can_admin(request):
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    if not dept:
        return redirect('core:admin_dashboard')
    subject_name = request.GET.get('subject', '').strip()
    batch_id = request.GET.get('batch_id', 'all')
    if batch_id == 'all':
        students = list(Student.objects.filter(department=dept).select_related('batch', 'mentor'))
    else:
        students = list(Student.objects.filter(department=dept, batch_id=batch_id).select_related('batch', 'mentor'))
    students.sort(key=lambda s: (s.batch.name if s.batch else '', _roll_sort_key(s)))
    student_data = _mark_analytics_build_data_from_students(students, dept)
    risk_data = _mark_analytics_build_risk_data(student_data)
    if subject_name:
        risk_data = [r for r in risk_data if r['subject_name'] == subject_name]
    if not risk_data:
        messages.error(request, 'No at-risk data for this selection.')
        return redirect('core:admin_mark_analytics')
    from django.http import HttpResponse
    wb = Workbook()
    ws = wb.active
    ws.title = subject_name or 'At Risk'
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    red_fill = PatternFill(start_color='FECACA', end_color='FECACA', fill_type='solid')
    for c, h in enumerate(['Roll No', 'Mentor', 'Name', 'Batch', 'Enrollment', 'Phase', 'Marks'], 1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    row = 2
    for subj_data in risk_data:
        for entry in subj_data['at_risk']:
            s = entry['student']
            vals = [s.roll_no, s.mentor.short_name if s.mentor else '', s.name, s.batch.name if s.batch else '', s.enrollment_no or '', entry['phase_name'], entry['marks']]
            for c, val in enumerate(vals, 1):
                cell = ws.cell(row, c, str(val) if val is not None else '')
                cell.border = thin_border
                if c == 7 and val is not None and str(val) != '' and float(val) < 9:
                    cell.fill = red_fill
            row += 1
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="at_risk_{(subject_name or "all").replace(" ", "_")}.xlsx"'
    wb.save(resp)
    return resp


@login_required
def mark_analytics_risk_all_excel(request):
    """Admin: Download ALL batches, ALL at-risk students. Phase selector. Two-row header: Phase (merged), subject names. Red for < 9."""
    if not user_can_admin(request):
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    if not dept:
        return redirect('core:admin_dashboard')
    phase_param = request.GET.get('phase', 'all')
    students = list(Student.objects.filter(department=dept).select_related('batch', 'mentor'))
    students.sort(key=lambda s: (s.batch.name if s.batch else '', _roll_sort_key(s)))
    student_data = _mark_analytics_build_data_from_students(students, dept)
    risk_data = _mark_analytics_build_risk_data(student_data)
    exam_phases = list(ExamPhase.objects.filter(department=dept).order_by('name'))
    phase_subjects = {ep.name: [eps.subject.name for eps in ExamPhaseSubject.objects.filter(exam_phase=ep).select_related('subject').order_by('subject__name')] for ep in exam_phases}
    selected_phases = [p for p in exam_phases if p.name == phase_param] if phase_param and phase_param != 'all' else exam_phases
    marks_by_student_phase_subj = defaultdict(lambda: defaultdict(dict))
    for item in student_data:
        for ph in item['phase_wise']:
            for s in ph['subjects']:
                if s.get('marks') is not None:
                    marks_by_student_phase_subj[item['student'].id][ph['phase_name']][s['name']] = float(s['marks'])
    at_risk_ids = set()
    for r in risk_data:
        for entry in r['at_risk']:
            at_risk_ids.add(entry['student'].id)
    from django.http import HttpResponse
    wb = Workbook()
    ws = wb.active
    ws.title = 'All At Risk'
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
    phase_fill = PatternFill(start_color='B91C1C', end_color='B91C1C', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    red_fill = PatternFill(start_color='FECACA', end_color='FECACA', fill_type='solid')
    col = 1
    for h in ['Mentor', 'Roll No', 'Enrollment', 'Name', 'Batch']:
        cell = ws.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        col += 1
    for ep in selected_phases:
        subs = phase_subjects.get(ep.name, [])
        if not subs:
            continue
        start_col = col
        for sub in subs:
            cell = ws.cell(2, col, sub)
            cell.font = header_font
            cell.fill = phase_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            col += 1
        cell = ws.cell(1, start_col, ep.name)
        cell.font = header_font
        cell.fill = phase_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=col - 1)
    row = 3
    for item in student_data:
        if item['student'].id not in at_risk_ids:
            continue
        s = item['student']
        cols = [s.mentor.short_name if s.mentor else '', s.roll_no, s.enrollment_no or '', s.name, s.batch.name if s.batch else '']
        for ep in selected_phases:
            for subj in phase_subjects.get(ep.name, []):
                m = marks_by_student_phase_subj[s.id].get(ep.name, {}).get(subj)
                if m is not None and float(m) < 9:
                    cols.append(str(m))
                else:
                    cols.append('-')
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row, c, val)
            cell.border = thin_border
            if c >= 6 and val != '-' and val:
                try:
                    if float(val) < 9:
                        cell.fill = red_fill
                except ValueError:
                    pass
        row += 1
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="all_batches_all_students_at_risk_{phase_param or "all"}.xlsx"'
    wb.save(resp)
    return resp


@login_required
def faculty_mark_analytics_risk_excel(request):
    """Faculty: Download at-risk Excel for a specific subject (mentorship students only)."""
    if not user_can_faculty(request):
        return redirect('core:faculty_mark_analytics')
    blocked = _faculty_portal_guard_redirect(request, 'faculty_mark_analytics_risk_excel')
    if blocked:
        return blocked
    faculty = get_faculty_user(request)
    if not faculty:
        return redirect('accounts:logout')
    dept = faculty.department
    subject_name = request.GET.get('subject', '').strip()
    batch_id = request.GET.get('batch_id', 'all')
    qs = Student.objects.filter(mentor=faculty, department=dept).select_related('batch', 'mentor')
    if batch_id and batch_id != 'all':
        qs = qs.filter(batch_id=batch_id)
    students = list(qs)
    students.sort(key=lambda s: (s.batch.name if s.batch else '', _roll_sort_key(s)))
    student_data = _mark_analytics_build_data_from_students(students, dept)
    risk_data = _mark_analytics_build_risk_data(student_data)
    if subject_name:
        risk_data = [r for r in risk_data if r['subject_name'] == subject_name]
    if not risk_data:
        messages.error(request, 'No at-risk data for this selection.')
        return redirect('core:faculty_mark_analytics')
    from django.http import HttpResponse
    wb = Workbook()
    ws = wb.active
    ws.title = subject_name or 'At Risk'
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    red_fill = PatternFill(start_color='FECACA', end_color='FECACA', fill_type='solid')
    for c, h in enumerate(['Roll No', 'Mentor', 'Name', 'Batch', 'Enrollment', 'Phase', 'Marks'], 1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    row = 2
    for subj_data in risk_data:
        for entry in subj_data['at_risk']:
            s = entry['student']
            vals = [s.roll_no, s.mentor.short_name if s.mentor else '', s.name, s.batch.name if s.batch else '', s.enrollment_no or '', entry['phase_name'], entry['marks']]
            for c, val in enumerate(vals, 1):
                cell = ws.cell(row, c, str(val) if val is not None else '')
                cell.border = thin_border
                if c == 7 and val is not None and str(val) != '' and float(val) < 9:
                    cell.fill = red_fill
            row += 1
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="at_risk_{(subject_name or "all").replace(" ", "_")}.xlsx"'
    wb.save(resp)
    return resp


@login_required
def faculty_mark_analytics_risk_all_excel(request):
    """Faculty: Download all mentorship students at risk. Phase selector. Two-row header: Phase (merged), subject names only. Red for < 9."""
    if not user_can_faculty(request):
        return redirect('core:faculty_mark_analytics')
    blocked = _faculty_portal_guard_redirect(request, 'faculty_mark_analytics_risk_all_excel')
    if blocked:
        return blocked
    faculty = get_faculty_user(request)
    if not faculty:
        return redirect('accounts:logout')
    dept = faculty.department
    phase_param = request.GET.get('phase', 'all')
    students = list(Student.objects.filter(mentor=faculty, department=dept).select_related('batch', 'mentor'))
    students.sort(key=lambda s: (s.batch.name if s.batch else '', _roll_sort_key(s)))
    student_data = _mark_analytics_build_data_from_students(students, dept)
    risk_data = _mark_analytics_build_risk_data(student_data)
    exam_phases = list(ExamPhase.objects.filter(department=dept).order_by('name'))
    phase_subjects = {ep.name: [eps.subject.name for eps in ExamPhaseSubject.objects.filter(exam_phase=ep).select_related('subject').order_by('subject__name')] for ep in exam_phases}
    selected_phases = [p for p in exam_phases if p.name == phase_param] if phase_param and phase_param != 'all' else exam_phases
    marks_by_student_phase_subj = defaultdict(lambda: defaultdict(dict))
    for item in student_data:
        for ph in item['phase_wise']:
            for s in ph['subjects']:
                if s.get('marks') is not None:
                    marks_by_student_phase_subj[item['student'].id][ph['phase_name']][s['name']] = float(s['marks'])
    at_risk_ids = set()
    for r in risk_data:
        for entry in r['at_risk']:
            at_risk_ids.add(entry['student'].id)
    from django.http import HttpResponse
    wb = Workbook()
    ws = wb.active
    ws.title = 'All At Risk'
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
    phase_fill = PatternFill(start_color='B91C1C', end_color='B91C1C', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    red_fill = PatternFill(start_color='FECACA', end_color='FECACA', fill_type='solid')
    col = 1
    for h in ['Mentor', 'Roll No', 'Enrollment', 'Name', 'Batch']:
        cell = ws.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        col += 1
    for ep in selected_phases:
        subs = phase_subjects.get(ep.name, [])
        if not subs:
            continue
        start_col = col
        for sub in subs:
            cell = ws.cell(2, col, sub)
            cell.font = header_font
            cell.fill = phase_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            col += 1
        cell = ws.cell(1, start_col, ep.name)
        cell.font = header_font
        cell.fill = phase_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=col - 1)
    row = 3
    for item in student_data:
        if item['student'].id not in at_risk_ids:
            continue
        s = item['student']
        cols = [s.mentor.short_name if s.mentor else '', s.roll_no, s.enrollment_no or '', s.name, s.batch.name if s.batch else '']
        for ep in selected_phases:
            for subj in phase_subjects.get(ep.name, []):
                m = marks_by_student_phase_subj[s.id].get(ep.name, {}).get(subj)
                if m is not None and float(m) < 9:
                    cols.append(str(m))
                else:
                    cols.append('-')
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row, c, val)
            cell.border = thin_border
            if c >= 6 and val != '-' and val:
                try:
                    if float(val) < 9:
                        cell.fill = red_fill
                except ValueError:
                    pass
        row += 1
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="all_mentorship_at_risk_{phase_param or "all"}.xlsx"'
    wb.save(resp)
    return resp


def _mark_analytics_report_excel_impl(request, is_admin):
    """Generate marks report Excel. Phase headers with subjects underneath, red for marks < 9."""
    if is_admin:
        if not user_can_admin(request):
            return redirect('core:admin_dashboard')
        dept = get_admin_department(request)
        if not dept:
            return redirect('core:admin_dashboard')
        students = Student.objects.filter(department=dept).select_related('batch', 'mentor')
    else:
        if not user_can_faculty(request):
            return redirect('core:faculty_mark_analytics')
        blocked = _faculty_portal_guard_redirect(request, 'faculty_mark_analytics_report_excel')
        if blocked:
            return blocked
        faculty = get_faculty_user(request)
        if not faculty:
            return redirect('accounts:logout')
        dept = faculty.department
        students = Student.objects.filter(mentor=faculty, department=dept).select_related('batch', 'mentor')
    batch_id = request.GET.get('batch_id', 'all')
    if batch_id and batch_id != 'all':
        students = students.filter(batch_id=batch_id)
    phase_combo = request.GET.get('phase_combo', '')
    subject_ids = request.GET.getlist('subject_ids')
    if not phase_combo:
        messages.error(request, 'Select a phase.')
        return redirect('core:admin_marks_report' if is_admin else 'core:faculty_marks_report')
    students = list(students)
    students.sort(key=lambda s: (s.batch.name if s.batch else '', _roll_sort_key(s)))
    exam_phases = list(ExamPhase.objects.filter(department=dept).order_by('name'))
    phase_names = [ep.name for ep in exam_phases]
    if phase_combo == 'all':
        selected_phases = phase_names
    elif phase_combo.startswith('combo_'):
        n = int(phase_combo.replace('combo_', '')) if phase_combo[6:].isdigit() else 2
        selected_phases = phase_names[:min(n, len(phase_names))]
    else:
        selected_phases = [phase_combo] if phase_combo in phase_names else [p for p in phase_names if p == phase_combo]
    if not selected_phases:
        selected_phases = [phase_combo]
    phase_subjects = {}
    for ep in ExamPhase.objects.filter(department=dept, name__in=selected_phases).order_by('name'):
        subs = list(ExamPhaseSubject.objects.filter(exam_phase=ep).select_related('subject').order_by('subject__name'))
        if subject_ids:
            try:
                sid_set = {int(x) for x in subject_ids if str(x).strip().isdigit()}
                if sid_set:
                    subs = [s for s in subs if s.subject_id in sid_set]
            except (ValueError, TypeError):
                pass
        phase_subjects[ep.name] = [eps.subject.name for eps in subs]
    student_data = _mark_analytics_build_data_from_students(students, dept)
    marks_map = defaultdict(lambda: defaultdict(dict))
    for item in student_data:
        for ph in item['phase_wise']:
            for s in ph['subjects']:
                if s.get('marks') is not None:
                    marks_map[item['student'].id][ph['phase_name']][s['name']] = s['marks']
    from django.http import HttpResponse
    wb = Workbook()
    ws = wb.active
    ws.title = 'Marks Report'
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    phase_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    red_fill = PatternFill(start_color='FECACA', end_color='FECACA', fill_type='solid')
    col = 1
    for h in ['Roll No', 'Name', 'Batch', 'Enrollment', 'Mentor']:
        cell = ws.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        col += 1
    for phase_name in selected_phases:
        subs = phase_subjects.get(phase_name, [])
        if not subs:
            continue
        start_col = col
        for sub in subs:
            cell = ws.cell(2, col, sub)
            cell.font = header_font
            cell.fill = phase_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            col += 1
        cell = ws.cell(1, start_col, phase_name)
        cell.font = header_font
        cell.fill = phase_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=col - 1)
    data_row = 3
    for item in student_data:
        s = item['student']
        col = 1
        for val in [s.roll_no, s.name, s.batch.name if s.batch else '', s.enrollment_no or '', s.mentor.short_name if s.mentor else '']:
            cell = ws.cell(data_row, col, str(val) if val is not None else '')
            cell.border = thin_border
            col += 1
        for phase_name in selected_phases:
            for sub in phase_subjects.get(phase_name, []):
                m = marks_map[s.id].get(phase_name, {}).get(sub)
                val = str(m) if m is not None else ''
                cell = ws.cell(data_row, col, val)
                cell.border = thin_border
                if m is not None and float(m) < 9:
                    cell.fill = red_fill
                col += 1
        data_row += 1
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="marks_report_{phase_combo}.xlsx"'
    wb.save(resp)
    return resp


@login_required
def mark_analytics_report_excel(request):
    """Admin: Download marks report - phase/subject/batch selection."""
    return _mark_analytics_report_excel_impl(request, is_admin=True)


@login_required
def faculty_mark_analytics_report_excel(request):
    """Faculty: Download marks report (mentorship students only)."""
    return _mark_analytics_report_excel_impl(request, is_admin=False)


@login_required
def compile_attendance_excel(request):
    """Download compile attendance: one sheet, all students, columns = Roll No, Name, Batch, W1 Held/Attended/%, W2 Cum Held/Attended/%, ..., Total Held, Total Attended, Total %."""
    if not user_can_admin(request):
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    phase = request.GET.get('phase')
    week_str = request.GET.get('week')
    if not dept or not phase:
        return redirect('core:compile_attendance')
    try:
        week_idx = int(week_str) if week_str is not None else 0
    except Exception:
        week_idx = 0
    weeks = _compile_phase_weeks_date_objects(dept, phase)
    if week_idx < 0 or week_idx >= len(weeks):
        return redirect('core:compile_attendance')
    phases_order = ['T1', 'T2', 'T3', 'T4']
    week_map = {p: _compile_phase_weeks_date_objects(dept, p) for p in phases_order}
    phase_week_offsets = _get_phase_week_offsets(week_map)
    offset = phase_week_offsets.get(phase, 0)
    global_week_num = offset + week_idx + 1
    # Date sets: week 1 only = weeks[0], week 2 cum = weeks[0]+weeks[1], ...
    cumulative_dates = []
    for i in range(week_idx + 1):
        cumulative_dates.append(set(weeks[i]) if i == 0 else cumulative_dates[-1] | set(weeks[i]))
    # All students, all batches, sorted by batch then roll_no (numeric ascending)
    students = list(Student.objects.filter(department=dept).select_related('batch', 'mentor').order_by('batch__name'))
    students.sort(key=lambda s: (s.batch.name, _roll_sort_key(s)))
    if not students:
        messages.error(request, 'No students in this department.')
        return redirect('core:compile_attendance')
    # Scheduled slots per batch: (date, time_slot) from timetable (ScheduleSlot) for all dates in selected range
    all_dates_in_range = cumulative_dates[week_idx]  # set of dates through selected week
    cancelled_set = get_cancelled_lectures_set(dept)
    batch_scheduled = defaultdict(set)
    for batch_id in {s.batch_id for s in students}:
        batch = next(b for b in students if b.batch_id == batch_id).batch
        _add_batch_schedule_pairs_for_attendance(dept, batch, all_dates_in_range, batch_scheduled[batch_id], cancelled_set)
    # Attendance: for each batch, (date, lecture_slot) -> set of absent roll numbers
    batch_att_map = defaultdict(lambda: defaultdict(set))
    for batch_id in {s.batch_id for s in students}:
        batch = next(b for b in students if b.batch_id == batch_id).batch
        for att in FacultyAttendance.objects.filter(batch=batch):
            key = (att.date, att.lecture_slot)
            batch_att_map[batch_id][key] = set(x.strip() for x in (att.absent_roll_numbers or '').split(',') if x.strip())
    # Per student, per week range: held = scheduled (date, slot) in date_set; attended = those where attendance was taken and student not in absent list
    def held_attended_for_dates(batch_id, str_roll, date_set):
        scheduled = batch_scheduled[batch_id]
        count_held = sum(1 for (d, slot) in scheduled if d in date_set)
        count_attended = 0
        for (d, slot) in scheduled:
            if d not in date_set:
                continue
            if (d, slot) not in batch_att_map[batch_id]:
                continue  # no attendance taken for this slot
            if str_roll not in batch_att_map[batch_id][(d, slot)]:
                count_attended += 1
        return count_held, count_attended
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font_white = Font(bold=True, color='FFFFFF')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Compile Attendance'
    col = 1
    ws.cell(1, col, 'Roll No').font = header_font_white
    ws.cell(1, col).fill = header_fill
    ws.cell(1, col).border = thin_border
    col += 1
    ws.cell(1, col, 'Student Name').font = header_font_white
    ws.cell(1, col).fill = header_fill
    ws.cell(1, col).border = thin_border
    col += 1
    ws.cell(1, col, 'Batch').font = header_font_white
    ws.cell(1, col).fill = header_fill
    ws.cell(1, col).border = thin_border
    col += 1
    ws.cell(1, col, 'Mentor Name').font = header_font_white
    ws.cell(1, col).fill = header_fill
    ws.cell(1, col).border = thin_border
    col += 1
    for i in range(week_idx + 1):
        gw = offset + i + 1
        if i == 0:
            label = f'Week {gw}'
        else:
            label = f'Week {gw} (Cum)'
        for suffix in ('Held', 'Attended', '%'):
            ws.cell(1, col, f'{label} {suffix}').font = header_font_white
            ws.cell(1, col).fill = header_fill
            ws.cell(1, col).border = thin_border
            col += 1
    total_col = col  # first of the three Total columns (Held, Attended, %)
    ws.cell(1, col, 'Total Held').font = header_font_white
    ws.cell(1, col).fill = header_fill
    ws.cell(1, col).border = thin_border
    col += 1
    ws.cell(1, col, 'Total Attended').font = header_font_white
    ws.cell(1, col).fill = header_fill
    ws.cell(1, col).border = thin_border
    col += 1
    ws.cell(1, col, 'Total %').font = header_font_white
    ws.cell(1, col).fill = header_fill
    ws.cell(1, col).border = thin_border
    col += 1
    for row_idx, s in enumerate(students, start=2):
        str_roll = str(s.roll_no)
        ws.cell(row_idx, 1, s.roll_no).border = thin_border
        ws.cell(row_idx, 2, s.name).border = thin_border
        ws.cell(row_idx, 3, s.batch.name).border = thin_border
        ws.cell(row_idx, 4, s.mentor.short_name if s.mentor else '').border = thin_border
        c = 5
        total_held = total_attended = 0
        for i in range(week_idx + 1):
            date_set = cumulative_dates[i]
            h, a = held_attended_for_dates(s.batch_id, str_roll, date_set)
            pct = round(a / h * 100, 2) if h else 0
            ws.cell(row_idx, c, h).border = thin_border
            c += 1
            ws.cell(row_idx, c, a).border = thin_border
            c += 1
            pct_cell = ws.cell(row_idx, c, f'{pct:.2f}%')
            pct_cell.border = thin_border
            if pct < 75 and h:
                pct_cell.font = Font(bold=True, color='FFFFFF')
                pct_cell.fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
            c += 1
            if i == week_idx:
                total_held, total_attended = h, a
        ws.cell(row_idx, total_col, total_held).border = thin_border
        ws.cell(row_idx, total_col + 1, total_attended).border = thin_border
        tpct = round(total_attended / total_held * 100, 2) if total_held else 0
        tpct_cell = ws.cell(row_idx, total_col + 2, f'{tpct:.2f}%')
        tpct_cell.border = thin_border
        if tpct < 75 and total_held:
            tpct_cell.font = Font(bold=True, color='FFFFFF')
            tpct_cell.fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 10
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f'Compile_Attendance_{phase}_through_week{global_week_num}.xlsx'
    resp = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename={fname}'
    return resp


@login_required
def overall_attendance(request):
    """Overall Attendance: download compiled sheet in format (DIV-A1)_WEEK-1_SY-1_SEM-IV_ COMPILED_ATTENDANCE SHEET."""
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    tp = TermPhase.objects.filter(department=dept).first()
    phases = ['T1', 'T2', 'T3', 'T4']
    week_map = {}
    for p in phases:
        start = getattr(tp, f'{p.lower()}_start', None) if tp else None
        end = getattr(tp, f'{p.lower()}_end', None) if tp else None
        if not start or not end:
            week_map[p] = []
            continue
        from core.schedule_utils import get_all_schedule_days
        days_set = get_all_schedule_days(dept) or _effective_day_set_for_dept(dept, datetime.now().date())
        days_set = {d.lower() for d in days_set if d}
        holidays = get_phase_holidays(dept, p)
        dates = []
        cur = start
        while cur <= end:
            if cur not in holidays and cur.strftime('%A').lower() in days_set:
                dates.append(cur)
            cur += timedelta(days=1)
        weeks = []
        week = []
        last_w = None
        for d in sorted(dates):
            w = d.isocalendar()[1]
            if last_w is not None and w != last_w and week:
                weeks.append(week)
                week = []
            week.append(d)
            last_w = w
        if week:
            weeks.append(week)
        week_map[p] = [[d.isoformat() for d in w] for w in weeks]
    batches = list(Batch.objects.filter(department=dept).order_by('name'))
    phase_week_offsets = _get_phase_week_offsets(week_map)
    ctx = {
        'department': dept,
        'phases': phases,
        'week_map': week_map,
        'week_map_json': json.dumps(week_map),
        'phase_week_offsets_json': json.dumps(phase_week_offsets),
        'batches': batches,
    }
    return render(request, 'core/admin/overall_attendance.html', ctx)


@login_required
def overall_attendance_excel(request):
    """Download Overall Attendance Excel: format (DIV-A1)_WEEK-1_SY-1_SEM-IV_ COMPILED_ATTENDANCE SHEET_2026.xlsx."""
    if not user_can_admin(request):
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    batch_id = request.GET.get('batch')
    phase = request.GET.get('phase')
    week_str = request.GET.get('week')
    fmt = request.GET.get('format', 'sheetwise')  # sheetwise or combined
    if not dept or not phase:
        return redirect('core:overall_attendance')
    try:
        week_idx = int(week_str) if week_str is not None else 0
    except Exception:
        return redirect('core:overall_attendance')
    weeks = _compile_phase_weeks_date_objects(dept, phase)
    if not weeks or week_idx < 0 or week_idx >= len(weeks):
        return redirect('core:overall_attendance')
    week_map = {p: _compile_phase_weeks_date_objects(dept, p) for p in ['T1', 'T2', 'T3', 'T4']}
    phase_week_offsets = _get_phase_week_offsets(week_map)
    global_week_num = phase_week_offsets.get(phase, 0) + week_idx + 1

    all_batches = batch_id == 'all'
    if all_batches:
        batches = list(Batch.objects.filter(department=dept).order_by('name'))
    else:
        batch = Batch.objects.filter(pk=batch_id, department=dept).first()
        if not batch:
            return redirect('core:overall_attendance')
        batches = [batch]

    if not batches:
        messages.error(request, 'No batches in this department.')
        return redirect('core:overall_attendance')

    def _build_subjectwise_from_student_analytics(batch):
        """Use same data as student analytics: _student_analytics_build_data returns subject_wise (held, attended, pct) per student.
        total_held/attended = cumulative. week_only_held/attended = selected week only (for Compiled Attendance of WEEK-N block)."""
        student_data, _ = _student_analytics_build_data(dept, phase, week_idx, batch.id)
        subjects_ordered = []
        rows = []
        for rec in student_data:
            s = rec['student']
            subject_wise = {sw['name']: {'held': sw['held'], 'attended': sw['attended']} for sw in rec['subject_wise']}
            if not subjects_ordered:
                subjects_ordered = [sw['name'] for sw in rec['subject_wise']]
            week_only = {'held': 0, 'attended': 0}
            week_wise = rec.get('week_wise', [])
            sel = next((w for w in week_wise if w.get('week') == global_week_num), None)
            if sel:
                week_only = {'held': sel['held'], 'attended': sel['attended']}
            rows.append({
                'student': s,
                'subject_wise': subject_wise,
                'total_held': rec['held'],
                'total_attended': rec['attended'],
                'week_only_held': week_only['held'],
                'week_only_attended': week_only['attended'],
                'mentor': s.mentor.short_name if s.mentor else '',
            })
        return subjects_ordered, rows

    year = datetime.now().year
    dept_label = (dept.name or 'SY-1').replace(' ', '-')[:20]
    if len(batches) == 1:
        batch_label = f'DIV-{batches[0].name}'
    else:
        names = [b.name for b in batches]
        batch_label = f'DIV-{names[0]} TO {names[-1]}' if names else 'DIV-ALL'
    fname_base = f'({batch_label})_WEEK-{global_week_num}_{dept_label}_SEM-IV_ COMPILED_ATTENDANCE SHEET_{year}.xlsx'

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    pct_red_fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
    title_sy_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    title_subwise_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')

    def _write_subjectwise_sheet(ws, batch, subjects_ordered, rows, combined=False):
        """Write subject-wise compiled format: Roll no, Div, Branch, Enrollment No, Name, [Compiled Attendance of WEEK-N: Total Attended, Total Lecture, Overall % - WEEK ONLY], [per subject: cumulative], OVERALL (cumulative), MENTOR NAME.
        The Compiled Attendance of WEEK-N block shows only that week's data (not cumulative). Subject-wise and OVERALL remain cumulative."""
        week_label = f'WEEK-{global_week_num}'
        title = f'SY ({dept.name}) Sem-IV {year} Compiled Attendance'
        subwise_title = f'Subjectwise Compiled Attendance upto Week-{global_week_num}'
        num_attendance_cols = 3 + 3 * len(subjects_ordered) + 3 + 1
        last_col = 4 + num_attendance_cols
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=4)
        cell = ws.cell(1, 1, title)
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.fill = title_sy_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.merge_cells(start_row=1, start_column=5, end_row=2, end_column=last_col)
        cell = ws.cell(1, 5, subwise_title)
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.fill = title_subwise_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        student_headers = ('Roll no.', 'Div', 'Enrollment No', 'Name')
        for c, label in enumerate(student_headers, start=1):
            cell = ws.cell(4, c, label)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.merge_cells(start_row=4, start_column=c, end_row=8, end_column=c)
        col = 5
        cell = ws.cell(4, col, f'Compiled Attendance of {week_label}')
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells(start_row=4, start_column=col, end_row=7, end_column=col + 2)
        col += 3
        for subj in subjects_ordered:
            cell = ws.cell(4, col, subj)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.merge_cells(start_row=4, start_column=col, end_row=7, end_column=col + 2)
            col += 3
        cell = ws.cell(4, col, 'OVERALL')
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells(start_row=4, start_column=col, end_row=7, end_column=col + 2)
        col += 3
        cell = ws.cell(4, col, 'MENTOR NAME')
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells(start_row=4, start_column=col, end_row=8, end_column=col)
        col = 5
        for _ in range(1 + len(subjects_ordered) + 1):
            for h in ('Total\nAttended', 'Total\nLecture', 'Overall\n%'):
                cell = ws.cell(8, col, h)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                col += 1
        for row_idx, r in enumerate(rows, start=9):
            s = r['student']
            ws.cell(row_idx, 1, s.roll_no).border = thin_border
            div_val = s.batch.name if combined else batch.name
            ws.cell(row_idx, 2, div_val).border = thin_border
            ws.cell(row_idx, 3, s.enrollment_no or '').border = thin_border
            ws.cell(row_idx, 4, s.name).border = thin_border
            c = 5
            wo_held, wo_attended = r.get('week_only_held', 0), r.get('week_only_attended', 0)
            ws.cell(row_idx, c, wo_attended).border = thin_border
            c += 1
            ws.cell(row_idx, c, wo_held).border = thin_border
            c += 1
            wo_pct = round(wo_attended / wo_held * 100, 2) if wo_held else 0
            wo_pct_cell = ws.cell(row_idx, c, f'{wo_pct:.2f}%' if wo_held else '')
            wo_pct_cell.border = thin_border
            if wo_held and wo_pct < 75:
                wo_pct_cell.font = Font(bold=True, color='FFFFFF')
                wo_pct_cell.fill = pct_red_fill
            c += 1
            for subj in subjects_ordered:
                sw = r['subject_wise'].get(subj, {'held': 0, 'attended': 0})
                held, attended = sw['held'], sw['attended']
                pct = round(attended / held * 100, 2) if held else 0
                ws.cell(row_idx, c, attended).border = thin_border
                c += 1
                ws.cell(row_idx, c, held).border = thin_border
                c += 1
                pct_cell = ws.cell(row_idx, c, f'{pct:.2f}%' if held else '')
                pct_cell.border = thin_border
                if held and pct < 75:
                    pct_cell.font = Font(bold=True, color='FFFFFF')
                    pct_cell.fill = pct_red_fill
                c += 1
            ws.cell(row_idx, c, r['total_attended']).border = thin_border
            c += 1
            ws.cell(row_idx, c, r['total_held']).border = thin_border
            c += 1
            tpct = round(r['total_attended'] / r['total_held'] * 100, 2) if r['total_held'] else 0
            tpct_cell = ws.cell(row_idx, c, f'{tpct:.2f}%' if r['total_held'] else '')
            tpct_cell.border = thin_border
            if r['total_held'] and tpct < 75:
                tpct_cell.font = Font(bold=True, color='FFFFFF')
                tpct_cell.fill = pct_red_fill
            c += 1
            ws.cell(row_idx, c, r['mentor']).border = thin_border
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 28
        ws.freeze_panes = 'E9'

    wb = Workbook()
    if fmt == 'combined' and len(batches) > 1:
        all_subjects = set()
        all_rows = []
        for batch in batches:
            subs, rows = _build_subjectwise_from_student_analytics(batch)
            all_subjects.update(subs)
            all_rows.extend(rows)
        subjects_ordered = sorted(all_subjects)
        ws = wb.active
        ws.title = 'Combined'[:31]
        _write_subjectwise_sheet(ws, batches[0], subjects_ordered, all_rows, combined=True)
    else:
        first = True
        for batch in batches:
            if first:
                ws = wb.active
                first = False
            else:
                ws = wb.create_sheet(title=(batch.name[:31] if batch.name else 'Sheet'))
            subjects_ordered, rows = _build_subjectwise_from_student_analytics(batch)
            _write_subjectwise_sheet(ws, batch, subjects_ordered, rows)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{fname_base}"'
    return resp


# ---------- Faculty: Attendance Entry ----------

@login_required
def faculty_attendance_entry(request):
    if not user_can_faculty(request):
        return redirect('accounts:role_redirect')
    faculty = get_faculty_user(request)
    if not faculty:
        return redirect('accounts:logout')
    dept = faculty.department
    tp = TermPhase.objects.filter(department=dept).first()

    def dates_for_faculty():
        if not tp:
            return []
        holidays = get_all_holiday_dates(dept)
        entries = _effective_slots_for_faculty_on_date(faculty, datetime.now().date())
        day_set = {e.day.lower() for e in entries if e.day}
        out = []
        for i in range(1, 5):
            start = getattr(tp, f't{i}_start', None)
            end = getattr(tp, f't{i}_end', None)
            if not start or not end:
                continue
            cur = start
            while cur <= end:
                if cur not in holidays and cur.strftime('%A').lower() in day_set:
                    out.append(cur)
                cur += timedelta(days=1)
        # Include dates where this faculty is substitute (LectureAdjustment.new_faculty), but not holidays
        adj_dates = LectureAdjustment.objects.filter(new_faculty=faculty).values_list('date', flat=True).distinct()
        for d in adj_dates:
            if d in holidays:
                continue
            for i in range(1, 5):
                start = getattr(tp, f't{i}_start', None)
                end = getattr(tp, f't{i}_end', None)
                if start and end and start <= d <= end:
                    out.append(d)
                    break
        # Include dates where this faculty has extra lectures
        extra_dates = ExtraLecture.objects.filter(faculty=faculty).values_list('date', flat=True).distinct()
        for d in extra_dates:
            if d in holidays:
                continue
            if tp:
                for i in range(1, 5):
                    start = getattr(tp, f't{i}_start', None)
                    end = getattr(tp, f't{i}_end', None)
                    if start and end and start <= d <= end:
                        out.append(d)
                        break
            else:
                out.append(d)
        return sorted(set(out))

    available_dates = dates_for_faculty()
    date_str = request.GET.get('date')
    selected_date = None
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            if selected_date not in available_dates:
                selected_date = available_dates[0] if available_dates else None
        except Exception:
            selected_date = available_dates[0] if available_dates else None
    else:
        # When no date in URL: auto-select today if available, else first available
        today = datetime.now().date()
        if available_dates and today in available_dates:
            selected_date = today
        else:
            selected_date = available_dates[0] if available_dates else None

    slots_by_batch = defaultdict(list)
    if selected_date:
        cancelled_set = get_cancelled_lectures_set(dept)
        weekday = selected_date.strftime('%A')
        # Exclude slots where this faculty was original but lecture was adjusted to another faculty
        excluded_by_adj = set(
            LectureAdjustment.objects.filter(date=selected_date, original_faculty=faculty).values_list('batch_id', 'time_slot')
        )
        faculty_slots = [s for s in _effective_slots_for_faculty_on_date(faculty, selected_date) if s.day == weekday]
        for s in sorted(faculty_slots, key=lambda x: x.time_slot or ''):
            if (s.batch_id, s.time_slot) in excluded_by_adj:
                continue
            if (selected_date, s.batch_id, s.time_slot) in cancelled_set:
                continue
            slots_by_batch[s.batch].append(s)
        # Add slots where this faculty is substitute (LectureAdjustment) for this date
        existing_pairs = {(b, sl.time_slot) for b, slots in slots_by_batch.items() for sl in slots}
        for adj in LectureAdjustment.objects.filter(date=selected_date, new_faculty=faculty).select_related('batch', 'new_subject', 'new_faculty'):
            if (adj.batch, adj.time_slot) in existing_pairs:
                continue
            if (selected_date, adj.batch_id, adj.time_slot) in cancelled_set:
                continue
            existing_pairs.add((adj.batch, adj.time_slot))
            virtual = type('Slot', (), {
                'batch': adj.batch, 'time_slot': adj.time_slot,
                'subject': adj.new_subject, 'faculty': adj.new_faculty,
            })()
            slots_by_batch[adj.batch].append(virtual)
        # Add slots where this faculty has extra lectures
        for ex in ExtraLecture.objects.filter(date=selected_date, faculty=faculty).select_related('batch', 'subject', 'faculty'):
            if (ex.batch, ex.time_slot) in existing_pairs:
                continue
            if (selected_date, ex.batch_id, ex.time_slot) in cancelled_set:
                continue
            existing_pairs.add((ex.batch, ex.time_slot))
            virtual = type('Slot', (), {
                'batch': ex.batch, 'time_slot': ex.time_slot,
                'subject': ex.subject, 'faculty': ex.faculty,
            })()
            slots_by_batch[ex.batch].append(virtual)
        # Keep slots ordered by time_slot per batch
        for b in slots_by_batch:
            slots_by_batch[b].sort(key=lambda s: s.time_slot or '')
        for batch, slots in slots_by_batch.items():
            for slot in slots:
                slot.is_extra_lecture = _is_extra_lecture_slot(
                    dept, selected_date, batch, (slot.time_slot or '').strip()
                )

    attendance_prefill = defaultdict(lambda: defaultdict(list))
    attendance_reasons = defaultdict(lambda: defaultdict(dict))  # batch_id -> lecture_slot -> {roll_no: reason}
    attendance_updated_at = {}  # (batch_id, lecture_slot) -> updated_at
    if selected_date:
        for a in FacultyAttendance.objects.filter(faculty=faculty, date=selected_date):
            attendance_prefill[a.batch.id][a.lecture_slot] = [x.strip() for x in (a.absent_roll_numbers or '').split(',') if x.strip()]
            attendance_updated_at[(a.batch.id, a.lecture_slot)] = a.updated_at
            try:
                reasons = json.loads(a.absent_reasons or '{}')
                attendance_reasons[a.batch.id][a.lecture_slot] = {k: v for k, v in reasons.items() if v}
            except Exception:
                pass

    batch_students_sorted = {}
    for batch, slots in slots_by_batch.items():
        sorted_students = sorted(batch.student_set.all(), key=_roll_sort_key)
        batch_students_sorted[batch.id] = sorted_students
        batch.students_sorted = sorted_students
        for slot in slots:
            slot.prefill_absent_set = set(attendance_prefill.get(batch.id, {}).get(slot.time_slot, []))
            reasons = attendance_reasons.get(batch.id, {}).get(slot.time_slot, {})
            slot.prefill_reasons = reasons
            slot.students_with_reasons = [(s, reasons.get(str(s.roll_no), 'general')) for s in sorted_students]
            slot.last_updated = attendance_updated_at.get((batch.id, slot.time_slot))
            if selected_date:
                fac, subj = get_faculty_subject_for_slot(selected_date, batch, slot.time_slot)
                slot.display_subject_name = subj.name if subj else (slot.subject.name if slot.subject else 'N/A')
                slot.display_faculty_name = fac.short_name if fac else (slot.faculty.short_name if slot.faculty else '—')

    attendance_locked = selected_date and _is_attendance_locked_for_date(selected_date)
    lock_time_warning = None
    if selected_date and not attendance_locked:
        try:
            lock = AttendanceLockSetting.objects.filter(pk=1).first()
            if lock and lock.enabled:
                lock_time_warning = f'{lock.lock_hour:02d}:{lock.lock_minute:02d} IST'
        except OperationalError:
            pass
    ctx = {
        'faculty': faculty,
        'available_dates': available_dates,
        'selected_date': selected_date,
        'slots_by_batch': dict(slots_by_batch),
        'batch_students_sorted': batch_students_sorted,
        'attendance_locked': attendance_locked,
        'lock_time_warning': lock_time_warning,
    }
    return render(request, 'core/faculty/attendance_entry.html', ctx)


@login_required
def faculty_attendance_save(request):
    if not request.method == 'POST' or not user_can_faculty(request):
        return redirect('core:faculty_attendance_entry')
    faculty = get_faculty_user(request)
    if not faculty:
        return redirect('accounts:logout')
    date_str = request.POST.get('date')
    if not date_str:
        messages.error(request, 'Missing data.')
        return redirect('core:faculty_attendance_entry')
    try:
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        messages.error(request, 'Invalid date.')
        return redirect('core:faculty_attendance_entry')
    if _is_attendance_locked_for_date(selected_date):
        messages.error(request, 'Attendance is locked for this date. Contact admin to update via Manual Attendance.')
        url = reverse('core:faculty_attendance_entry') + f'?date={date_str}'
        return redirect(url)
    batch_id = request.POST.get('batch_id')
    lecture_slot = request.POST.get('lecture_slot', '').strip()
    absent_list = request.POST.getlist('absent_roll_numbers')
    if not batch_id:
        messages.error(request, 'Missing data.')
        return redirect('core:faculty_attendance_entry')
    batch = Batch.objects.filter(pk=batch_id, department=faculty.department).first()
    if not batch:
        messages.error(request, 'Invalid batch.')
        return redirect('core:faculty_attendance_entry')
    absent_roll_numbers = ','.join(x.strip() for x in absent_list if x.strip())
    absent_reasons = {}
    for r in absent_list:
        r = str(r).strip()
        if not r:
            continue
        reason = request.POST.get(f'absent_reason_{r}', 'general').strip() or 'general'
        if reason not in ('general', 'washroom', 'playing game', 'others'):
            reason = 'general'
        absent_reasons[r] = reason
    FacultyAttendance.objects.update_or_create(
        faculty=faculty, date=selected_date, batch=batch, lecture_slot=lecture_slot,
        defaults={
            'absent_roll_numbers': absent_roll_numbers,
            'absent_reasons': json.dumps(absent_reasons) if absent_reasons else '',
        }
    )
    sync_faculty_combine_cache_for_attendance(
        faculty.department, faculty, selected_date, batch, lecture_slot,
    )
    messages.success(request, 'Attendance saved.')
    url = reverse('core:faculty_attendance_entry') + f'?date={date_str}'
    return redirect(url)


def _doubt_date_phase_week(dept, d):
    """Return (phase_str, week_index_within_phase) or (None, None) if date outside term weeks."""
    if not dept or not d:
        return None, None
    for phase in ('T1', 'T2', 'T3', 'T4'):
        weeks = _compile_phase_weeks_date_objects(dept, phase)
        for wi, week_dates in enumerate(weeks):
            if d in week_dates:
                return phase, wi
    return None, None


def _doubt_global_week_for_date(dept, d):
    return _build_date_to_week_map(dept).get(d)


def _faculty_doubt_effective_breakdown(dept, faculty):
    """Phase × week summary for accepted DS (effective h, session count, lecture date span)."""
    accepted = list(
        FacultyDoubtRequest.objects.filter(
            faculty=faculty,
            department=dept,
            status=FacultyDoubtRequest.STATUS_ACCEPTED,
        )
    )
    hours_by_pw = defaultdict(float)
    counts_by_pw = defaultdict(int)
    for dr in accepted:
        phase, wi = _doubt_date_phase_week(dept, dr.date)
        if phase is None or wi is None:
            continue
        key = (phase, wi)
        hours_by_pw[key] += dr.nominal_ds_hours()
        counts_by_pw[key] += 1

    breakdown = []
    scheduled_total = 0.0
    for phase in ('T1', 'T2', 'T3', 'T4'):
        weeks = _compile_phase_weeks_date_objects(dept, phase)
        if not weeks:
            continue
        week_rows = []
        phase_total = 0.0
        for wi, week_dates in enumerate(weeks):
            if not week_dates:
                continue
            dmin, dmax = min(week_dates), max(week_dates)
            key = (phase, wi)
            eff = hours_by_pw.get(key, 0.0)
            scount = counts_by_pw.get(key, 0)
            week_rows.append({
                'week_num': wi + 1,
                'date_min': dmin,
                'date_max': dmax,
                'lecture_days': len(week_dates),
                'session_count': scount,
                'effective_hours': round(eff, 2),
            })
            phase_total += eff
        breakdown.append({
            'phase': phase,
            'weeks': week_rows,
            'phase_total': round(phase_total, 2),
        })
        scheduled_total += phase_total

    outside_h = 0.0
    outside_n = 0
    for dr in accepted:
        phase, wi = _doubt_date_phase_week(dept, dr.date)
        if phase is None:
            outside_h += dr.nominal_ds_hours()
            outside_n += 1

    return {
        'phases': breakdown,
        'scheduled_total': round(scheduled_total, 2),
        'outside_phase_hours': round(outside_h, 2),
        'outside_session_count': outside_n,
    }


def _time_slot_duration_hours(ts):
    """Clock hours from timetable string 'HH:MM-HH:MM'; default 1.0 if unparseable."""
    ts = (ts or '').strip().replace('–', '-').replace('—', '-')
    if not ts:
        return 1.0
    m = re.match(r'^(\d{1,2}):(\d{2})\s*-\s*(\d{1,2}):(\d{2})$', ts.replace(' ', ''))
    if m:
        h1, mi1, h2, mi2 = map(int, m.groups())
        t1 = h1 * 60 + mi1
        t2 = h2 * 60 + mi2
        delta = t2 - t1
        if delta <= 0:
            delta += 24 * 60
        return delta / 60.0
    return 1.0


def _faculty_teaching_on_date(dept, faculty, d):
    """Scheduled teaching (incl. adjustments & extra) for this faculty on date; returns (hours, row dicts)."""
    total = 0.0
    rows = []
    for batch in Batch.objects.filter(department=dept).order_by('name'):
        for slot_obj in _dr_slots_for_batch_on_date(dept, batch, d):
            ts = (slot_obj.time_slot or '').strip()
            if not ts:
                continue
            fac, subj = get_faculty_subject_for_slot(d, batch, ts)
            if not fac or fac.id != faculty.id:
                continue
            h = _time_slot_duration_hours(ts)
            total += h
            rows.append({
                'batch': batch.name,
                'time_slot': ts,
                'subject': (subj.name if subj else '—'),
                'hours': round(h, 2),
            })
    return round(total, 4), rows


def _user_can_faculty_load_report(request):
    return is_hod(request) or is_super_admin(request)


def _dr_lecture_count_for_faculty_on_date(dept, faculty, d):
    """Number of real (non-blank) Daily Report lecture rows for this faculty on date."""
    rows = _dr_collect_rows_for_date(dept, d)
    return sum(1 for r in rows if r['faculty'].id == faculty.id and not r.get('is_blank'))


def _dates_cumulative_upto_phase_week(dept, end_phase, end_week_1based):
    """All lecture-calendar dates from T1 week 1 through end of given phase/week (inclusive)."""
    phases = ('T1', 'T2', 'T3', 'T4')
    if end_phase not in phases:
        end_phase = 'T1'
    try:
        end_week_1based = int(end_week_1based)
    except (TypeError, ValueError):
        end_week_1based = 1
    end_week_1based = max(1, end_week_1based)
    end_idx = phases.index(end_phase)
    out = []
    for pi, phase in enumerate(phases):
        if pi > end_idx:
            break
        weeks = _compile_phase_weeks_date_objects(dept, phase)
        if not weeks:
            continue
        if pi < end_idx:
            iter_weeks = range(len(weeks))
        else:
            iter_weeks = range(min(end_week_1based, len(weeks)))
        for wi in iter_weeks:
            out.extend(weeks[wi])
    return sorted(set(out))


def sync_faculty_combine_cache_for_attendance(dept, faculty, day, batch, lecture_slot):
    """Align FacultyCombineDrCache with saved attendance (DR weekly combine rules)."""
    slot = (lecture_slot or '').strip()
    FacultyCombineDrCache.objects.filter(
        faculty=faculty, date=day, batch=batch, lecture_slot=slot,
    ).delete()
    cancelled = get_cancelled_lectures_set(dept)
    if (day, batch.id, slot) in cancelled:
        return
    rf, rs = get_faculty_subject_for_slot(day, batch, slot)
    if not rf or not rs or rf.id != faculty.id:
        return
    att = FacultyAttendance.objects.filter(
        date=day, batch=batch, lecture_slot=slot, faculty=faculty,
    ).first() or FacultyAttendance.objects.filter(date=day, batch=batch, lecture_slot=slot).first()
    if not att:
        return
    tot = Student.objects.filter(batch=batch).count()
    absent_n = len([x for x in (att.absent_roll_numbers or '').split(',') if x.strip()])
    present = tot - absent_n
    is_extra = _is_extra_lecture_slot(dept, day, batch, slot)
    eff = _dr_slot_effective_load(is_extra, True, present)
    FacultyCombineDrCache.objects.update_or_create(
        faculty=faculty,
        date=day,
        batch=batch,
        lecture_slot=slot,
        defaults={'department': dept, 'subject': rs, 'effective_load': eff},
    )


def rebuild_faculty_combine_cache_dept(dept):
    """Rebuild cache from all attendance rows. Run after deploy or if timetable changes."""
    FacultyCombineDrCache.objects.filter(department=dept).delete()
    qs = FacultyAttendance.objects.filter(batch__department=dept).select_related('faculty', 'batch')
    for att in qs.iterator(chunk_size=1000):
        sync_faculty_combine_cache_for_attendance(
            dept, att.faculty, att.date, att.batch, att.lecture_slot,
        )


def _ensure_combine_dr_cache_populated(dept):
    if FacultyCombineDrCache.objects.filter(department=dept).exists():
        return
    if FacultyAttendance.objects.filter(batch__department=dept).exists():
        rebuild_faculty_combine_cache_dept(dept)


def _faculty_upto_week_choices(dept):
    choices = []
    for phase in ('T1', 'T2', 'T3', 'T4'):
        weeks = _compile_phase_weeks_date_objects(dept, phase)
        for wi, week_dates in enumerate(weeks):
            if not week_dates:
                continue
            d0, d1 = min(week_dates), max(week_dates)
            choices.append({
                'phase': phase,
                'week_num': wi + 1,
                'value': f'{phase}|{wi + 1}',
                'label': f'{phase} — Week {wi + 1}: {d0.strftime("%d %b")} – {d1.strftime("%d %b %Y")}',
            })
    return choices


def _build_faculty_dr_weekly_combine_upto(dept, faculty, end_phase, end_week_1based):
    """Week rows (combine sheet rules) from FacultyCombineDrCache (updated on attendance save)."""
    cum_dates_set = set(_dates_cumulative_upto_phase_week(dept, end_phase, end_week_1based))
    phases = ('T1', 'T2', 'T3', 'T4')
    end_idx = phases.index(end_phase) if end_phase in phases else 0
    try:
        end_week_1based = max(1, int(end_week_1based))
    except (TypeError, ValueError):
        end_week_1based = 1

    week_specs = []
    for pi, phase in enumerate(phases):
        if pi > end_idx:
            break
        weeks = _compile_phase_weeks_date_objects(dept, phase)
        if not weeks:
            continue
        if pi < end_idx:
            wi_range = range(len(weeks))
        else:
            wi_range = range(min(end_week_1based, len(weeks)))
        for wi in wi_range:
            week_specs.append((phase, wi, weeks[wi]))
        if pi == end_idx:
            break

    date_to_week = {}
    for phase, wi, week_dates in week_specs:
        for d0 in week_dates:
            date_to_week[d0] = (phase, wi)

    _ensure_combine_dr_cache_populated(dept)

    week_trip_counts = defaultdict(lambda: defaultdict(int))
    week_eff_sum = defaultdict(float)
    date_batch_counts = defaultdict(Counter)
    date_totals = defaultdict(int)

    for row in FacultyCombineDrCache.objects.filter(
        department=dept, faculty=faculty, date__in=cum_dates_set,
    ).select_related('batch', 'subject'):
        wk = date_to_week.get(row.date)
        if wk is None:
            continue
        week_trip_counts[wk][(row.batch_id, row.subject_id)] += 1
        week_eff_sum[wk] += float(row.effective_load)
        date_batch_counts[row.date][row.batch.name] += 1
        date_totals[row.date] += 1

    week_rows = []
    sr = 0
    cum_lec = 0
    cum_eff_raw = 0.0
    for phase, wi, week_dates in week_specs:
        wk = (phase, wi)
        counts = week_trip_counts[wk]
        n_lec = sum(counts.values())
        eff_raw = week_eff_sum[wk]
        cum_lec += n_lec
        cum_eff_raw += eff_raw
        sr += 1
        date_details = []
        for d0 in sorted(week_dates):
            bc = date_batch_counts[d0]
            batch_list = [{'batch': name, 'count': n} for name, n in sorted(bc.items())]
            date_details.append({
                'date': d0,
                'total': date_totals[d0],
                'batches': batch_list,
            })
        week_rows.append({
            'sr_no': sr,
            'phase': phase,
            'week_num': wi + 1,
            'week_label': f'{phase} — Week {wi + 1}',
            'date_min': min(week_dates),
            'date_max': max(week_dates),
            'lecture_days': len(week_dates),
            'total_lectures': n_lec,
            'total_effective': round(eff_raw, 2),
            'date_details': date_details,
        })

    return {
        'week_rows': week_rows,
        'cumulative_lectures': cum_lec,
        'cumulative_effective': round(cum_eff_raw, 2),
        'cumulative_calendar_days': len(cum_dates_set),
    }


def _parse_doubt_time(s):
    s = (s or '').strip()
    if not s:
        return None
    for fmt in ('%H:%M:%S', '%H:%M'):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return None


@login_required
def faculty_doubt_students_data(request):
    """JSON: students in batch (roll, name, phone) for doubt-solving form."""
    if not user_can_faculty(request):
        return JsonResponse({'error': 'Forbidden'}, status=403)
    faculty = get_faculty_user(request)
    if not faculty:
        return JsonResponse({'error': 'Forbidden'}, status=403)
    if not faculty_portal_feature_allowed(faculty, 'faculty_doubt_students_data'):
        return JsonResponse({'error': 'Forbidden'}, status=403)
    batch_ids = [x for x in request.GET.getlist('batch_id') if str(x).strip().isdigit()]
    if not batch_ids:
        return JsonResponse({'students': []})
    batch_qs = Batch.objects.filter(pk__in=batch_ids, department=faculty.department).order_by('name')
    found_ids = set(batch_qs.values_list('pk', flat=True))
    data = []
    for batch in batch_qs:
        studs = (
            Student.objects.filter(batch=batch, department=faculty.department)
            .annotate(roll_no_int=Cast('roll_no', IntegerField()))
            .order_by('roll_no_int', 'roll_no')
        )
        for s in studs:
            data.append({
                'id': s.id,
                'batch_id': batch.id,
                'batch_name': batch.name,
                'roll_no': s.roll_no,
                'name': s.name,
                'phone': (s.student_phone_number or '').strip(),
            })
    return JsonResponse({'students': data, 'valid_batch_ids': sorted(found_ids)})


@login_required
def faculty_doubt_solving(request):
    if not user_can_faculty(request):
        return redirect('accounts:role_redirect')
    blocked = _faculty_portal_guard_redirect(request, 'faculty_doubt_solving')
    if blocked:
        return blocked
    faculty = get_faculty_user(request)
    if not faculty:
        return redirect('accounts:logout')
    dept = faculty.department
    batches = list(Batch.objects.filter(department=dept).order_by('name'))

    if request.method == 'POST':
        date_str = request.POST.get('date', '').strip()
        batch_ids = [x for x in request.POST.getlist('batch_ids') if str(x).strip().isdigit()]
        student_ids = request.POST.getlist('student_ids')
        start_s = request.POST.get('start_time', '').strip()
        end_s = request.POST.get('end_time', '').strip()
        if not all([date_str, batch_ids, start_s, end_s]):
            messages.error(request, 'Please select at least one batch, and fill date and times.')
            return redirect('core:faculty_doubt_solving')
        if not student_ids:
            messages.error(request, 'Select at least one student.')
            return redirect('core:faculty_doubt_solving')
        try:
            sess_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, 'Invalid date.')
            return redirect('core:faculty_doubt_solving')
        batch_list = list(Batch.objects.filter(pk__in=batch_ids, department=dept).order_by('name'))
        if len(batch_list) != len(set(batch_ids)):
            messages.error(request, 'Invalid batch selection.')
            return redirect('core:faculty_doubt_solving')
        allowed_batch_ids = {b.id for b in batch_list}
        t_start = _parse_doubt_time(start_s)
        t_end = _parse_doubt_time(end_s)
        if not t_start or not t_end:
            messages.error(request, 'Invalid start or end time.')
            return redirect('core:faculty_doubt_solving')
        req = FacultyDoubtRequest.objects.create(
            faculty=faculty,
            department=dept,
            batch=None,
            date=sess_date,
            start_time=t_start,
            end_time=t_end,
            status=FacultyDoubtRequest.STATUS_PENDING,
        )
        req.batches.set(batch_list)
        added = 0
        for sid in student_ids:
            st = Student.objects.filter(pk=sid, department=dept, batch_id__in=allowed_batch_ids).first()
            if st:
                FacultyDoubtRequestStudent.objects.get_or_create(request=req, student=st)
                added += 1
        if not added:
            req.delete()
            messages.error(request, 'No valid students selected.')
            return redirect('core:faculty_doubt_solving')
        messages.success(request, f'Request submitted to HOD with {added} student(s). Status: Pending.')
        q = '&'.join(f'batch_id={b.id}' for b in batch_list)
        return redirect(f"{reverse('core:faculty_doubt_solving')}?{q}")

    batch_ids_get = [x for x in request.GET.getlist('batch_id') if str(x).strip().isdigit()]
    selected_batches = list(Batch.objects.filter(pk__in=batch_ids_get, department=dept).order_by('name'))

    requests_qs = (
        FacultyDoubtRequest.objects.filter(faculty=faculty)
        .select_related('batch')
        .prefetch_related('batches', 'student_lines__student')
        .order_by('-date', '-start_time', '-pk')
    )
    accepted_qs = requests_qs.filter(status=FacultyDoubtRequest.STATUS_ACCEPTED)
    total_hours = round(sum(r.nominal_ds_hours() for r in accepted_qs), 2)
    history = list(requests_qs[:200])

    ctx = {
        'faculty': faculty,
        'batches': batches,
        'selected_batches': selected_batches,
        'selected_batch_ids': [b.id for b in selected_batches],
        'history': history,
        'total_hours': total_hours,
        'students_json_url': reverse('core:faculty_doubt_students_data'),
    }
    return render(request, 'core/faculty/doubt_solving.html', ctx)


@login_required
def faculty_dr_load(request):
    """Faculty DR weekly combine (0.75×); same cache-backed builder as HOD Faculty load."""
    if not user_can_faculty(request):
        return redirect('accounts:role_redirect')
    blocked = _faculty_portal_guard_redirect(request, 'faculty_dr_load')
    if blocked:
        return blocked
    faculty = get_faculty_user(request)
    if not faculty:
        return redirect('accounts:logout')
    dept = faculty.department
    choices = _faculty_upto_week_choices(dept)
    sel_phase, sel_w = 'T1', 1
    upto = (request.GET.get('upto') or '').strip()
    if '|' in upto:
        p, w = upto.split('|', 1)
        if p in ('T1', 'T2', 'T3', 'T4') and w.strip().isdigit():
            sel_phase, sel_w = p, int(w.strip())
    if choices:
        if f'{sel_phase}|{sel_w}' not in {c['value'] for c in choices}:
            last = choices[-1]
            sel_phase, sel_w = last['phase'], last['week_num']
    selected_value = f'{sel_phase}|{sel_w}'
    selected_label = next((c['label'] for c in choices if c['value'] == selected_value), '')
    detail = (
        _build_faculty_dr_weekly_combine_upto(dept, faculty, sel_phase, sel_w)
        if choices
        else None
    )
    ctx = {
        'faculty': faculty,
        'department': dept,
        'upto_choices': choices,
        'selected_upto': selected_value,
        'selected_label': selected_label,
        'detail': detail,
    }
    return render(request, 'core/faculty/dr_load.html', ctx)


@login_required
def hod_doubt_requests(request):
    if not user_can_admin(request) or not is_hod(request):
        messages.error(request, 'Only HOD can access doubt requests.')
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    if not dept:
        return redirect('core:admin_dashboard')
    pending = list(
        FacultyDoubtRequest.objects.filter(department=dept, status=FacultyDoubtRequest.STATUS_PENDING)
        .select_related('faculty', 'batch')
        .prefetch_related('batches', 'student_lines__student')
        .order_by('date', 'start_time', 'pk')
    )
    recent = list(
        FacultyDoubtRequest.objects.filter(department=dept)
        .exclude(status=FacultyDoubtRequest.STATUS_PENDING)
        .select_related('faculty', 'batch', 'reviewed_by')
        .prefetch_related('batches', 'student_lines__student')
        .order_by('-reviewed_at', '-pk')[:80]
    )
    ctx = {
        'department': dept,
        'pending': pending,
        'recent': recent,
    }
    return render(request, 'core/admin/hod_doubt_requests.html', ctx)


@login_required
def admin_faculty_teaching_ds_load(request):
    """HOD/superadmin: same DR weekly combine report as faculty 'My DR weekly load', with faculty selector."""
    if not user_can_admin(request) or not _user_can_faculty_load_report(request):
        messages.error(request, 'Only HOD or super admin can view faculty load.')
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    faculties = list(Faculty.objects.filter(department=dept).order_by('full_name'))
    selected = None
    raw_fid = request.GET.get('faculty_id')
    if raw_fid and str(raw_fid).isdigit():
        selected = next((f for f in faculties if f.id == int(raw_fid)), None)
    if not selected and faculties:
        myf = get_faculty_user(request)
        if myf and myf.department_id == dept.id:
            selected = next((f for f in faculties if f.id == myf.id), None)
        if not selected:
            selected = faculties[0]

    choices = _faculty_upto_week_choices(dept)
    sel_phase, sel_w = 'T1', 1
    upto = (request.GET.get('upto') or '').strip()
    if '|' in upto:
        p, w = upto.split('|', 1)
        if p in ('T1', 'T2', 'T3', 'T4') and w.strip().isdigit():
            sel_phase, sel_w = p, int(w.strip())
    if choices:
        if f'{sel_phase}|{sel_w}' not in {c['value'] for c in choices}:
            last = choices[-1]
            sel_phase, sel_w = last['phase'], last['week_num']
    selected_value = f'{sel_phase}|{sel_w}'
    selected_label = next((c['label'] for c in choices if c['value'] == selected_value), '')
    detail = (
        _build_faculty_dr_weekly_combine_upto(dept, selected, sel_phase, sel_w)
        if selected and choices
        else None
    )
    ctx = {
        'department': dept,
        'faculties': faculties,
        'selected_faculty': selected,
        'upto_choices': choices,
        'selected_upto': selected_value,
        'selected_label': selected_label,
        'detail': detail,
    }
    return render(request, 'core/admin/faculty_teaching_ds_load.html', ctx)


@login_required
def hod_doubt_request_review(request, pk):
    if request.method != 'POST' or not user_can_admin(request) or not is_hod(request):
        return redirect('core:hod_doubt_requests')
    dept = get_admin_department(request)
    if not dept:
        return redirect('core:hod_doubt_requests')
    dr = FacultyDoubtRequest.objects.filter(pk=pk, department=dept).first()
    if not dr or dr.status != FacultyDoubtRequest.STATUS_PENDING:
        messages.error(request, 'Request not found or already reviewed.')
        return redirect('core:hod_doubt_requests')
    action = request.POST.get('action')
    notes = (request.POST.get('notes') or '').strip()[:500]
    if action == 'accept':
        dr.status = FacultyDoubtRequest.STATUS_ACCEPTED
        dr.reviewed_at = timezone.now()
        dr.reviewed_by = request.user
        upd = ['status', 'reviewed_at', 'reviewed_by']
        if notes:
            dr.review_notes = notes
            upd.append('review_notes')
        dr.save(update_fields=upd)
        messages.success(request, 'Request accepted. It now counts as doubt solving.')
    elif action == 'reject':
        dr.status = FacultyDoubtRequest.STATUS_REJECTED
        dr.reviewed_at = timezone.now()
        dr.reviewed_by = request.user
        dr.review_notes = notes
        dr.save(update_fields=['status', 'reviewed_at', 'reviewed_by', 'review_notes'])
        messages.success(request, 'Request rejected.')
    else:
        messages.error(request, 'Invalid action.')
    return redirect('core:hod_doubt_requests')


@login_required
def hod_doubt_request_delete(request, pk):
    if request.method != 'POST' or not user_can_admin(request) or not is_hod(request):
        return redirect('core:hod_doubt_requests')
    dept = get_admin_department(request)
    if not dept:
        return redirect('core:hod_doubt_requests')
    dr = FacultyDoubtRequest.objects.filter(pk=pk, department=dept).first()
    if not dr:
        messages.error(request, 'Request not found.')
        return redirect('core:hod_doubt_requests')
    if dr.status == FacultyDoubtRequest.STATUS_PENDING:
        messages.error(
            request, 'Pending requests cannot be deleted here—use Accept or Reject.',
        )
        return redirect('core:hod_doubt_requests')
    dr.delete()
    messages.success(
        request,
        'Record removed permanently. It no longer appears in totals or Excel exports.',
    )
    return redirect('core:hod_doubt_requests')


@login_required
def hod_doubt_reports_excel(request):
    if not user_can_admin(request) or not is_hod(request):
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    if not dept:
        return redirect('core:admin_dashboard')
    report = (request.GET.get('report') or 'full').lower()
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    hdr_font = Font(bold=True)

    def _detail_sheet(ws):
        ws.append([
            'Request ID', 'Status', 'Faculty', 'Date', 'Batches', 'Phase', 'Week in phase', 'Global week',
            'Start', 'End', 'Effective hours (DS)', '# Students', 'Roll numbers', 'Names', 'Phones',
            'Reviewed at', 'Notes',
        ])
        for c in ws[1]:
            c.font = hdr_font
            c.border = thin
        qs = (
            FacultyDoubtRequest.objects.filter(department=dept)
            .select_related('faculty', 'batch')
            .prefetch_related('batches', 'student_lines__student')
            .order_by('-date', '-pk')
        )
        for dr in qs:
            phase, wi = _doubt_date_phase_week(dept, dr.date)
            gw = _doubt_global_week_for_date(dept, dr.date)
            lines = list(dr.student_lines.all())
            rolls = ', '.join(x.student.roll_no for x in lines)
            names = ', '.join(x.student.name for x in lines)
            phones = ', '.join((x.student.student_phone_number or '').strip() or '—' for x in lines)
            ws.append([
                dr.pk,
                dr.get_status_display(),
                dr.faculty.short_name,
                dr.date.isoformat(),
                dr.batches_label(),
                phase or '—',
                (wi + 1) if wi is not None else '—',
                gw if gw is not None else '—',
                dr.start_time.strftime('%H:%M'),
                dr.end_time.strftime('%H:%M'),
                round(dr.nominal_ds_hours(), 2),
                len(lines),
                rolls,
                names,
                phones,
                dr.reviewed_at.strftime('%Y-%m-%d %H:%M') if dr.reviewed_at else '',
                dr.review_notes or '',
            ])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for c in row:
                c.border = thin
        for col in range(1, 18):
            ws.column_dimensions[get_column_letter(col)].width = 14
        ws.column_dimensions['N'].width = 28

    def _doubt_accepted_qs():
        return list(
            FacultyDoubtRequest.objects.filter(
                department=dept, status=FacultyDoubtRequest.STATUS_ACCEPTED
            ).select_related('faculty').prefetch_related('batches', 'student_lines__student')
        )

    def _write_doubt_detail_block(ws, r, title_text, week_dates, accepted_all, thin_local, hdr_fill_blue, hdr_w, title_f, sub_f):
        """Write one week’s detail table; return (next_row, week_effective_sum)."""
        detail_ncol = 11
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=detail_ncol)
        c0 = ws.cell(r, 1, value=title_text)
        c0.font = title_f
        c0.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        r += 1
        headers = [
            'Faculty', 'Date', 'Batches', 'Start', 'End', 'Clock (min)', 'Effective hours',
            'Roll numbers', 'Student names', 'Phone numbers', '# Students',
        ]
        for col, h in enumerate(headers, 1):
            c = ws.cell(r, col, value=h)
            c.font = hdr_w
            c.fill = hdr_fill_blue
            c.border = thin_local
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        r += 1
        week_set = set(week_dates)
        week_reqs = [dr for dr in accepted_all if dr.date in week_set]
        week_reqs.sort(key=lambda x: (x.faculty.short_name.lower(), x.date, x.pk))
        week_sum = 0.0
        for dr in week_reqs:
            lines = list(dr.student_lines.all())
            rolls = ', '.join(x.student.roll_no for x in lines)
            names = ', '.join(x.student.name for x in lines)
            phones = ', '.join((x.student.student_phone_number or '').strip() or '—' for x in lines)
            nh = dr.nominal_ds_hours()
            week_sum += nh
            cm = round(dr.duration_minutes(), 1)
            vals = [
                dr.faculty.short_name,
                dr.date.isoformat(),
                dr.batches_label(),
                dr.start_time.strftime('%H:%M'),
                dr.end_time.strftime('%H:%M'),
                cm,
                round(nh, 2),
                rolls,
                names,
                phones,
                len(lines),
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(r, col, value=val)
                c.border = thin_local
                if col in (6, 7, 11):
                    c.alignment = Alignment(horizontal='right', vertical='center')
            r += 1
        if not week_reqs:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=detail_ncol)
            c = ws.cell(r, 1, value='No accepted doubt sessions in this week.')
            c.font = Font(italic=True, color='666666')
            r += 1
        for col in range(1, detail_ncol + 1):
            ws.cell(r, col).border = thin_local
        ws.cell(r, 1, value='Week subtotal — effective hours (DS):').font = sub_f
        ws.cell(r, 7, value=round(week_sum, 2)).font = sub_f
        r += 2
        return r, week_sum

    def _phase_sheet(ws, phase):
        weeks = _compile_phase_weeks_date_objects(dept, phase)
        if not weeks:
            ws.cell(1, 1, f'No lecture weeks in {phase} — set term phase dates for this department.')
            return
        accepted = _doubt_accepted_qs()
        hdr_fill_blue = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        hdr_w = Font(bold=True, color='FFFFFF', size=10)
        title_f = Font(bold=True, size=11)
        sub_f = Font(bold=True)
        r = 1
        for wi, week_dates in enumerate(weeks):
            if not week_dates:
                continue
            wn = wi + 1
            dmin, dmax = min(week_dates), max(week_dates)
            title = (
                f'Week {wn} — lecture days: {dmin.strftime("%d-%b-%Y")} → {dmax.strftime("%d-%b-%Y")} '
                f'({len(week_dates)} day(s)) — detail (60 min clock = 0.5 effective h)'
            )
            r, _ = _write_doubt_detail_block(ws, r, title, week_dates, accepted, thin, hdr_fill_blue, hdr_w, title_f, sub_f)
        date_to_wi = {}
        for wi, dates in enumerate(weeks):
            for d0 in dates:
                date_to_wi[d0] = wi
        nweeks = len(weeks)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max(nweeks + 2, 3))
        ws.cell(r, 1, value=f'Summary — {phase}: faculty × week (effective hours)').font = title_f
        r += 1
        header = ['Faculty'] + [f'Week {i + 1}' for i in range(nweeks)] + [f'Total ({phase})']
        for col, h in enumerate(header, 1):
            c = ws.cell(r, col, value=h)
            c.font = hdr_font
            c.border = thin
        r += 1
        facs = list(Faculty.objects.filter(department=dept).order_by('full_name'))
        by_fac = defaultdict(list)
        for dr in accepted:
            by_fac[dr.faculty_id].append(dr)
        col_totals = [0.0] * nweeks
        for fac in facs:
            wsum = [0.0] * nweeks
            for dr in by_fac.get(fac.id, []):
                wi = date_to_wi.get(dr.date)
                if wi is not None and 0 <= wi < nweeks:
                    h = dr.nominal_ds_hours()
                    wsum[wi] += h
            for i in range(nweeks):
                col_totals[i] += wsum[i]
            ftot = round(sum(wsum), 2)
            fac_row = [fac.short_name] + [round(x, 2) if x else '' for x in wsum] + [ftot]
            for col, val in enumerate(fac_row, 1):
                ws.cell(r, col, value=val).border = thin
            r += 1
        for col in range(1, nweeks + 3):
            ws.cell(r, col).border = thin
        ws.cell(r, 1, value='ALL FACULTIES — week totals (effective h)').font = sub_f
        for i in range(nweeks):
            ws.cell(r, i + 2, value=round(col_totals[i], 2)).font = sub_f
        ws.cell(r, nweeks + 2, value=round(sum(col_totals), 2)).font = sub_f
        r += 1
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 22
        ws.column_dimensions['I'].width = 28
        ws.column_dimensions['J'].width = 22

    def _semester_sheet(ws):
        dmap = _build_date_to_week_map(dept)
        if not dmap:
            ws.cell(1, 1, 'Set term phases to generate semester summary.')
            return
        max_gw = max(dmap.values())
        gw_to_dates = defaultdict(list)
        for d0, gw in dmap.items():
            gw_to_dates[gw].append(d0)
        for gw in gw_to_dates:
            gw_to_dates[gw] = sorted(gw_to_dates[gw])
        accepted = _doubt_accepted_qs()
        hdr_fill_blue = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        hdr_w = Font(bold=True, color='FFFFFF', size=10)
        title_f = Font(bold=True, size=11)
        sub_f = Font(bold=True)
        r = 1
        for gw in range(1, max_gw + 1):
            dates = gw_to_dates.get(gw, [])
            if not dates:
                continue
            dmin, dmax = dates[0], dates[-1]
            title = (
                f'Global week W{gw} — {dmin.strftime("%d-%b-%Y")} → {dmax.strftime("%d-%b-%Y")} '
                f'({len(dates)} lecture day(s)) — detail (60 min clock = 0.5 effective h)'
            )
            r, _ = _write_doubt_detail_block(ws, r, title, dates, accepted, thin, hdr_fill_blue, hdr_w, title_f, sub_f)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max(max_gw + 2, 3))
        ws.cell(r, 1, value='Summary — full semester: faculty × global week (effective hours)').font = title_f
        r += 1
        header = ['Faculty'] + [f'W{i}' for i in range(1, max_gw + 1)] + ['Total DS effective']
        for col, h in enumerate(header, 1):
            c = ws.cell(r, col, value=h)
            c.font = hdr_font
            c.border = thin
        r += 1
        facs = list(Faculty.objects.filter(department=dept).order_by('full_name'))
        by_fac = defaultdict(list)
        for dr in accepted:
            by_fac[dr.faculty_id].append(dr)
        col_totals = [0.0] * max_gw
        for fac in facs:
            parts = [0.0] * max_gw
            for dr in by_fac.get(fac.id, []):
                gw = dmap.get(dr.date)
                if gw and 1 <= gw <= max_gw:
                    h = dr.nominal_ds_hours()
                    parts[gw - 1] += h
            for i in range(max_gw):
                col_totals[i] += parts[i]
            tot = round(sum(parts), 2)
            fac_row = [fac.short_name] + [round(x, 2) if x else '' for x in parts] + [tot]
            for col, val in enumerate(fac_row, 1):
                ws.cell(r, col, value=val).border = thin
            r += 1
        for col in range(1, max_gw + 3):
            ws.cell(r, col).border = thin
        ws.cell(r, 1, value='ALL FACULTIES — week totals (effective h)').font = sub_f
        for i in range(max_gw):
            ws.cell(r, i + 2, value=round(col_totals[i], 2)).font = sub_f
        ws.cell(r, max_gw + 2, value=round(sum(col_totals), 2)).font = sub_f
        r += 1
        for col_letter, w in zip(
            'ABCDEFGHIJK',
            (12, 12, 14, 8, 8, 10, 12, 22, 28, 22, 10),
        ):
            ws.column_dimensions[col_letter].width = w

    wb = Workbook()
    if report == 'semester':
        wb.remove(wb.active)
        ws = wb.create_sheet('Semester_DS')
        _semester_sheet(ws)
    elif report == 'phase':
        ph = (request.GET.get('phase') or 'T1').upper()
        if ph not in ('T1', 'T2', 'T3', 'T4'):
            ph = 'T1'
        wb.remove(wb.active)
        ws = wb.create_sheet(f'DS_{ph}'[:31])
        _phase_sheet(ws, ph)
    else:
        ws0 = wb.active
        ws0.title = 'Detail'
        _detail_sheet(ws0)
        for ph in ('T1', 'T2', 'T3', 'T4'):
            wsp = wb.create_sheet(f'DS_{ph}'[:31])
            _phase_sheet(wsp, ph)
        wss = wb.create_sheet('Semester_DS')
        _semester_sheet(wss)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    safe_dept = re.sub(r'[^\w\-.]+', '_', (dept.name or 'dept'))[:60]
    fname = f'Doubt_Solving_{safe_dept}_{report}_{timezone.localdate().isoformat()}.xlsx'
    resp = HttpResponse(
        bio.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@login_required
def faculty_report_excel(request):
    """Export attendance sheet for one date + batch in same format as old project (Roll No, Name, date row, Lect row, P/A with red A, borders)."""
    if not user_can_faculty(request):
        return redirect('core:faculty_dashboard')
    faculty = get_faculty_user(request)
    if not faculty:
        return redirect('accounts:logout')
    date_str = request.GET.get('date')
    batch_id = request.GET.get('batch_id')
    if not date_str or not batch_id:
        messages.error(request, 'Select date and batch.')
        return redirect('core:faculty_attendance_entry')
    try:
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        return redirect('core:faculty_attendance_entry')
    batch = Batch.objects.filter(pk=batch_id, department=faculty.department).first()
    if not batch:
        return redirect('core:faculty_attendance_entry')
    weekday = selected_date.strftime('%A')
    cancelled_set = get_cancelled_lectures_set(faculty.department)
    all_slots = [s for s in _effective_slots_for_faculty_on_date(faculty, selected_date) if s.batch_id == batch.id and s.day == weekday]
    all_slots = sorted(all_slots, key=lambda s: s.time_slot or '')
    slots = [s for s in all_slots if (selected_date, batch.id, s.time_slot) not in cancelled_set]
    atts = FacultyAttendance.objects.filter(faculty=faculty, date=selected_date, batch=batch).order_by('lecture_slot')
    att_map = {a.lecture_slot: set(x.strip() for x in (a.absent_roll_numbers or '').split(',') if x.strip()) for a in atts}
    students = list(Student.objects.filter(department=faculty.department, batch=batch).annotate(roll_no_int=Cast('roll_no', IntegerField())).order_by('roll_no_int', 'roll_no'))

    wb = Workbook()
    ws = wb.active
    ws.title = (f'{batch.name} {date_str}')[:31]

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    red_font = Font(color='FF0000')
    # Use explicit style instances (not cell.fill/cell.font) to avoid openpyxl StyleProxy unhashable error
    date_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    date_font = Font(bold=True, color='FFFFFF')
    date_align = Alignment(horizontal='center', vertical='center')
    header_font = Font(bold=True)

    # Row 1: Roll No, Student Name, then date (merged) over all lecture columns
    ws.cell(row=1, column=1, value='Roll No').font = header_font
    ws.cell(row=1, column=2, value='Student Name').font = header_font
    n_lec = max(len(slots), 1)
    if n_lec > 1:
        ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=2 + n_lec)
    for c in range(1, 3):
        ws.cell(1, c).border = thin_border
    for c in range(3, 3 + n_lec):
        cell = ws.cell(row=1, column=c, value=selected_date.strftime('%d-%b') if c == 3 else None)
        cell.border = thin_border
        cell.fill = date_fill
        cell.font = date_font
        cell.alignment = date_align

    # Row 2: blank A,B; actual time_slot label + subject per column
    lect_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    lect_font = Font(bold=True, color='FFFFFF')
    lect_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for c in range(1, 3):
        ws.cell(row=2, column=c, value='').border = thin_border
    for i, slot in enumerate(slots, start=1):
        fac, subj = get_faculty_subject_for_slot(selected_date, batch, slot.time_slot)
        subj_name = subj.name if subj else (slot.subject.name if slot.subject else 'N/A')
        lect_label = _lecture_label_for_slot(faculty.department, batch, selected_date, slot.time_slot, i)
        cell = ws.cell(row=2, column=2 + i, value=f'{lect_label}\n{subj_name}')
        cell.alignment = lect_align
        cell.fill = lect_fill
        cell.font = lect_font
        cell.border = thin_border
    if not slots:
        ws.cell(row=2, column=3, value='').border = thin_border

    # Data rows: roll_no, name, then P/A per lecture
    for idx, s in enumerate(students, start=3):
        ws.cell(row=idx, column=1, value=s.roll_no).border = thin_border
        ws.cell(row=idx, column=2, value=s.name).border = thin_border
        str_roll = str(s.roll_no)
        for i, slot in enumerate(slots, start=1):
            is_absent = str_roll in att_map.get(slot.time_slot, set())
            cell = ws.cell(row=idx, column=2 + i, value='A' if is_absent else 'P')
            cell.border = thin_border
            if is_absent:
                cell.font = red_font

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    for c in range(3, 3 + n_lec):
        ws.column_dimensions[get_column_letter(c)].width = 12
    ws.freeze_panes = 'C3'

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename=Attendance_{batch.name}_{date_str}.xlsx'
    return resp


@login_required
def faculty_batchwise_attendance_excel(request):
    """Export batchwise attendance for faculty: all dates up to today, lecture-wise, faculty's lectures only."""
    if not user_can_faculty(request):
        return redirect('core:faculty_dashboard')
    faculty = get_faculty_user(request)
    if not faculty:
        return redirect('accounts:logout')
    batch_id = request.GET.get('batch')
    if not batch_id:
        messages.error(request, 'Select a batch.')
        return redirect('core:faculty_dashboard')
    batch = Batch.objects.filter(pk=batch_id, department=faculty.department).first()
    if not batch:
        messages.error(request, 'Invalid batch.')
        return redirect('core:faculty_dashboard')
    # Verify faculty teaches this batch
    if not ScheduleSlot.objects.filter(faculty=faculty, batch=batch).exists():
        messages.error(request, 'You do not teach this batch.')
        return redirect('core:faculty_dashboard')
    from datetime import date as date_type
    today = date_type.today()
    atts = FacultyAttendance.objects.filter(
        faculty=faculty, batch=batch, date__lte=today
    ).order_by('date', 'lecture_slot')
    if not atts.exists():
        messages.error(request, f'No attendance marked yet for batch {batch.name}.')
        return redirect('core:faculty_dashboard')
    # Build date -> [(lecture_slot, subject_name), ...] ordered by slot
    date_slots = defaultdict(list)
    seen = set()
    for a in atts:
        key = (a.date, a.lecture_slot)
        if key in seen:
            continue
        seen.add(key)
        fac, subj = get_faculty_subject_for_slot(a.date, batch, a.lecture_slot)
        subj_name = subj.name if subj else 'N/A'
        date_slots[a.date].append((a.lecture_slot, subj_name))
    for d in date_slots:
        date_slots[d].sort(key=lambda x: x[0] or '')
    dates_sorted = sorted(date_slots.keys())
    att_map = {}
    for a in atts:
        key = (a.date, a.lecture_slot)
        att_map[key] = set(x.strip() for x in (a.absent_roll_numbers or '').split(',') if x.strip())
    students = list(
        Student.objects.filter(department=faculty.department, batch=batch)
        .select_related('mentor')
        .annotate(roll_no_int=Cast('roll_no', IntegerField()))
        .order_by('roll_no_int', 'roll_no')
    )
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    red_font = Font(color='FF0000', bold=True)
    absent_fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')  # light red
    date_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    date_font = Font(bold=True, color='FFFFFF')
    date_align = Alignment(horizontal='center', vertical='center')
    header_font = Font(bold=True)
    lect_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    lect_font = Font(bold=True, color='FFFFFF')
    lect_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    week_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')  # green for week row
    date_to_week = _build_date_to_week_map(faculty.department)
    wb = Workbook()
    ws = wb.active
    ws.title = (f'{batch.name} Attendance')[:31]
    base_row = 2
    ws.cell(row=base_row, column=1, value='Roll No').font = header_font
    ws.cell(row=base_row, column=2, value='Student Name').font = header_font
    ws.cell(row=base_row, column=3, value='Enrollment').font = header_font
    ws.cell(row=base_row, column=4, value='Mentor Name').font = header_font
    for c in range(1, 5):
        ws.cell(base_row, c).border = thin_border
    col = 5
    col_ranges = []
    for d in dates_sorted:
        slots = date_slots[d]
        n = len(slots)
        if n == 0:
            continue
        start_col = col
        end_col = col + n - 1
        col_ranges.append((d, slots, start_col, end_col))
        if n > 1:
            ws.merge_cells(start_row=base_row, start_column=start_col, end_row=base_row, end_column=end_col)
        cell = ws.cell(row=base_row, column=start_col, value=d.strftime('%d-%b'))
        cell.border = thin_border
        cell.fill = date_fill
        cell.font = date_font
        cell.alignment = date_align
        for i in range(1, n):
            ws.cell(base_row, start_col + i).border = thin_border
        col = end_col + 1
    for c in range(1, 5):
        ws.cell(1, c, value='').border = thin_border
    week_spans = []
    curr_week, curr_start, curr_end = None, None, None
    for d, slots, start_col, end_col in col_ranges:
        w = date_to_week.get(d) or 0
        if curr_week == w and curr_end is not None:
            curr_end = end_col
        else:
            if curr_week is not None:
                week_spans.append((curr_week, curr_start, curr_end))
            curr_week, curr_start, curr_end = w, start_col, end_col
    if curr_week is not None:
        week_spans.append((curr_week, curr_start, curr_end))
    for w, sc, ec in week_spans:
        if sc < ec:
            ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)
        cell = ws.cell(row=1, column=sc, value=f'Week {w}' if w else '')
        cell.border, cell.fill, cell.font, cell.alignment = thin_border, week_fill, date_font, date_align
        for c in range(sc, ec + 1):
            ws.cell(1, c).border = thin_border
    for c in range(1, 5):
        ws.cell(row=3, column=c, value='').border = thin_border
    for d, slots, start_col, end_col in col_ranges:
        for i, (slot, subj_name) in enumerate(slots):
            c = start_col + i
            lect_label = _lecture_label_for_slot(faculty.department, batch, d, slot, i + 1)
            cell = ws.cell(row=3, column=c, value=f'{lect_label}\n{subj_name}')
            cell.alignment = lect_align
            cell.fill = lect_fill
            cell.font = lect_font
            cell.border = thin_border
    data_start = 4
    for idx, s in enumerate(students, start=data_start):
        ws.cell(row=idx, column=1, value=s.roll_no).border = thin_border
        ws.cell(row=idx, column=2, value=s.name).border = thin_border
        ws.cell(row=idx, column=3, value=s.enrollment_no or '').border = thin_border
        ws.cell(row=idx, column=4, value=(s.mentor.short_name if s.mentor else '') or '').border = thin_border
        str_roll = str(s.roll_no)
        for d, slots, start_col, end_col in col_ranges:
            for i, (slot, _) in enumerate(slots):
                c = start_col + i
                is_absent = str_roll in att_map.get((d, slot), set())
                cell = ws.cell(row=idx, column=c, value='' if is_absent else s.roll_no)
                cell.border = thin_border
                if is_absent:
                    cell.font = red_font
                    cell.fill = absent_fill
    total_row = data_start + len(students)
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color='E8F4EA', end_color='E8F4EA', fill_type='solid')  # light green
    ws.cell(row=total_row, column=1, value='Total Present').font = total_font
    for c in range(2, 5):
        ws.cell(total_row, c, value='').font = total_font
    for c in range(1, 5):
        ws.cell(total_row, c).border = thin_border
        ws.cell(total_row, c).fill = total_fill
    for d, slots, start_col, end_col in col_ranges:
        for i, (slot, _) in enumerate(slots):
            c = start_col + i
            present_count = sum(
                1 for s in students
                if str(s.roll_no) not in att_map.get((d, slot), set())
            )
            cell = ws.cell(row=total_row, column=c, value=present_count)
            cell.border = thin_border
            cell.fill = total_fill
            cell.font = total_font
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    for d, slots, start_col, end_col in col_ranges:
        for c in range(start_col, end_col + 1):
            ws.column_dimensions[get_column_letter(c)].width = 10
    ws.freeze_panes = 'E4'
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f'Batchwise_Attendance_{batch.name}_{today:%Y-%m-%d}.xlsx'
    resp = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename={fname}'
    return resp


def _write_batchwise_subject_sheet(ws, batch, subject_name, date_slots_list, students, att_map, styles, date_to_week=None):
    """Write one subject sheet: Roll No, Student Name, Enrollment, Mentor Name, date columns with Lect 1/2..., P/A, Total Present.
    If date_to_week: {date: global_week_num}, adds a Week header row above dates (Week 1, Week 2, ... merged across that week's columns)."""
    thin_border, red_font, absent_fill, date_fill, date_font, date_align = styles['thin_border'], styles['red_font'], styles['absent_fill'], styles['date_fill'], styles['date_font'], styles['date_align']
    header_font, lect_fill, lect_font, lect_align = styles['header_font'], styles['lect_fill'], styles['lect_font'], styles['lect_align']
    total_font, total_fill = styles['total_font'], styles['total_fill']
    week_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')  # green for week row
    row_offset = 1 if date_to_week else 0
    base_row = 1 + row_offset
    ws.cell(row=base_row, column=1, value='Roll No').font = header_font
    ws.cell(row=base_row, column=2, value='Student Name').font = header_font
    ws.cell(row=base_row, column=3, value='Enrollment').font = header_font
    ws.cell(row=base_row, column=4, value='Mentor Name').font = header_font
    for c in range(1, 5):
        ws.cell(base_row, c).border = thin_border
    col = 5
    col_ranges = []
    for d, slots in date_slots_list:
        n = len(slots)
        if n == 0:
            continue
        start_col, end_col = col, col + n - 1
        col_ranges.append((d, slots, start_col, end_col))
        if n > 1:
            ws.merge_cells(start_row=base_row, start_column=start_col, end_row=base_row, end_column=end_col)
        cell = ws.cell(row=base_row, column=start_col, value=d.strftime('%d-%b'))
        cell.border, cell.fill, cell.font, cell.alignment = thin_border, date_fill, date_font, date_align
        for i in range(1, n):
            ws.cell(base_row, start_col + i).border = thin_border
        col = end_col + 1
    if date_to_week:
        for c in range(1, 5):
            ws.cell(1, c, value='').border = thin_border
        week_spans = []
        curr_week, curr_start, curr_end = None, None, None
        for d, slots, start_col, end_col in col_ranges:
            w = date_to_week.get(d)
            if w is None:
                w = 0
            if curr_week == w and curr_end is not None:
                curr_end = end_col
            else:
                if curr_week is not None:
                    week_spans.append((curr_week, curr_start, curr_end))
                curr_week, curr_start, curr_end = w, start_col, end_col
        if curr_week is not None:
            week_spans.append((curr_week, curr_start, curr_end))
        for w, sc, ec in week_spans:
            if sc < ec:
                ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)
            cell = ws.cell(row=1, column=sc, value=f'Week {w}' if w else '')
            cell.border, cell.fill, cell.font, cell.alignment = thin_border, week_fill, date_font, date_align
            for c in range(sc, ec + 1):
                ws.cell(1, c).border = thin_border
    for c in range(1, 5):
        ws.cell(row=base_row + 1, column=c, value='').border = thin_border
    dept_for_batch = batch.department
    for d, slots, start_col, end_col in col_ranges:
        for i, (slot, _) in enumerate(slots):
            c = start_col + i
            lect_label = _lecture_label_for_slot(dept_for_batch, batch, d, slot, i + 1)
            cell = ws.cell(row=base_row + 1, column=c, value=f'{lect_label}\n{subject_name}')
            cell.alignment, cell.fill, cell.font, cell.border = lect_align, lect_fill, lect_font, thin_border
    data_start = base_row + 2
    for idx, s in enumerate(students, start=data_start):
        ws.cell(row=idx, column=1, value=s.roll_no).border = thin_border
        ws.cell(row=idx, column=2, value=s.name).border = thin_border
        ws.cell(row=idx, column=3, value=s.enrollment_no or '').border = thin_border
        ws.cell(row=idx, column=4, value=(s.mentor.short_name if s.mentor else '') or '').border = thin_border
        str_roll = str(s.roll_no)
        for d, slots, start_col, end_col in col_ranges:
            for i, (slot, _) in enumerate(slots):
                c = start_col + i
                is_absent = str_roll in att_map.get((d, slot), set())
                cell = ws.cell(row=idx, column=c, value='' if is_absent else s.roll_no)
                cell.border = thin_border
                if is_absent:
                    cell.font, cell.fill = red_font, absent_fill
    total_row = data_start + len(students)
    ws.cell(row=total_row, column=1, value='Total Present').font = total_font
    for c in range(2, 5):
        ws.cell(total_row, c, value='').font = total_font
    for c in range(1, 5):
        ws.cell(total_row, c).border, ws.cell(total_row, c).fill = thin_border, total_fill
    for d, slots, start_col, end_col in col_ranges:
        for i, (slot, _) in enumerate(slots):
            c = start_col + i
            present_count = sum(1 for s in students if str(s.roll_no) not in att_map.get((d, slot), set()))
            cell = ws.cell(row=total_row, column=c, value=present_count)
            cell.border, cell.fill, cell.font = thin_border, total_fill, total_font
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    for d, slots, start_col, end_col in col_ranges:
        for c in range(start_col, end_col + 1):
            ws.column_dimensions[get_column_letter(c)].width = 10
    ws.freeze_panes = f'E{data_start + 1}'


@login_required
def admin_batchwise_attendance_manager(request):
    """Admin: Batchwise Attendance - select batch, download Excel with one sheet per subject."""
    if not user_can_admin(request):
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    batches = list(Batch.objects.filter(department=dept).order_by('name'))
    ctx = {'batches': batches}
    return render(request, 'core/admin/batchwise_attendance.html', ctx)


@login_required
def admin_batchwise_attendance_excel(request):
    """Admin: Download batchwise Excel - one sheet per subject, same format as faculty batchwise."""
    if not user_can_admin(request):
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first.')
        return redirect('core:admin_dashboard')
    batch_id = request.GET.get('batch')
    if not batch_id:
        messages.error(request, 'Select a batch.')
        return redirect('core:admin_batchwise_attendance_manager')
    batch = Batch.objects.filter(pk=batch_id, department=dept).first()
    if not batch:
        messages.error(request, 'Invalid batch.')
        return redirect('core:admin_batchwise_attendance_manager')
    from datetime import date as date_type
    today = date_type.today()
    atts = FacultyAttendance.objects.filter(batch=batch, date__lte=today).order_by('date', 'lecture_slot')
    if not atts.exists():
        messages.error(request, f'No attendance marked yet for batch {batch.name}.')
        return redirect('core:admin_batchwise_attendance_manager')
    subject_to_dateslots = defaultdict(list)
    seen = set()
    for a in atts:
        key = (a.date, a.lecture_slot)
        if key in seen:
            continue
        seen.add(key)
        fac, subj = get_faculty_subject_for_slot(a.date, batch, a.lecture_slot)
        subj_name = subj.name if subj else 'Other'
        subject_to_dateslots[subj_name].append((a.date, a.lecture_slot))
    for subj_name in subject_to_dateslots:
        subject_to_dateslots[subj_name].sort(key=lambda x: (x[0], x[1] or ''))
    att_map = {}
    for a in atts:
        key = (a.date, a.lecture_slot)
        att_map[key] = set(x.strip() for x in (a.absent_roll_numbers or '').split(',') if x.strip())
    students = list(
        Student.objects.filter(department=dept, batch=batch)
        .select_related('mentor')
        .annotate(roll_no_int=Cast('roll_no', IntegerField()))
        .order_by('roll_no_int', 'roll_no')
    )
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    styles = {
        'thin_border': thin_border,
        'red_font': Font(color='FF0000', bold=True),
        'absent_fill': PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid'),
        'date_fill': PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid'),
        'date_font': Font(bold=True, color='FFFFFF'),
        'date_align': Alignment(horizontal='center', vertical='center'),
        'header_font': Font(bold=True),
        'lect_fill': PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid'),
        'lect_font': Font(bold=True, color='FFFFFF'),
        'lect_align': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'total_font': Font(bold=True),
        'total_fill': PatternFill(start_color='E8F4EA', end_color='E8F4EA', fill_type='solid'),
    }
    def _safe_sheet_name(name):
        s = str(name)[:31]
        for c in '\\/:*?[]':
            s = s.replace(c, '_')
        return s or 'Sheet'
    date_to_week = _build_date_to_week_map(dept)
    wb = Workbook()
    first = True
    for subj_name in sorted(subject_to_dateslots.keys()):
        pairs = subject_to_dateslots[subj_name]
        dates_sorted = sorted(set(p[0] for p in pairs))
        date_slots_list = []
        for d in dates_sorted:
            slots = [(slot, subj_name) for pd, slot in pairs if pd == d]
            slots.sort(key=lambda x: x[0] or '')
            date_slots_list.append((d, slots))
        if not date_slots_list:
            continue
        safe_name = _safe_sheet_name(subj_name)
        ws = wb.active if first else wb.create_sheet(title=safe_name)
        if first:
            ws.title = safe_name
            first = False
        _write_batchwise_subject_sheet(ws, batch, subj_name, date_slots_list, students, att_map, styles, date_to_week=date_to_week)
    if first:
        ws = wb.active
        ws.title = 'No data'
        ws.cell(1, 1, value='No attendance data for this batch.')
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f'Batchwise_Attendance_{batch.name}_{today:%Y-%m-%d}.xlsx'
    resp = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename={fname}'
    return resp


# ---------- Admin: Manual Attendance (mark on behalf of faculty) ----------

def _dates_for_department(dept):
    """Return sorted list of dates that have lectures in this department (from term phases, excluding holidays)."""
    tp = TermPhase.objects.filter(department=dept).first()
    if not tp:
        return []
    holidays = get_all_holiday_dates(dept)
    day_set = _effective_day_set_for_dept(dept, datetime.now().date())
    day_set = {d.lower() for d in day_set if d}
    out = []
    for i in range(1, 5):
        start = getattr(tp, f't{i}_start', None)
        end = getattr(tp, f't{i}_end', None)
        if not start or not end:
            continue
        cur = start
        while cur <= end:
            if cur not in holidays and cur.strftime('%A').lower() in day_set:
                out.append(cur)
            cur += timedelta(days=1)
    adj_dates = LectureAdjustment.objects.filter(batch__department=dept).values_list('date', flat=True).distinct()
    for d in adj_dates:
        if d in holidays:
            continue
        for i in range(1, 5):
            start = getattr(tp, f't{i}_start', None)
            end = getattr(tp, f't{i}_end', None)
            if start and end and start <= d <= end:
                out.append(d)
                break
    extra_dates = ExtraLecture.objects.filter(batch__department=dept).values_list('date', flat=True).distinct()
    for d in extra_dates:
        if d in holidays:
            continue
        for i in range(1, 5):
            start = getattr(tp, f't{i}_start', None)
            end = getattr(tp, f't{i}_end', None)
            if start and end and start <= d <= end:
                out.append(d)
                break
    return sorted(set(out))


def _faculties_for_date(dept, selected_date):
    """Return faculties who have lectures on this date in this department."""
    effective_slots = _effective_slots_for_date(dept, selected_date)
    faculty_ids = {s.faculty_id for s in effective_slots if s.faculty_id}
    for adj in LectureAdjustment.objects.filter(date=selected_date, batch__department=dept).select_related('new_faculty', 'original_faculty'):
        if adj.new_faculty_id:
            faculty_ids.add(adj.new_faculty_id)
        if adj.original_faculty_id:
            faculty_ids.add(adj.original_faculty_id)
    for ex in ExtraLecture.objects.filter(date=selected_date, batch__department=dept).values_list('faculty_id', flat=True):
        if ex:
            faculty_ids.add(ex)
    return Faculty.objects.filter(pk__in=faculty_ids).order_by('short_name')


@login_required
def admin_manual_attendance(request):
    """Admin marks attendance on behalf of faculty. Step 1: select date. Step 2: select faculty. Then show same attendance page."""
    if not user_can_admin(request):
        return redirect('accounts:role_redirect')
    dept = get_admin_department(request)
    if not dept:
        messages.error(request, 'Select a department first from Dashboard.')
        return redirect('core:admin_dashboard')

    available_dates = _dates_for_department(dept)
    date_str = request.GET.get('date')
    faculty_id = request.GET.get('faculty_id')
    selected_date = None
    selected_faculty = None
    faculties_for_date = []

    if date_str:
        try:
            parsed = datetime.strptime(date_str, '%Y-%m-%d').date()
            selected_date = parsed
            if selected_date not in available_dates:
                available_dates = sorted(set(available_dates) | {selected_date})
        except Exception:
            pass
        if selected_date:
            faculties_for_date = _faculties_for_date(dept, selected_date)
            if faculty_id:
                selected_faculty = faculties_for_date.filter(pk=faculty_id).first()
                if not selected_faculty:
                    selected_faculty = faculties_for_date.first() if faculties_for_date else None

    faculty = selected_faculty
    slots_by_batch = defaultdict(list)
    if selected_date and faculty:
        cancelled_set = get_cancelled_lectures_set(dept)
        weekday = selected_date.strftime('%A')
        excluded_by_adj = set(
            LectureAdjustment.objects.filter(date=selected_date, original_faculty=faculty).values_list('batch_id', 'time_slot')
        )
        faculty_slots = [s for s in _effective_slots_for_faculty_on_date(faculty, selected_date) if s.day == weekday]
        for s in sorted(faculty_slots, key=lambda x: x.time_slot or ''):
            if (s.batch_id, s.time_slot) in excluded_by_adj:
                continue
            if (selected_date, s.batch_id, s.time_slot) in cancelled_set:
                continue
            slots_by_batch[s.batch].append(s)
        for adj in LectureAdjustment.objects.filter(date=selected_date, new_faculty=faculty).select_related('batch', 'new_subject', 'new_faculty'):
            existing_pairs = {(b, sl.time_slot) for b, slots in slots_by_batch.items() for sl in slots}
            if (adj.batch, adj.time_slot) in existing_pairs:
                continue
            if (selected_date, adj.batch_id, adj.time_slot) in cancelled_set:
                continue
            virtual = type('Slot', (), {
                'batch': adj.batch, 'time_slot': adj.time_slot,
                'subject': adj.new_subject, 'faculty': adj.new_faculty,
            })()
            slots_by_batch[adj.batch].append(virtual)
        for ex in ExtraLecture.objects.filter(date=selected_date, faculty=faculty).select_related('batch', 'subject', 'faculty'):
            existing_pairs = {(b, sl.time_slot) for b, slots in slots_by_batch.items() for sl in slots}
            if (ex.batch, ex.time_slot) in existing_pairs:
                continue
            if (selected_date, ex.batch_id, ex.time_slot) in cancelled_set:
                continue
            virtual = type('Slot', (), {
                'batch': ex.batch, 'time_slot': ex.time_slot,
                'subject': ex.subject, 'faculty': ex.faculty,
            })()
            slots_by_batch[ex.batch].append(virtual)
        for b in slots_by_batch:
            slots_by_batch[b].sort(key=lambda s: s.time_slot or '')
        for batch, slots in slots_by_batch.items():
            for slot in slots:
                slot.is_extra_lecture = _is_extra_lecture_slot(
                    dept, selected_date, batch, (slot.time_slot or '').strip()
                )

    attendance_prefill = defaultdict(lambda: defaultdict(list))
    attendance_reasons = {}  # (batch_id, lecture_slot) -> {roll_no: reason}
    attendance_updated_at = {}
    if selected_date and faculty:
        for a in FacultyAttendance.objects.filter(faculty=faculty, date=selected_date):
            attendance_prefill[a.batch.id][a.lecture_slot] = [x.strip() for x in (a.absent_roll_numbers or '').split(',') if x.strip()]
            attendance_updated_at[(a.batch.id, a.lecture_slot)] = a.updated_at
            try:
                reasons = json.loads(a.absent_reasons or '{}')
                attendance_reasons[(a.batch.id, a.lecture_slot)] = {k: v for k, v in reasons.items() if v}
            except Exception:
                attendance_reasons[(a.batch.id, a.lecture_slot)] = {}

    batch_students_sorted = {}
    if faculty:
        for batch, slots in slots_by_batch.items():
            sorted_students = sorted(batch.student_set.all(), key=_roll_sort_key)
            batch_students_sorted[batch.id] = sorted_students
            batch.students_sorted = sorted_students
            for slot in slots:
                slot.prefill_absent_set = set(attendance_prefill.get(batch.id, {}).get(slot.time_slot, []))
                reasons = attendance_reasons.get((batch.id, slot.time_slot), {})
                slot.prefill_reasons = reasons
                slot.students_with_reasons = [(s, reasons.get(str(s.roll_no), 'general')) for s in sorted_students]
                slot.last_updated = attendance_updated_at.get((batch.id, slot.time_slot))
                if selected_date:
                    fac, subj = get_faculty_subject_for_slot(selected_date, batch, slot.time_slot)
                    slot.display_subject_name = subj.name if subj else (slot.subject.name if slot.subject else 'N/A')
                    slot.display_faculty_name = fac.short_name if fac else (slot.faculty.short_name if slot.faculty else '—')

    admin_manual_locked = bool(
        selected_date
        and user_can_admin(request)
        and not is_hod(request)
        and not is_super_admin(request)
        and _is_admin_manual_locked_by_hod(dept, selected_date)
    )
    ctx = {
        'available_dates': available_dates,
        'selected_date': selected_date,
        'faculties_for_date': faculties_for_date,
        'selected_faculty': selected_faculty,
        'faculty': faculty,
        'slots_by_batch': dict(slots_by_batch),
        'batch_students_sorted': batch_students_sorted,
        'is_admin_manual': True,
        'admin_manual_locked': admin_manual_locked,
    }
    return render(request, 'core/admin/manual_attendance.html', ctx)


@login_required
def admin_manual_attendance_save(request):
    """Save attendance on behalf of faculty. Departmental admin blocked if that week is locked for the dept; HOD and super admin may still save."""
    if not request.method == 'POST' or not user_can_admin(request):
        return redirect('core:admin_manual_attendance')
    dept = get_admin_department(request)
    if not dept:
        return redirect('core:admin_dashboard')
    date_str = request.POST.get('date')
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            if (
                not is_hod(request)
                and not is_super_admin(request)
                and _is_admin_manual_locked_by_hod(dept, selected_date)
            ):
                messages.error(
                    request,
                    'This week is locked for your department. Only HOD or super admin can change manual attendance for these dates.',
                )
                return redirect('core:admin_manual_attendance')
        except Exception:
            pass
    faculty_id = request.POST.get('faculty_id')
    batch_id = request.POST.get('batch_id')
    lecture_slot = request.POST.get('lecture_slot', '').strip()
    date_str = request.POST.get('date')
    absent_list = request.POST.getlist('absent_roll_numbers')
    if not faculty_id or not batch_id or not date_str:
        messages.error(request, 'Missing data.')
        return redirect('core:admin_manual_attendance')
    faculty = Faculty.objects.filter(pk=faculty_id, department=dept).first()
    if not faculty:
        messages.error(request, 'Invalid faculty.')
        return redirect('core:admin_manual_attendance')
    try:
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        messages.error(request, 'Invalid date.')
        return redirect('core:admin_manual_attendance')
    batch = Batch.objects.filter(pk=batch_id, department=dept).first()
    if not batch:
        messages.error(request, 'Invalid batch.')
        return redirect('core:admin_manual_attendance')
    absent_roll_numbers = ','.join(x.strip() for x in absent_list if x.strip())
    absent_reasons = {}
    for r in absent_list:
        r = str(r).strip()
        if not r:
            continue
        reason = request.POST.get(f'absent_reason_{r}', 'general').strip() or 'general'
        if reason not in ('general', 'washroom', 'playing game', 'others'):
            reason = 'general'
        absent_reasons[r] = reason
    FacultyAttendance.objects.update_or_create(
        faculty=faculty, date=selected_date, batch=batch, lecture_slot=lecture_slot,
        defaults={
            'absent_roll_numbers': absent_roll_numbers,
            'absent_reasons': json.dumps(absent_reasons) if absent_reasons else '',
        }
    )
    sync_faculty_combine_cache_for_attendance(
        dept, faculty, selected_date, batch, lecture_slot,
    )
    messages.success(request, f'Attendance saved for {faculty.short_name}.')
    url = reverse('core:admin_manual_attendance') + f'?date={date_str}&faculty_id={faculty_id}'
    return redirect(url)


@login_required
def admin_manual_attendance_excel(request):
    """Export attendance Excel for admin manual attendance (on behalf of faculty)."""
    if not user_can_admin(request):
        return redirect('core:admin_dashboard')
    dept = get_admin_department(request)
    if not dept:
        return redirect('core:admin_dashboard')
    date_str = request.GET.get('date')
    batch_id = request.GET.get('batch_id')
    faculty_id = request.GET.get('faculty_id')
    if not date_str or not batch_id or not faculty_id:
        messages.error(request, 'Select date, faculty and batch.')
        return redirect('core:admin_manual_attendance')
    faculty = Faculty.objects.filter(pk=faculty_id, department=dept).first()
    if not faculty:
        return redirect('core:admin_manual_attendance')
    try:
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        return redirect('core:admin_manual_attendance')
    batch = Batch.objects.filter(pk=batch_id, department=dept).first()
    if not batch:
        return redirect('core:admin_manual_attendance')
    if (
        not is_hod(request)
        and not is_super_admin(request)
        and _is_admin_manual_locked_by_hod(dept, selected_date)
    ):
        messages.error(request, 'This week is locked. Manual attendance export is disabled for departmental admins.')
        return redirect('core:admin_manual_attendance')
    weekday = selected_date.strftime('%A')
    cancelled_set = get_cancelled_lectures_set(dept)
    all_slots = [s for s in _effective_slots_for_faculty_on_date(faculty, selected_date) if s.batch_id == batch.id and s.day == weekday]
    seen_slots = {s.time_slot for s in all_slots if s.time_slot}
    for ex in ExtraLecture.objects.filter(date=selected_date, faculty=faculty, batch=batch).select_related('subject', 'faculty'):
        if (selected_date, batch.id, ex.time_slot) in cancelled_set or ex.time_slot in seen_slots:
            continue
        seen_slots.add(ex.time_slot)
        virtual = type('Slot', (), {'time_slot': ex.time_slot, 'subject': ex.subject, 'faculty': ex.faculty})()
        all_slots.append(virtual)
    all_slots = sorted(all_slots, key=lambda s: s.time_slot or '')
    slots = [s for s in all_slots if (selected_date, batch.id, s.time_slot) not in cancelled_set]
    atts = FacultyAttendance.objects.filter(faculty=faculty, date=selected_date, batch=batch).order_by('lecture_slot')
    att_map = {a.lecture_slot: set(x.strip() for x in (a.absent_roll_numbers or '').split(',') if x.strip()) for a in atts}
    students = list(Student.objects.filter(department=dept, batch=batch).annotate(roll_no_int=Cast('roll_no', IntegerField())).order_by('roll_no_int', 'roll_no'))

    wb = Workbook()
    ws = wb.active
    ws.title = (f'{batch.name} {date_str}')[:31]
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    red_font = Font(color='FF0000')
    date_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    date_font = Font(bold=True, color='FFFFFF')
    date_align = Alignment(horizontal='center', vertical='center')
    header_font = Font(bold=True)
    lect_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    lect_font = Font(bold=True, color='FFFFFF')
    lect_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.cell(row=1, column=1, value='Roll No').font = header_font
    ws.cell(row=1, column=2, value='Student Name').font = header_font
    n_lec = max(len(slots), 1)
    if n_lec > 1:
        ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=2 + n_lec)
    for c in range(1, 3):
        ws.cell(1, c).border = thin_border
    for c in range(3, 3 + n_lec):
        cell = ws.cell(row=1, column=c, value=selected_date.strftime('%d-%b') if c == 3 else None)
        cell.border = thin_border
        cell.fill = date_fill
        cell.font = date_font
        cell.alignment = date_align
    for c in range(1, 3):
        ws.cell(row=2, column=c, value='').border = thin_border
    for i, slot in enumerate(slots, start=1):
        fac, subj = get_faculty_subject_for_slot(selected_date, batch, slot.time_slot)
        subj_name = subj.name if subj else (slot.subject.name if slot.subject else 'N/A')
        lect_label = _lecture_label_for_slot(dept, batch, selected_date, slot.time_slot, i)
        cell = ws.cell(row=2, column=2 + i, value=f'{lect_label}\n{subj_name}')
        cell.alignment = lect_align
        cell.fill = lect_fill
        cell.font = lect_font
        cell.border = thin_border
    if not slots:
        ws.cell(row=2, column=3, value='').border = thin_border
    for idx, s in enumerate(students, start=3):
        ws.cell(row=idx, column=1, value=s.roll_no).border = thin_border
        ws.cell(row=idx, column=2, value=s.name).border = thin_border
        str_roll = str(s.roll_no)
        for i, slot in enumerate(slots, start=1):
            is_absent = str_roll in att_map.get(slot.time_slot, set())
            cell = ws.cell(row=idx, column=2 + i, value='A' if is_absent else 'P')
            cell.border = thin_border
            if is_absent:
                cell.font = red_font
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    for c in range(3, 3 + n_lec):
        ws.column_dimensions[get_column_letter(c)].width = 12
    ws.freeze_panes = 'C3'
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename=Attendance_{batch.name}_{date_str}.xlsx'
    return resp


@login_required
def faculty_mentorship(request):
    """Faculty view: mentorship students with week-wise attendance and at-risk list."""
    if not user_can_faculty(request):
        return redirect('accounts:role_redirect')
    faculty = get_faculty_user(request)
    if not faculty:
        return redirect('accounts:logout')
    dept = faculty.department
    mentorship_students = list(
        Student.objects.filter(mentor=faculty, department=dept)
        .select_related('batch')
        .annotate(roll_no_int=Cast('roll_no', IntegerField()))
        .order_by('batch__name', 'roll_no_int', 'roll_no')
    )
    if not mentorship_students:
        ctx = {'faculty': faculty, 'mentorship_students': [], 'student_stats': [], 'at_risk': [], 'phase': 'T1', 'phases': ['T1', 'T2', 'T3', 'T4'], 'week_options': [], 'selected_week': 'all', 'selected_week_global_num': None}
        return render(request, 'core/faculty/mentorship.html', ctx)
    tp = TermPhase.objects.filter(department=dept).first()
    phase = request.GET.get('phase', 'T1')
    week_param = request.GET.get('week', 'all')
    phases = ['T1', 'T2', 'T3', 'T4']
    week_map, _, phase_dates = _student_phase_weeks_and_dates(dept, mentorship_students[0].batch)
    phase_week_offsets = _get_phase_week_offsets(week_map)
    weeks = week_map.get(phase, [])
    week_idx = None
    if week_param and week_param != 'all':
        try:
            week_idx = int(week_param)
            if week_idx < 0 or week_idx >= len(weeks):
                week_idx = None
        except ValueError:
            week_idx = None
    # Cumulative phase dates: T2 = T1+T2, T3 = T1+T2+T3, T4 = T1+T2+T3+T4
    phase_order_idx = phases.index(phase) if phase in phases else 0
    phase_dates_set = set()
    for i in range(phase_order_idx + 1):
        phase_dates_set.update(phase_dates.get(phases[i], []))
    if week_idx is not None and weeks:
        # Limit to previous phases + weeks 0..week_idx of current phase
        cum_dates = set()
        for i in range(phase_order_idx):
            cum_dates.update(phase_dates.get(phases[i], []))
        for i in range(week_idx + 1):
            cum_dates.update(weeks[i])
        phase_dates_set = cum_dates
    prev_dates_list = [set(phase_dates.get(phases[i], [])) for i in range(phase_order_idx)]
    batch_cache = {}
    student_stats = []
    at_risk = []
    for s in mentorship_students:
        bid = s.batch_id
        if bid not in batch_cache:
            cancelled_set = get_cancelled_lectures_set(dept)
            batch_scheduled = set()
            _add_batch_schedule_pairs_for_attendance(dept, s.batch, phase_dates_set, batch_scheduled, cancelled_set)
            batch_att_map = {}
            for att in FacultyAttendance.objects.filter(batch=s.batch, date__in=phase_dates_set).only('date', 'lecture_slot', 'absent_roll_numbers'):
                batch_att_map[(att.date, att.lecture_slot)] = set(x.strip() for x in (att.absent_roll_numbers or '').split(',') if x.strip())
            slot_subj = _build_slot_subject_cache(s.batch, phase_dates_set, batch_scheduled)
            batch_cache[bid] = (batch_scheduled, batch_att_map, slot_subj)
        batch_scheduled, batch_att_map, slot_subj = batch_cache[bid]
        held = len(batch_scheduled)
        str_roll = str(s.roll_no)
        attended = sum(1 for (d, slot) in batch_scheduled if (d, slot) in batch_att_map and str_roll not in batch_att_map[(d, slot)])
        pct = round(attended / held * 100, 2) if held else 0
        week_wise = []
        cum_held = cum_attended = 0
        for prev_idx in range(phase_order_idx):
            prev_dates = prev_dates_list[prev_idx]
            prev_held = sum(1 for (d, slot) in batch_scheduled if d in prev_dates)
            prev_attended = sum(1 for (d, slot) in batch_scheduled if d in prev_dates and (d, slot) in batch_att_map and str_roll not in batch_att_map[(d, slot)])
            prev_pct = round(prev_attended / prev_held * 100, 2) if prev_held else 0
            cum_held += prev_held
            cum_attended += prev_attended
            cum_pct = round(cum_attended / cum_held * 100, 2) if cum_held else 0
            week_wise.append({'label': f'{phases[prev_idx]} Overall', 'held': prev_held, 'attended': prev_attended, 'pct': prev_pct, 'cum_held': cum_held, 'cum_attended': cum_attended, 'cum_pct': cum_pct})
        weeks_to_show = range(len(weeks)) if week_idx is None else range(min(week_idx + 1, len(weeks)))
        offset = phase_week_offsets.get(phase, 0)
        for i in weeks_to_show:
            week_set = set(weeks[i])
            w_held = sum(1 for (d, slot) in batch_scheduled if d in week_set)
            w_attended = sum(1 for (d, slot) in batch_scheduled if d in week_set and (d, slot) in batch_att_map and str_roll not in batch_att_map[(d, slot)])
            w_pct = round(w_attended / w_held * 100, 2) if w_held else 0
            cum_held += w_held
            cum_attended += w_attended
            cum_pct = round(cum_attended / cum_held * 100, 2) if cum_held else 0
            gw = offset + i + 1
            week_wise.append({'label': f'Week {gw}', 'week': gw, 'held': w_held, 'attended': w_attended, 'pct': w_pct, 'cum_held': cum_held, 'cum_attended': cum_attended, 'cum_pct': cum_pct})
        subject_wise = defaultdict(lambda: {'held': 0, 'attended': 0})
        for (d, slot) in batch_scheduled:
            subj_name = slot_subj.get((d, slot), 'N/A')
            subject_wise[subj_name]['held'] += 1
            if (d, slot) in batch_att_map and str_roll not in batch_att_map[(d, slot)]:
                subject_wise[subj_name]['attended'] += 1
        subj_list = [{'name': n, 'held': t['held'], 'attended': t['attended'], 'pct': round(t['attended'] / t['held'] * 100, 2) if t['held'] else 0} for n, t in sorted(subject_wise.items())]
        student_stats.append({
            'student': s, 'held': held, 'attended': attended, 'pct': pct,
            'week_wise': week_wise, 'subject_wise': subj_list,
        })
        if held and pct < 75:
            at_risk.append({'student': s, 'held': held, 'attended': attended, 'pct': pct, 'week_wise': week_wise, 'subject_wise': subj_list})
    week_options = [(i, phase_week_offsets.get(phase, 0) + i + 1) for i in range(len(weeks))]
    selected_week_global_num = (phase_week_offsets.get(phase, 0) + week_idx + 1) if week_idx is not None and weeks else None
    ctx = {
        'faculty': faculty,
        'mentorship_students': mentorship_students,
        'student_stats': student_stats,
        'at_risk': sorted(at_risk, key=lambda x: x['pct']),
        'phase': phase,
        'phases': phases,
        'week_options': week_options,
        'selected_week': week_param,
        'selected_week_global_num': selected_week_global_num,
    }
    return render(request, 'core/faculty/mentorship.html', ctx)


# ---------- Student: Attendance Summary ----------

def _student_lecture_records(student, batch, dept, start_date, end_date):
    """Return list of dicts: date, time_slot, subject_name, attended (bool) for all lectures in date range."""
    from datetime import date as date_type
    atts = FacultyAttendance.objects.filter(
        batch=batch, date__gte=start_date, date__lte=end_date
    ).order_by('date', 'lecture_slot')
    str_roll = str(student.roll_no)
    records = []
    for att in atts:
        absents = set(x.strip() for x in (att.absent_roll_numbers or '').split(',') if x.strip())
        attended = str_roll not in absents
        fac, subj = get_faculty_subject_for_slot(att.date, batch, att.lecture_slot)
        subject_name = subj.name if subj else 'N/A'
        records.append({
            'date': att.date,
            'time_slot': att.lecture_slot,
            'subject_name': subject_name,
            'attended': attended,
        })
    return records


def _student_phase_weeks_and_dates(dept, batch):
    """Return (week_map with date objects, available_dates list, phase_dates dict phase -> list of dates). Excludes holidays.
    Uses get_all_schedule_days so phase dates include all weekdays that have lectures in ANY timetable version."""
    tp = TermPhase.objects.filter(department=dept).first()
    phases = ['T1', 'T2', 'T3', 'T4']
    from core.schedule_utils import get_all_schedule_days
    days_set = get_all_schedule_days(dept)
    if not days_set and batch:
        days_set = _effective_day_set_for_batch(batch, datetime.now().date())
    if not days_set:
        days_set = _effective_day_set_for_dept(dept, datetime.now().date())
    if not days_set:
        days_set = {'monday', 'tuesday', 'wednesday', 'thursday', 'friday'}
    days_set = {d.lower() for d in days_set if d}
    week_map = {}
    phase_dates = {}
    all_dates = []
    for p in phases:
        start = getattr(tp, f'{p.lower()}_start', None) if tp else None
        end = getattr(tp, f'{p.lower()}_end', None) if tp else None
        if not start or not end:
            week_map[p] = []
            phase_dates[p] = []
            continue
        holidays = get_phase_holidays(dept, p)
        dates = []
        cur = start
        while cur <= end:
            if cur not in holidays and cur.strftime('%A').lower() in days_set:
                dates.append(cur)
            cur += timedelta(days=1)
        dates = sorted(dates)
        phase_dates[p] = dates
        weeks = []
        week = []
        last_w = None
        for d in dates:
            w = d.isocalendar()[1]
            if last_w is not None and w != last_w and week:
                weeks.append(week)
                week = []
            week.append(d)
            last_w = w
        if week:
            weeks.append(week)
        week_map[p] = weeks
        all_dates.extend(dates)
    available_dates = sorted(set(all_dates))
    return week_map, available_dates, phase_dates


@login_required
def student_attendance_summary(request):
    """Redirect to student dashboard (My Attendance section removed)."""
    if not user_can_student(request):
        return redirect('accounts:role_redirect')
    return redirect('core:student_dashboard')


@login_required
def student_attendance_analytics(request):
    """Full analytics: held = scheduled (timetable, excluding holidays), attended = where taken and present.
    Totals respect filters and only count lectures up to today (real date-wise)."""
    if not user_can_student(request):
        return redirect('accounts:role_redirect')
    student = get_student_user(request)
    if not student:
        return redirect('accounts:logout')
    dept = student.department
    batch = student.batch
    tp = TermPhase.objects.filter(department=dept).first()
    today = datetime.now().date()
    week_map, available_dates, phase_dates = _student_phase_weeks_and_dates(dept, batch)
    phase_week_offsets = _get_phase_week_offsets(week_map)
    str_roll = str(student.roll_no)
    # Build batch_scheduled and batch_att_map for all phase dates
    phase_dates_all = set()
    for dates in phase_dates.values():
        phase_dates_all.update(dates)
    cancelled_set = get_cancelled_lectures_set(dept)
    batch_scheduled = set()
    _add_batch_schedule_pairs_for_attendance(dept, batch, phase_dates_all, batch_scheduled, cancelled_set)
    batch_scheduled_upto_today = batch_scheduled  # Use all scheduled (include future for demo/test data)
    batch_att_map = {}
    for att in FacultyAttendance.objects.filter(batch=batch, date__in=phase_dates_all):
        batch_att_map[(att.date, att.lecture_slot)] = set(x.strip() for x in (att.absent_roll_numbers or '').split(',') if x.strip())

    period_type = request.GET.get('period_type', 'date')
    date_str = request.GET.get('date')
    phase = request.GET.get('phase', 'T1')
    week_str = request.GET.get('week')
    selected_date = None
    selected_week_idx = None
    if period_type == 'date' and date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except Exception:
            pass
    if period_type in ('week', 'phase') and week_str is not None:
        try:
            selected_week_idx = int(week_str)
        except Exception:
            pass

    # --- Date-wise: selected day P/A and day percentage (only count if date <= today)
    day_slots = []
    day_held = day_attended = 0
    if period_type == 'date' and selected_date:
        weekday = selected_date.strftime('%A')
        slots = [s for s in _effective_slots_for_date(dept, selected_date, extra_filters={'batch': batch}) if s.day == weekday]
        slots = sorted(slots, key=lambda s: s.time_slot or '')
        for s in slots:
            key = (selected_date, s.time_slot)
            attended = None
            if key in batch_att_map:
                attended = str_roll not in batch_att_map[key]
            if key in batch_scheduled:
                day_held += 1
                if attended is not None and attended:
                    day_attended += 1
            day_slots.append({
                'time_slot': s.time_slot,
                'subject': s.subject.name if s.subject else 'N/A',
                'faculty': s.faculty.short_name if s.faculty else 'N/A',
                'attended': attended,
                'status': 'Present' if attended else 'Absent' if attended is False else '—',
            })
        day_pct = round(day_attended / day_held * 100, 2) if day_held else None
    else:
        day_pct = None

    # --- Week-wise: held = scheduled in week (up to today), attended = where taken and present
    weeks_summary = []
    cumulative_held = cumulative_attended = 0
    selected_week_summary = None
    phase_start_date = None
    phase_end_date = None
    if period_type in ('week', 'phase') and tp and week_map.get(phase):
        phase_start_date = getattr(tp, f'{phase.lower()}_start', None)
        phase_end_date = getattr(tp, f'{phase.lower()}_end', None)
        weeks_list = week_map[phase]
        week_dates_sets = [set(w) for w in weeks_list]
        for i, week_dates in enumerate(weeks_list):
            week_set = week_dates_sets[i]
            week_held = sum(1 for (d, slot) in batch_scheduled_upto_today if d in week_set)
            week_attended = sum(1 for (d, slot) in batch_scheduled_upto_today if d in week_set and (d, slot) in batch_att_map and str_roll not in batch_att_map[(d, slot)])
            week_pct = round(week_attended / week_held * 100, 2) if week_held else 0
            cumulative_held += week_held
            cumulative_attended += week_attended
            cum_pct = round(cumulative_attended / cumulative_held * 100, 2) if cumulative_held else 0
            weeks_summary.append({
                'week_num': phase_week_offsets.get(phase, 0) + i + 1,
                'dates': week_dates,
                'held': week_held,
                'attended': week_attended,
                'pct': week_pct,
                'cum_held': cumulative_held,
                'cum_attended': cumulative_attended,
                'cum_pct': cum_pct,
            })
            if selected_week_idx is not None and i == selected_week_idx:
                selected_week_summary = weeks_summary[-1]
        if period_type == 'phase':
            selected_week_summary = weeks_summary[-1] if weeks_summary else None
            selected_week_idx = len(weeks_summary) - 1 if weeks_summary else None

    # --- Subject-wise: use batch_scheduled + batch_att_map, group by subject
    subject_stats = defaultdict(lambda: {'held': 0, 'attended': 0})
    period_dates_set = set()
    if period_type == 'date' and selected_date:
        period_dates_set = {selected_date}
    elif period_type == 'week' and selected_week_idx is not None and week_map.get(phase):
        weeks_list = week_map[phase]
        for i in range(min(selected_week_idx + 1, len(weeks_list))):
            period_dates_set.update(weeks_list[i])
    elif period_type == 'phase' and phase_dates.get(phase):
        period_dates_set = set(phase_dates[phase])
    else:
        period_dates_set = set(phase_dates.get(phase, []) or phase_dates_all)
    for (d, slot) in batch_scheduled_upto_today:
        if d not in period_dates_set:
            continue
        fac, subj = get_faculty_subject_for_slot(d, batch, slot)
        subj_name = subj.name if subj else 'N/A'
        subject_stats[subj_name]['held'] += 1
        if (d, slot) in batch_att_map and str_roll not in batch_att_map[(d, slot)]:
            subject_stats[subj_name]['attended'] += 1
    subject_wise = []
    for name in sorted(subject_stats.keys()):
        s = subject_stats[name]
        pct = round(s['attended'] / s['held'] * 100, 2) if s['held'] else 0
        subject_wise.append({'name': name, 'held': s['held'], 'attended': s['attended'], 'pct': pct})

    # Summary cards: respect filters
    if period_type == 'date' and selected_date:
        total_held = day_held
        total_attended = day_attended
        overall_pct = day_pct if day_pct is not None else 0
    elif period_type in ('week', 'phase') and selected_week_summary is not None:
        total_held = selected_week_summary['cum_held']
        total_attended = selected_week_summary['cum_attended']
        overall_pct = selected_week_summary['cum_pct']
    else:
        # No selection: show selected phase (or all phases)
        phase_dates_set = set(phase_dates.get(phase, []) or phase_dates_all)
        scheduled_in_period = {(d, slot) for (d, slot) in batch_scheduled_upto_today if d in phase_dates_set}
        total_held = len(scheduled_in_period)
        total_attended = sum(1 for (d, slot) in scheduled_in_period if (d, slot) in batch_att_map and str_roll not in batch_att_map[(d, slot)])
        overall_pct = round(total_attended / total_held * 100, 2) if total_held else 0

    phase_weeks = week_map.get(phase, [])
    week_options = [(i, phase_week_offsets.get(phase, 0) + i + 1) for i in range(len(phase_weeks))]

    ctx = {
        'student': student,
        'period_type': period_type,
        'selected_date': selected_date,
        'phase': phase,
        'phases': ['T1', 'T2', 'T3', 'T4'],
        'selected_week_idx': selected_week_idx,
        'week_map': week_map,
        'phase_weeks': phase_weeks,
        'week_options': week_options,
        'available_dates': available_dates,
        'day_slots': day_slots,
        'day_held': day_held,
        'day_attended': day_attended,
        'day_pct': day_pct,
        'weeks_summary': weeks_summary,
        'selected_week_summary': selected_week_summary,
        'phase_start_date': phase_start_date,
        'phase_end_date': phase_end_date,
        'subject_wise': subject_wise,
        'total_held': total_held,
        'total_attended': total_attended,
        'overall_pct': overall_pct,
    }
    return render(request, 'core/student/attendance_analytics.html', ctx)
