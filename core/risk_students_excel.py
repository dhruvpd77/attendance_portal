"""Multi-sheet Excel: low attendance cumulative through each WEEK-n + failed marks; thresholds from risk_policy."""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from io import BytesIO

from django.db.models import IntegerField
from django.db.models.functions import Cast
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.models import (
    Department,
    ExamPhase,
    ExamPhaseSubject,
    FacultyAttendance,
    RiskStudentMentorLog,
    Student,
    StudentMark,
)
from core.risk_policy import attendance_risk_min_percent, mark_fail_below_threshold

PHASES = ('T1', 'T2', 'T3', 'T4')

HEADER_BLUE = '9BC2E6'
HEADER_YELLOW = 'FFFF99'
MENTOR_PINK = 'F8CBAD'
THIN = Side(style='thin')


def _cum_dates_through_phase_week(week_map: dict, phase_ph: str, local_week_idx: int) -> set[date]:
    """Lecture dates from T1 start through end of local_week_idx within phase_ph (prior phases fully included). Matches faculty_mentorship."""
    phases = list(PHASES)
    phase_ph = (phase_ph or 'T1').upper()
    if phase_ph not in phases:
        return set()
    end_pi = phases.index(phase_ph)
    cum: set[date] = set()
    for pi in range(end_pi + 1):
        ph = phases[pi]
        weeks = week_map.get(ph, [])
        if pi < end_pi:
            for w in weeks:
                cum.update(w)
        else:
            for i in range(min(local_week_idx + 1, len(weeks))):
                cum.update(weeks[i])
            break
    return cum


def _mentor_div_roll_sort_key(row: dict) -> tuple:
    return (
        (row.get('mentor') or '').upper(),
        str(row.get('div') or ''),
        str(row.get('roll') or ''),
    )


def _exam_phase_for_term(dept: Department, term_key: str) -> ExamPhase | None:
    k = (term_key or '').upper()
    ep = ExamPhase.objects.filter(department=dept, name__iexact=k).first()
    if not ep and k == 'T4':
        ep = ExamPhase.objects.filter(department=dept, name__iexact='SEE').first()
    return ep


def _ordered_students(dept: Department) -> list[Student]:
    return list(
        Student.objects.filter(department=dept)
        .select_related('batch', 'mentor', 'department')
        .annotate(roll_no_int=Cast('roll_no', IntegerField()))
        .order_by('batch__name', 'roll_no_int', 'roll_no')
    )


def _batch_att_map_for_dates(batch_ids: set[int], dates: set[date]) -> dict[int, dict[tuple, set]]:
    out: dict[int, dict[tuple, set]] = defaultdict(dict)
    if not dates or not batch_ids:
        return out
    for att in FacultyAttendance.objects.filter(batch_id__in=batch_ids, date__in=dates):
        key = (att.date, att.lecture_slot)
        out[att.batch_id][key] = {x.strip() for x in (att.absent_roll_numbers or '').split(',') if x.strip()}
    return out


def _cumulative_attendance_risk_rows(
    dept: Department,
    students: list[Student],
    cum_dates: set[date],
) -> list[dict]:
    """Students with overall % below dept attendance risk threshold over scheduled slots on cum_dates."""
    from core.views import _add_batch_schedule_pairs_for_attendance, get_cancelled_lectures_set

    if not cum_dates:
        return []
    att_lim = float(attendance_risk_min_percent(dept))
    ds = cum_dates
    cancelled = get_cancelled_lectures_set(dept)
    batch_scheduled: dict[int, set] = defaultdict(set)
    batch_ids = {s.batch_id for s in students}
    for bid in batch_ids:
        batch = next(s.batch for s in students if s.batch_id == bid)
        _add_batch_schedule_pairs_for_attendance(dept, batch, ds, batch_scheduled[bid], cancelled)
    att_map = _batch_att_map_for_dates(batch_ids, ds)
    rows = []
    for s in students:
        scheduled = {p for p in batch_scheduled.get(s.batch_id, set()) if p[0] in ds}
        if not scheduled:
            continue
        roll = str(s.roll_no).strip()
        held = len(scheduled)
        attended = sum(
            1
            for (d, slot) in scheduled
            if (d, slot) in att_map[s.batch_id] and roll not in att_map[s.batch_id][(d, slot)]
        )
        pct = round(attended / held * 100, 2) if held else 0.0
        if pct < att_lim:
            m = s.mentor
            rows.append(
                {
                    'student_id': s.id,
                    'roll': s.roll_no,
                    'div': s.batch.name if s.batch else '',
                    'branch': dept.name,
                    'enrollment': s.enrollment_no or '',
                    'name': s.name,
                    'attended': attended,
                    'held': held,
                    'pct': pct,
                    'mentor': (m.short_name if m else '') or (m.full_name if m else ''),
                }
            )
    rows.sort(key=_mentor_div_roll_sort_key)
    return rows


def _phase_fail_rows_long(dept: Department, term_key: str) -> list[dict]:
    """One row per (student, subject) where mark is below the phase cutoff. Sorted by subject, mentor, div, roll."""
    ep = _exam_phase_for_term(dept, term_key)
    if not ep:
        return []
    fail_lt = float(mark_fail_below_threshold(dept, term_key))
    subjs = list(
        ExamPhaseSubject.objects.filter(exam_phase=ep)
        .select_related('subject')
        .order_by('subject__name')
    )
    subj_ids = [ps.subject_id for ps in subjs]
    marks_qs = StudentMark.objects.filter(exam_phase=ep, subject_id__in=subj_ids).select_related(
        'student', 'student__batch', 'student__mentor', 'subject'
    )
    rows = []
    for m in marks_qs:
        if m.marks_obtained is None:
            continue
        val = float(m.marks_obtained)
        if val >= fail_lt:
            continue
        s = m.student
        if s.department_id != dept.id:
            continue
        ment = s.mentor
        rows.append(
            {
                'student_id': s.id,
                'roll': s.roll_no,
                'div': s.batch.name if s.batch else '',
                'branch': dept.name,
                'enrollment': s.enrollment_no or '',
                'name': s.name,
                'subject': m.subject.name,
                'mark': val,
                'mentor': (ment.short_name if ment else '') or (ment.full_name if ment else ''),
            }
        )
    rows.sort(
        key=lambda r: (
            r['subject'].upper(),
            (r['mentor'] or '').upper(),
            str(r['div']),
            str(r['roll']),
        )
    )
    return rows


def mentee_phase_fail_rows(dept: Department, mentee_ids: list[int], term_key: str) -> list[dict]:
    """Fail rows (below phase cutoff) restricted to given mentee student ids."""
    if not mentee_ids:
        return []
    want = set(mentee_ids)
    return [r for r in _phase_fail_rows_long(dept, term_key) if r.get('student_id') in want]


def _export_contact_cells(log: RiskStudentMentorLog | None) -> tuple[str, str, str, str]:
    if not log:
        return '', '', '', ''
    d = log.call_date.strftime('%d-%m-%Y') if log.call_date else ''
    t = log.call_time.strftime('%H:%M') if log.call_time else ''
    return (log.contact_person or '', d, t, (log.remarks or '').strip())


@dataclass
class SheetSpec:
    kind: str
    phase: str
    week_index: int | None = None
    global_week_num: int | None = None


def build_sheet_specs(end_phase: str, end_week_0based: int, week_map: dict) -> list[SheetSpec]:
    from core.views import _get_phase_week_offsets

    phases = list(PHASES)
    end_phase = (end_phase or 'T1').upper()
    if end_phase not in phases:
        end_phase = 'T1'
    end_pi = phases.index(end_phase)
    offsets = _get_phase_week_offsets(week_map)
    specs: list[SheetSpec] = []
    for pi in range(end_pi + 1):
        ph = phases[pi]
        weeks = week_map.get(ph, [])
        if pi < end_pi:
            for wi in range(len(weeks)):
                gw = offsets.get(ph, 0) + wi + 1
                specs.append(SheetSpec('week', ph, wi, gw))
            specs.append(SheetSpec('phase', ph))
        else:
            last_w = min(end_week_0based + 1, len(weeks)) if weeks else 0
            for wi in range(last_w):
                gw = offsets.get(ph, 0) + wi + 1
                specs.append(SheetSpec('week', ph, wi, gw))
            specs.append(SheetSpec('phase', ph))
    return specs


def _apply_week_sheet(
    ws,
    dept: Department,
    semester_label: str,
    global_week_num: int,
    data_rows: list[dict],
    log_by_student: dict[int, RiskStudentMentorLog] | None = None,
    *,
    att_risk_pct: float,
) -> None:
    thin = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    title_fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type='solid')
    yellow_fill = PatternFill(start_color=HEADER_YELLOW, end_color=HEADER_YELLOW, fill_type='solid')
    mentor_fill = PatternFill(start_color=MENTOR_PINK, end_color=MENTOR_PINK, fill_type='solid')
    title_font = Font(bold=True, size=14)
    sub_font = Font(bold=True, size=11)

    ws.merge_cells('A1:I1')
    c1 = ws.cell(row=1, column=1, value=f'{dept.name} {semester_label} Compiled Attendance'.strip())
    c1.fill = title_fill
    c1.font = title_font
    c1.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('J1:M1')
    c2 = ws.cell(
        row=1,
        column=10,
        value=f'WEEK-{global_week_num}: cumulative attendance through this week — students below {att_risk_pct:g}%',
    )
    c2.fill = yellow_fill
    c2.font = sub_font
    c2.alignment = Alignment(horizontal='center', vertical='center')

    headers = [
        'Roll no',
        'Div',
        'BRANCH',
        'Enrollment No',
        'Name',
        'Total Attended',
        'Total Lecture',
        'Overall %',
        'MENTOR NAME',
        'CONTACT PERSON (FATHER/MOTHER)',
        'DATE OF PHONE CALL',
        'TIME',
        'REMARK(reason for Less attendance)',
    ]
    hr = 3
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=col, value=h)
        cell.font = Font(bold=True)
        cell.border = thin
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if col == 9:
            cell.fill = mentor_fill

    row = hr + 1
    log_by_student = log_by_student or {}
    for r in data_rows:
        sid = r.get('student_id')
        cont = _export_contact_cells(log_by_student.get(sid) if sid is not None else None)
        vals = [
            r['roll'],
            r['div'],
            r['branch'],
            r['enrollment'],
            r['name'],
            r['attended'],
            r['held'],
            r['pct'],
            r['mentor'],
            cont[0],
            cont[1],
            cont[2],
            cont[3],
        ]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin
            if col == 9:
                cell.fill = mentor_fill
        row += 1

    widths = [10, 8, 14, 16, 28, 14, 14, 10, 12, 22, 16, 12, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _phase_sheet_column_widths(ws) -> None:
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 32
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 14
    for c in range(9, 13):
        ws.column_dimensions[get_column_letter(c)].width = 22


def _apply_phase_sheet(
    ws,
    dept: Department,
    semester_label: str,
    term_key: str,
    data_rows: list[dict],
    log_by_marks_key: dict[tuple[int, str], RiskStudentMentorLog] | None = None,
    *,
    mark_fail_below: float,
) -> None:
    """Subject-grouped fail sheet: merged SUBJECT column, MARKS column, alternating band fill (like FAIL_STUDENT template)."""
    thin = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    title_fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type='solid')
    band_fills = (
        PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
        PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid'),
    )
    ncols = 12
    end_letter = get_column_letter(ncols)
    ws.merge_cells(f'A1:{end_letter}1')
    c1 = ws.cell(
        row=1,
        column=1,
        value=f'{dept.name} {semester_label} — {term_key}: fail data (marks < {mark_fail_below:g})',
    )
    c1.fill = title_fill
    c1.font = Font(bold=True, size=14)
    c1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = [
        'Roll no.',
        'Div',
        'BRANCH',
        'Enrollment No',
        'Name',
        'SUBJECT',
        'MARKS',
        'MENTOR NAME',
        'CONTACT PERSON (FATHER/MOTHER)',
        'DATE OF PHONE CALL',
        'TIME',
        'REMARK(REASON FOR FAIL IN EXAM)',
    ]
    hr = 3
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=col, value=h)
        cell.font = Font(bold=True)
        cell.border = thin
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    _phase_sheet_column_widths(ws)
    if not data_rows:
        return

    log_by_marks_key = log_by_marks_key or {}
    row = hr + 1
    idx = 0
    group_ord = 0
    while idx < len(data_rows):
        subj = data_rows[idx]['subject']
        grp_fill = band_fills[group_ord % 2]
        group_ord += 1
        start_row = row
        while idx < len(data_rows) and data_rows[idx]['subject'] == subj:
            r = data_rows[idx]
            sid = r.get('student_id')
            sk = (sid, r.get('subject') or '') if sid is not None else None
            cont = _export_contact_cells(log_by_marks_key.get(sk) if sk else None)
            for c, v in enumerate(
                [
                    r['roll'],
                    r['div'],
                    r['branch'],
                    r['enrollment'],
                    r['name'],
                    None,
                    r['mark'],
                    r['mentor'],
                    cont[0],
                    cont[1],
                    cont[2],
                    cont[3],
                ],
                start=1,
            ):
                if c == 6:
                    continue
                cell = ws.cell(row=row, column=c, value=v)
                cell.border = thin
                cell.fill = grp_fill
                cell.alignment = Alignment(horizontal='left' if c == 5 else 'center', vertical='center', wrap_text=True)
                if c == 7 and isinstance(v, float) and v == int(v):
                    cell.value = int(v)
            idx += 1
            row += 1
        end_row = row - 1
        ws.merge_cells(start_row=start_row, start_column=6, end_row=end_row, end_column=6)
        sc = ws.cell(row=start_row, column=6, value=subj)
        sc.font = Font(bold=True)
        sc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sc.border = thin
        sc.fill = grp_fill
        for rr in range(start_row + 1, end_row + 1):
            c6 = ws.cell(row=rr, column=6)
            c6.border = thin
            c6.fill = grp_fill
        if idx < len(data_rows):
            row += 1


def build_risk_students_workbook(dept: Department, end_phase: str, end_week_0based: int) -> Workbook:
    from core.models import Batch
    from core.views import _compile_phase_weeks_date_objects, _student_phase_weeks_and_dates

    batch = Batch.objects.filter(department=dept).order_by('name').first()
    if batch:
        week_map, _, _ = _student_phase_weeks_and_dates(dept, batch)
    else:
        week_map = {p: _compile_phase_weeks_date_objects(dept, p) for p in PHASES}
    specs = build_sheet_specs(end_phase, end_week_0based, week_map)
    students = _ordered_students(dept)
    sem = dept.institute_semester.label if dept.institute_semester_id else ''
    att_pct = float(attendance_risk_min_percent(dept))

    wb = Workbook()
    first = True
    used_titles: set[str] = set()

    for spec in specs:
        if spec.kind == 'week':
            assert spec.week_index is not None and spec.global_week_num is not None
            weeks = week_map.get(spec.phase, [])
            if spec.week_index >= len(weeks):
                continue
            cum = _cum_dates_through_phase_week(week_map, spec.phase, spec.week_index)
            rows = _cumulative_attendance_risk_rows(dept, students, cum)
            title = f'WEEK-{spec.global_week_num}'[:31]
            sids = {r['student_id'] for r in rows}
            logs = list(
                RiskStudentMentorLog.objects.filter(
                    department=dept,
                    kind=RiskStudentMentorLog.KIND_ATTENDANCE_WEEK,
                    phase=spec.phase,
                    week_index=spec.week_index,
                    student_id__in=sids,
                )
            )
            log_by = {lg.student_id: lg for lg in logs}
        else:
            rows = _phase_fail_rows_long(dept, spec.phase)
            title = spec.phase[:31]
            sids = {r['student_id'] for r in rows}
            logs = list(
                RiskStudentMentorLog.objects.filter(
                    department=dept,
                    kind=RiskStudentMentorLog.KIND_MARKS_SUBJECT,
                    phase=spec.phase,
                    student_id__in=sids,
                )
            )
            log_by_marks = {(lg.student_id, lg.subject_name or ''): lg for lg in logs}

        base_title = title
        n = 2
        while title in used_titles:
            suffix = f'_{n}'
            title = (base_title[: 31 - len(suffix)] + suffix)[:31]
            n += 1
        used_titles.add(title)

        if first:
            ws = wb.active
            ws.title = title
            first = False
        else:
            ws = wb.create_sheet(title=title)

        if spec.kind == 'week':
            _apply_week_sheet(ws, dept, sem, spec.global_week_num, rows, log_by_student=log_by, att_risk_pct=att_pct)
        else:
            _apply_phase_sheet(
                ws,
                dept,
                sem,
                spec.phase,
                rows,
                log_by_marks_key=log_by_marks,
                mark_fail_below=float(mark_fail_below_threshold(dept, spec.phase)),
            )

    if first:
        ws = wb.active
        ws.cell(row=1, column=1, value='No term weeks configured — set Term Phases (T1–T4 dates) first.')

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_faculty_risk_studentwise_workbook(
    dept: Department,
    end_phase: str,
    end_week_0based: int,
    mentee_ids: list[int],
    *,
    mentor_label: str = '',
) -> Workbook:
    """Single sheet: each at-risk mentee block with column A merged (name), rows = week-wise attendance + phase fail marks + call fields."""
    from core.models import Batch
    from core.views import _compile_phase_weeks_date_objects, _student_phase_weeks_and_dates

    want = set(int(x) for x in mentee_ids)
    if not want:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Student_wise'[:31]
        ws.cell(row=1, column=1, value='No mentees selected for this export.')
        return wb

    batch = Batch.objects.filter(department=dept).order_by('name').first()
    if batch:
        week_map, _, _ = _student_phase_weeks_and_dates(dept, batch)
    else:
        week_map = {p: _compile_phase_weeks_date_objects(dept, p) for p in PHASES}
    specs = build_sheet_specs(end_phase, end_week_0based, week_map)
    students = [s for s in _ordered_students(dept) if s.id in want]
    if not students:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Student_wise'[:31]
        ws.cell(row=1, column=1, value='No students found for this export.')
        return wb

    sem = dept.institute_semester.label if dept.institute_semester_id else ''
    blocks: dict[int, list[dict]] = defaultdict(list)

    for spec in specs:
        if spec.kind == 'week':
            assert spec.week_index is not None and spec.global_week_num is not None
            weeks = week_map.get(spec.phase, [])
            if spec.week_index >= len(weeks):
                continue
            cum = _cum_dates_through_phase_week(week_map, spec.phase, spec.week_index)
            rows = _cumulative_attendance_risk_rows(dept, students, cum)
            sids = {r['student_id'] for r in rows}
            logs = list(
                RiskStudentMentorLog.objects.filter(
                    department=dept,
                    kind=RiskStudentMentorLog.KIND_ATTENDANCE_WEEK,
                    phase=spec.phase,
                    week_index=spec.week_index,
                    student_id__in=sids,
                )
            )
            log_by = {lg.student_id: lg for lg in logs}
            for r in rows:
                sid = r['student_id']
                cont = _export_contact_cells(log_by.get(sid))
                wk_local = spec.week_index + 1
                blocks[sid].append(
                    {
                        'kind': 'attendance',
                        'period': f'WEEK-{spec.global_week_num} ({spec.phase} week {wk_local})',
                        'roll': r['roll'],
                        'div': r['div'],
                        'mentor': r['mentor'],
                        'attended': r['attended'],
                        'held': r['held'],
                        'pct': r['pct'],
                        'subject': '',
                        'mark': '',
                        'contact': cont[0],
                        'date': cont[1],
                        'time': cont[2],
                        'remarks': cont[3],
                    }
                )
        else:
            rows = [x for x in _phase_fail_rows_long(dept, spec.phase) if x.get('student_id') in want]
            sids = {r['student_id'] for r in rows}
            logs = list(
                RiskStudentMentorLog.objects.filter(
                    department=dept,
                    kind=RiskStudentMentorLog.KIND_MARKS_SUBJECT,
                    phase=spec.phase,
                    student_id__in=sids,
                )
            )
            log_by_marks = {(lg.student_id, lg.subject_name or ''): lg for lg in logs}
            for r in rows:
                sid = r['student_id']
                subj = r.get('subject') or ''
                sk = (sid, subj)
                cont = _export_contact_cells(log_by_marks.get(sk))
                mk = r['mark']
                if isinstance(mk, float) and mk == int(mk):
                    mk = int(mk)
                mk_thr = float(mark_fail_below_threshold(dept, spec.phase))
                blocks[sid].append(
                    {
                        'kind': 'marks',
                        'period': f'{spec.phase} (marks < {mk_thr:g})',
                        'roll': r['roll'],
                        'div': r['div'],
                        'mentor': r['mentor'],
                        'attended': '',
                        'held': '',
                        'pct': '',
                        'subject': subj,
                        'mark': mk,
                        'contact': cont[0],
                        'date': cont[1],
                        'time': cont[2],
                        'remarks': cont[3],
                    }
                )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Student_wise'[:31]
    thin = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    title_fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type='solid')
    yellow_fill = PatternFill(start_color=HEADER_YELLOW, end_color=HEADER_YELLOW, fill_type='solid')
    mentor_fill = PatternFill(start_color=MENTOR_PINK, end_color=MENTOR_PINK, fill_type='solid')
    title_font = Font(bold=True, size=14)
    sub_font = Font(bold=True, size=11)

    line_a = f'{dept.name} {sem}'.strip() if sem else dept.name
    if mentor_label:
        line_a = f'{line_a} · {mentor_label}'
    att_line = float(attendance_risk_min_percent(dept))
    line_b = (
        f'Through {end_phase} week {end_week_0based + 1} — at-risk only: attendance cumulative < {att_line:g}% '
        f'and/or marks below each phase cutoff; merged names, one row per week or failed subject.'
    )

    ws.merge_cells('A1:O1')
    c1 = ws.cell(row=1, column=1, value=line_a)
    c1.fill = title_fill
    c1.font = title_font
    c1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.merge_cells('A2:O2')
    c2 = ws.cell(row=2, column=1, value=line_b)
    c2.fill = yellow_fill
    c2.font = sub_font
    c2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = [
        'Student name',
        'Roll',
        'Div',
        'Mentor',
        'Type',
        'Period / subject context',
        'Attended',
        'Held',
        'Overall %',
        'Subject',
        'Mark',
        'Contact person',
        'Date of call',
        'Time',
        'Remarks',
    ]
    hr = 3
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=col, value=h)
        cell.font = Font(bold=True)
        cell.border = thin
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if col == 4:
            cell.fill = mentor_fill

    row = hr + 1
    for s in students:
        br = blocks.get(s.id) or []
        if not br:
            continue
        start_merge = row
        for item in br:
            typ = 'Attendance' if item['kind'] == 'attendance' else 'Marks'
            vals = [
                None,
                item['roll'],
                item['div'],
                item['mentor'],
                typ,
                item['period'],
                item['attended'],
                item['held'],
                item['pct'],
                item['subject'],
                item['mark'],
                item['contact'],
                item['date'],
                item['time'],
                item['remarks'],
            ]
            for col, v in enumerate(vals, start=1):
                if col == 1:
                    continue
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = thin
                if col == 4:
                    cell.fill = mentor_fill
                cell.alignment = Alignment(
                    horizontal='left' if col in (5, 6, 10, 15) else 'center',
                    vertical='center',
                    wrap_text=True,
                )
            row += 1
        end_merge = row - 1
        top = ws.cell(row=start_merge, column=1, value=s.name)
        top.border = thin
        top.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        top.font = Font(bold=True)
        if end_merge > start_merge:
            ws.merge_cells(start_row=start_merge, start_column=1, end_row=end_merge, end_column=1)
        for rr in range(start_merge + 1, end_merge + 1):
            ws.cell(row=rr, column=1).border = thin

    widths = [26, 10, 8, 14, 12, 28, 10, 10, 10, 22, 8, 14, 14, 10, 36]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    if row == hr + 1:
        ws.cell(row=hr + 1, column=1, value='No at-risk rows through this phase/week for these mentees.')

    return wb
