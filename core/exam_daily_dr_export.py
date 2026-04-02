"""
Daily DR Excel: clone institute daily format (exam.xlsx) and fill supervision
from completed SupervisionDuty rows. Supports one or many exam dates + Compile sheet.
"""
from __future__ import annotations

import re
from copy import copy
from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from core.models import (
    DepartmentExamCreditRule,
    Faculty,
    PaperCheckingCompletionRequest,
    PaperSettingCompletionRequest,
    SupervisionDuty,
)
from core.paper_checking_credits import (
    aggregate_paper_check_credits_for_compile,
    paper_eval_approved_by_faculty_date,
)
from core.paper_setting_credits import (
    aggregate_paper_setting_for_compile,
    paper_setting_approved_by_faculty_date,
    paper_setting_dr_column_for_bucket,
)

DATA_START_ROW = 8
LAST_DATA_COL = 57

_DAILY_SHEET_RE = re.compile(
    r'^\d{1,2}\s+[A-Za-z]{3}\s+[A-Za-z]{3,9}\s*$',
    re.I,
)


def _template_path() -> Path:
    base = Path(settings.BASE_DIR)
    for rel in (
        base / 'core' / 'data' / 'exam.xlsx',
        base / 'exam.xlsx',
        base / 'core' / 'data' / 'exam_daily_dr_template.xlsx',
    ):
        if rel.is_file():
            return rel
    for alt in base.glob('EXAM_Daily_UG_SYALL*.xlsx'):
        return alt
    return base / 'core' / 'data' / 'exam.xlsx'


def _pick_source_sheet(wb) -> str:
    skip_substr = (
        'test eval',
        'imp rules',
        'imp points',
        'sheet2',
        'wk 14',
        '(wk',
        'compile',
    )
    candidates = []
    for n in wb.sheetnames:
        nl = (n or '').lower().strip()
        if any(s in nl for s in skip_substr):
            continue
        if _DAILY_SHEET_RE.match((n or '').strip()):
            candidates.append(n)
    if candidates:
        without_paren = [x for x in candidates if '(' not in x]
        return sorted(without_paren or candidates)[0]
    if wb.sheetnames:
        return wb.sheetnames[0]
    raise ValueError('No sheet found in DR template workbook.')


def _safe_sheet_title(dt: date) -> str:
    t = dt.strftime('%d %b %a')
    return re.sub(r'[\[\]\:\*\?\/\\]', '-', t)[:31]


def _session_label(time_slot: str) -> str:
    s = (time_slot or '').upper()
    if 'NOON' in s or 'AFTERNOON' in s or ' PM' in f' {s}' or any(
        s.startswith(x) for x in ('13', '14', '15', '16', '17', '18')
    ):
        return 'Afternoon'
    if 'MORNING' in s or ' AM' in f' {s}' or 'MORNING' in (time_slot or ''):
        return 'Morning'
    if 'EVENING' in s:
        return 'Evening'
    return 'Morning'


def _phase_supervision_bucket(phase_name: str) -> str:
    n = (phase_name or '').upper()
    if 'REM' in n:
        return 'L'
    if 'SEE' in n or 'T4' in n:
        return 'K'
    return 'J'


def _faculty_row_label(fac: Faculty) -> str:
    parts = [fac.full_name.strip(), f'[{fac.department.name}]']
    if fac.email:
        parts.append(f'Mb./Email: {fac.email}')
    return ' '.join(parts)


def _completed_for_date(
    report_date: date,
    *,
    hub_coordinator_id: int | None = None,
    hub_institute_semester_id: int | None = None,
    duty_phase_semester_ids: list[int] | None = None,
) -> list[SupervisionDuty]:
    qs = SupervisionDuty.objects.filter(
        supervision_date=report_date,
        completion_status=SupervisionDuty.COMPLETED,
        faculty__isnull=False,
    )
    if hub_coordinator_id is not None:
        qs = qs.filter(phase__hub_coordinator_id=hub_coordinator_id)
        if hub_institute_semester_id is not None:
            qs = qs.filter(phase__institute_semester_id=hub_institute_semester_id)
    elif duty_phase_semester_ids:
        from core.semester_scope import q_supervision_duty_phase_in_semesters

        qs = qs.filter(q_supervision_duty_phase_in_semesters(duty_phase_semester_ids))
    return list(
        qs.select_related('faculty', 'faculty__department', 'original_faculty', 'phase').order_by(
            'faculty__full_name', 'time_slot', 'pk'
        )
    )


def _group_by_faculty(duties: list[SupervisionDuty]) -> dict[int, list[SupervisionDuty]]:
    by_f: dict[int, list[SupervisionDuty]] = defaultdict(list)
    for d in duties:
        by_f[d.faculty_id].append(d)
    return by_f


def _clear_data_rows(ws) -> None:
    max_r = ws.max_row or DATA_START_ROW
    for r in range(DATA_START_ROW, max_r + 1):
        for c in range(1, LAST_DATA_COL + 1):
            ws.cell(r, c).value = None


def _snapshot_cell_style(cell) -> dict:
    """Shallow copy of style parts (safe to reuse on other cells)."""
    return {
        'font': copy(cell.font) if cell.font else None,
        'border': copy(cell.border) if cell.border else None,
        'fill': copy(cell.fill) if cell.fill else None,
        'number_format': cell.number_format,
        'alignment': copy(cell.alignment) if cell.alignment else None,
    }


def _apply_cell_style(cell, snap: dict) -> None:
    if snap.get('font') is not None:
        cell.font = snap['font']
    if snap.get('border') is not None:
        cell.border = snap['border']
    if snap.get('fill') is not None:
        cell.fill = snap['fill']
    if snap.get('number_format') is not None:
        cell.number_format = snap['number_format']
    if snap.get('alignment') is not None:
        cell.alignment = snap['alignment']


def _snapshot_faculty_row_styles(ws, last_col: int) -> tuple[list[list[dict]], list[float | None]]:
    """Capture row 8 / 9 (alternating band) before values are cleared; row heights too."""
    template_rows = [DATA_START_ROW]
    second = DATA_START_ROW + 1
    if (ws.max_row or 0) >= second:
        template_rows.append(second)
    snaps: list[list[dict]] = []
    heights: list[float | None] = []
    for tr in template_rows:
        row_snaps = [_snapshot_cell_style(ws.cell(tr, c)) for c in range(1, last_col + 1)]
        snaps.append(row_snaps)
        heights.append(ws.row_dimensions[tr].height)
    if len(snaps) == 1:
        tr0 = template_rows[0]
        snaps.append(
            [_snapshot_cell_style(ws.cell(tr0, c)) for c in range(1, last_col + 1)]
        )
        heights.append(heights[0])
    return snaps, heights


def _apply_faculty_row_format(
    ws,
    row: int,
    row_snaps: list[list[dict]],
    heights: list[float | None],
    last_col: int,
) -> None:
    """Reapply template banding/format to one faculty data row (dynamic row count)."""
    i = (row - DATA_START_ROW) % len(row_snaps)
    snap_row = row_snaps[i]
    h = heights[i] if i < len(heights) else None
    if h is not None:
        ws.row_dimensions[row].height = h
    for c in range(1, last_col + 1):
        _apply_cell_style(ws.cell(row, c), snap_row[c - 1])


def _set_daily_title(ws, report_date: date) -> None:
    ws['A1'].value = f"DAILY REPORT (EXAM)  DATE :  {report_date.strftime('%d %b %a')}"


def _fmt_block_room(val) -> str:
    """One block or room as string; prefer numeric look when possible."""
    s = (str(val).strip() if val is not None else '') or '-'
    try:
        f = float(s.replace(',', '.'))
        if f == int(f):
            return str(int(f))
        return str(f)
    except (TypeError, ValueError):
        return s


def _write_supervision_for_faculty_row(
    ws,
    row: int,
    serial: int,
    fac: Faculty,
    duties_today: list[SupervisionDuty],
) -> None:
    ws.cell(row, 1).value = float(serial)
    ws.cell(row, 2).value = _faculty_row_label(fac)
    ws.cell(row, 3).value = fac.short_name.strip()

    if not duties_today:
        return

    ws.cell(row, 4).value = float(len(duties_today))
    # One value per completed supervision, comma-separated (same order as time_slot / import).
    ws.cell(row, 5).value = ', '.join(_session_label(d.time_slot) for d in duties_today)
    ws.cell(row, 6).value = ', '.join(_fmt_block_room(d.block_no) for d in duties_today)
    ws.cell(row, 7).value = ', '.join(_fmt_block_room(d.room_no) for d in duties_today)
    ws.cell(row, 8).value = ', '.join('Proxy' if d.is_proxy else 'Assigned' for d in duties_today)
    proxy_parts: list[str] = []
    for d in duties_today:
        if d.is_proxy and d.original_faculty_id:
            proxy_parts.append(d.original_faculty.short_name.strip())
        else:
            proxy_parts.append('-')
    ws.cell(row, 9).value = ', '.join(proxy_parts)

    cj, ck, cl = 0, 0, 0
    for d in duties_today:
        b = _phase_supervision_bucket(d.phase.name if d.phase else '')
        if b == 'J':
            cj += 1
        elif b == 'K':
            ck += 1
        else:
            cl += 1
    if cj:
        ws.cell(row, 10).value = float(cj)
    if ck:
        ws.cell(row, 11).value = float(ck)
    if cl:
        ws.cell(row, 12).value = float(cl)


def _apply_paper_eval_to_row(ws, row: int, paper_info: dict | None) -> None:
    """Merge approved paper-checking credits into Answerbook Evaluation + activity text (exam.xlsx)."""
    if not paper_info:
        return
    for col_key in ('23', '24', '25'):
        v = paper_info.get(col_key)
        if v is not None and float(v) != 0.0:
            col = int(col_key)
            cur = ws.cell(row, col).value
            try:
                cur_f = float(cur) if cur is not None else 0.0
            except (TypeError, ValueError):
                cur_f = 0.0
            ws.cell(row, col).value = cur_f + float(v)
    lines = paper_info.get('lines') or []
    if lines:
        prev = ws.cell(row, 42).value
        chunk = '\n'.join(lines)
        ws.cell(row, 42).value = f'{prev}\n{chunk}' if prev else chunk


def _apply_paper_setting_to_row(ws, row: int, setting_info: dict | None) -> None:
    """Merge approved paper-setting credits into Question Paper Setting columns (see settings.EXAM_DR_PAPER_SETTING_COLUMNS)."""
    if not setting_info:
        return
    for col_key, v in setting_info.items():
        if col_key in ('lines', 'activity_lines'):
            continue
        try:
            col = int(col_key)
        except (TypeError, ValueError):
            continue
        if v is None or float(v) == 0.0:
            continue
        cur = ws.cell(row, col).value
        try:
            cur_f = float(cur) if cur is not None else 0.0
        except (TypeError, ValueError):
            cur_f = 0.0
        ws.cell(row, col).value = cur_f + float(v)
    # Same "Any other activity / not defined" narrative column as paper-check (col 42).
    act = setting_info.get('activity_lines') or []
    if act:
        prev = ws.cell(row, 42).value
        chunk = '\n'.join(act)
        ws.cell(row, 42).value = f'{prev}\n{chunk}' if prev else chunk


def _write_daily_sheet(
    ws,
    report_date: date,
    faculties,
    by_faculty: dict[int, list[SupervisionDuty]],
    *,
    faculty_id_filter: set[int] | None = None,
    hub_coordinator_id: int | None = None,
    hub_institute_semester_id: int | None = None,
    duty_phase_semester_ids: list[int] | None = None,
) -> None:
    row_snaps, row_heights = _snapshot_faculty_row_styles(ws, LAST_DATA_COL)
    _set_daily_title(ws, report_date)
    _clear_data_rows(ws)
    paper_by_f = paper_eval_approved_by_faculty_date(
        report_date,
        faculty_id_filter=faculty_id_filter,
        hub_coordinator_id=hub_coordinator_id,
        hub_institute_semester_id=hub_institute_semester_id,
        duty_phase_semester_ids=duty_phase_semester_ids,
    )
    setting_by_f = paper_setting_approved_by_faculty_date(
        report_date,
        faculty_id_filter=faculty_id_filter,
        hub_coordinator_id=hub_coordinator_id,
        hub_institute_semester_id=hub_institute_semester_id,
        duty_phase_semester_ids=duty_phase_semester_ids,
    )
    row = DATA_START_ROW
    for sr, fac in enumerate(faculties, start=1):
        dtlist = by_faculty.get(fac.id, [])
        _write_supervision_for_faculty_row(ws, row, sr, fac, dtlist)
        _apply_paper_setting_to_row(ws, row, setting_by_f.get(fac.id))
        _apply_paper_eval_to_row(ws, row, paper_by_f.get(fac.id))
        _apply_faculty_row_format(ws, row, row_snaps, row_heights, LAST_DATA_COL)
        row += 1


def _aggregate_compile(
    dates: list[date],
    *,
    hub_coordinator_id: int | None = None,
    hub_institute_semester_id: int | None = None,
    duty_phase_semester_ids: list[int] | None = None,
) -> dict[int, dict]:
    """Totals per faculty across dates (completed duties only)."""
    agg: dict[int, dict] = defaultdict(lambda: {'n': 0, 'j': 0, 'k': 0, 'l': 0})
    qs = SupervisionDuty.objects.filter(
        supervision_date__in=dates,
        completion_status=SupervisionDuty.COMPLETED,
        faculty__isnull=False,
    )
    if hub_coordinator_id is not None:
        qs = qs.filter(phase__hub_coordinator_id=hub_coordinator_id)
        if hub_institute_semester_id is not None:
            qs = qs.filter(phase__institute_semester_id=hub_institute_semester_id)
    elif duty_phase_semester_ids:
        from core.semester_scope import q_supervision_duty_phase_in_semesters

        qs = qs.filter(q_supervision_duty_phase_in_semesters(duty_phase_semester_ids))
    qs = qs.select_related('phase')
    for d in qs:
        ent = agg[d.faculty_id]
        ent['n'] += 1
        b = _phase_supervision_bucket(d.phase.name if d.phase else '')
        if b == 'J':
            ent['j'] += 1
        elif b == 'K':
            ent['k'] += 1
        else:
            ent['l'] += 1
    return dict(agg)


def _add_compile_sheet(
    wb,
    dates: list[date],
    faculties,
    *,
    faculty_id_filter: set[int] | None = None,
    hub_coordinator_id: int | None = None,
    hub_institute_semester_id: int | None = None,
    duty_phase_semester_ids: list[int] | None = None,
) -> None:
    ws = wb.create_sheet(title='Compile')
    ws['A1'] = 'COMPILE — supervision (completed) + paper eval + paper setting credits (approved)'
    period = f'{dates[0].isoformat()} to {dates[-1].isoformat()}' if len(dates) > 1 else dates[0].isoformat()
    ws['A2'] = (
        f'Period: {period}  |  {len(dates)} day sheet(s). '
        'Paper-check credits = approved completions by decision date (DR eval cols). '
        'Paper-setting credits = approved completions by decision date (DR setting cols).'
    )
    agg = _aggregate_compile(
        dates,
        hub_coordinator_id=hub_coordinator_id,
        hub_institute_semester_id=hub_institute_semester_id,
        duty_phase_semester_ids=duty_phase_semester_ids,
    )
    pcr = aggregate_paper_check_credits_for_compile(
        dates,
        faculty_id_filter,
        hub_coordinator_id=hub_coordinator_id,
        hub_institute_semester_id=hub_institute_semester_id,
        duty_phase_semester_ids=duty_phase_semester_ids,
    )
    pset = aggregate_paper_setting_for_compile(
        dates,
        faculty_id_filter,
        hub_coordinator_id=hub_coordinator_id,
        hub_institute_semester_id=hub_institute_semester_id,
        duty_phase_semester_ids=duty_phase_semester_ids,
    )
    bold = Font(bold=True)
    headers = [
        'Sr',
        'Faculty name',
        'Initial',
        'Department',
        'Total supervisions',
        '(T1–T3) sup',
        '(T4 SEE) sup',
        '(REM) sup',
        'Setting (T1–T3) cr',
        'Setting (SEE) cr',
        'Setting (REM) cr',
        'Setting (FT) cr',
        'Setting total cr',
        'Eval (T1–T3) cr',
        'Eval (SEE) cr',
        'Eval (REM) cr',
        'Eval total cr',
    ]
    hr = 4
    for c, h in enumerate(headers, 1):
        cell = ws.cell(hr, c)
        cell.value = h
        cell.font = bold
    row = hr + 1
    b0 = DepartmentExamCreditRule.BUCKET_T1_T3
    b1 = DepartmentExamCreditRule.BUCKET_SEE
    b2 = DepartmentExamCreditRule.BUCKET_REMEDIAL
    b3 = DepartmentExamCreditRule.BUCKET_FAST_TRACK
    c0, c1, c2, c3 = (
        str(paper_setting_dr_column_for_bucket(b0)),
        str(paper_setting_dr_column_for_bucket(b1)),
        str(paper_setting_dr_column_for_bucket(b2)),
        str(paper_setting_dr_column_for_bucket(b3)),
    )
    for sr, fac in enumerate(faculties, start=1):
        a = agg.get(fac.id, {'n': 0, 'j': 0, 'k': 0, 'l': 0})
        ps = pset.get(
            fac.id,
            {c0: 0, c1: 0, c2: 0, c3: 0, 'total': 0},
        )
        pc = pcr.get(fac.id, {'23': 0, '24': 0, '25': 0, 'total': 0})
        ws.cell(row, 1).value = sr
        ws.cell(row, 2).value = fac.full_name
        ws.cell(row, 3).value = fac.short_name
        ws.cell(row, 4).value = fac.department.name
        ws.cell(row, 5).value = a['n']
        ws.cell(row, 6).value = a['j']
        ws.cell(row, 7).value = a['k']
        ws.cell(row, 8).value = a['l']
        ws.cell(row, 9).value = float(ps.get(c0, 0))
        ws.cell(row, 10).value = float(ps.get(c1, 0))
        ws.cell(row, 11).value = float(ps.get(c2, 0))
        ws.cell(row, 12).value = float(ps.get(c3, 0))
        ws.cell(row, 13).value = float(ps.get('total', 0))
        ws.cell(row, 14).value = float(pc['23'])
        ws.cell(row, 15).value = float(pc['24'])
        ws.cell(row, 16).value = float(pc['25'])
        ws.cell(row, 17).value = float(pc['total'])
        row += 1
    for col in range(1, 18):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions['B'].width = 34


def _dates_with_supervision_in_range(
    d0: date,
    d1: date,
    hub_coordinator_id: int | None = None,
    hub_institute_semester_id: int | None = None,
    duty_phase_semester_ids: list[int] | None = None,
) -> list[date]:
    qs = SupervisionDuty.objects.filter(supervision_date__gte=d0, supervision_date__lte=d1)
    if hub_coordinator_id is not None:
        qs = qs.filter(phase__hub_coordinator_id=hub_coordinator_id)
        if hub_institute_semester_id is not None:
            qs = qs.filter(phase__institute_semester_id=hub_institute_semester_id)
    elif duty_phase_semester_ids:
        from core.semester_scope import q_supervision_duty_phase_in_semesters

        qs = qs.filter(q_supervision_duty_phase_in_semesters(duty_phase_semester_ids))
    raw = qs.values_list('supervision_date', flat=True).distinct().order_by('supervision_date')
    return list(raw)


def build_exam_daily_dr_workbook(
    dates: list[date],
    *,
    department_ids: list[int] | None = None,
    hub_coordinator_id: int | None = None,
    hub_institute_semester_id: int | None = None,
    duty_phase_semester_ids: list[int] | None = None,
) -> HttpResponse:
    if not dates:
        raise ValueError('No exam dates to export.')

    path = _template_path()
    if not path.is_file():
        raise FileNotFoundError(
            f'Missing daily DR template. Add core/data/exam.xlsx (or exam.xlsx in project root). Path tried: {path}'
        )

    # Hub: same full roster as exam section; only supervision / paper cells are hub-scoped (not fewer rows).
    if hub_coordinator_id is not None:
        faculties = list(
            Faculty.objects.select_related('department').order_by('department__name', 'full_name')
        )
        faculty_id_filter: set[int] | None = None
    elif duty_phase_semester_ids:
        from core.models import Department

        dept_ids = list(
            Department.objects.filter(institute_semester_id__in=duty_phase_semester_ids).values_list(
                'pk', flat=True
            )
        )
        fac_qs = (
            Faculty.objects.filter(department_id__in=dept_ids)
            .select_related('department')
            .order_by('department__name', 'full_name')
        )
        faculties = list(fac_qs)
        faculty_id_filter = {f.id for f in faculties}
    else:
        fac_qs = Faculty.objects.select_related('department').order_by('department__name', 'full_name')
        if department_ids is not None:
            fac_qs = fac_qs.filter(department_id__in=department_ids)
        faculties = list(fac_qs)
        faculty_id_filter = (
            {f.id for f in faculties} if department_ids is not None else None
        )

    wb = load_workbook(path)
    src_name = _pick_source_sheet(wb)
    src_ws = wb[src_name]

    filled = []
    for dt in dates:
        ws = wb.copy_worksheet(src_ws)
        ws.title = _safe_sheet_title(dt)
        completed = _completed_for_date(
            dt,
            hub_coordinator_id=hub_coordinator_id,
            hub_institute_semester_id=hub_institute_semester_id,
            duty_phase_semester_ids=duty_phase_semester_ids,
        )
        by_f = _group_by_faculty(completed)
        _write_daily_sheet(
            ws,
            dt,
            faculties,
            by_f,
            faculty_id_filter=faculty_id_filter,
            hub_coordinator_id=hub_coordinator_id,
            hub_institute_semester_id=hub_institute_semester_id,
            duty_phase_semester_ids=duty_phase_semester_ids,
        )
        filled.append(ws)

    for nm in list(wb.sheetnames):
        if wb[nm] not in filled:
            wb.remove(wb[nm])

    _add_compile_sheet(
        wb,
        dates,
        faculties,
        faculty_id_filter=faculty_id_filter,
        hub_coordinator_id=hub_coordinator_id,
        hub_institute_semester_id=hub_institute_semester_id,
        duty_phase_semester_ids=duty_phase_semester_ids,
    )

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    dept_tag = ''
    if department_ids and len(department_ids) == 1:
        dept_tag = f"_dept{department_ids[0]}"
    hub_tag = ''
    if hub_coordinator_id is not None:
        hub_tag = f'_hub{hub_coordinator_id}'
    if len(dates) == 1:
        fname = f"exam_daily_dr_{dates[0]:%Y-%m-%d}{dept_tag}{hub_tag}.xlsx"
    else:
        fname = f"exam_daily_dr_{dates[0]:%Y%m%d}_{dates[-1]:%Y%m%d}{dept_tag}{hub_tag}.xlsx"
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


def build_exam_daily_dr_excel(report_date: date) -> HttpResponse:
    return build_exam_daily_dr_workbook([report_date])


def _dates_with_paper_approval_in_range(
    d0: date,
    d1: date,
    hub_coordinator_id: int | None = None,
    hub_institute_semester_id: int | None = None,
    duty_phase_semester_ids: list[int] | None = None,
) -> list[date]:
    qs = PaperCheckingCompletionRequest.objects.filter(
        status=PaperCheckingCompletionRequest.APPROVED,
        decided_at__date__gte=d0,
        decided_at__date__lte=d1,
    )
    if hub_coordinator_id is not None:
        qs = qs.filter(duty__phase__hub_coordinator_id=hub_coordinator_id)
        if hub_institute_semester_id is not None:
            qs = qs.filter(duty__phase__institute_semester_id=hub_institute_semester_id)
    elif duty_phase_semester_ids:
        from core.semester_scope import q_completion_duty_phase_in_semesters

        qs = qs.filter(q_completion_duty_phase_in_semesters(duty_phase_semester_ids))
    raw = qs.values_list('decided_at__date', flat=True).distinct().order_by('decided_at__date')
    return list(raw)


def _dates_with_paper_setting_approval_in_range(
    d0: date,
    d1: date,
    hub_coordinator_id: int | None = None,
    hub_institute_semester_id: int | None = None,
    duty_phase_semester_ids: list[int] | None = None,
) -> list[date]:
    qs = PaperSettingCompletionRequest.objects.filter(
        status=PaperSettingCompletionRequest.APPROVED,
        decided_at__date__gte=d0,
        decided_at__date__lte=d1,
    )
    if hub_coordinator_id is not None:
        qs = qs.filter(duty__phase__hub_coordinator_id=hub_coordinator_id)
        if hub_institute_semester_id is not None:
            qs = qs.filter(duty__phase__institute_semester_id=hub_institute_semester_id)
    elif duty_phase_semester_ids:
        from core.semester_scope import q_completion_duty_phase_in_semesters

        qs = qs.filter(q_completion_duty_phase_in_semesters(duty_phase_semester_ids))
    raw = qs.values_list('decided_at__date', flat=True).distinct().order_by('decided_at__date')
    return list(raw)


def parse_dates_from_request(
    date_str: str | None,
    date_to_str: str | None,
    *,
    hub_coordinator_id: int | None = None,
    hub_institute_semester_id: int | None = None,
    duty_phase_semester_ids: list[int] | None = None,
) -> list[date]:
    """Single day (date only) or range with supervision and/or paper-check approval dates."""
    if not (date_str or '').strip():
        raise ValueError('Missing start date.')
    d0 = datetime.strptime(date_str.strip(), '%Y-%m-%d').date()
    if not (date_to_str or '').strip():
        sup0 = set(
            _dates_with_supervision_in_range(
                d0,
                d0,
                hub_coordinator_id,
                hub_institute_semester_id,
                duty_phase_semester_ids=duty_phase_semester_ids,
            )
        )
        pap0 = set(
            _dates_with_paper_approval_in_range(
                d0,
                d0,
                hub_coordinator_id,
                hub_institute_semester_id,
                duty_phase_semester_ids=duty_phase_semester_ids,
            )
        )
        pset0 = set(
            _dates_with_paper_setting_approval_in_range(
                d0,
                d0,
                hub_coordinator_id,
                hub_institute_semester_id,
                duty_phase_semester_ids=duty_phase_semester_ids,
            )
        )
        if not (sup0 | pap0 | pset0):
            raise ValueError(
                f'No supervision or approved paper activity on {d0.isoformat()}. Pick another date.'
            )
        return [d0]
    d1 = datetime.strptime(date_to_str.strip(), '%Y-%m-%d').date()
    if d1 < d0:
        d0, d1 = d1, d0
    sup = set(
        _dates_with_supervision_in_range(
            d0,
            d1,
            hub_coordinator_id,
            hub_institute_semester_id,
            duty_phase_semester_ids=duty_phase_semester_ids,
        )
    )
    pap = set(
        _dates_with_paper_approval_in_range(
            d0,
            d1,
            hub_coordinator_id,
            hub_institute_semester_id,
            duty_phase_semester_ids=duty_phase_semester_ids,
        )
    )
    pset = set(
        _dates_with_paper_setting_approval_in_range(
            d0,
            d1,
            hub_coordinator_id,
            hub_institute_semester_id,
            duty_phase_semester_ids=duty_phase_semester_ids,
        )
    )
    dates = sorted(sup | pap | pset)
    if not dates:
        raise ValueError(
            f'No supervision or approved paper rows between {d0.isoformat()} and {d1.isoformat()}. '
            'Widen the range or complete duties.'
        )
    return dates
